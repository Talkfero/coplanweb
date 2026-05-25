# -*- coding: utf-8 -*-
"""Mixin de dominio "tecnico" da CoplanApi (extraido de main_web.py).

Nao instanciar diretamente: compoe backend.api.CoplanApi via heranca.
"""
from __future__ import annotations

import getpass  # noqa: F401
import hashlib  # noqa: F401
import os  # noqa: F401
import re  # noqa: F401
import sys  # noqa: F401
import threading  # noqa: F401
from datetime import datetime  # noqa: F401
from pathlib import Path  # noqa: F401
from typing import Any, Callable  # noqa: F401

from backend._state import (  # noqa: F401
    APP_VERSION,
    HERE,
    _OP_LOCK,
    _OP_STATE,
    _op_check_cancel,
    _op_finish,
    _op_reset,
    _op_set_progress,
    _op_snapshot,
)


class TecnicoMixin:

    # ------------------------------------------------------------------
    # Bloco 4 (Auditoria #41 + #42): Snapshot tecnico real.
    # Port direto de tecnico_snapshot_mixin.py (desktop):
    #   - _compute_file_token(path)        - sha1 de path|mtime|size
    #   - _compute_folder_token(folder, [required_files]) - sha1 do dir
    #     incluindo cada arquivo esperado (mtime|size ou "missing")
    #   - tecnico_snapshot()               - token completo
    #     (db + apoio + ganhos + tecnico_paths)
    #   - tecnico_check_dirty()            - compara token atual vs
    #     config['tecnico_last_token']; marca dirty automaticamente
    #     quando muda (RB-1.1 do desktop, fallback simples).
    # ------------------------------------------------------------------
    @staticmethod
    def _compute_file_token(path: str) -> str:
        """Hash sha1 baseado em path/mtime/tamanho. Vazio em erro."""
        try:
            info = os.stat(path)
            raw = f"{os.path.abspath(path)}|{info.st_mtime}|{info.st_size}"
            return hashlib.sha1(raw.encode("utf-8")).hexdigest()
        except Exception:  # noqa: BLE001
            return ""

    @staticmethod
    def _compute_folder_token(folder: str, required: list[str]) -> str:
        """Hash sha1 baseado no folder + cada arquivo esperado (mtime|size
        ou 'missing'). Paridade com _compute_folder_token desktop."""
        parts = [os.path.abspath(folder)]
        for name in required:
            path = os.path.join(folder, name)
            try:
                info = os.stat(path)
                parts.append(f"{name}:{info.st_mtime}:{info.st_size}")
            except Exception:  # noqa: BLE001
                parts.append(f"{name}:missing")
        return hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()

    def _tecnico_paths(self) -> dict[str, str]:
        """Resolve os 4 paths usados no token tecnico:
        db, apoio, ganhos (folder), tecnico_txt (folder, falls back
        para ganhos)."""
        cfg = self._config or {}
        db_path = ""
        try:
            db = getattr(self, "_db_manager", None)
            if db is not None:
                db_path = str(getattr(db, "db_path", "") or "")
        except Exception:  # noqa: BLE001
            db_path = ""
        if not db_path:
            db_path = str(cfg.get("obras") or "")
        apoio_path = str(cfg.get("apoio") or "")
        ganhos_path = str(cfg.get("caminho_pasta_ganhos") or "")
        # tecnico_txt nao tem path proprio no web -> usa pasta de ganhos.
        tecnico_path = ganhos_path
        return {
            "db":      db_path,
            "apoio":   apoio_path,
            "ganhos":  ganhos_path,
            "tecnico": tecnico_path,
        }

    def tecnico_snapshot(self) -> dict[str, Any]:
        """Snapshot tecnico completo (token + timestamp + src descritivo).

        Token = sha1 de db|apoio|ganhos|tecnico (cada um com seu hash
        proprio via _compute_file_token / _compute_folder_token).
        Paridade com _compute_tecnico_snapshot_token do desktop."""
        try:
            from runtime.config import (  # type: ignore[import-not-found]
                TECNICO_REQUIRED_FILES,
            )
        except Exception:  # noqa: BLE001
            TECNICO_REQUIRED_FILES = [
                "FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT",
            ]
        paths = self._tecnico_paths()
        parts: list[str] = []
        if paths["db"]:
            parts.append(f"db:{self._compute_file_token(paths['db'])}")
        if paths["apoio"]:
            parts.append(f"apoio:{self._compute_file_token(paths['apoio'])}")
        if paths["ganhos"]:
            parts.append(
                f"ganhos:{self._compute_folder_token(paths['ganhos'], TECNICO_REQUIRED_FILES)}"
            )
        if paths["tecnico"] and paths["tecnico"] != paths["ganhos"]:
            parts.append(
                f"txt:{self._compute_folder_token(paths['tecnico'], TECNICO_REQUIRED_FILES)}"
            )
        raw = "|".join(parts)
        token = hashlib.sha1(raw.encode("utf-8")).hexdigest() if raw else ""

        # src descritivo (paridade _get_tecnico_snapshot_source)
        src_parts: list[str] = []
        if paths["apoio"]:
            src_parts.append(f"Apoio:{os.path.basename(paths['apoio'])}")
        if paths["ganhos"]:
            src_parts.append(f"Ganhos:{os.path.basename(paths['ganhos'])}")
        if paths["tecnico"] and paths["tecnico"] != paths["ganhos"]:
            src_parts.append(f"TXT:{os.path.basename(paths['tecnico'])}")
        src = " | ".join(src_parts) if src_parts else "N/D"

        ts = datetime.now().strftime("%d/%m/%y %H:%M")
        # Sempre devolve count atual (UI usa pra render pill warn).
        dirty_count = 0
        try:
            db, _err = self._ensure_db_connected()
            if db is not None:
                dirty_count = int(db.count_tecnico_dirty() or 0)
        except Exception:  # noqa: BLE001
            dirty_count = 0
        return {
            "ok": True,
            "token": token,
            "ts": ts,
            "src": src,
            "paths": paths,
            "dirty_count": dirty_count,
            # Mantem chave 'tecnico_dirty' por compat com consumers
            # antigos do M029 STUB (cadastro_mixin testa esse campo).
            "tecnico_dirty": "SIM" if dirty_count > 0 else "NÃO",
            "error": "",
        }

    def tecnico_check_dirty(self) -> dict[str, Any]:
        """Compara token tecnico atual vs config['tecnico_last_token'].
        Quando diferente E ha obras no banco, marca TODAS como
        tecnico_dirty='SIM' (fallback simples do _apply_tecnico_token_change_db
        desktop, que tem logica de escopo mais complexa).
        Persiste o novo token no config para a proxima checagem.

        Retorna {ok, token_changed, dirty_applied, count, token, error}.
        JS chama no boot e em coplan:state events para detectar mudancas
        nas fontes (db/apoio/ganhos)."""
        snap = self.tecnico_snapshot()
        token_now = str(snap.get("token") or "")
        if not token_now:
            return {
                "ok": True,
                "token_changed": False,
                "dirty_applied": False,
                "count": int(snap.get("dirty_count") or 0),
                "token": "",
                "error": "",
            }
        # Le ultimo token salvo
        last_token = ""
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager,
            )
            cfg = ConfigManager.load_config() or {}
            last_token = str(cfg.get("tecnico_last_token") or "")
        except Exception:  # noqa: BLE001
            cfg = {}
            last_token = ""
        token_changed = bool(last_token) and (last_token != token_now)
        dirty_applied = False
        # Se token mudou (e nao e' a primeira execucao), marca dirty
        db, err = self._ensure_db_connected()
        if token_changed and db is not None:
            try:
                # Conta obras antes - so marca se houver
                count_obras = 0
                try:
                    cursor = db._get_cursor()
                    if cursor is not None:
                        cursor.execute("SELECT COUNT(*) FROM obras")
                        row = cursor.fetchone()
                        count_obras = int(row[0]) if row else 0
                except Exception:  # noqa: BLE001
                    count_obras = 0
                if count_obras > 0:
                    db.mark_tecnico_dirty_all()
                    dirty_applied = True
            except Exception as exc:  # noqa: BLE001
                return {
                    "ok": False,
                    "token_changed": True,
                    "dirty_applied": False,
                    "count": int(snap.get("dirty_count") or 0),
                    "token": token_now,
                    "error": f"mark_dirty: {exc}",
                }
        # Persiste novo token (mesmo se nao houve mudanca - garante
        # que o "primeiro contato" salve o baseline).
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager,
            )
            ConfigManager.save_config({"tecnico_last_token": token_now})
            self._config = None
        except Exception:  # noqa: BLE001
            pass
        # Recount apos eventual mark_dirty_all
        count_after = int(snap.get("dirty_count") or 0)
        if dirty_applied and db is not None:
            try:
                count_after = int(db.count_tecnico_dirty() or 0)
            except Exception:  # noqa: BLE001
                pass
        return {
            "ok": True,
            "token_changed": token_changed,
            "dirty_applied": dirty_applied,
            "count": count_after,
            "token": token_now,
            "first_seen": (not last_token),
            "error": "",
        }

    @classmethod
    def _list_ganhos_in(cls, folder: str) -> list[dict[str, Any]]:
        """Lista arquivos com extensoes relevantes em ``folder``."""
        items: list[dict[str, Any]] = []
        if not folder or not os.path.isdir(folder):
            return items
        try:
            for name in sorted(os.listdir(folder)):
                full = os.path.join(folder, name)
                if not os.path.isfile(full):
                    continue
                if not name.lower().endswith(cls.GANHOS_EXTS):
                    continue
                try:
                    st = os.stat(full)
                    size = int(st.st_size)
                    mtime = float(st.st_mtime)
                except OSError:
                    size, mtime = 0, 0.0
                items.append({
                    "name": name,
                    "path": full,
                    "size": size,
                    "mtime": mtime,
                })
        except OSError:
            pass
        return items

    @staticmethod
    def _resolve_ganhos_folder(base: str, alimentador: str) -> tuple[str, str]:
        """Tenta resolver subpasta do alimentador dentro da base.

        Procura nesta ordem:
          base/ALIM
          base/<ano>/ALIM (qualquer subpasta de ano que exista)
          base                       (fallback)
        Retorna (path, label_used).
        """
        base = (base or "").strip()
        alim = (alimentador or "").strip().upper()
        if not base or not os.path.isdir(base):
            return base, ""
        if not alim:
            return base, ""
        # base/ALIM
        direct = os.path.join(base, alim)
        if os.path.isdir(direct):
            return direct, alim
        # base/<ano>/ALIM
        try:
            for sub in sorted(os.listdir(base)):
                year_dir = os.path.join(base, sub)
                if not os.path.isdir(year_dir):
                    continue
                cand = os.path.join(year_dir, alim)
                if os.path.isdir(cand):
                    return cand, f"{sub}/{alim}"
        except OSError:
            pass
        return base, ""

    def list_ganhos_files(self, alimentador: Any = "") -> dict[str, Any]:
        """Retorna {ok, base, folder, alim, files, error}.

        - ``base``  = config.caminho_pasta_ganhos (raiz)
        - ``folder``= pasta efetivamente listada (pode ser raiz ou subpasta
                      do alimentador; ver _resolve_ganhos_folder)
        - ``alim``  = label resolvido (ex.: '2026/ATB-204') ou ''
        - ``files`` = lista de {name, path, size, mtime}
        """
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {
                "ok": False, "base": "", "folder": "", "alim": "",
                "files": [], "error": f"config: {exc}",
            }

        base = str(cfg.get("caminho_pasta_ganhos") or "").strip()
        if not base:
            return {
                "ok": False, "base": "", "folder": "", "alim": "",
                "files": [],
                "error": "caminho_pasta_ganhos nao configurado em config.json",
            }
        if not os.path.isdir(base):
            return {
                "ok": False, "base": base, "folder": base, "alim": "",
                "files": [],
                "error": f"pasta nao encontrada: {base}",
            }
        folder, label = self._resolve_ganhos_folder(base, str(alimentador or ""))
        files = self._list_ganhos_in(folder)
        return {
            "ok": True, "base": base, "folder": folder, "alim": label,
            "files": files, "error": "",
        }

    @staticmethod
    def _wv_dialog_const(kind: str) -> Any:
        """Resolve a constante de dialog do pywebview no runtime.

        Pywebview >= 5 usa ``FileDialog.OPEN``/``SAVE``/``FOLDER`` em
        vez de ``OPEN_DIALOG``/``SAVE_DIALOG``/``FOLDER_DIALOG``. Tenta a
        forma nova e cai pra antiga (silencia deprecation warnings)."""
        import webview  # type: ignore[import-not-found]
        kind = kind.upper()
        FileDialog = getattr(webview, "FileDialog", None)
        if FileDialog is not None:
            mapping = {
                "OPEN": getattr(FileDialog, "OPEN", None),
                "SAVE": getattr(FileDialog, "SAVE", None),
                "FOLDER": getattr(FileDialog, "FOLDER", None),
            }
            v = mapping.get(kind)
            if v is not None:
                return v
        # Fallback antigo (pywebview < 5).
        legacy = {
            "OPEN":   getattr(webview, "OPEN_DIALOG", None),
            "SAVE":   getattr(webview, "SAVE_DIALOG", None),
            "FOLDER": getattr(webview, "FOLDER_DIALOG", None),
        }
        return legacy.get(kind)

    def pick_ganhos_folder(self) -> dict[str, Any]:
        """Abre o file dialog do pywebview (folder) e atualiza
        config.caminho_pasta_ganhos com o caminho escolhido. Retorna o
        novo estado de list_ganhos_files."""
        try:
            import webview  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "base": "", "folder": "", "alim": "",
                    "files": [], "error": f"pywebview indisponivel: {exc}"}
        try:
            wins = webview.windows
            if not wins:
                return {"ok": False, "base": "", "folder": "", "alim": "",
                        "files": [], "error": "janela pywebview nao encontrada"}
            dlg = self._wv_dialog_const("FOLDER")
            if dlg is None:
                return {"ok": False, "base": "", "folder": "", "alim": "",
                        "files": [], "error": "FOLDER dialog indisponivel"}
            result = wins[0].create_file_dialog(dlg)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "base": "", "folder": "", "alim": "",
                    "files": [], "error": f"file dialog: {exc}"}
        if not result:
            # Usuario cancelou; devolve estado atual sem mudar config.
            return self.list_ganhos_files("")
        path = result[0] if isinstance(result, (list, tuple)) else str(result)
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
            cfg["caminho_pasta_ganhos"] = path
            ConfigManager.save_config(cfg)
        except Exception as exc:  # noqa: BLE001
            self._data_state_set(
                "ganhos", "INVALIDADO", path=path,
                error=f"save_config: {exc}")
            return {"ok": False, "base": path, "folder": path, "alim": "",
                    "files": [], "error": f"save_config: {exc}"}
        # Recarrega config interno
        self._config = None
        self._ensure_managers()
        # Hook estado: pasta de ganhos selecionada com sucesso.
        # Validacao "tem arquivos esperados" acontece em read_ganhos_file
        # ou no calc_*; aqui marcamos VALIDADO se a pasta existe.
        if os.path.isdir(path):
            self._data_state_set(
                "ganhos", "CARREGADO_VALIDADO", path=path)
        else:
            self._data_state_set(
                "ganhos", "INVALIDADO", path=path,
                error="pasta nao existe")
        return self.list_ganhos_files("")

    # ------------------------------------------------------------------
    # Passo 5.2 (Ganhos / read_ganhos_file): leitor generico de XLSX/CSV
    # /TXT que devolve cabecalho + linhas para o JS popular a tabela
    # "Parametros de Ganhos". A logica avancada do desktop (parse de
    # FlowMT.TXT/Topologia.TXT/Confiabilidade.TXT por alimentador) e'
    # complexa demais para um wrapper unico; aqui fazemos preview
    # tabular + heuristica para extrair pares Antes/Depois quando o
    # arquivo tem 2-3 colunas no formato "param;antes;depois".
    # ------------------------------------------------------------------
    def read_ganhos_file(self, path: Any, max_rows: Any = 200) -> dict[str, Any]:
        path_s = str(path or "").strip()
        if not path_s:
            return {"ok": False, "error": "path vazio", "headers": [],
                    "rows": [], "parametros": [], "total_rows": 0}
        if not os.path.isfile(path_s):
            return {"ok": False, "error": f"arquivo nao encontrado: {path_s}",
                    "headers": [], "rows": [], "parametros": [], "total_rows": 0}

        try:
            limit = int(max_rows) if max_rows not in (None, "", 0) else 200
        except (TypeError, ValueError):
            limit = 200

        ext = os.path.splitext(path_s)[1].lower()
        headers: list[str] = []
        rows: list[list[str]] = []

        try:
            if ext in (".xlsx", ".xlsm"):
                from openpyxl import load_workbook  # type: ignore[import-not-found]
                wb = load_workbook(path_s, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for r_i, row in enumerate(ws.iter_rows(values_only=True)):
                        cells = [
                            "" if c is None else str(c) for c in row
                        ]
                        if r_i == 0:
                            headers = cells
                        else:
                            rows.append(cells)
                        if len(rows) >= limit:
                            break
                finally:
                    wb.close()
            elif ext in (".xls",):
                # xlrd nao e' garantido; tenta pandas se disponivel.
                try:
                    import pandas as pd  # type: ignore[import-not-found]
                    df = pd.read_excel(path_s, header=0, nrows=limit)
                    headers = [str(c) for c in df.columns]
                    rows = [
                        ["" if (c != c) else str(c) for c in r]  # NaN check
                        for r in df.values.tolist()
                    ]
                except Exception as exc:  # noqa: BLE001
                    return {
                        "ok": False, "error": f"xls (pandas falhou): {exc}",
                        "headers": [], "rows": [], "parametros": [], "total_rows": 0,
                    }
            elif ext in (".csv", ".txt", ".tsv"):
                import csv as _csv
                # Detecta delimitador a partir de uma amostra.
                with open(path_s, encoding="utf-8", errors="replace") as f:
                    sample = f.read(4096)
                if "\t" in sample:
                    delim = "\t"
                elif sample.count(";") > sample.count(","):
                    delim = ";"
                elif "," in sample:
                    delim = ","
                else:
                    delim = ";"
                with open(path_s, encoding="utf-8", errors="replace", newline="") as f:
                    reader = _csv.reader(f, delimiter=delim)
                    for i, row in enumerate(reader):
                        if i == 0:
                            headers = list(row)
                        else:
                            rows.append(list(row))
                        if len(rows) >= limit:
                            break
            else:
                return {
                    "ok": False, "error": f"extensao nao suportada: {ext}",
                    "headers": [], "rows": [], "parametros": [], "total_rows": 0,
                }
        except Exception as exc:  # noqa: BLE001
            return {
                "ok": False, "error": f"leitura falhou: {exc}",
                "headers": [], "rows": [], "parametros": [], "total_rows": 0,
            }

        # Heuristica: extrai (label, antes, depois) das primeiras colunas.
        # Util quando o arquivo tem 2-3 colunas estruturadas.
        parametros: list[dict[str, str]] = []
        for r in rows:
            if not r:
                continue
            label = str(r[0] if len(r) > 0 else "").strip()
            if not label:
                continue
            antes = str(r[1] if len(r) > 1 else "").strip()
            depois = str(r[2] if len(r) > 2 else "").strip()
            parametros.append({"label": label, "a": antes, "d": depois})
            if len(parametros) >= 30:
                break

        # Hook estado: se o arquivo lido for um dos 3 tecnicos
        # (FlowMT.TXT, Topologia.TXT, Confiabilidade.TXT), marca a fonte
        # tecnico_txt como CARREGADO_PARCIAL. Validacao completa (3 juntos)
        # e' feita em validate_tecnico_files / antes de calc_*.
        try:
            base = os.path.basename(path_s).lower()
            if base in ("flowmt.txt", "topologia.txt", "confiabilidade.txt"):
                self._data_state_set(
                    "tecnico_txt", "CARREGADO_PARCIAL",
                    path=os.path.dirname(path_s),
                    version_token=str(int(os.path.getmtime(path_s))))
        except OSError:
            pass

        return {
            "ok": True, "error": "",
            "headers": headers, "rows": rows[:limit],
            "parametros": parametros,
            "total_rows": len(rows),
            "path": path_s,
            "ext": ext,
        }

    def validate_tecnico_files(self, pasta: Any = "") -> dict[str, Any]:
        """Verifica se os 3 arquivos tecnicos obrigatorios existem na pasta.
        Marca tecnico_txt como CARREGADO_VALIDADO (ok) ou INVALIDADO.
        Equivalente ao _validate_ganhos_files do desktop, sem o passo
        de tentar ler os arquivos (custo alto)."""
        path = str(pasta or "").strip()
        if not path:
            try:
                from codigo5_coplan import ConfigManager  # noqa: PLC0415
                cfg = ConfigManager.load_config() or {}
                path = str(cfg.get("caminho_pasta_ganhos") or "").strip()
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "error": f"config: {exc}",
                        "missing": []}
        if not path:
            return {"ok": False, "error": "pasta nao configurada",
                    "missing": []}
        if not os.path.isdir(path):
            self._data_state_set(
                "tecnico_txt", "INVALIDADO", path=path,
                error="pasta nao existe")
            return {"ok": False, "error": f"pasta nao existe: {path}",
                    "missing": []}
        required = ["FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT"]
        missing = [
            f for f in required
            if not os.path.isfile(os.path.join(path, f))
        ]
        if missing:
            self._data_state_set(
                "tecnico_txt", "INVALIDADO", path=path,
                error=f"arquivos ausentes: {', '.join(missing)}")
            return {"ok": False, "missing": missing,
                    "error": f"faltam: {', '.join(missing)}",
                    "pasta": path}
        # Token: hash dos mtimes dos 3 arquivos.
        try:
            tokens = [
                str(int(os.path.getmtime(os.path.join(path, f))))
                for f in required
            ]
            version_token = "-".join(tokens)
        except OSError:
            version_token = ""
        self._data_state_set(
            "tecnico_txt", "CARREGADO_VALIDADO", path=path,
            version_token=version_token)
        return {"ok": True, "missing": [], "pasta": path,
                "version_token": version_token}

    def tecnico_snapshot_update(self, cods: Any = None) -> dict[str, Any]:
        """Atualiza snapshot tecnico para um conjunto de cods (limpa
        tecnico_dirty para essas obras + grava token/timestamp/src).
        Equivalente a atualizar_snapshot_tecnico_selecionados do desktop:
        usado pelo botao 'Atualizar snapshot tecnico' apos refrescar
        FlowMT/Topologia/Confiabilidade.

        Token e' derivado dos mtimes dos 3 .TXT da pasta de ganhos
        (mesma logica de validate_tecnico_files). snapshot_src guarda
        o path da pasta para auditoria."""
        cods_list = [str(c).strip() for c in (cods or []) if str(c or "").strip()]
        if not cods_list:
            return {"ok": False, "atualizadas": 0, "error": "cods vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "atualizadas": 0,
                    "error": err or "db indisponivel"}
        # Token: mesma estrategia de validate_tecnico_files (3 .TXT
        # mtimes).
        cfg = self._config or {}
        pasta = str(cfg.get("caminho_pasta_ganhos") or "").strip()
        if not pasta or not os.path.isdir(pasta):
            return {"ok": False, "atualizadas": 0,
                    "error": "pasta de ganhos nao configurada"}
        required = ["FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT"]
        try:
            tokens = [
                str(int(os.path.getmtime(os.path.join(pasta, f))))
                for f in required if os.path.isfile(os.path.join(pasta, f))
            ]
            if len(tokens) < len(required):
                return {"ok": False, "atualizadas": 0,
                        "error": "arquivos tecnicos ausentes"}
            token = "-".join(tokens)
        except OSError as exc:
            return {"ok": False, "atualizadas": 0,
                    "error": f"mtime: {exc}"}
        snapshot_at = datetime.now().strftime("%d/%m/%y %H:%M")
        snapshot_src = pasta
        try:
            db.update_tecnico_snapshot_for_cods(
                cods_list, token, snapshot_at, snapshot_src,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atualizadas": 0,
                    "error": f"update_snapshot: {exc}"}
        return {
            "ok": True,
            "atualizadas": len(cods_list),
            "token": token,
            "snapshot_at": snapshot_at,
            "snapshot_src": snapshot_src,
            "error": "",
        }

    def _read_tecnico_files(self, alimentador: Any = "") -> dict[str, Any]:
        """Le os 3 arquivos tecnicos (FlowMT/Topologia/Confiabilidade)
        da pasta tecnica configurada. Retorna {flow_mt, topologia, confiabilidade}.

        A pasta e' resolvida pelo helper `_ganhos_resolve_pasta`, que tenta
        `caminho_pasta_ganhos` (chave canonica usada pela UI de Ganhos)
        e cai pra `caminho_pasta_arquivos` como fallback historico.
        """
        pasta = self._ganhos_resolve_pasta()
        if not pasta:
            return {"flow_mt": [], "topologia": [], "confiabilidade": [],
                    "error": "pasta tecnica nao configurada"}
        try:
            from runtime.file_io import carregar_arquivos
            dados = carregar_arquivos(
                pasta, ["FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT"]
            )
        except Exception as exc:  # noqa: BLE001
            return {"flow_mt": [], "topologia": [], "confiabilidade": [],
                    "error": f"carregar: {exc}"}
        return {
            "flow_mt": list(dados.get("FlowMT.TXT") or []),
            "topologia": list(dados.get("Topologia.TXT") or []),
            "confiabilidade": list(dados.get("Confiabilidade.TXT") or []),
            "error": "",
        }
