# -*- coding: utf-8 -*-
"""Mixin de dominio "obras" da CoplanApi (extraido de main_web.py).

Nao instanciar diretamente: compoe backend.api.CoplanApi via heranca.
"""
from __future__ import annotations

import getpass  # noqa: F401
import hashlib  # noqa: F401
import os  # noqa: F401
import re  # noqa: F401
import sys  # noqa: F401
import threading  # noqa: F401
from datetime import datetime  # noqa: F401
from pathlib import Path  # noqa: F401
from typing import Any, Callable  # noqa: F401

from backend._state import (  # noqa: F401
    APP_VERSION,
    HERE,
    _OP_LOCK,
    _OP_STATE,
    _op_check_cancel,
    _op_finish,
    _op_reset,
    _op_set_progress,
    _op_snapshot,
)


class ObrasMixin:

    # ------------------------------------------------------------------
    # Passo 3.1 (Visualizar / list_obras): le o banco real via
    # DatabaseManager.fetch_all e mapeia colunas (ORDERED_COLUMNS) para o
    # shape esperado pelo template do mock (cod, ano, pi, projeto, alim,
    # se, regional, pacote, valor, aprovada, passou, tecAtual).
    # ------------------------------------------------------------------
    @staticmethod
    def _row_to_dict(row: Any, cols: list[str]) -> dict[str, Any]:
        """Converte uma row do fetch_all (lista posicional) em dict por
        nome de coluna -- usado pelos servicos do core que esperam dict."""
        out: dict[str, Any] = {}
        for i, name in enumerate(cols):
            out[name] = row[i] if i < len(row) else ""
        return out

    @staticmethod
    def _fmt_pi(pi_base: Any, item: Any) -> str:
        a = str(pi_base or "").strip()
        b = str(item or "").strip()
        if a and b:
            return f"{a}-{b}"
        return a or b

    def list_obras(self, limit: Any = None) -> dict[str, Any]:
        """Retorna ``{ok, rows, total, error}``. ``rows`` no formato JS do mock.

        Cenarios (Sprint A): quando cenario_ativo != '', filtra raw_rows
        para apenas obras em cenarios_obras E aplica overrides de
        cenario_obras_overrides + ano_final de cenarios_obras."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "rows": [], "total": 0, "error": err}

        try:
            raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "rows": [], "total": 0, "error": f"fetch_all: {exc}"}

        # Cenarios: filtro + overrides
        cen_nome = self._cenario_active_name()
        if cen_nome:
            cod_set, cen_info = self._cenario_cod_set(db, cen_nome)
            ovmap = self._cenario_overrides_map(db, cen_nome)
            try:
                idx_cod_filter = cols.index("cod")
            except ValueError:
                idx_cod_filter = -1
            if idx_cod_filter >= 0 and cod_set:
                # Filtra
                raw_rows = [
                    r for r in raw_rows
                    if str(r[idx_cod_filter] if idx_cod_filter < len(r)
                           else "").strip() in cod_set
                ]
                # Aplica overrides + ano_final
                applied: list[Any] = []
                for r in raw_rows:
                    cod_r = str(r[idx_cod_filter] if idx_cod_filter < len(r)
                                else "").strip()
                    applied.append(self._cenario_apply_to_row(
                        r, cols, cod_r,
                        cen_info.get(cod_r) or {},
                        ovmap.get(cod_r) or {},
                    ))
                raw_rows = applied
            elif cod_set == set():
                # Cenario com 0 obras (ou tabelas inexistentes): retorna vazio
                raw_rows = []

        # ---- Passo 3.5: computa "passou" via verificar_criterios_v2 -----
        # Mesma regra V2 usada pela MainWindow desktop (cor unica por
        # projeto). Se faltam colunas obrigatorias o servico devolve
        # [True, ...] (tratado como "tudo ok"). Falhas de import nao
        # sao fatais: caem para todos True.
        passou_list: list[bool] = [True] * len(raw_rows)
        try:
            from core.services.relatorio_criterios_service import (  # type: ignore[import-not-found]
                verificar_criterios_v2,
            )
            criterios = (self._config or {}).get("criterios_planejamento") or {}
            if criterios and raw_rows:
                raw_passou = verificar_criterios_v2(
                    raw_rows, cols, criterios=criterios
                )
                passou_list = [
                    True if v is None else bool(v) for v in raw_passou
                ]
        except Exception as exc:  # noqa: BLE001
            print(f"[main_web] verificar_criterios_v2 falhou: {exc}", file=sys.stderr)

        # Resolve indices uma vez (rows pode ser milhares).
        def idx(name: str) -> int:
            try:
                return cols.index(name)
            except ValueError:
                return -1

        i_cod = idx("cod")
        i_ano = idx("ano_")
        i_pi_base = idx("pi_base")
        i_item = idx("codigo_item")
        i_projeto = idx("nome_projeto")
        i_alim = idx("alimentador_principal")
        i_alim_benef = idx("alimentadores_beneficiados")  # [B6]
        i_se = idx("subestacao")
        i_regional = idx("nome_regional")
        i_super = idx("nome_superintendencia")  # [B6/superintendencia]
        i_pacote = idx("tipo_pacote")
        i_valor = idx("valor_obra")
        i_aprovada = idx("obra_aprovada")
        i_dirty = idx("tecnico_dirty")

        max_rows: int | None
        try:
            max_rows = int(limit) if limit not in (None, "", 0) else None
        except (TypeError, ValueError):
            max_rows = None

        out: list[dict[str, Any]] = []
        raw_out: list[list[Any]] = []
        passou_out: list[bool] = []
        for row_i, r in enumerate(raw_rows):
            def g(i: int, default: Any = "") -> Any:
                return r[i] if 0 <= i < len(r) else default

            ano_raw = g(i_ano)
            try:
                ano: Any = int(str(ano_raw).strip()) if str(ano_raw).strip() else ""
            except (TypeError, ValueError):
                ano = ano_raw
            try:
                valor = float(str(g(i_valor) or 0).replace(",", "."))
            except (TypeError, ValueError):
                valor = 0.0
            aprovada = str(g(i_aprovada) or "").strip().upper() == "SIM"
            # tecnico_dirty == 'SIM' -> snapshot desatualizado;
            # exibimos o oposto (tecAtual True quando NAO dirty).
            tec_atual = str(g(i_dirty) or "").strip().upper() != "SIM"
            row_passou = passou_list[row_i] if row_i < len(passou_list) else True

            # Curado (compat com codigo antigo dos passos 3.x).
            out.append({
                "cod": str(g(i_cod) or ""),
                "ano": ano,
                "pi": self._fmt_pi(g(i_pi_base), g(i_item)),
                "projeto": str(g(i_projeto) or ""),
                "alim": str(g(i_alim) or ""),
                "alim_benef": str(g(i_alim_benef) or ""),  # [B6]
                "se": str(g(i_se) or ""),
                "regional": str(g(i_regional) or ""),
                "superintendencia": str(g(i_super) or ""),  # [B6]
                "pacote": str(g(i_pacote) or ""),
                "valor": valor,
                "aprovada": aprovada,
                "passou": row_passou,
                "tecAtual": tec_atual,
            })
            # Raw: mesmas linhas que o desktop usa em
            # MainWindow.load_obras_into_table (todas as colunas, ordem
            # de get_column_names()).
            raw_row: list[Any] = []
            for i in range(len(cols)):
                v = r[i] if i < len(r) else ""
                # Serializa pra string -- pywebview JSON nao aceita
                # tipos exoticos do sqlite (bytes, datetime).
                if v is None:
                    raw_row.append("")
                elif isinstance(v, (bytes, bytearray)):
                    try:
                        raw_row.append(v.decode("utf-8", "replace"))
                    except Exception:  # noqa: BLE001
                        raw_row.append(repr(v))
                else:
                    raw_row.append(v if isinstance(v, (int, float, str, bool))
                                     else str(v))
            raw_out.append(raw_row)
            passou_out.append(row_passou)
            if max_rows is not None and len(out) >= max_rows:
                break

        return {
            "ok": True, "error": "",
            "rows": out,
            "total": len(raw_rows),
            # Dados crus para a renderizacao "fiel" do desktop:
            # todas as colunas em get_column_names(), na mesma ordem.
            "columns": cols,
            "raw_rows": raw_out,
            "passou_per_row": passou_out,
        }

    # ------------------------------------------------------------------
    # Passo 3.2 (Visualizar / stat cards): agregados rapidos para a faixa
    # de 4 cards no topo da aba Visualizar (Obras, Aprovadas, Pendentes,
    # Valor planejado). Roda no banco direto (SQL) quando possivel para
    # nao precisar trazer a lista inteira; cai no fetch_all como fallback.
    # ------------------------------------------------------------------
    def format_pagination_label(
        self,
        current_page: Any = 1,
        total_pages: Any = 1,
        total_items: Any = 0,
    ) -> dict[str, Any]:
        """[D6] Wrapper sobre visualizar_pagination.format_pagination_label
        para o JS reutilizar a logica do desktop sem duplicar string."""
        try:
            from visualizar_pagination import format_pagination_label as _impl
            label = _impl(int(current_page or 1),
                          int(total_pages or 1),
                          int(total_items or 0))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "label": "", "error": f"format: {exc}"}
        return {"ok": True, "label": str(label or ""), "error": ""}

    def get_obras_stats(self) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {
                "ok": False, "error": err,
                "total": 0, "aprovadas": 0, "pendentes": 0,
                "valor_total": 0.0, "ano_dominante": None,
            }

        # Tenta SQL agregado (rapido em bancos grandes); se falhar usa
        # fetch_all e itera em Python.
        try:
            cursor = db._get_cursor()  # acessor interno do DatabaseManager
            if cursor is not None:
                cursor.execute(
                    "SELECT "
                    " COUNT(*),"
                    " SUM(CASE WHEN UPPER(COALESCE(obra_aprovada,''))='SIM' THEN 1 ELSE 0 END),"
                    " COALESCE(SUM(CAST(REPLACE(REPLACE(COALESCE(valor_obra,'0'),'.',''),',','.') AS REAL)), 0)"
                    " FROM obras"
                )
                row = cursor.fetchone()
                total = int(row[0] or 0)
                aprovadas = int(row[1] or 0)
                valor_total = float(row[2] or 0.0)
                # Ano dominante via SQL separado (ordem por count).
                cursor.execute(
                    "SELECT ano_, COUNT(*) c FROM obras WHERE ano_ IS NOT NULL "
                    "AND TRIM(ano_)<>'' GROUP BY ano_ ORDER BY c DESC LIMIT 1"
                )
                ano_row = cursor.fetchone()
                ano_dominante = (
                    str(ano_row[0]).strip() if ano_row and ano_row[0] is not None else None
                )
                return {
                    "ok": True, "error": "",
                    "total": total,
                    "aprovadas": aprovadas,
                    "pendentes": max(0, total - aprovadas),
                    "valor_total": valor_total,
                    "ano_dominante": ano_dominante,
                }
        except Exception:  # noqa: BLE001
            # Cai no fallback abaixo.
            pass

        # Fallback: agregacao em Python.
        try:
            rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {
                "ok": False, "error": f"fetch_all: {exc}",
                "total": 0, "aprovadas": 0, "pendentes": 0,
                "valor_total": 0.0, "ano_dominante": None,
            }

        try:
            i_apr = cols.index("obra_aprovada")
        except ValueError:
            i_apr = -1
        try:
            i_val = cols.index("valor_obra")
        except ValueError:
            i_val = -1
        try:
            i_ano = cols.index("ano_")
        except ValueError:
            i_ano = -1

        total = len(rows)
        aprovadas = 0
        valor_total = 0.0
        anos: dict[str, int] = {}
        for r in rows:
            if i_apr >= 0 and i_apr < len(r):
                if str(r[i_apr] or "").strip().upper() == "SIM":
                    aprovadas += 1
            if i_val >= 0 and i_val < len(r):
                try:
                    valor_total += float(str(r[i_val] or 0).replace(",", "."))
                except (TypeError, ValueError):
                    pass
            if i_ano >= 0 and i_ano < len(r):
                ano_v = str(r[i_ano] or "").strip()
                if ano_v:
                    anos[ano_v] = anos.get(ano_v, 0) + 1
        ano_dominante = max(anos.items(), key=lambda kv: kv[1])[0] if anos else None
        return {
            "ok": True, "error": "",
            "total": total,
            "aprovadas": aprovadas,
            "pendentes": max(0, total - aprovadas),
            "valor_total": valor_total,
            "ano_dominante": ano_dominante,
        }

    # ------------------------------------------------------------------
    # Passo 3.3 (Visualizar / search + filtros): aplica busca textual
    # multi-termo (separador ;,) sobre todos os campos visiveis e filtros
    # estruturados vindos do modal de filtros avancados. Reutiliza
    # ui_helpers.matches_filter_value / matches_cod_terms ja existentes.
    # ------------------------------------------------------------------
    @staticmethod
    def _split_terms(s: str) -> list[str]:
        return [t.strip() for t in re.split(r"[;,]", str(s or "")) if t.strip()]

    def search_obras(
        self,
        query: Any = "",
        filters: Any = None,
    ) -> dict[str, Any]:
        base = self.list_obras()
        if not base.get("ok"):
            return base
        rows: list[dict[str, Any]] = base.get("rows") or []
        # Mantemos referencias paralelas para filtrar raw_rows e
        # passou_per_row em sintonia com o array curado.
        raw_all = list(base.get("raw_rows") or [])
        passou_all = list(base.get("passou_per_row") or [])
        cols = list(base.get("columns") or [])
        # Indexa cada row curado pela sua posicao original; o filtro
        # remove tanto do curado quanto dos paralelos.
        for i, o in enumerate(rows):
            if isinstance(o, dict):
                o["__src_idx"] = i

        # --- 1) Busca textual global ---
        q_terms = self._split_terms(str(query or "").lower())
        if q_terms:
            def _haystack(o: dict[str, Any]) -> str:
                return " ".join(str(o.get(k, "")) for k in (
                    "cod", "ano", "pi", "projeto", "alim",
                    "alim_benef", "se", "regional", "pacote",
                )).lower()
            rows = [o for o in rows if any(t in _haystack(o) for t in q_terms)]

        # --- 2) Filtros estruturados ---
        f = filters if isinstance(filters, dict) else {}

        def fv(key: str) -> str:
            return str(f.get(key, "") or "").strip()

        try:
            from ui_helpers import (  # type: ignore[import-not-found]
                matches_cod_terms,
                matches_filter_value,
            )
        except Exception:  # noqa: BLE001
            matches_cod_terms = lambda v, p: True  # noqa: E731
            matches_filter_value = lambda v, p: True  # noqa: E731

        if fv("cod"):
            rows = [o for o in rows if matches_cod_terms(str(o.get("cod", "")), fv("cod"))]
        if fv("ano"):
            anos = self._split_terms(fv("ano"))
            rows = [o for o in rows if str(o.get("ano", "")).strip() in anos]
        for key, field in (
            ("pi", "pi"),
            ("projeto", "projeto"),
            ("alim", "alim"),
            ("alim_benef", "alim_benef"),  # [B6] coluna propria agora exposta em list_obras
            ("se", "se"),
        ):
            if fv(key):
                rows = [
                    o for o in rows
                    if matches_filter_value(str(o.get(field, "")), fv(key))
                ]
        # Selects: ignoram placeholders "Todas" / "Todos" / "—".
        sentinels_select = {"todas", "todos", "—", "-", "todos os pacotes", ""}

        def _multi(key: str) -> list[str]:
            """Split ;-separated values do front (multi-select). Filtra
            sentinels e vazios. Retorna lista normalizada uppercase."""
            raw = fv(key)
            if not raw or raw.lower() in sentinels_select:
                return []
            parts = [p.strip() for p in raw.replace(",", ";").split(";")]
            return [p.upper() for p in parts
                    if p and p.lower() not in sentinels_select]

        regs = _multi("regional")
        if regs:
            rows = [o for o in rows
                    if str(o.get("regional", "")).upper() in regs]
        # Superintendencia ainda nao exposta em list_obras (placeholder
        # ate Passo de mapping incluir o atributo).
        sups = _multi("superintendencia")
        if sups:
            rows = [o for o in rows
                    if str(o.get("superintendencia", "")).upper() in sups]
        pacs = _multi("pacote")
        if pacs:
            rows = [o for o in rows
                    if str(o.get("pacote", "")).strip().upper() in pacs]
        # Aprovada / Tecnico Atualizado / Criterios usam as flags ja
        # mapeadas em list_obras.
        apr = fv("aprovada").upper().replace("Ã", "A")
        if apr in ("SIM", "NAO"):
            want = apr == "SIM"
            rows = [o for o in rows if bool(o.get("aprovada")) == want]
        tec = fv("tecnico").upper().replace("Ã", "A")
        if tec in ("SIM", "NAO"):
            want = tec == "SIM"
            rows = [o for o in rows if bool(o.get("tecAtual")) == want]
        crit = fv("criterios").lower().replace("ã", "a").replace("ç", "c")
        if crit == "atenderam":
            rows = [o for o in rows if o.get("passou", True)]
        elif crit == "falharam":
            rows = [o for o in rows if not o.get("passou", True)]
        elif crit == "aprovadas":
            rows = [o for o in rows if o.get("aprovada")]
        elif crit in ("nao aprovadas", "nao_aprovadas"):
            rows = [o for o in rows if not o.get("aprovada")]

        # Reconstroi os arrays paralelos com base nos __src_idx que
        # sobreviveram aos filtros.
        kept_idx = [int(o.get("__src_idx", -1)) for o in rows
                    if isinstance(o, dict)]
        kept_idx = [i for i in kept_idx if 0 <= i < len(raw_all)]
        filtered_raw = [raw_all[i] for i in kept_idx]
        filtered_passou = [passou_all[i] for i in kept_idx]
        for o in rows:
            if isinstance(o, dict):
                o.pop("__src_idx", None)
        return {
            "ok": True, "error": "",
            "rows": rows,
            "raw_rows": filtered_raw,
            "passou_per_row": filtered_passou,
            "columns": cols,
            "total": len(rows),
        }

    # ------------------------------------------------------------------
    # Passo 3.6 (Visualizar / acoes da toolbar): delete + export real
    # para os botoes Excluir/Detalhamento. Atualizar e' apenas JS
    # (chama coplanLoadObras). Relatorio de Criterios e Nota de Colapso
    # ficam como stubs ate Passos 5.x/6.x trazerem os geradores
    # adequados (dependem de templates e ganhos parsing).
    # ------------------------------------------------------------------
    @staticmethod
    def _default_export_dir() -> Path:
        target = Path.home() / "Downloads"
        try:
            if not target.exists():
                target = Path.home()
            target.mkdir(parents=True, exist_ok=True)
        except OSError:
            target = Path.home()
        return target

    # ------------------------------------------------------------------
    # Gating de obras aprovadas (RB-2 do desktop, FiltrosPaginacaoMixin):
    # Replica _gate_aprovadas_for_action + _confirmar_exclusao_excepcional
    # + _registrar_exclusao_excepcional. Antes de qualquer acao destrutiva
    # ou de mutacao em obras, o JS deve consultar
    # `gate_aprovadas_for_action(cods)` para descobrir quais estao aprovadas.
    # Se houver aprovadas e o usuario nao marcou 'incluir aprovadas',
    # essas devem ser ignoradas.
    # ------------------------------------------------------------------
    def gate_aprovadas_for_action(
        self, cods: Any = None, include_aprovadas: Any = False,
    ) -> dict[str, Any]:
        """Filtra obras pela coluna ``obra_aprovada`` ('SIM'/'NAO').
        Retorna 3 listas:
          * ``targets``: cods que podem ser processados nesta acao
          * ``aprovadas``: cods que estao aprovados (so entrariam em
            ``targets`` se include_aprovadas=True)
          * ``inexistentes``: cods que nao foram encontrados no banco

        O JS usa isso pra mostrar dialog do tipo 'X obra(s) aprovada(s)
        foram ignoradas' antes de chamar delete_obras/marcar_correcao."""
        cods_list = [str(c).strip() for c in (cods or []) if str(c or "").strip()]
        include = bool(include_aprovadas)
        if not cods_list:
            return {"ok": False, "targets": [], "aprovadas": [],
                    "inexistentes": [], "error": "cods vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "targets": [], "aprovadas": [],
                    "inexistentes": [], "error": err or "db indisponivel"}
        try:
            cols = list(db.get_column_names() or [])
            i_cod = cols.index("cod") if "cod" in cols else -1
            i_aprov = (cols.index("obra_aprovada")
                       if "obra_aprovada" in cols else -1)
            rows = list(db.fetch_by_cods(cods_list) or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "targets": [], "aprovadas": [],
                    "inexistentes": [], "error": f"fetch: {exc}"}
        achados: dict[str, str] = {}
        for row in rows:
            cod_v = (str(row[i_cod]).strip()
                     if 0 <= i_cod < len(row) else "")
            aprov = (str(row[i_aprov]).strip()
                     if 0 <= i_aprov < len(row) else "").upper()
            if cod_v:
                achados[cod_v] = aprov
        targets: list[str] = []
        aprovadas: list[str] = []
        inexistentes: list[str] = []
        for c in cods_list:
            if c not in achados:
                inexistentes.append(c)
                continue
            if achados[c] == "SIM":
                aprovadas.append(c)
                if include:
                    targets.append(c)
            else:
                targets.append(c)
        return {
            "ok": True,
            "targets": targets,
            "aprovadas": aprovadas,
            "inexistentes": inexistentes,
            "include_aprovadas": include,
            "error": "",
        }

    def register_exclusao_excepcional(
        self, cod: Any, motivo: Any = "",
    ) -> dict[str, Any]:
        """Auditoria de exclusao excepcional de obra aprovada.
        Replica _registrar_exclusao_excepcional do desktop: anexa nota
        em ``observacoes_gerais`` (ou ``observacoes`` ou ``ultima_acao``)
        com timestamp + usuario + motivo, e emite warning no log.
        Chamar SEMPRE antes de delete_obras quando o cod estiver na
        lista de 'aprovadas'."""
        cod_s = str(cod or "").strip()
        if not cod_s:
            return {"ok": False, "error": "cod vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel"}
        try:
            usuario = getpass.getuser()
        except Exception:  # noqa: BLE001
            usuario = "?"
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
        motivo_s = str(motivo or "").strip()
        nota = f"EXCLUSAO EXCEPCIONAL em {timestamp} por {usuario}"
        if motivo_s:
            nota += f" -- motivo: {motivo_s}"
        try:
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"cols: {exc}"}
        coluna_log = None
        for col in ("ultima_acao", "observacoes_gerais",
                    "observacoes", "log"):
            if col in cols:
                coluna_log = col
                break
        if not coluna_log:
            # Sem coluna de log; ainda registra no stderr.
            print(f"[main_web] {nota} (cod={cod_s}, sem coluna de log)",
                  file=sys.stderr)
            return {"ok": True, "logged_to_db": False, "nota": nota}
        try:
            existing = db.fetch_by_cod(cod_s)
            if not existing:
                return {"ok": False, "error": f"cod {cod_s} nao encontrado"}
            i_log = cols.index(coluna_log)
            atual = str(existing[i_log] or "").strip() if 0 <= i_log < len(existing) else ""
            novo = f"{atual}\n{nota}" if atual else nota
            db.update_obra({coluna_log: novo}, cod_s, skip_blank=True)
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            err_s = friendly or f"update: {exc}"
            print(f"[main_web] register_exclusao falhou cod={cod_s}: {err_s}",
                  file=sys.stderr)
            out: dict[str, Any] = {
                "ok": False, "error": err_s,
                "logged_to_db": False, "nota": nota,
            }
            if friendly:
                out["blocked"] = "db_busy"
            return out
        print(f"[main_web] AUDIT: {nota} cod={cod_s}", file=sys.stderr)
        return {"ok": True, "logged_to_db": True, "nota": nota,
                "coluna": coluna_log}

    def delete_obras(self, cods: Any) -> dict[str, Any]:
        """Deleta obras por COD. Usa DatabaseManager.delete_obra (ja
        protegido por lock + transacao). Retorna {ok, deleted, errors}.
        BLOQUEADO quando cenario_ativo != '' (cenario nao deve apagar
        obras em obras)."""
        cen_nome = self._cenario_active_name()
        if cen_nome:
            return {
                "ok": False, "deleted": 0,
                "errors": [(f"Operacao bloqueada: cenario '{cen_nome}'"
                            f" ativo. Saia do cenario para excluir obras.")],
                "blocked": "cenario_active",
            }
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "deleted": 0, "errors": [err or "db indisponivel"]}
        if not isinstance(cods, (list, tuple)) or not cods:
            return {"ok": False, "deleted": 0, "errors": ["lista de cods vazia"]}
        deleted = 0
        errors: list[str] = []
        busy_msg = ""
        for cod in cods:
            cod_s = str(cod or "").strip()
            if not cod_s:
                continue
            try:
                db.delete_obra(cod_s)
                deleted += 1
            except Exception as exc:  # noqa: BLE001
                friendly = self._friendly_busy_error(exc)
                if friendly:
                    # Busy/locked: para o loop (nao adianta tentar mais
                    # cods se o banco esta em uso por outro usuario).
                    busy_msg = friendly
                    errors.append(f"{cod_s}: {friendly}")
                    break
                errors.append(f"{cod_s}: {exc}")
        out: dict[str, Any] = {
            "ok": not errors, "deleted": deleted, "errors": errors,
        }
        if busy_msg:
            out["blocked"] = "db_busy"
            out["error"] = busy_msg
        return out

    def export_detalhamento(self, cods: Any = None) -> dict[str, Any]:
        """Exporta as obras selecionadas (ou todas) para XLSX em
        ~/Downloads. Retorna {ok, path, count, error, cenario}."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count": 0, "error": err or "db indisponivel"}
        try:
            cods_list = [str(c).strip() for c in (cods or []) if str(c).strip()]
            if cods_list:
                raw_rows = db.fetch_by_cods(cods_list) or []
            else:
                raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"fetch: {exc}"}

        # Cenario ativo: restringe + aplica overrides (paridade get_obras).
        # Loga para diagnostico: cenario={X}, rows antes/depois.
        cen_nome = self._cenario_active_name()
        before_n = len(raw_rows)
        raw_rows = self._apply_cenario_to_rows(db, raw_rows, cols)
        print(
            f"[main_web] export_detalhamento: cenario={cen_nome!r}, "
            f"cods_in={len(cods_list)}, rows_before={before_n}, "
            f"rows_after={len(raw_rows)}",
            file=sys.stderr,
        )

        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            from openpyxl.styles import Font, PatternFill  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"openpyxl indisponivel: {exc}"}

        target = self._default_export_dir()
        fname = f"coplan_detalhamento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = target / fname
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Obras"
            ws.append(cols)
            header_fill = PatternFill(start_color="2A3460", end_color="2A3460", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for r in raw_rows:
                ws.append([
                    (r[i] if i < len(r) else "") for i in range(len(cols))
                ])
            ws.freeze_panes = "A2"
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"openpyxl save: {exc}"}

        return {
            "ok": True, "path": str(path),
            "count": len(raw_rows), "error": "",
            "cenario": cen_nome,
        }

    # ------------------------------------------------------------------
    # Fase A7 (resumo_service.montar_resumo_detalhamento):
    # exporta XLSX agrupando por (nome_projeto, ano, pacote) com
    # antes/depois por alimentador. Equivalente a
    # MainWindow._montar_resumo_detalhamento_excel + _exportar_obras
    # (sub-modo "Detalhamento por Regional").
    # ------------------------------------------------------------------
    def export_resumo_detalhamento(self, cods: Any = None) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count": 0,
                    "error": err or "db indisponivel"}
        try:
            cods_list = [str(c).strip() for c in (cods or []) if str(c).strip()]
            if cods_list:
                raw_rows = db.fetch_by_cods(cods_list) or []
            else:
                raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"fetch: {exc}"}
        raw_rows = self._apply_cenario_to_rows(db, raw_rows, cols)
        if not raw_rows:
            return {"ok": False, "path": "", "count": 0,
                    "error": "sem obras para resumir"}
        try:
            from core.services.resumo_service import (  # type: ignore[import-not-found]
                montar_resumo_detalhamento,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"import: {exc}"}
        try:
            df = montar_resumo_detalhamento(raw_rows, cols)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"montar: {exc}"}
        if df is None or df.empty:
            return {"ok": False, "path": "", "count": 0,
                    "error": ("sem dados (faltam colunas obrigatorias ou"
                              " nenhum alimentador resolvido)")}
        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            from openpyxl.styles import (  # type: ignore[import-not-found]
                Font, PatternFill,
            )
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumo Detalhamento"
            cols_out = list(df.columns)
            ws.append(cols_out)
            header_fill = PatternFill(start_color="2A3460",
                                      end_color="2A3460", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for _, row in df.iterrows():
                ws.append([row.get(c, "") for c in cols_out])
            ws.freeze_panes = "A2"
            target = self._default_export_dir()
            fname = (
                f"coplan_resumo_detalhamento_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            path = target / fname
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"xlsx: {exc}"}
        return {"ok": True, "path": str(path),
                "count": int(len(df)), "error": ""}

    def export_relatorio_criterios(self, cods: Any = None) -> dict[str, Any]:
        """Gera planilha das obras que NAO atenderam criterios.
        Visualizar Sprint 1 (Auditoria #5): aceita filtro `cods` opcional
        (lista de cods para escopo 'filtradas'/'selecionadas')."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count": 0, "error": err or "db indisponivel"}
        try:
            raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"fetch: {exc}"}
        # Cenario ativo: restringe + overrides ANTES do filtro de cods
        raw_rows = self._apply_cenario_to_rows(db, raw_rows, cols)
        # Filtra por cods se fornecido (#5)
        if isinstance(cods, (list, tuple)) and cods:
            cod_set = {str(c).strip() for c in cods if str(c or "").strip()}
            if cod_set:
                try:
                    idx_cod = cols.index("cod")
                    raw_rows = [
                        r for r in raw_rows
                        if str(r[idx_cod] if idx_cod < len(r) else "").strip()
                        in cod_set
                    ]
                except ValueError:
                    pass  # 'cod' nao esta nas colunas - ignora filtro
        try:
            from core.services.relatorio_criterios_service import (  # type: ignore[import-not-found]
                verificar_criterios_v2,
            )
            criterios = (self._config or {}).get("criterios_planejamento") or {}
            verdict = verificar_criterios_v2(raw_rows, cols, criterios=criterios)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"verificar_criterios_v2: {exc}"}

        falhas = [
            r for r, ok in zip(raw_rows, verdict)
            if ok is False  # None = indefinido (nao conta como falha)
        ]
        if not falhas:
            return {"ok": True, "path": "", "count": 0,
                    "error": "Todas as obras atenderam aos criterios."}
        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            wb = Workbook()
            ws = wb.active
            ws.title = "Falhas"
            ws.append(cols)
            for r in falhas:
                ws.append([r[i] if i < len(r) else "" for i in range(len(cols))])
            ws.freeze_panes = "A2"
            target = self._default_export_dir()
            fname = (
                f"coplan_relatorio_criterios_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            path = target / fname
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"xlsx: {exc}"}
        return {"ok": True, "path": str(path), "count": len(falhas), "error": ""}

    # ------------------------------------------------------------------
    # Fase B2 (excluir_obra_mixin.marcar_obras_correcao):
    # marca COD(s) como DESPACHO_STATUS = "CORRECAO". Usado depois que
    # uma obra DESPACHADA precisa ser alterada (Fase A9 bloqueia o save
    # ate ela voltar a CORRECAO).
    # ------------------------------------------------------------------
    def marcar_obras_correcao(
        self, cods: Any = None, motivo: Any = "",
    ) -> dict[str, Any]:
        cen_nome = self._cenario_active_name()
        if cen_nome:
            return {
                "ok": False, "error":
                (f"Operacao bloqueada: cenario '{cen_nome}' ativo."
                 f" Saia do cenario para marcar obras como CORRECAO."),
                "blocked": "cenario_active",
                "marcadas": 0, "falhas": [],
            }
        if not isinstance(cods, (list, tuple)) or not cods:
            return {"ok": False, "error": "cods vazio",
                    "marcadas": 0, "falhas": []}
        motivo_s = str(motivo or "").strip()
        # [FIX] Aceita motivo vazio com placeholder. UX: o user nao
        # precisa digitar motivo ANTES de editar (era pedido em prompt
        # + de novo no salvamento). Agora marca como pendente; o
        # motivo real e' capturado em cad-input-motivo no save_obra.
        if not motivo_s:
            motivo_s = "PENDENTE - informar no salvamento"
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "marcadas": 0, "falhas": []}
        now_iso = datetime.now().isoformat(timespec="seconds")
        despacho_ref = f"CORRECAO: {motivo_s}"
        falhas: list[str] = []
        marcadas = 0
        busy_msg = ""
        for c in cods:
            cod_s = str(c or "").strip()
            if not cod_s:
                continue
            try:
                db.update_obra({
                    "despacho_status": "CORRECAO",
                    "despacho_em": now_iso,
                    "despacho_ref": despacho_ref,
                }, cod_s, skip_blank=True)
                marcadas += 1
            except Exception as exc:  # noqa: BLE001
                friendly = self._friendly_busy_error(exc)
                if friendly:
                    busy_msg = friendly
                    falhas.append(f"{cod_s}: {friendly}")
                    break
                falhas.append(f"{cod_s}: {exc}")
        return {
            "ok": (marcadas > 0 and not busy_msg),
            "error": busy_msg,
            "blocked": "db_busy" if busy_msg else "",
            "marcadas": marcadas,
            "falhas": falhas,
            "motivo": motivo_s,
        }

    # ------------------------------------------------------------------
    # Fase A12 (relatorio_criterios_service.montar_relatorio_criterios_por_projeto):
    # Gera XLSX 2-sheet (Projetos + Alimentadores) com avaliacao
    # detalhada de criterios por projeto. Equivalente a
    # MainWindow.montar_relatorio_criterios_por_projeto do desktop.
    # ------------------------------------------------------------------
    def export_relatorio_criterios_projeto(
        self, cods: Any = None,
    ) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": err or "db indisponivel"}
        try:
            cods_list = [str(c).strip() for c in (cods or []) if str(c).strip()]
            if cods_list:
                raw_rows = db.fetch_by_cods(cods_list) or []
            else:
                raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": f"fetch: {exc}"}
        raw_rows = self._apply_cenario_to_rows(db, raw_rows, cols)
        if not raw_rows:
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0,
                    "error": "sem obras para analisar"}
        try:
            from core.services.relatorio_criterios_service import (  # type: ignore[import-not-found]
                montar_relatorio_criterios_por_projeto,
            )
            from codigo5_coplan import (  # type: ignore[import-not-found]
                DEFAULT_CRITERIOS,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": f"import: {exc}"}

        criterios = dict(DEFAULT_CRITERIOS)
        criterios.update((self._config or {}).get("criterios_planejamento") or {})

        try:
            rel = montar_relatorio_criterios_por_projeto(
                raw_rows, cols, criterios=criterios,
            )
        except ValueError as exc:
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": str(exc)}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": f"montar: {exc}"}

        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            from openpyxl.styles import (  # type: ignore[import-not-found]
                Font, PatternFill,
            )
            wb = Workbook()
            header_fill = PatternFill(start_color="2A3460",
                                      end_color="2A3460", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # Sheet 1: Projetos
            ws1 = wb.active
            ws1.title = "Projetos"
            cols1 = list(rel.df_projetos.columns)
            ws1.append(cols1)
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
            for _, row in rel.df_projetos.iterrows():
                ws1.append([row.get(c, "") for c in cols1])
            ws1.freeze_panes = "A2"

            # Sheet 2: Alimentadores
            ws2 = wb.create_sheet("Alimentadores")
            cols2 = list(rel.df_alimentadores.columns)
            ws2.append(cols2)
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
            for _, row in rel.df_alimentadores.iterrows():
                ws2.append([row.get(c, "") for c in cols2])
            ws2.freeze_panes = "A2"

            target = self._default_export_dir()
            fname = (
                f"coplan_relatorio_criterios_projeto_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            path = target / fname
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": f"xlsx: {exc}"}

        return {
            "ok": True, "path": str(path), "error": "",
            "count_projetos": int(len(rel.df_projetos)),
            "count_alimentadores": int(len(rel.df_alimentadores)),
        }

    # ------------------------------------------------------------------
    # Fase A3 (core/services/nota_colapso_service):
    # exporta nota de colapso via core (sem Qt). Substitui o stub.
    # Para cada COD: monta Obra + PIMetadata, chama calcular_nota_colapso
    # e produz XLSX com cod/nota/criterio/valores_considerados.
    # ------------------------------------------------------------------
    def export_nota_colapso(self, cods: Any = None) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count": 0,
                    "error": err or "db indisponivel"}
        try:
            from core.models import Obra  # type: ignore[import-not-found]
            from core.services.nota_colapso_service import (  # type: ignore[import-not-found]
                calcular_nota_colapso,
            )
            from core.services.pi_metadata_service import (  # type: ignore[import-not-found]
                buscar_pi_metadata,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"import: {exc}"}

        cods_list: list[str] = []
        if isinstance(cods, (list, tuple)):
            cods_list = [str(c).strip() for c in cods if str(c or "").strip()]
        try:
            cols = list(db.get_column_names() or [])
            if cods_list:
                rows = db.fetch_by_cods(cods_list) or []
            else:
                rows = db.fetch_all() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"fetch: {exc}"}

        rows = self._apply_cenario_to_rows(db, rows, cols)

        cfg = self._config or {}
        resultados: list[dict[str, Any]] = []
        for r in rows:
            row_dict = {col: r[i] if i < len(r) else ""
                        for i, col in enumerate(cols)}
            try:
                obra = Obra.from_row_dict(row_dict)
            except Exception as exc:  # noqa: BLE001
                resultados.append({
                    "cod": str(row_dict.get("cod") or ""),
                    "valor": None,
                    "criterio": f"Erro ao montar Obra: {exc}",
                    "valores": {},
                })
                continue
            pi_md = None
            try:
                pi_md = buscar_pi_metadata(
                    str(obra.ident.projeto_investimento or ""), cfg,
                )
            except Exception:  # noqa: BLE001
                pi_md = None
            try:
                nota = calcular_nota_colapso(obra, pi_md)
            except Exception as exc:  # noqa: BLE001
                resultados.append({
                    "cod": obra.ident.cod,
                    "valor": None,
                    "criterio": f"Erro no calculo: {exc}",
                    "valores": {},
                })
                continue
            resultados.append({
                "cod": obra.ident.cod,
                "pi": obra.ident.projeto_investimento,
                "valor": nota.valor,
                "criterio": nota.criterio,
                "valores": dict(nota.valores_considerados or {}),
            })

        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            wb = Workbook()
            ws = wb.active
            ws.title = "Nota Colapso"
            ws.append([
                "COD", "PI", "Nota", "Criterio",
                "Carreg Inicial", "Carreg Max",
                "Tensao Min Inicial", "Tensao Max Inicial",
                "Tmin Registrada", "Tmax Registrada",
            ])
            for item in resultados:
                v = item.get("valores") or {}
                ws.append([
                    item.get("cod") or "",
                    item.get("pi") or "",
                    item.get("valor") if item.get("valor") is not None else "",
                    item.get("criterio") or "",
                    v.get("carreg_inicial", ""),
                    v.get("carreg_max", ""),
                    v.get("tensao_min_inicial", ""),
                    v.get("tensao_max_inicial", ""),
                    v.get("tmin_registrada", ""),
                    v.get("tmax_registrada", ""),
                ])
            ws.freeze_panes = "A2"
            target = self._default_export_dir()
            fname = (
                f"coplan_nota_colapso_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            path = target / fname
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"xlsx: {exc}"}

        return {
            "ok": True, "path": str(path),
            "count": len(resultados), "error": "",
        }

    # ------------------------------------------------------------------
    # Passo 4.1 (Cadastro / get_obra): le UMA obra completa por COD para
    # popular o formulario quando o usuario decide editar uma existente.
    # Retorna o dict cru (todas as colunas de ORDERED_COLUMNS) + alguns
    # campos derivados que o JS precisa (ex: alimentadores_beneficiados
    # ja como lista).
    # ------------------------------------------------------------------
    def get_obra(self, cod: Any) -> dict[str, Any]:
        cod_s = str(cod or "").strip()
        if not cod_s:
            return {"ok": False, "obra": None, "error": "cod vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "obra": None, "error": err or "db indisponivel"}
        try:
            row = db.fetch_by_cod(cod_s)
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "obra": None, "error": f"fetch_by_cod: {exc}"}
        if not row:
            return {"ok": False, "obra": None, "error": f"obra nao encontrada: {cod_s}"}

        # Cenarios (Sprint A): aplica ano_final + overrides quando ativo
        cen_nome = self._cenario_active_name()
        if cen_nome:
            cod_set, cen_info = self._cenario_cod_set(db, cen_nome)
            if cod_s in cod_set:
                ovmap = self._cenario_overrides_map(db, cen_nome)
                row = tuple(self._cenario_apply_to_row(
                    list(row), cols, cod_s,
                    cen_info.get(cod_s) or {},
                    ovmap.get(cod_s) or {},
                ))
        obra = self._row_to_dict(row, cols)
        # Deriva lista de alimentadores beneficiados (separador ; comum
        # no banco). JS pode chunkar como chips diretamente.
        alim_benef_raw = str(obra.get("alimentadores_beneficiados") or "")
        alim_benef = [
            x.strip() for x in re.split(r"[;,]", alim_benef_raw) if x.strip()
        ]
        # Deriva subestacoes a partir do prefixo dos alimentadores
        # beneficiados (igual ao desktop: ATB-204 -> ATB).
        ses = []
        for a in alim_benef:
            prefix = re.split(r"[-_/]", a, 1)[0].strip().upper()
            if prefix and prefix not in ses:
                ses.append(prefix)

        return {
            "ok": True,
            "obra": obra,
            "alim_benef": alim_benef,
            "ses_derivadas": ses,
            "error": "",
        }

    # ------------------------------------------------------------------
    # Passo 4.2 (Cadastro / save_obra): persiste o dict do formulario.
    # Decide entre INSERT e UPDATE com base na existencia do COD na base.
    # Reusa DatabaseManager.insert_obra / update_obra que ja aplicam:
    #   * sanitizacao de alimentador (sem '_')
    #   * derivacao de pi_base via get_pi_base
    #   * data_criacao / data_modificacao / criado_por / modificado_por
    #   * defaults (obra_aprovada=NAO, tecnico_dirty=NAO)
    #   * empresa + cod_pep tail
    #   * lock + transacao + retry-on-busy
    # ------------------------------------------------------------------
    def save_obra(self, payload: Any) -> dict[str, Any]:
        if not isinstance(payload, dict):
            return {"ok": False, "cod": "", "mode": "", "error": "payload nao e dict"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "cod": "", "mode": "", "error": err or "db indisponivel"}

        cod = str(payload.get("cod") or "").strip()

        # Limpeza basica: remove None, normaliza string. Mantem numeros
        # como estao (insert_obra._normalize_decimal trata).
        cleaned: dict[str, Any] = {}
        cols = list(db.get_column_names() or [])
        for key, value in payload.items():
            # Ignora chaves que nao sao colunas reais (defensivo).
            if key not in cols and key != "cod":
                continue
            if value is None:
                cleaned[key] = ""
            elif isinstance(value, (int, float, bool)):
                cleaned[key] = value
            else:
                cleaned[key] = str(value).strip()

        # Fallback de tensao_operacao -> nivel_tensao_obra (regra do
        # build_obra_dados / save_data legado).
        if "tensao_operacao" in cols and not str(cleaned.get("tensao_operacao") or "").strip():
            niv = cleaned.get("nivel_tensao_obra") or ""
            if niv:
                cleaned["tensao_operacao"] = niv
        # Toda obra recem-salva parte como NAO dirty (mesma regra do core).
        if "tecnico_dirty" in cols:
            cleaned["tecnico_dirty"] = "NÃO"

        # Paridade desktop validar_campos_obrigatorios: defense-in-depth.
        # JS ja valida no form, mas se o payload chegar sem algum
        # obrigatorio, abortamos antes de gastar insert/update.
        faltam = self._validar_campos_obrigatorios(cleaned)
        if faltam:
            return {
                "ok": False, "cod": cod, "mode": "",
                "error": ("Campos obrigatorios vazios: "
                          + ", ".join(faltam)),
                "campos_obrigatorios_vazios": faltam,
            }

        try:
            existing = db.fetch_by_cod(cod) if cod else None
        except Exception:  # noqa: BLE001
            existing = None

        # Fase A8: avalia diff + anexa historico. Bloqueio por DESPACHADA
        # foi REMOVIDO (2026-05-08) a pedido do usuario: o operador pode
        # editar obras despachadas livremente. O status da nota e seu
        # numero ficam visiveis no card "Status da Nota" da aba Cadastro
        # (informativo, sem bloqueio).
        old_map: dict[str, Any] = {}
        if existing:
            for i, c in enumerate(cols):
                old_map[c] = existing[i] if i < len(existing) else ""

        diff_decision = None
        try:
            from core.services.salvar_obra_service import (  # type: ignore[import-not-found]
                aplicar_historico_ao_dict, avaliar_diff,
            )
            diff_decision = avaliar_diff(cleaned, old_map, db_columns=cols)
        except Exception as exc:  # noqa: BLE001
            print(f"[main_web] avaliar_diff falhou: {exc}", file=sys.stderr)

        if diff_decision is not None:
            # Anexa historico se houver mudancas + a coluna existe.
            try:
                cleaned = aplicar_historico_ao_dict(
                    cleaned, diff_decision, motivo="",
                )
            except Exception as exc:  # noqa: BLE001
                print(f"[main_web] aplicar_historico falhou: {exc}",
                      file=sys.stderr)

        # Auto-gera COD pra INSERT (paridade com codigo5_coplan.py L1127):
        # CalculationManager.gerar_cod a partir dos campos do form.
        if not existing and not cod:
            cm = self._ensure_calc_manager()
            if cm is None:
                return {"ok": False, "cod": "", "mode": "",
                        "error": "calc indisponivel para gerar COD"}
            try:
                cod = cm.gerar_cod(
                    str(cleaned.get("tipo_pacote") or ""),
                    str(cleaned.get("alimentador_principal") or ""),
                    str(cleaned.get("projeto_investimento") or ""),
                    str(cleaned.get("quantidade_material") or ""),
                    str(cleaned.get("caracteristicas_material") or ""),
                    str(cleaned.get("coordenada_fim") or ""),
                    pi_base=str(cleaned.get("pi_base") or "") or None,
                )
            except ValueError as exc:
                return {"ok": False, "cod": "", "mode": "",
                        "error": f"gerar_cod: {exc}"}
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "cod": "", "mode": "",
                        "error": f"gerar_cod: {exc}"}
            cleaned["cod"] = cod

        # Cenarios (Sprint A): redireciona save para cenario_obras_overrides
        # quando ha cenario ativo. NUNCA toca a tabela obras.
        cen_nome = self._cenario_active_name()
        if cen_nome:
            if not existing:
                return {
                    "ok": False, "cod": cod, "mode": "",
                    "error": ("Cenario ativo nao cria obras novas."
                              " Saia do cenario ('Sair do cenario' no"
                              " banner) para criar uma nova obra."),
                    "blocked": "cenario_no_create",
                    "cenario": cen_nome,
                }
            # Verifica se a obra esta no escopo do cenario
            cod_set, cen_info = self._cenario_cod_set(db, cen_nome)
            if cod and cod not in cod_set:
                return {
                    "ok": False, "cod": cod, "mode": "",
                    "error": (f"Obra {cod} nao faz parte do cenario "
                              f"'{cen_nome}'. Saia do cenario para editar."),
                    "blocked": "cenario_out_of_scope",
                    "cenario": cen_nome,
                }
            # Computa diff: para cada coluna em cleaned cujo valor
            # difere do valor atual em obras (ou da view do cenario,
            # i.e., considerando overrides anteriores), grava em
            # cenario_obras_overrides.
            ovmap = self._cenario_overrides_map(db, cen_nome)
            existing_with_overrides: dict[str, Any] = {}
            for i, c in enumerate(cols):
                base_val = existing[i] if i < len(existing) else ""
                existing_with_overrides[c] = (
                    (ovmap.get(cod) or {}).get(c, base_val)
                )
            # ano_final do CAPEX (se houver) tambem entra no baseline
            ano_final = (cen_info.get(cod) or {}).get("ano_final")
            if ano_final is not None and "ano_" in cols:
                if "ano_" not in (ovmap.get(cod) or {}):
                    existing_with_overrides["ano_"] = str(ano_final)
            diff_pairs: list[tuple[str, Any]] = []
            for col_name, new_val in cleaned.items():
                if col_name == "cod":
                    continue
                if col_name not in cols:
                    continue
                cur_val = existing_with_overrides.get(col_name, "")
                # Comparacao tolerante (string trim, case-sensitive
                # exceto whitespace).
                if str(cur_val or "").strip() != str(new_val or "").strip():
                    diff_pairs.append((col_name, new_val))
            wrote = self._cenario_save_overrides(
                db, cen_nome, cod, diff_pairs)
            return {
                "ok": True, "cod": cod, "mode": "cenario_override",
                "error": "",
                "cenario": cen_nome,
                "campos_alterados_no_cenario": [p[0] for p in diff_pairs],
                "overrides_salvos": wrote,
            }

        try:
            if existing:
                db.update_obra(cleaned, cod)
                mode = "update"
            else:
                db.insert_obra(cleaned)
                mode = "insert"
        except PermissionError as exc:
            return {"ok": False, "cod": cod, "mode": "", "error": f"permissao: {exc}"}
        except ValueError as exc:
            return {"ok": False, "cod": cod, "mode": "", "error": f"validacao: {exc}"}
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            if friendly:
                return {"ok": False, "cod": cod, "mode": "",
                        "blocked": "db_busy", "error": friendly}
            return {"ok": False, "cod": cod, "mode": "", "error": str(exc)}

        out: dict[str, Any] = {"ok": True, "cod": cod, "mode": mode, "error": ""}
        if diff_decision is not None:
            out["campos_alterados"] = list(diff_decision.campos_alterados)
            out["campos_criticos_alterados"] = list(diff_decision.campos_criticos_alterados)
        return out

    # ------------------------------------------------------------------
    # [M027] Detecao semantica de duplicada (alim+pi+ano+municipio+
    # descricao). Espelha find_duplicate_in_db do desktop. Retorna lista
    # (0 ou 1 match) para o JS abrir o modal "Obra similar encontrada".
    # ------------------------------------------------------------------
    def obras_por_codigo_semelhante(self, payload: Any) -> dict[str, Any]:
        """Procura obras semelhantes por chave semantica (definida em
        runtime/row_helpers.find_duplicate_in_db / core.repositories.
        obra_query_repo.find_duplicate). Retorna {ok, matches:list[dict]}.
        Lista pode estar vazia."""
        if not isinstance(payload, dict):
            return {"ok": False, "matches": [], "error": "payload nao eh dict"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "matches": [],
                    "error": err or "db indisponivel"}
        try:
            from runtime.row_helpers import (  # type: ignore[import-not-found]
                find_duplicate_in_db,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "matches": [],
                    "error": f"import find_duplicate_in_db: {exc}"}
        try:
            dup = find_duplicate_in_db(db, dict(payload))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "matches": [],
                    "error": f"find_duplicate: {exc}"}
        matches: list[dict[str, Any]] = []
        if dup:
            # Garante chaves uteis para o modal (cod, descricao, alim, ano).
            cod = str(dup.get("cod") or "").strip()
            matches.append({
                "cod": cod,
                "alimentador": str(dup.get("alimentador_principal")
                                   or dup.get("alimentador") or "").strip(),
                "ano": str(dup.get("ano_") or dup.get("ano") or "").strip(),
                "projeto_investimento": str(
                    dup.get("projeto_investimento") or "").strip(),
                "pi_base": str(dup.get("pi_base") or "").strip(),
                "nome_projeto": str(dup.get("nome_projeto") or "").strip(),
                "descricao_obra": str(dup.get("descricao_obra") or "").strip(),
                "municipio": str(dup.get("municipio") or "").strip(),
                # Devolve o dict original tambem para o JS poder fazer
                # merge sem perder colunas extras.
                "raw": dup,
            })
        return {"ok": True, "matches": matches, "error": ""}

    # ------------------------------------------------------------------
    # Passo 4.3 (Cadastro / gerar_cod_pep): constroi o COD da obra no
    # formato <SIGLA>-<YY>-<PI>-<ITEM> (ex.: MA-26-DI-047).
    # Importante: este "COD_PEP gerado" do mock e' o identificador da
    # obra (coluna `cod`), NAO o COD_PEP sequencial SSSS-AAA gerado pelo
    # cod_pep() do legado (que so existe pos-aprovacao).
    # ------------------------------------------------------------------
    def gerar_cod_pep(
        self,
        projeto_investimento: Any = "",
        ano: Any = "",
        item: Any = "",
        pi_base: Any = "",
    ) -> dict[str, Any]:
        # Sigla: prefere o que esta no config.json (mantem paridade com
        # desktop). Default 'MA' se nao configurado.
        cfg: dict[str, Any] = {}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception:  # noqa: BLE001
            pass
        sigla = str(cfg.get("empresa_sigla") or "MA").strip().upper()

        # YY: ultimos 2 digitos do ano informado.
        ano_s = str(ano or "").strip()
        digits = "".join(c for c in ano_s if c.isdigit())
        yy = digits[-2:] if len(digits) >= 2 else ""

        # PI: usa o pi_base passado ou deriva via legado.
        pi = str(pi_base or "").strip().upper()
        if not pi and projeto_investimento:
            try:
                from codigo5_coplan import get_pi_base  # type: ignore[import-not-found]
                pi = (get_pi_base(str(projeto_investimento), prompt_user=False) or "").upper()
            except Exception:  # noqa: BLE001
                # Fallback grosseiro: 2 primeiras letras nao-espaco do nome.
                only_letters = "".join(
                    c for c in str(projeto_investimento) if c.isalpha()
                )
                pi = only_letters[:2].upper()

        # Item: zero-pad em 3 digitos quando totalmente numerico.
        item_s = str(item or "").strip()
        if item_s.isdigit():
            item_s = item_s.zfill(3)

        parts = [sigla, yy, pi, item_s]
        complete = all(p for p in parts)
        cod = "-".join(p for p in parts if p)
        return {
            "ok": complete,
            "cod": cod,
            "sigla": sigla,
            "ano_yy": yy,
            "pi": pi,
            "item": item_s,
            "missing": [
                name for name, val in zip(
                    ("sigla", "ano", "pi", "item"), parts
                ) if not val
            ],
        }
