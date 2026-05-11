"""Mixin Exportar Excel -- 11 metodos.

Extraido de ``codigo5_coplan.py::MainWindow`` na Etapa C (Tier 3).
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PySide6 import QtCore, QtWidgets
from PySide6.QtCore import Slot


class ExportarExcelMixin:
    """Espera estar misturado em ``MainWindow(QMainWindow, ...)``."""

    def _export_qtable_to_sheet(
        self, table: QtWidgets.QTableWidget, ws
    ) -> None:
        """Exporta uma QTableWidget para uma planilha do Excel."""

        def _max_line_length(text: str) -> int:
            if not text:
                return 0
            return max(len(parte) for parte in str(text).splitlines())

        column_count = table.columnCount()
        row_count = table.rowCount()
        max_widths = [0] * column_count

        for col in range(column_count):
            header_item = table.horizontalHeaderItem(col)
            header_text = header_item.text() if header_item else ""
            cell = ws.cell(row=1, column=col + 1, value=header_text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            max_widths[col] = max(max_widths[col], _max_line_length(header_text))

        ws.freeze_panes = "A2"

        for row in range(row_count):
            for col in range(column_count):
                item = table.item(row, col)
                value = item.text() if item is not None else ""
                ws.cell(row=row + 2, column=col + 1, value=value)
                max_widths[col] = max(max_widths[col], _max_line_length(value))

        for col, width in enumerate(max_widths, start=1):
            column_width = min(max(width + 2, 10), 60)
            ws.column_dimensions[get_column_letter(col)].width = column_width

    def _close_export_progress_dialog(self) -> None:
        progress = getattr(self, "_export_progress_dialog", None)
        if progress is None:
            return
        try:
            progress.close()
        finally:
            progress.deleteLater()
            self._export_progress_dialog = None

    @Slot(int, int, str)
    def _on_export_worker_progress(self, value: int, total: int, label: str) -> None:
        progress = getattr(self, "_export_progress_dialog", None)
        if progress is None:
            return
        max_value = max(int(total), 1)
        if progress.maximum() != max_value:
            progress.setRange(0, max_value)
        progress.setValue(max(0, min(int(value), max_value)))
        if label:
            progress.setLabelText(str(label))

    @Slot(str)
    def _on_export_worker_success(self, file_path: str) -> None:
        self._close_export_progress_dialog()
        self._export_worker = None
        self.export_success(file_path)

    @Slot(str)
    def _on_export_worker_error(self, error_msg: str) -> None:
        self._close_export_progress_dialog()
        self._export_worker = None
        self.export_error(error_msg)

    def export_to_excel(self):
        from ui.main_window import legacy_module
        legacy = legacy_module()
        DataStateManager = legacy.DataStateManager
        require_ganhos_ok_or_confirm = legacy.require_ganhos_ok_or_confirm
        get_selected_or_visible_rows = legacy.get_selected_or_visible_rows
        ExportExcelWorker = legacy.ExportExcelWorker

        if not self.require_state(
            "Exportar Excel", {"db": DataStateManager.CARREGADO_VALIDADO}
        ):
            return
        # Captura apenas as linhas visíveis atualmente selecionadas
        visible_indices = self.table_obras.selectedVisibleRows()

        # Se nenhuma linha foi selecionada (ou nenhuma selecionada está visível)
        if not visible_indices:
            reply = QtWidgets.QMessageBox.question(
                self,
                "Exportar",
                "Nenhuma obra foi selecionada. Deseja exportar todas as obras visíveis?",
                QtWidgets.QMessageBox.StandardButton.Yes
                | QtWidgets.QMessageBox.StandardButton.No,
                QtWidgets.QMessageBox.StandardButton.No,
            )
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                col_cod = self.col_index("cod")
                visible_indices = [
                    QtCore.QModelIndex(self.table_obras.model().index(row, col_cod))
                    for row in range(self.table_obras.rowCount())
                    if not self.table_obras.isRowHidden(row)
                ]
            else:
                return
        if not require_ganhos_ok_or_confirm(
            self,
            get_selected_or_visible_rows(self.table_obras),
            "Exportar Excel",
        ):
            return

        idx_cod = self.col_index("cod")
        selected_ids = []
        for index in visible_indices:
            item = self.table_obras.item(index.row(), idx_cod) if idx_cod >= 0 else None
            selected_ids.append(item.text() if item else "")
        selected_ids = [str(cod).strip() for cod in selected_ids if str(cod).strip()]
        if not selected_ids:
            QtWidgets.QMessageBox.information(
                self,
                "Exportar Excel",
                "Nenhuma obra válida foi encontrada para exportação.",
            )
            return

        export_mode = self._prompt_export_columns_mode()
        if export_mode is None:
            return

        cols_db = self._get_visualizar_column_names()
        if not cols_db:
            QtWidgets.QMessageBox.warning(
                self,
                "Aviso",
                "Não foi possível obter as colunas do banco para exportação.",
            )
            return

        if export_mode == "all_db":
            cols_export = cols_db
        else:
            cols_export = self._get_visible_db_columns_from_table()
            if not cols_export:
                QtWidgets.QMessageBox.information(
                    self,
                    "Exportar Excel",
                    "Nenhuma coluna visível foi detectada. O sistema exportará todas as colunas do banco.",
                )
                cols_export = cols_db

        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Exportar para Excel", "", "Excel Files (*.xlsx)")
        if file_path:
            self._close_export_progress_dialog()
            self._export_progress_dialog = QtWidgets.QProgressDialog(
                "Preparando exportação...",
                "",
                0,
                0,
                self,
            )
            self._export_progress_dialog.setWindowTitle("Exportação em andamento")
            self._export_progress_dialog.setWindowModality(
                QtCore.Qt.WindowModality.ApplicationModal
            )
            self._export_progress_dialog.setMinimumDuration(0)
            self._export_progress_dialog.setAutoClose(False)
            self._export_progress_dialog.setAutoReset(False)
            self._export_progress_dialog.setValue(0)
            self._export_progress_dialog.show()
            QtWidgets.QApplication.processEvents(
                QtCore.QEventLoop.ProcessEventsFlag.AllEvents
            )

            worker = ExportExcelWorker(
                self.config["obras"],
                file_path,
                selected_ids,
                selected_columns=cols_export,
            )
            worker.export_progress.connect(
                self._on_export_worker_progress,
                QtCore.Qt.ConnectionType.QueuedConnection,
            )
            worker.export_success.connect(
                self._on_export_worker_success,
                QtCore.Qt.ConnectionType.QueuedConnection,
            )
            worker.export_error.connect(
                self._on_export_worker_error,
                QtCore.Qt.ConnectionType.QueuedConnection,
            )
            self._export_worker = worker
            QtCore.QThreadPool.globalInstance().start(worker)

    def _prompt_export_columns_mode(self) -> str | None:
        from ui.main_window import legacy_module
        legacy = legacy_module()
        ConfigManager = legacy.ConfigManager
        DEFAULT_EXPORT_PROFILES = legacy.DEFAULT_EXPORT_PROFILES

        msg = QtWidgets.QMessageBox(self)
        msg.setIcon(QtWidgets.QMessageBox.Icon.Question)
        msg.setWindowTitle("Exportar Excel")
        msg.setText("Quais colunas deseja exportar?")

        config = ConfigManager.load_config()
        profiles = config.get("export_profiles", DEFAULT_EXPORT_PROFILES)
        if not isinstance(profiles, list) or not profiles:
            profiles = DEFAULT_EXPORT_PROFILES

        button_map: dict[QtWidgets.QAbstractButton, str] = {}
        for profile in profiles:
            if not isinstance(profile, dict):
                continue
            label = str(profile.get("label") or profile.get("id") or "").strip()
            mode = str(profile.get("mode") or profile.get("id") or "").strip()
            if not label or not mode:
                continue
            btn = msg.addButton(label, QtWidgets.QMessageBox.ButtonRole.AcceptRole)
            button_map[btn] = mode

        btn_cancel = msg.addButton(QtWidgets.QMessageBox.StandardButton.Cancel)

        msg.exec()

        clicked = msg.clickedButton()
        if clicked == btn_cancel:
            return None
        return button_map.get(clicked)

    def _get_visible_db_columns_from_table(self) -> list[str]:
        tabela = getattr(self, "table_obras", None)
        if tabela is None or tabela.columnCount() == 0:
            return []

        cols_db = self._get_visualizar_column_names()
        if not cols_db:
            return []

        visible: list[str] = []
        max_idx = min(len(cols_db), tabela.columnCount())
        for logical_idx in range(max_idx):
            if not tabela.isColumnHidden(logical_idx):
                visible.append(cols_db[logical_idx])

        return visible

    def exportar_relatorio_criterios_excel(self) -> None:
        import pandas as pd
        from ui.main_window import legacy_module
        legacy = legacy_module()
        DataStateManager = legacy.DataStateManager
        show_user_error = legacy.show_user_error
        LOGGER = legacy.LOGGER

        if not self.require_state(
            "Exportar Relatório de Critérios", {"db": DataStateManager.CARREGADO_VALIDADO}
        ):
            return
        tabela = getattr(self, "table_obras", None)
        has_hidden = False
        if tabela is not None and tabela.rowCount() > 0:
            has_hidden = any(tabela.isRowHidden(row) for row in range(tabela.rowCount()))
        selected_rows = self._get_visualizar_scope_rows("selected")
        scope = self._prompt_relatorio_criterios_scope(bool(selected_rows), has_hidden)
        if scope is None:
            return
        cods: list[str] | None
        if scope == "selected":
            cods = self._get_visualizar_scope_cods(selected_rows)
        elif scope == "visible":
            cods = self._get_visualizar_scope_cods(self._get_visualizar_scope_rows("visible"))
        else:
            cods = None
        if cods is not None and not cods:
            QtWidgets.QMessageBox.information(
                self, "Exportar Relatório de Critérios", "Nenhuma obra encontrada no escopo."
            )
            return
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Exportar Relatório de Critérios",
            "",
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            df_projetos, df_alims = self.montar_relatorio_criterios_por_projeto(cods)
            empty_series = pd.Series([""] * len(df_projetos))
            status_series = (
                df_projetos["projeto_atende"].replace({"SIM": "ATENDE", "NÃO": "NÃO ATENDE"})
                if "projeto_atende" in df_projetos.columns
                else empty_series
            )
            df_relatorio = pd.DataFrame(
                {
                    "nome_projeto": df_projetos["nome_projeto"]
                    if "nome_projeto" in df_projetos.columns
                    else empty_series,
                    "status_projeto": status_series,
                    "alimentadores_analisados": df_projetos["alimentadores_projeto"]
                    if "alimentadores_projeto" in df_projetos.columns
                    else empty_series,
                    "alimentadores_nao_atenderam": df_projetos["alimentadores_falha"]
                    if "alimentadores_falha" in df_projetos.columns
                    else empty_series,
                    "motivos": df_projetos["motivos_falha"]
                    if "motivos_falha" in df_projetos.columns
                    else empty_series,
                    "ganhos_utilizados": df_projetos["resumo_origem_dados"]
                    if "resumo_origem_dados" in df_projetos.columns
                    else empty_series,
                    "observacoes": empty_series,
                }
            )
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df_relatorio.to_excel(writer, index=False, sheet_name="Relatorio_Alimentadores")
                df_projetos.to_excel(writer, index=False, sheet_name="Projetos")
                df_alims.to_excel(writer, index=False, sheet_name="Alimentadores")

                for sheet_name in ("Relatorio_Alimentadores", "Projetos", "Alimentadores"):
                    ws = writer.sheets.get(sheet_name)
                    if ws is None:
                        continue
                    ws.freeze_panes = "A2"
                    ws.auto_filter.ref = ws.dimensions

            QtWidgets.QMessageBox.information(
                self,
                "Sucesso",
                f"Relatório exportado para {file_path}",
            )
        except ValueError as exc:
            show_user_error(
                "Erro ao exportar relatório",
                str(exc),
                "Verifique se o banco contém todas as colunas necessárias.",
                parent=self,
            )
        except Exception as exc:
            LOGGER.exception("Erro ao exportar relatório de critérios.")
            show_user_error(
                "Erro ao exportar relatório",
                f"Erro ao exportar relatório: {exc}",
                "Verifique o caminho de destino e tente novamente.",
                parent=self,
            )

    @Slot(str)
    def export_success(self, file_path):
        from ui.main_window import legacy_module
        DataStateManager = legacy_module().DataStateManager

        QtWidgets.QMessageBox.information(
            self,
            "Sucesso",
            f"Dados exportados para {file_path}"
        )
        self._set_data_state(
            "db",
            DataStateManager.CARREGADO_VALIDADO,
            path=self.db_manager.db_path or "",
            version_token=self._compute_file_token(self.db_manager.db_path or ""),
        )

    @Slot(str)
    def export_error(self, error_msg):
        from ui.main_window import legacy_module
        legacy = legacy_module()
        LOGGER = legacy.LOGGER
        show_user_error = legacy.show_user_error
        DataStateManager = legacy.DataStateManager

        LOGGER.error("Erro na exportação: %s", error_msg)
        show_user_error(
            "Erro na exportação",
            error_msg,
            "Verifique o caminho de destino e o banco de dados.",
            parent=self,
        )
        self._set_data_state(
            "db",
            DataStateManager.CARREGADO_VALIDADO,
            path=self.db_manager.db_path or "",
            version_token=self._compute_file_token(self.db_manager.db_path or ""),
        )
