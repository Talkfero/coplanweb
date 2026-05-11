"""Mixin Resumo / Volumetria Financeira -- 12 metodos.

Extraido de ``codigo5_coplan.py::MainWindow`` na Etapa C (Tier 4).
"""
from __future__ import annotations

import datetime
import logging
import os
import subprocess
import tempfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtWidgets import QSizePolicy, QSpacerItem
from PySide6.QtCore import Qt


class ResumoVolumetriaMixin:
    """Espera estar misturado em ``MainWindow(QMainWindow, ...)``."""

    def _obter_obras_visiveis_resumo(self):
        if not self._visualizar_filtered_rows:
            return []

        indices = {
            "ano_": self.col_index("ano_"),
            "projeto_investimento": self.col_index("projeto_investimento"),
            "valor_obra": self.col_index("valor_obra"),
            "quantidade_material": self.col_index("quantidade_material"),
        }

        if any(idx < 0 for idx in indices.values()):
            return []

        obras = []
        for row_data, _atende in self._visualizar_filtered_rows:
            registro = {}
            for chave, idx in indices.items():
                registro[chave] = str(row_data[idx]).strip() if idx < len(row_data) else ""
            obras.append(registro)

        return obras

    def _init_resumo_grid(self):
        """Inicializa o grid de blocos na aba de Resumo."""
        self.scroll_resumo = QtWidgets.QScrollArea(self.tab_resumo)
        self.scroll_resumo.setWidgetResizable(True)
        self.resumo_container = QtWidgets.QWidget(self.scroll_resumo)
        self.scroll_resumo.setWidget(self.resumo_container)

        self.layout_resumo_grid = QtWidgets.QGridLayout(self.resumo_container)
        self.layout_resumo_grid.setContentsMargins(8, 8, 8, 8)
        self.layout_resumo_grid.setHorizontalSpacing(10)
        self.layout_resumo_grid.setVerticalSpacing(10)

        self._resumo_blocks: list[QtWidgets.QWidget] = []
        self._resumo_cols = 2

        raiz = QtWidgets.QVBoxLayout(self.tab_resumo)
        # === HEADER TOPO COMPACTO E ALINHADO BEGIN ===
        header_widget = QtWidgets.QWidget()
        header_layout = QtWidgets.QHBoxLayout(header_widget)
        header_layout.setContentsMargins(6, 0, 6, 0)
        header_layout.setSpacing(6)
        header_widget.setFixedHeight(48)
        header_widget.setStyleSheet(
            "QLabel { margin: 0px; padding: 0px; }"
            "QToolButton { margin: 0px; padding: 2px; }"
        )

        left_widget = QtWidgets.QWidget()
        left_layout = QtWidgets.QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(0)
        self.label_tecnico_status = QtWidgets.QLabel(
            "Dados técnicos atualizados após consolidação: N/D"
        )
        left_layout.addWidget(self.label_tecnico_status)

        right_widget = QtWidgets.QWidget()
        right_layout = QtWidgets.QHBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(6)
        self.btn_atualizar_snapshot_tecnico = QtWidgets.QPushButton(
            "Atualizar snapshot técnico dos selecionados"
        )
        self.btn_atualizar_snapshot_tecnico.setAutoDefault(False)
        self.btn_atualizar_snapshot_tecnico.clicked.connect(
            self.atualizar_snapshot_tecnico_selecionados
        )
        right_layout.addWidget(self.btn_atualizar_snapshot_tecnico)
        self.btn_exportar_resumo = QtWidgets.QPushButton("Exportar Resumo")
        self.btn_exportar_resumo.setAutoDefault(False)
        self.btn_exportar_resumo.clicked.connect(self.on_exportar_resumo)
        right_layout.addWidget(self.btn_exportar_resumo)

        header_layout.addWidget(
            left_widget, 0, QtCore.Qt.AlignmentFlag.AlignVCenter
        )
        header_layout.addItem(
            QSpacerItem(0, 0, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
        )
        header_layout.addWidget(
            right_widget, 0, QtCore.Qt.AlignmentFlag.AlignVCenter
        )

        raiz.addWidget(header_widget)
        # === HEADER TOPO COMPACTO E ALINHADO END ===
        raiz.addWidget(self.scroll_resumo)

    def add_resumo_bloco(
        self, titulo: str, widget_central: QtWidgets.QWidget | None = None
    ):
        """Cria um bloco de resumo com título e adiciona ao próximo quadrante."""
        card = QtWidgets.QGroupBox(titulo, self.resumo_container)
        card_layout = QtWidgets.QVBoxLayout(card)
        card_layout.setContentsMargins(8, 8, 8, 8)

        if widget_central is None:
            tabela = QtWidgets.QTableWidget(card)
            tabela.setColumnCount(1)
            tabela.setRowCount(0)
            tabela.setHorizontalHeaderLabels(["(sem dados)"])
            tabela.horizontalHeader().setStretchLastSection(True)
            widget_central = tabela

        card_layout.addWidget(widget_central)
        card.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Preferred,
            QtWidgets.QSizePolicy.Policy.Preferred,
        )
        card.setMinimumSize(320, 200)

        idx = len(self._resumo_blocks)
        row, col = divmod(idx, self._resumo_cols)
        self.layout_resumo_grid.addWidget(card, row, col)
        self._resumo_blocks.append(card)

    def setup_tab_resumo(self) -> None:
        # Inicializa o grid de blocos na aba de resumo
        self._init_resumo_grid()

        # Tabela exibida no bloco "Resumo de Ganhos"
        self.tabelaResumo = QtWidgets.QTableWidget(0, 4)
        headers = ["Alimentador", "Carregamento", "Tensão (MIN | MÁX)", "Cliente"]
        self.tabelaResumo.setHorizontalHeaderLabels(headers)
        header = self.tabelaResumo.horizontalHeader()
        font = header.font()
        font.setBold(True)
        header.setFont(font)
        header.setDefaultAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.tabelaResumo.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers
        )
        self.tabelaResumo.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.SingleSelection
        )
        self.tabelaResumo.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows
        )

        # Tabela exibida no bloco "Resumo de Ganhos / Projeto"
        self.tabelaResumoProjeto = QtWidgets.QTableWidget(0, 4)
        self.tabelaResumoProjeto.setHorizontalHeaderLabels(headers)
        header_proj = self.tabelaResumoProjeto.horizontalHeader()
        header_proj.setFont(font)
        header_proj.setDefaultAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.tabelaResumoProjeto.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers
        )
        self.tabelaResumoProjeto.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.SingleSelection
        )
        self.tabelaResumoProjeto.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows
        )

        # Tabela exibida no bloco "Volumetria e Financeiro"
        self.tabelaVolumetriaFinanceiro = QtWidgets.QTableWidget(0, 0)
        self.tabelaVolumetriaFinanceiro.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers
        )
        self.tabelaVolumetriaFinanceiro.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.NoSelection
        )
        self.tabelaVolumetriaFinanceiro.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectItems
        )
        header_vol = self.tabelaVolumetriaFinanceiro.horizontalHeader()
        header_vol.setDefaultAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        header_vol.setStretchLastSection(False)
        header_vol.setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeMode.Interactive
        )
        self.tabelaVolumetriaFinanceiro.verticalHeader().setVisible(False)

        # Tabela exibida no bloco "Regional / Subestação / Tensão / Quantidade"
        self.tabelaResumoRegionalSe = QtWidgets.QTableWidget(0, 8)
        self.tabelaResumoRegionalSe.setHorizontalHeaderLabels(
            [
                "PI",
                "Regional",
                "Alimentador",
                "Subestação",
                "Tensão",
                "Quantidade",
                "Coordenadas para",
                "Observação",
            ]
        )
        self.tabelaResumoRegionalSe.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers
        )
        self.tabelaResumoRegionalSe.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.NoSelection
        )
        header_regional = self.tabelaResumoRegionalSe.horizontalHeader()
        header_regional.setDefaultAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.tabelaResumoRegionalSe.verticalHeader().setVisible(False)
        self.tabelaResumoRegionalSe.setContextMenuPolicy(
            QtCore.Qt.ContextMenuPolicy.CustomContextMenu
        )
        self.tabelaResumoRegionalSe.customContextMenuRequested.connect(
            self.mostrar_menu_export_resumo_regional
        )

        # Adiciona blocos iniciais na grade
        self.add_resumo_bloco("Resumo de Ganhos", self.tabelaResumo)
        self.add_resumo_bloco("Resumo de Ganhos / Projeto", self.tabelaResumoProjeto)
        self.add_resumo_bloco(
            "Volumetria e Financeiro", self.tabelaVolumetriaFinanceiro
        )
        self.add_resumo_bloco(
            "Regional / Subestação / Tensão / Quantidade",
            self.tabelaResumoRegionalSe,
        )

    def on_exportar_resumo(self) -> None:
        from ui.main_window import legacy_module
        _l = legacy_module()
        open_file = _l.open_file
        require_tecnico_clean_or_confirm = _l.require_tecnico_clean_or_confirm
        if not require_tecnico_clean_or_confirm(
            self, self.db_manager, "Exportar Resumo"
        ):
            return
        if not self.require_export_sources("Exportar resumo"):  # [RB-4]
            return
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Exportar Resumo",
            "resumo.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return

        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            tabelas = [
                (getattr(self, "tabelaResumo", None), "Resumo de Ganhos"),
                (
                    getattr(self, "tabelaResumoProjeto", None),
                    "Resumo de Ganhos - Projeto",
                ),
                (
                    getattr(self, "tabelaVolumetriaFinanceiro", None),
                    "Volumetria e Financeiro",
                ),
                (
                    getattr(self, "tabelaResumoRegionalSe", None),
                    "Regional - SE",
                ),
            ]

            workbook = Workbook()
            sheet = workbook.active
            if sheet is None:
                sheet = workbook.create_sheet()
            first_sheet = True

            for tabela, titulo in tabelas:
                if tabela is None:
                    continue
                if first_sheet:
                    ws = sheet
                    ws.title = titulo
                    first_sheet = False
                else:
                    ws = workbook.create_sheet(title=titulo)
                self._export_qtable_to_sheet(tabela, ws)

            workbook.save(file_path)
            open_file(file_path)
        except Exception as exc:  # pragma: no cover - interação com disco/GUI
            QtWidgets.QMessageBox.critical(
                self, "Erro", f"Erro ao exportar o resumo: {exc}"
            )

    def popular_volumetria_financeiro(self, obras_visiveis):
        # Logica pura extraida para core/services/resumo_service (Etapa B.3).
        # A UI continua responsavel por popular o QTableWidget.
        from core.services.resumo_service import montar_volumetria_financeiro

        tabela = getattr(self, "tabelaVolumetriaFinanceiro", None)
        if tabela is None:
            return

        tabela.clear()
        tabela.setRowCount(0)
        tabela.setColumnCount(0)

        out = montar_volumetria_financeiro(obras_visiveis or [])
        if not out.linhas:
            return

        tabela.setRowCount(len(out.linhas))
        tabela.setColumnCount(len(out.cabecalhos))
        tabela.setHorizontalHeaderLabels(out.cabecalhos)

        for linha_idx, linha in enumerate(out.linhas):
            # Coluna 0: PI (alinhamento default)
            item_pi = QtWidgets.QTableWidgetItem(linha[0])
            item_pi.setFlags(item_pi.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tabela.setItem(linha_idx, 0, item_pi)
            # Demais colunas: alinhadas a direita
            for col_idx in range(1, len(linha)):
                item = QtWidgets.QTableWidgetItem(linha[col_idx])
                item.setTextAlignment(
                    QtCore.Qt.AlignmentFlag.AlignRight
                    | QtCore.Qt.AlignmentFlag.AlignVCenter
                )
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                tabela.setItem(linha_idx, col_idx, item)

        tabela.resizeColumnsToContents()

    def popular_quadro_resumo_from_ganhos_depois(
        self,
        alimentador_principal: str,
        alimentadores_beneficiados: str,
    ) -> None:
        from ui.main_window import legacy_module
        _l = legacy_module()
        DEFAULT_CRITERIOS = _l.DEFAULT_CRITERIOS
        """Preenche tabelaResumo a partir de ``ganhos_totais_depois``.

        Logica pura extraida para core/services/resumo_service (Etapa B.3).
        A UI continua responsavel por QTableWidget/QColor/QBrush.
        """
        from core.services.resumo_service import montar_quadro_resumo_from_ganhos

        criterios = self.config.get("criterios_planejamento", DEFAULT_CRITERIOS)
        ganhos_text = self.field_ganhos_totais_depois.text().strip()
        linhas_resumo = montar_quadro_resumo_from_ganhos(
            alimentador_principal=alimentador_principal,
            alimentadores_beneficiados=alimentadores_beneficiados,
            ganhos_totais_depois=ganhos_text,
            criterios=criterios,
        )

        self.tabelaResumo.setRowCount(0)
        for linha in linhas_resumo:
            row = self.tabelaResumo.rowCount()
            self.tabelaResumo.insertRow(row)

            item_alim = QtWidgets.QTableWidgetItem(linha.alimentador)
            item_alim.setFlags(item_alim.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.tabelaResumo.setItem(row, 0, item_alim)

            self._preencher_celula_criterio(self.tabelaResumo, row, 1, linha.carregamento)
            self._preencher_celula_criterio(self.tabelaResumo, row, 2, linha.tensao)

            item_clientes = QtWidgets.QTableWidgetItem(linha.clientes_text)
            item_clientes.setFlags(item_clientes.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.tabelaResumo.setItem(row, 3, item_clientes)

        self.tabelaResumo.resizeColumnsToContents()

    def popular_resumo_ganhos_projeto(self, nome_projeto: str) -> None:
        from ui.main_window import legacy_module
        _l = legacy_module()
        DEFAULT_CRITERIOS = _l.DEFAULT_CRITERIOS
        """Consolida ganhos do projeto e popula tabelaResumoProjeto.

        Logica pura extraida para core/services/resumo_service (Etapa B.3).
        A UI continua responsavel por ensure_db, fetch_by_project e
        QTableWidget/QColor.
        """
        from core.services.resumo_service import montar_resumo_ganhos_projeto

        self.tabelaResumoProjeto.setRowCount(0)
        if not nome_projeto or not self.db_manager:
            return
        if not self.ensure_db_connected():
            return
        rows = self.db_manager.fetch_by_project(nome_projeto)
        if not rows:
            return

        cols = self.db_manager.get_column_names()
        criterios = self.config.get("criterios_planejamento", DEFAULT_CRITERIOS)
        linhas_resumo = montar_resumo_ganhos_projeto(
            rows=rows, cols=cols, criterios=criterios,
        )

        for linha in linhas_resumo:
            row = self.tabelaResumoProjeto.rowCount()
            self.tabelaResumoProjeto.insertRow(row)

            item_alim = QtWidgets.QTableWidgetItem(linha.alimentador)
            item_alim.setFlags(item_alim.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.tabelaResumoProjeto.setItem(row, 0, item_alim)

            self._preencher_celula_criterio(self.tabelaResumoProjeto, row, 1, linha.carregamento)
            self._preencher_celula_criterio(self.tabelaResumoProjeto, row, 2, linha.tensao)

            item_clientes = QtWidgets.QTableWidgetItem(linha.clientes_text)
            item_clientes.setFlags(item_clientes.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.tabelaResumoProjeto.setItem(row, 3, item_clientes)

        self.tabelaResumoProjeto.resizeColumnsToContents()

    def gerar_detalhamento(self):
        from ui.main_window import legacy_module
        _l = legacy_module()
        DataStateManager = _l.DataStateManager
        open_file = _l.open_file
        _selected_ids_from_view = _l._selected_ids_from_view
        get_row_value_by_key = _l.get_row_value_by_key
        sort_processing_rows = _l.sort_processing_rows
        if not self.require_state(
            "Gerar VT/Resumo", {"db": DataStateManager.CARREGADO_VALIDADO}
        ):  # [RB-RESTORE-OLD]
            return
        targets, include_aprovadas, _ignoradas_aprovadas = self._gate_aprovadas_for_action(
            "Despacho", self.table_obras
        )
        if not targets:
            return
        obras = []
        if not self.ensure_db_connected():
            return

        selected_ids = _selected_ids_from_view(self.table_obras)
        rows_sorted = sort_processing_rows(targets, selected_ids, include_aprovadas)

        ids = []
        for row in rows_sorted:
            cod = get_row_value_by_key(row, "cod")
            if cod:
                ids.append(cod)
        ids = list(dict.fromkeys(ids))
        if ids:
            obras = self.db_manager.fetch_by_cods(ids)

        if obras and not include_aprovadas:
            cols = self.db_manager.get_column_names()
            if "obra_aprovada" in cols:
                idx_aprovada = cols.index("obra_aprovada")
                obras = [
                    obra
                    for obra in obras
                    if str(obra[idx_aprovada] or "").strip().upper() != "SIM"
                ]

        if not obras:
            QtWidgets.QMessageBox.information(
                self, "Despacho", "Nenhuma obra encontrada para a seleção."
            )
            return

        # ------------------ DESPACHO TXT (COMO JÁ ERA) ------------------
        despacho = self.calc_manager.calcular_despacho_vt(obras)
        despacho_success = False
        despacho_ref = ""
        try:
            with tempfile.NamedTemporaryFile(
                mode="w", delete=False, suffix=".txt", encoding="utf-8"
            ) as tmp:
                tmp.write(despacho)
                temp_path = tmp.name
            os.startfile(temp_path)
            despacho_success = True
            despacho_ref = os.path.basename(temp_path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "Erro", f"Erro ao gerar arquivo de despacho: {e}"
            )

        # ------------------ QUADRINHO RESUMO EM EXCEL ------------------
        try:
            resumo_df = self._montar_resumo_detalhamento_excel(obras)
            if resumo_df is not None and not resumo_df.empty:

                # 🔵 AQUI: garantir ALIMENTADOR em maiúsculo no Excel
                if "Alimentador" in resumo_df.columns:
                    resumo_df["Alimentador"] = (
                        resumo_df["Alimentador"].astype(str).str.upper()
                    )

                with tempfile.NamedTemporaryFile(
                    delete=False, suffix=".xlsx"
                ) as tmp_xlsx:
                    tmp_xlsx_path = tmp_xlsx.name

                with pd.ExcelWriter(tmp_xlsx_path) as writer:
                    resumo_df.to_excel(
                        writer, index=False, sheet_name="Resumo"
                    )

                # abre o arquivo XLSX temporário
                open_file(tmp_xlsx_path)
        except Exception as e:
            logging.error(
                "Erro ao gerar resumo em Excel no Detalhamento de obras: %s", e
            )
        if despacho_success and ids:
            now_iso = datetime.datetime.now().isoformat(timespec="seconds")
            erros = []
            for cod in ids:
                try:
                    self.db_manager.update_obra(
                        {
                            "despacho_status": "DESPACHADA",
                            "despacho_em": now_iso,
                            "despacho_ref": despacho_ref,
                        },
                        cod,
                        skip_blank=True,
                    )
                except Exception as exc:
                    erros.append(f"{cod}: {exc}")
            if erros:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Aviso",
                    "Despacho gerado, mas houve erro ao atualizar status:\n"
                    + "\n".join(erros),
                )

    def _parse_ganhos_totais_resumo(self, ganhos_str):
        """Wrapper -- delega ao core/services/resumo_service (parser permissivo V3)."""
        from core.services.resumo_service import parse_ganhos_demand_max as _impl
        return _impl(ganhos_str)

    def _formatar_decimal_resumo(self, valor):
        """Wrapper -- delega ao core/services/resumo_service."""
        from core.services.resumo_service import formatar_decimal_resumo as _impl
        return _impl(valor)

    def _montar_resumo_detalhamento_excel(self, obras):
        """Wrapper -- delega ao core/services/resumo_service.

        Logica pura extraida na Etapa B.3. Devolve o mesmo
        ``pd.DataFrame | None`` do legado.
        """
        from core.services.resumo_service import montar_resumo_detalhamento as _impl
        return _impl(obras, self.db_manager.get_column_names())
