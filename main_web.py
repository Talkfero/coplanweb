# -*- coding: utf-8 -*-
"""COPLAN main_web -- entry point alternativo (web/pywebview).

Carrega o mock visual ``Coplan UI.html`` em uma janela pywebview e (nas
proximas etapas da Section 6 do HANDOFF) liga cada aba ao banco real e
aos mixins ja existentes em ``codigo5_coplan.py`` / ``core/`` / ``ui/``.

Restricoes (impostas pelo usuario):
  * ESTE e' o UNICO arquivo novo. Nao reescrever nem editar nada que ja
    existia (codigo5_coplan.py, ui/*, core/*, Coplan UI.html, etc.).
  * O HTML do mock e' lido como string em memoria; modificacoes (injecao
    de JS bridge) acontecem apenas no buffer Python e sao servidas para o
    pywebview -- o arquivo em disco nunca e' tocado.
  * Quando precisar de dados ou logica de negocio, importar dos modulos
    existentes (DatabaseManager, ConfigManager, services, etc.) sem
    duplicar.

Uso:
    pip install pywebview pandas openpyxl
    python main_web.py

Mapeamento Section 6 -> blocos neste arquivo:
    Passo 1 (Tokens e tema)  -> bootstrap; CSS vem do proprio HTML.
    Passo 2 (Shell)          -> CoplanApi.get_app_state() + JS injetado.
    Passo 3 (Visualizar)     -> CoplanApi.list_obras / stats / search.
    Passo 4 (Cadastro)       -> CoplanApi.save_obra / get_obra / cod_pep.
    Passo 5 (Ganhos)         -> CoplanApi.list_ganhos / read_ganhos.
    Passo 6 (Resumo)         -> CoplanApi.resumo_volumetria / regional.
    Passo 7 (Configuracoes)  -> CoplanApi.get_config / save_config.
"""

from __future__ import annotations

import getpass
import hashlib
import os
import re
import sys
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

HERE = Path(__file__).resolve().parent
FRONTEND_DIR = HERE / "frontend"
HTML_FILE = FRONTEND_DIR / "index.html"
BRIDGE_JS_FILE = FRONTEND_DIR / "js" / "coplan_bridge.js"

# Versao do main_web (independente da versao do desktop). Bumpa quando
# uma feature web visivel pro usuario muda.
APP_VERSION = "web-0.2.0"

# Garante que codigo5_coplan e core/ sejam importaveis quando rodando
# tanto via "python main_web.py" quanto via PyInstaller frozen.
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))


# ----------------------------------------------------------------------
# Bloco 5 (Auditoria #44): Sistema de progress + cancel para operacoes
# longas. Singleton por processo - uma operacao de cada vez. JS faz
# polling de progress_state() e pode cancelar via progress_cancel().
# Op longa roda em threading.Thread (daemon) para nao bloquear o bridge.
# ----------------------------------------------------------------------
_OP_LOCK = threading.Lock()
_OP_STATE: dict[str, Any] = {
    "id": "",
    "label": "",
    "processed": 0,
    "total": 0,
    "started_at": 0.0,
    "finished": True,
    "cancel_requested": False,
    "result": None,
    "error": "",
}


def _op_reset(label: str) -> str:
    """Reinicia _OP_STATE para uma nova operacao. Retorna op_id novo."""
    import time as _t
    import uuid as _u
    op_id = _u.uuid4().hex[:12]
    with _OP_LOCK:
        _OP_STATE.update({
            "id":               op_id,
            "label":            str(label or ""),
            "processed":        0,
            "total":            0,
            "started_at":       _t.time(),
            "finished":         False,
            "cancel_requested": False,
            "result":           None,
            "error":            "",
        })
    return op_id


def _op_set_progress(processed: int, total: int, label: str = "") -> None:
    """Atualiza progresso. Chamado pela worker thread."""
    with _OP_LOCK:
        _OP_STATE["processed"] = int(processed)
        _OP_STATE["total"] = int(total)
        if label:
            _OP_STATE["label"] = str(label)


def _op_check_cancel() -> bool:
    """Le flag de cancel (set por progress_cancel())."""
    with _OP_LOCK:
        return bool(_OP_STATE.get("cancel_requested"))


def _op_finish(result: Any = None, error: str = "") -> None:
    """Marca a operacao como finalizada (pela worker thread)."""
    with _OP_LOCK:
        _OP_STATE["finished"] = True
        _OP_STATE["result"] = result
        _OP_STATE["error"] = str(error or "")


def _op_snapshot() -> dict[str, Any]:
    """Devolve copia thread-safe do estado atual (para o JS pollar)."""
    with _OP_LOCK:
        return dict(_OP_STATE)


class CoplanApi:
    """API exposta ao JS via ``window.pywebview.api.<metodo>``.

    Os managers (DatabaseManager, SupportFileManager, CalculationManager)
    sao importados sob demanda na primeira chamada que precise deles, para
    nao pagar o custo de inicializar Qt/SQLite no boot do main_web.
    """

    def __init__(self) -> None:
        self._db_manager: Any = None
        self._support_manager: Any = None
        self._calc_manager: Any = None
        self._config: dict[str, Any] | None = None
        # Locks para serializar inicializacao e conexao quando o JS
        # dispara varias chamadas API em paralelo (caso contrario:
        # race em add_column_if_missing -> "duplicate column name").
        self._managers_lock = threading.Lock()
        self._connect_lock = threading.Lock()
        # Cache do ultimo erro de conexao para o JS oferecer correcao
        # de path em Configuracoes sem repetir tentativas barulhentas.
        self._last_connect_error: str = ""
        self._last_connect_path: str = ""
        # Conjunto de paths ja inicializados (full migration + cache).
        # DatabaseManager fecha a conexao no final de cada `_with_connection`,
        # mas o `data_access_layer` continua valido para queries. Sem este
        # cache, cada API call do JS chamava `db.connect()` de novo (WAL +
        # migration + audit_columns + weekly_backup), ate 14x por boot.
        self._connected_paths: set[str] = set()
        # Cache do ultimo dict retornado por SupportFileManager.load_support_file
        # (alimentadores, dados_alimentador, projetos_investimento, etc.)
        # Populado por _load_apoio_into_manager.
        self._apoio_cache: dict[str, Any] = {}
        self._apoio_path_loaded: str = ""
        # Estado de fontes (RB-1.1 / RB-5 do desktop): rastreia se db,
        # apoio, ganhos e tecnico_txt estao VALIDADO/INVALIDADO/NAO_CARREGADO,
        # com timestamps + version_token. Usado por require_state pra gateara
        # acoes que dependem de fonte carregada e por update_reliability_labels
        # pra pintar chips no header. Lazy-init: instanciado na primeira vez
        # que algum hook ou API toca pra evitar custo de import no boot.
        self._data_state: Any = None
        # Marca do ultimo refresh do db (data_modificacao + usuario), pra
        # detectar updates externos (outro usuario gravou) entre 2 leituras.
        self._last_db_refresh_data: str = ""
        self._last_db_refresh_user: str = ""
        self._last_db_modification_warned: str = ""

    def _reload_config(self) -> None:
        """Recarrega self._config do disco. Chamado automaticamente
        sempre que algum save_* invalida o cache (self._config = None)."""
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            self._config = ConfigManager.load_config()
        except Exception as exc:  # noqa: BLE001
            print(f"[main_web] reload config falhou: {exc}", file=sys.stderr)
            if self._config is None:
                self._config = {}

    def _get_data_state(self) -> Any:
        """Lazy-init do DataStateManager. Importa do runtime/config.py."""
        if self._data_state is None:
            try:
                from runtime.config import DataStateManager  # noqa: PLC0415
                self._data_state = DataStateManager()
            except Exception as exc:  # noqa: BLE001
                print(f"[main_web] DataStateManager indisponivel: {exc}",
                      file=sys.stderr)
                return None
        return self._data_state

    def _data_state_set(
        self,
        source: str,
        state: str,
        path: str = "",
        error: str = "",
        version_token: str = "",
    ) -> None:
        """Atalho para chamar update_source com tolerancia a erros.
        Hooks chamam isto sem se preocupar se o manager esta disponivel."""
        ds = self._get_data_state()
        if ds is None:
            return
        try:
            ds.update_source(
                source,
                state,
                path=path or None,
                error=error or None,
                version_token=version_token or None,
            )
        except Exception as exc:  # noqa: BLE001
            print(f"[main_web] data_state.update_source falhou: {exc}",
                  file=sys.stderr)

    # ------------------------------------------------------------------
    # _write_op_log: persiste um TXT em <HERE>/logs/<op>_<ts>.txt com os
    # detalhes de uma operacao que NAO foi 100% bem-sucedida. Chamado
    # antes de retornar o result dict de operacoes (atualizar, export,
    # delete, etc.). Garante que o usuario sempre tem o arquivo mesmo
    # se nao clicar em "Salvar TXT" no modal de detalhes.
    #
    # Devolve o path ('' em caso de falha de escrita). NUNCA levanta.
    # ------------------------------------------------------------------
    def _write_op_log(
        self, op: str, result: dict[str, Any],
        meta: dict[str, Any] | None = None,
    ) -> str:
        try:
            logs_dir = HERE / "logs"
            try:
                logs_dir.mkdir(exist_ok=True)
            except Exception:  # noqa: BLE001
                pass
            ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            op_slug = (op or "log").lower()
            import re as _re_op
            op_slug = _re_op.sub(r"[^a-z0-9_-]+", "_", op_slug)
            fname = f"{op_slug}_{ts}.txt"
            path = logs_dir / fname
            lines: list[str] = []
            sep = "=" * 60
            lines.append(sep)
            lines.append(f"Operacao: {op}")
            lines.append(f"Data/hora: {datetime.now().isoformat()}")
            if meta:
                for k, v in meta.items():
                    lines.append(f"{k}: {v}")
            lines.append(sep)
            lines.append("")
            # Resumo dos contadores comuns.
            counters = []
            for k in ("ok", "total", "processadas_ok", "atualizadas",
                     "falhas_total", "deleted", "imported", "merged",
                     "skipped", "count", "blocked", "cancelled"):
                if k in result:
                    counters.append(f"{k}={result.get(k)}")
            if counters:
                lines.append("Contadores: " + " | ".join(counters))
                lines.append("")
            err = str(result.get("error") or "").strip()
            if err:
                lines.append("-- Erro --")
                lines.append(err)
                lines.append("")
            errors = result.get("errors") or []
            if errors:
                lines.append(f"-- Erros ({len(errors)}) --")
                for e in errors:
                    lines.append(str(e))
                lines.append("")
            falhas = result.get("falhas") or []
            if falhas:
                lines.append(
                    f"-- Falhas ({result.get('falhas_total', len(falhas))}) --"
                )
                for f in falhas:
                    lines.append(str(f))
                lines.append("")
            chaves = result.get("chaves_inexistentes") or []
            if chaves:
                lines.append(f"-- Chaves inexistentes ({len(chaves)}) --")
                for c in chaves:
                    lines.append(str(c))
                lines.append("")
            preservadas_msgs = result.get("preservadas_msgs") or []
            if preservadas_msgs:
                lines.append(
                    f"-- Valores preservados ({len(preservadas_msgs)}) --"
                )
                lines.append(
                    "Obras com calculo parcial (chave extra ausente ou "
                    "sem valor) que JA TINHAM valor no banco: o valor "
                    "antigo foi mantido."
                )
                for m in preservadas_msgs:
                    lines.append(str(m))
                lines.append("")
            diag = result.get("diagnostico") or []
            if diag:
                lines.append(
                    f"-- Diagnostico por obra falha ({len(diag)}) --"
                )
                lines.append(
                    "Listing input values + chave tentada para CADA obra "
                    "que falhou. Use isto para comparar com o calculo "
                    "feito na aba Cadastro (botao Calcular)."
                )
                for d in diag:
                    lines.append(str(d))
                lines.append("")
            diag_all_l = result.get("diagnostico_todas") or []
            if diag_all_l:
                lines.append(
                    f"-- Breakdown por obra ({len(diag_all_l)}) --"
                )
                lines.append(
                    "Para CADA obra processada (sucesso ou falha): "
                    "inputs, chave montada, extras resolvidos, valor "
                    "calculado. Compare 'extras_resolvidos' com o que o "
                    "Cadastro/Calcular usou para detectar key mismatch "
                    "no last_pi_extra_map."
                )
                for d in diag_all_l:
                    lines.append(str(d))
                lines.append("")
            extra_map_snap = result.get("last_pi_extra_map") or {}
            if extra_map_snap:
                lines.append(
                    f"-- Snapshot last_pi_extra_map ({len(extra_map_snap)} chaves) --"
                )
                lines.append(
                    "Conteudo de cfg['last_pi_extra_map'] no momento "
                    "do Atualizar. Compare a chave deste PI com o "
                    "pi_base que aparece no Breakdown acima -- se nao "
                    "for IDENTICA (case-sensitive, acentuacao), o "
                    "lookup falha e os extras manuais nao sao aplicados."
                )
                for k in sorted(extra_map_snap.keys()):
                    lines.append(f"  {k!r} -> {extra_map_snap[k]!r}")
                lines.append("")
            duplicadas = result.get("duplicadas") or []
            if duplicadas:
                lines.append(f"-- Duplicadas ({len(duplicadas)}) --")
                for d in duplicadas:
                    if isinstance(d, dict):
                        lines.append(
                            f"linha {d.get('linha', '?')} - "
                            f"COD excel={d.get('cod_excel', '?')} / "
                            f"dup COD={d.get('dup_cod', '?')}"
                        )
                    else:
                        lines.append(str(d))
                lines.append("")
            missing = result.get("missing_columns") or []
            if missing:
                lines.append(f"-- Colunas faltantes ({len(missing)}) --")
                for c in missing:
                    lines.append(str(c))
                lines.append("")
            with open(path, "w", encoding="utf-8", newline="") as fh:
                fh.write("\n".join(lines))
            print(f"[main_web] log salvo: {path}", file=sys.stderr)
            return str(path)
        except Exception as exc:  # noqa: BLE001
            print(f"[main_web] _write_op_log falhou: {exc}",
                  file=sys.stderr)
            return ""

    @staticmethod
    def _result_has_errors(result: dict[str, Any]) -> bool:
        """True se o result indica qualquer coisa diferente de sucesso
        pleno: !ok, falhas_total>0, chaves_inexistentes nao vazio,
        errors nao vazio, duplicadas nao vazio, cancelled."""
        if not result:
            return True
        if not result.get("ok"):
            return True
        if result.get("cancelled"):
            return True
        if int(result.get("falhas_total") or 0) > 0:
            return True
        if (result.get("chaves_inexistentes") or []):
            return True
        if (result.get("errors") or []):
            return True
        if (result.get("duplicadas") or []):
            return True
        return False

    # ------------------------------------------------------------------
    # Paridade desktop: mensagem amigavel quando o banco esta ocupado
    # por outro usuario / processo. Substitui str(exc) cru por uma
    # mensagem com user/maquina/desde lida do .lock ao lado do .db.
    # ------------------------------------------------------------------
    def _friendly_busy_error(self, exc: Exception) -> str | None:
        """Se ``exc`` for busy/locked, devolve build_database_busy_message.
        Caso contrario, devolve None (chamador propaga str(exc)).
        """
        try:
            from runtime.database import (  # noqa: PLC0415
                DatabaseBusyError, DatabaseLockedError,
                build_database_busy_message, get_lock_info_path,
                is_sqlite_busy_error, read_lock_info,
            )
        except Exception:  # noqa: BLE001
            return None
        is_busy = isinstance(exc, (DatabaseBusyError, DatabaseLockedError))
        if not is_busy:
            try:
                import sqlite3 as _sq  # noqa: PLC0415
                if isinstance(exc, _sq.OperationalError) and \
                        is_sqlite_busy_error(exc):
                    is_busy = True
            except Exception:  # noqa: BLE001
                pass
        if not is_busy:
            return None
        lock_info: dict[str, Any] | None = None
        try:
            db_path = ""
            if self._db_manager is not None:
                db_path = str(
                    getattr(self._db_manager, "db_path", "") or "")
            if db_path:
                lock_info = read_lock_info(get_lock_info_path(db_path))
        except Exception:  # noqa: BLE001
            lock_info = None
        try:
            return build_database_busy_message(lock_info)
        except Exception:  # noqa: BLE001
            return None

    # ------------------------------------------------------------------
    # Paridade desktop validar_campos_obrigatorios
    # (ui/main_window/cadastro_mixin.py:580):
    # campos obrigatorios no save de uma obra. JS ja valida no form,
    # mas duplicamos no backend como defense-in-depth (payload pode
    # vir incompleto se o front bypassar a validacao).
    # ------------------------------------------------------------------
    _CAMPOS_OBRIGATORIOS_SAVE = (
        # (coluna_db, label_amigavel)
        ("ano_",                    "Ano"),
        ("projeto_investimento",    "Projeto de Investimentos"),
        ("alimentador_principal",   "Alimentador Obra"),
        ("quantidade_material",     "Quantidade"),
        ("coordenada_fim",          "Coordenadas Para"),
        ("tipo_pacote",             "Pacote"),
        ("caracteristicas_material","Caracteristicas"),
        ("manobra",                 "Manobra"),
    )

    def _validar_campos_obrigatorios(
        self, cleaned: dict[str, Any],
    ) -> list[str]:
        """Retorna labels dos campos obrigatorios vazios. Vazio => OK."""
        faltam: list[str] = []
        for col, label in self._CAMPOS_OBRIGATORIOS_SAVE:
            val = cleaned.get(col)
            if val is None or str(val).strip() == "":
                faltam.append(label)
        # Para PI base = DISTRIBUICAO* o nome_projeto e' obrigatorio
        # (paridade: ver cadastro_mixin.py:602-606).
        try:
            from runtime.text_utils import normalize_key  # noqa: PLC0415
            from codigo5_coplan import get_pi_base  # noqa: PLC0415
        except Exception:  # noqa: BLE001
            normalize_key = lambda s: str(s or "").strip().upper()  # noqa: E731
            get_pi_base = lambda pi, prompt_user=False: ""  # noqa: E731
        pi = str(cleaned.get("projeto_investimento") or "")
        try:
            pi_base = get_pi_base(pi, prompt_user=False) or ""
        except Exception:  # noqa: BLE001
            pi_base = ""
        distrib = {"DISTRIBUICAO", "DISTRIBUICAO LD 34,5 KV"}
        if normalize_key(pi) in distrib or normalize_key(pi_base) in distrib:
            if not str(cleaned.get("nome_projeto") or "").strip():
                faltam.append("Projeto")
        return faltam

    def _ensure_managers(self) -> None:
        # Cache de config invalidado por save_config_empresa / save_criterios
        # / etc. Sem este reload, todo cfg = self._config or {} cai em {}
        # apos a primeira invalidacao -- causando "banco nao configurado"
        # mesmo com config.json correto no disco.
        if self._config is None:
            self._reload_config()
        if self._db_manager is not None:
            return
        # So uma thread inicializa os managers; as outras esperam
        # e veem self._db_manager preenchido na re-checagem.
        with self._managers_lock:
            if self._db_manager is not None:
                return
            try:
                from codigo5_coplan import (  # type: ignore[import-not-found]
                    DatabaseManager,
                    SupportFileManager,
                    CalculationManager,
                    ConfigManager,
                )
                self._db_manager = DatabaseManager()
                self._support_manager = SupportFileManager()
                self._calc_manager = CalculationManager(
                    self._support_manager, prompt_pi_base=False
                )
                if self._config is None:
                    self._config = ConfigManager.load_config()
                # Auto-connect do banco no boot: se config['obras']
                # aponta para arquivo valido, conecta ja para a pill
                # ficar verde e bridges nao precisarem reconectar.
                try:
                    db_path_boot = str(
                        (self._config or {}).get("obras") or "").strip()
                    if db_path_boot and os.path.isfile(db_path_boot):
                        self._ensure_db_connected()
                except Exception as exc:  # noqa: BLE001
                    print(f"[main_web] auto-connect db falhou: {exc}",
                          file=sys.stderr)
                # Auto-load do apoio: DB-backed only. Hidrata
                # _apoio_cache lendo as tabelas apoio_* do banco.
                # Nao toca xlsx (use 'Atualizar apoio' nas Configuracoes
                # para reimportar).
                try:
                    self._load_apoio_into_manager("")
                except Exception as exc:  # noqa: BLE001
                    print(f"[main_web] auto-load apoio (db) falhou: {exc}",
                          file=sys.stderr)
                # [FIX] Auto-validate dos 3 .TXT se caminho_pasta_ganhos
                # esta no config (paridade com desktop: ao iniciar com
                # pasta conhecida, ja marca tecnico_txt = CARREGADO_VALIDADO
                # se FlowMT/Topologia/Confiabilidade existem). Sem isso,
                # a pill 'Tecnico' do header fica eternamente NAO_CARREGADO
                # ate o user clicar em Selecionar Pasta na aba Ganhos.
                ganhos_path = (self._config or {}).get("caminho_pasta_ganhos") or ""
                if ganhos_path and os.path.isdir(str(ganhos_path)):
                    try:
                        self.validate_tecnico_files(str(ganhos_path))
                        # Tambem marca ganhos como CARREGADO_VALIDADO para
                        # destravar os botoes da aba Ganhos (G050).
                        self._data_state_set(
                            "ganhos", "CARREGADO_VALIDADO",
                            path=str(ganhos_path))
                    except Exception as exc:  # noqa: BLE001
                        print(f"[main_web] auto-validate tecnico falhou: {exc}",
                              file=sys.stderr)
            except Exception as exc:  # noqa: BLE001
                # Nao deixa exception subir pro pywebview (apaga o app).
                # Marca como erro e devolve managers possivelmente parciais.
                print(f"[main_web] _ensure_managers falhou: {exc}",
                      file=sys.stderr)
                if self._config is None:
                    self._config = {}

    # ------------------------------------------------------------------
    # Passo 1 (Tokens e tema): nenhum API necessario -- o CSS do
    # Coplan UI.html ja carrega tudo (Inter, JetBrains Mono via Google
    # Fonts; oklch tokens nas variaveis :root). Mantemos um ping de
    # health check para validar o canal Python<->JS.
    # ------------------------------------------------------------------
    def ping(self) -> dict[str, Any]:
        return {"ok": True, "msg": "coplan-bridge-ativo"}

    # ------------------------------------------------------------------
    # Estado de Fontes (RB-1.1 / RB-5 do desktop, EstadoFontesMixin):
    # rastreia se db, apoio, ganhos, tecnico_txt estao
    # NAO_CARREGADO / CARREGADO_VALIDADO / INVALIDADO. APIs publicas
    # consumidas pelo header (chips de confiabilidade) e pelos botoes
    # de export/relatorio (gating com "Ir para X").
    # ------------------------------------------------------------------
    def data_state_get(self) -> dict[str, Any]:
        """Devolve o estado das 4 fontes (db, apoio, ganhos, tecnico_txt)
        + label/timestamp formatados para o header. JS usa isto para
        pintar os chips depois de cada API que muda fonte."""
        # Lazy-disparo do auto-connect: _ensure_managers contem o
        # auto-connect do banco a partir de config['obras'] no boot.
        # Sem isto, na 1a chamada vinda do JS (antes de list_obras),
        # o DataStateManager devolve NAO_CARREGADO para 'db' mesmo com
        # o banco corretamente configurado, fazendo o tooltip do chip
        # dizer "Status: Nao carregado" enquanto o nome do arquivo
        # aparece no chip (vem de get_app_state, que le config direto).
        try:
            self._ensure_managers()
        except Exception:  # noqa: BLE001
            pass
        ds = self._get_data_state()
        if ds is None:
            return {"ok": False, "sources": {}, "error": "manager indisponivel"}
        out: dict[str, Any] = {}
        for source, info in ds.sources.items():
            ts = info.validated_at or info.loaded_at
            ts_str = ts.strftime("%d/%m %H:%M") if ts else ""
            out[source] = {
                "state": info.state,
                "path": info.path,
                "validated_at": ts_str,
                "error_last": info.error_last,
                "version_token": info.version_token,
            }
        # tecnico_dirty count: vem do banco quando conectado
        dirty = 0
        try:
            if self._db_manager is not None and self._connected_paths:
                dirty = int(self._db_manager.count_tecnico_dirty() or 0)
        except Exception:  # noqa: BLE001
            dirty = 0
        out["tecnico_dirty_count"] = dirty
        return {"ok": True, "sources": out}

    def data_state_require(
        self, action_name: Any = "", required: Any = None
    ) -> dict[str, Any]:
        """Verifica pre-requisitos. ``required`` e dict como
        ``{"db": "CARREGADO_VALIDADO", "ganhos": "CARREGADO_VALIDADO"}``.
        Retorna ``{ok: bool, pendencias: [{source, label, hint, detail}]}``.
        Se ``ok=False``, JS mostra dialog com botoes 'Ir para X'."""
        ds = self._get_data_state()
        if ds is None:
            return {"ok": False, "pendencias": [],
                    "error": "manager indisponivel"}
        try:
            from runtime.config import DataStateManager  # noqa: PLC0415
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "pendencias": [], "error": str(exc)}

        action = str(action_name or "")
        req = required if isinstance(required, dict) else {}
        if not req:
            req = {
                "db": DataStateManager.CARREGADO_VALIDADO,
            }

        labels = {
            "db": ("Banco de dados",
                   "Conecte ou crie um banco de dados."),
            "apoio": ("Planilha de apoio",
                      "Carregue a planilha de apoio."),
            "ganhos": ("Pasta de ganhos",
                       "Selecione a pasta dos arquivos de ganhos."),
            "tecnico_txt": ("Arquivos tecnicos (TXT)",
                            "Carregue os arquivos tecnicos."),
        }

        pendencias: list[dict[str, Any]] = []
        for source, min_state in req.items():
            if not ds.meets_required(source, min_state):
                info = ds.get_state(source)
                label, hint = labels.get(source, (source, ""))
                erro = f" Erro: {info.error_last}" if info.error_last else ""
                pendencias.append({
                    "source": source,
                    "label": label,
                    "hint": hint,
                    "state": info.state,
                    "detail": f"{label} ({info.state}).{erro} {hint}".strip(),
                })

        return {
            "ok": not pendencias,
            "action": action,
            "pendencias": pendencias,
        }

    def data_state_invalidate(self, source: Any, error: Any = "") -> dict[str, Any]:
        """Marca uma fonte como INVALIDADO. Usado quando o JS detecta
        condicao externa (arquivo apagado, banco fechado, etc.)."""
        try:
            from runtime.config import DataStateManager  # noqa: PLC0415
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": str(exc)}
        src = str(source or "").strip()
        if not src:
            return {"ok": False, "error": "source vazio"}
        self._data_state_set(
            src,
            DataStateManager.INVALIDADO,
            error=str(error or ""),
        )
        return {"ok": True}

    # ------------------------------------------------------------------
    # Passo 2.1 (Shell/Header): estado das fontes (Banco, Apoio, Tecnico)
    # + identificacao do usuario/versao para popular as src-pill do header.
    # ------------------------------------------------------------------
    @staticmethod
    def _file_state(p: str) -> dict[str, Any]:
        """Classifica um caminho de arquivo como ok/warn/err para as pills.

        ok    arquivo existe e e' legivel
        warn  caminho configurado mas arquivo nao encontrado / sem acesso
        err   nao configurado

        Tambem retorna ``size`` (bytes, 0 se indisponivel) usado pela
        status bar do Passo 2.2.
        """
        path = (p or "").strip()
        if not path:
            return {
                "status": "err", "label": "nao configurado",
                "name": "", "size": 0, "mtime": 0,
            }
        name = os.path.basename(path) or path
        size = 0
        mtime = 0.0
        try:
            if not os.path.exists(path):
                return {
                    "status": "warn", "label": "nao encontrado",
                    "name": name, "size": 0, "mtime": 0,
                }
            if not os.access(path, os.R_OK):
                return {
                    "status": "warn", "label": "sem acesso",
                    "name": name, "size": 0, "mtime": 0,
                }
            try:
                stat = os.stat(path)
                size = int(stat.st_size)
                mtime = float(stat.st_mtime)
            except OSError:
                pass
        except OSError:
            return {
                "status": "warn", "label": "erro de IO",
                "name": name, "size": 0, "mtime": 0,
            }
        return {
            "status": "ok", "label": "",
            "name": name, "size": size, "mtime": mtime,
        }

    def _ensure_db_connected(self) -> tuple[Any, str]:
        """Garante que o DatabaseManager esteja conectado ao banco do
        config.json e retorna ``(db_manager, "")`` ou ``(None, motivo)``.

        Idempotente: nao reconecta se ja apontando para o mesmo arquivo.

        SERIALIZADO via _connect_lock para evitar race em
        add_column_if_missing quando varias chamadas API do JS chegam
        em paralelo no boot. Nunca propaga excecao -- erros viram
        ``(None, motivo)`` para o JS poder tratar e o usuario corrigir
        o caminho via Configuracoes."""
        try:
            self._ensure_managers()
        except Exception as exc:  # noqa: BLE001
            return None, f"managers indisponiveis: {exc}"

        cfg = self._config or {}
        db_path = (cfg.get("obras") or "").strip()
        if not db_path:
            self._data_state_set(
                "db", "NAO_CARREGADO", error="banco nao configurado")
            return None, "banco nao configurado em config.json (chave 'obras')"
        if not os.path.exists(db_path):
            self._data_state_set(
                "db", "INVALIDADO", path=db_path,
                error="banco nao encontrado")
            return None, f"banco nao encontrado: {db_path}"
        db = self._db_manager
        if db is None:
            self._data_state_set(
                "db", "INVALIDADO", path=db_path,
                error="DatabaseManager indisponivel")
            return None, "DatabaseManager indisponivel"

        # Cache de erro: se o caminho mudou, limpa o cache + lista de
        # paths ja inicializados (vai precisar reconectar).
        if self._last_connect_path != db_path:
            self._last_connect_error = ""
            self._last_connect_path = db_path
            self._connected_paths.discard(db_path)

        # Caminho rapido: ja inicializamos este db_path com sucesso antes.
        # DatabaseManager.data_access_layer continua valido para queries
        # mesmo apos `_with_connection` fechar a conn fisica.
        if db_path in self._connected_paths:
            return db, ""

        with self._connect_lock:
            # Re-checa apos lock (outra thread pode ter conectado
            # enquanto esperavamos).
            if db_path in self._connected_paths:
                return db, ""
            try:
                db.connect(db_path)
                self._connected_paths.add(db_path)
                self._last_connect_error = ""
                # Hook estado: db conectado e migrado com sucesso.
                self._data_state_set(
                    "db", "CARREGADO_VALIDADO", path=db_path,
                    version_token=str(int(os.path.getmtime(db_path))))
                return db, ""
            except Exception as exc:  # noqa: BLE001
                msg = f"falha ao conectar: {exc}"
                self._last_connect_error = msg
                self._data_state_set(
                    "db", "INVALIDADO", path=db_path, error=msg)
                # Nao deixa o app travar: log no stderr e devolve erro.
                print(f"[main_web] {msg}", file=sys.stderr)
                return None, msg

    # ------------------------------------------------------------------
    # Identificacao do usuario - port direto de
    # cadastro_viabilidades.main_web._ad_display_name + get_user_info +
    # set_display_name. Resolve display_name via:
    #   1. config['ui_preferences_por_usuario'][<username>]['display_name']
    #   2. Active Directory Windows (GetUserNameExW NameDisplay)
    #   3. Fallback: username tratado (split _.- + Title Case)
    # Iniciais derivadas do display_name final (avatar mostra "AS" mesmo
    # quando username e' uma matricula tipo "12345").
    # ------------------------------------------------------------------
    def _ad_display_name(self) -> str:
        """Tenta resolver o nome amigavel do usuario via Active Directory
        no Windows (secur32.GetUserNameExW com NameDisplay = 3). Retorna
        '' se nao estiver em dominio ou se falhar."""
        if not sys.platform.startswith("win"):
            return ""
        try:
            import ctypes
            from ctypes import wintypes
        except Exception:  # noqa: BLE001
            return ""
        # EXTENDED_NAME_FORMAT.NameDisplay = 3
        NAME_DISPLAY = 3
        try:
            secur32 = ctypes.WinDLL("secur32", use_last_error=True)
            secur32.GetUserNameExW.argtypes = [
                wintypes.INT, wintypes.LPWSTR, ctypes.POINTER(wintypes.ULONG),
            ]
            secur32.GetUserNameExW.restype = wintypes.BOOLEAN
            size = wintypes.ULONG(256)
            buf = ctypes.create_unicode_buffer(size.value)
            ok = secur32.GetUserNameExW(NAME_DISPLAY, buf, ctypes.byref(size))
            if not ok:
                # Tenta com buffer maior se falhou por tamanho
                size_2 = wintypes.ULONG(size.value)
                buf = ctypes.create_unicode_buffer(size_2.value + 1)
                ok = secur32.GetUserNameExW(
                    NAME_DISPLAY, buf, ctypes.byref(size_2))
                if not ok:
                    return ""
            return (buf.value or "").strip()
        except Exception:  # noqa: BLE001
            return ""

    def get_user_info(self) -> dict[str, Any]:
        """Identifica o usuario logado e resolve um nome amigavel.

        Ordem de resolucao do nome de exibicao (display_name):
            1. config['ui_preferences_por_usuario'][<username>]['display_name']
            2. Active Directory (Windows): GetUserNameEx(NameDisplay).
            3. Username tratado (split por . _ - + Title Case).

        Sempre retorna `username` cru (chave do config — preferencias usam
        esse valor, nao o display)."""
        # 1. Resolve username cru
        name = ""
        try:
            import getpass as _gp
            name = (_gp.getuser() or "").strip()
        except Exception:  # noqa: BLE001
            name = ""
        if not name:
            try:
                name = (os.getlogin() or "").strip()
            except Exception:  # noqa: BLE001
                name = ""
        if not name:
            name = (
                os.environ.get("USER")
                or os.environ.get("USERNAME")
                or os.environ.get("LOGNAME")
                or ""
            ).strip()
        username = name or "default"

        # 2. Override por config (precedencia maxima)
        display_name = ""
        source = "fallback"
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager,
            )
            cfg = ConfigManager.load_config() or {}
            ui_prefs = cfg.get("ui_preferences_por_usuario") or {}
            if isinstance(ui_prefs, dict):
                bucket = ui_prefs.get(username) or {}
                if isinstance(bucket, dict):
                    manual = str(bucket.get("display_name") or "").strip()
                    if manual:
                        display_name = manual
                        source = "config"
        except Exception:  # noqa: BLE001
            pass

        # 3. Active Directory (so se nao tem override)
        ad_name = ""
        if not display_name:
            ad_name = self._ad_display_name()
            if ad_name:
                display_name = ad_name
                source = "ad"

        # 4. Fallback: username tratado
        if not display_name:
            base = username if username and username != "default" else "Usuário"
            parts = [p for p in re.split(r"[._\s-]+", base) if p]
            if parts:
                display_name = " ".join(p.capitalize() for p in parts)
            else:
                display_name = base
            source = "fallback"

        # Iniciais derivadas do display_name final
        parts = [p for p in re.split(r"[._\s-]+", display_name) if p]
        if parts:
            initials = "".join(p[0] for p in parts[:2]).upper()
        else:
            initials = (display_name[:2] or "US").upper()

        return {
            "ok": True,
            "username": username,
            "display_name": display_name,
            "initials": initials or "US",
            "source": source,
            "ad_suggestion": ad_name,
        }

    def set_display_name(self, name: Any = "") -> dict[str, Any]:
        """Override manual de display_name em
        config['ui_preferences_por_usuario'][<username>]['display_name'].
        ``name`` vazio remove o override (volta a usar AD ou fallback)."""
        info = self.get_user_info()
        username = info.get("username") or "default"
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager,
            )
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"config: {exc}"}
        ui_prefs = cfg.get("ui_preferences_por_usuario") or {}
        if not isinstance(ui_prefs, dict):
            ui_prefs = {}
        bucket = ui_prefs.get(username) or {}
        if not isinstance(bucket, dict):
            bucket = {}
        clean = str(name or "").strip()
        if clean:
            bucket["display_name"] = clean
        else:
            bucket.pop("display_name", None)
        ui_prefs[username] = bucket
        try:
            ConfigManager.save_config({"ui_preferences_por_usuario": ui_prefs})
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return {"ok": True, "user": self.get_user_info()}

    def get_app_state(self) -> dict[str, Any]:
        """Estado global para popular header (source pills) + status bar.

        Le caminhos do config.json (mesmo arquivo que o desktop usa),
        nao requer DatabaseManager inicializado para ser barato no boot.
        """
        # Carrega config sem instanciar managers (mais rapido no boot da UI).
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]

            cfg = ConfigManager.load_config()
        except Exception as exc:  # noqa: BLE001
            cfg = {}
            cfg_error = str(exc)
        else:
            cfg_error = ""

        # Caminho do banco: chave "obras" e' a usada pelo desktop atual.
        db_path = str(cfg.get("obras") or "").strip()
        # Apoio DB-backed (2026-05-07): pill "Apoio" reflete tabelas no
        # banco. last_path em apoio_meta indica de qual xlsx os dados
        # foram importados (sem precisar abrir xlsx para hidratar).
        apoio_path = ""
        apoio_db_state: dict[str, Any] | None = None
        try:
            db_for_meta = getattr(self, "_db_manager", None)
            if db_for_meta is not None and db_path:
                meta = self._apoio_meta_dict(db_for_meta)
                if meta:
                    apoio_path = str(meta.get("last_path") or "")
                    if meta.get("last_imported_at"):
                        apoio_db_state = {
                            "status": "ok",
                            "label": (
                                f"{meta.get('sheet_count') or 0} aba(s)"
                                f" · banco · "
                                f"{meta.get('last_imported_at') or ''}"),
                            "path": apoio_path,
                        }
        except Exception:  # noqa: BLE001
            apoio_db_state = None
        # [FIX] Tecnico_txt: agora le caminho_pasta_ganhos do config
        # e checa presenca dos 3 arquivos obrigatorios (paridade com
        # validate_tecnico_files). Antes mostrava warn so com base no
        # apoio_path, ignorando totalmente a pasta de ganhos configurada.
        ganhos_path = str(cfg.get("caminho_pasta_ganhos") or "").strip()
        tecnico_files = ("FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT")
        if ganhos_path and os.path.isdir(ganhos_path):
            faltantes = [
                f for f in tecnico_files
                if not any(
                    os.path.exists(os.path.join(ganhos_path, name))
                    for name in (f, f.lower(), f.upper())
                )
            ]
            if not faltantes:
                tecnico = {"status": "ok",
                           "label": "3 arquivos validados",
                           "path": ganhos_path}
            else:
                tecnico = {
                    "status": "warn",
                    "label": "Faltam: " + ", ".join(faltantes),
                    "path": ganhos_path,
                }
        else:
            tecnico = {"status": "warn",
                       "label": "Pasta de ganhos nao configurada"}

        # User: usa get_user_info() para resolver display_name amigavel
        # (override config -> Active Directory Windows -> fallback Title Case).
        # Port da logica de cadastro_viabilidades.
        try:
            uinfo = self.get_user_info()
            user = uinfo.get("display_name") or uinfo.get("username") or "?"
            user_initials = uinfo.get("initials") or "?"
        except Exception:  # noqa: BLE001
            try:
                user = getpass.getuser()
            except Exception:  # noqa: BLE001
                user = "?"
            user_initials = "?"

        # Apoio: prioriza DB-backed state; cai para arquivo so se ainda
        # nao houve importacao (banco virgem - mostra warn pedindo import).
        apoio_state = apoio_db_state or {
            "status": "warn",
            "label": "Apoio nao importado - use 'Atualizar apoio'",
            "path": "",
        }
        return {
            "app": {
                "version": APP_VERSION,
                "user": user,
                "user_initials": user_initials,
                "config_error": cfg_error,
            },
            "sources": {
                "db": self._file_state(db_path) | {"path": db_path},
                "apoio": apoio_state,
                "tecnico": tecnico,
            },
        }

    # ------------------------------------------------------------------
    # Passo 3.1 (Visualizar / list_obras): le o banco real via
    # DatabaseManager.fetch_all e mapeia colunas (ORDERED_COLUMNS) para o
    # shape esperado pelo template do mock (cod, ano, pi, projeto, alim,
    # se, regional, pacote, valor, aprovada, passou, tecAtual).
    # ------------------------------------------------------------------
    @staticmethod
    def _row_to_dict(row: Any, cols: list[str]) -> dict[str, Any]:
        """Converte uma row do fetch_all (lista posicional) em dict por
        nome de coluna -- usado pelos servicos do core que esperam dict."""
        out: dict[str, Any] = {}
        for i, name in enumerate(cols):
            out[name] = row[i] if i < len(row) else ""
        return out

    @staticmethod
    def _fmt_pi(pi_base: Any, item: Any) -> str:
        a = str(pi_base or "").strip()
        b = str(item or "").strip()
        if a and b:
            return f"{a}-{b}"
        return a or b

    def list_obras(self, limit: Any = None) -> dict[str, Any]:
        """Retorna ``{ok, rows, total, error}``. ``rows`` no formato JS do mock.

        Cenarios (Sprint A): quando cenario_ativo != '', filtra raw_rows
        para apenas obras em cenarios_obras E aplica overrides de
        cenario_obras_overrides + ano_final de cenarios_obras."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "rows": [], "total": 0, "error": err}

        try:
            raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "rows": [], "total": 0, "error": f"fetch_all: {exc}"}

        # Cenarios: filtro + overrides
        cen_nome = self._cenario_active_name()
        if cen_nome:
            cod_set, cen_info = self._cenario_cod_set(db, cen_nome)
            ovmap = self._cenario_overrides_map(db, cen_nome)
            try:
                idx_cod_filter = cols.index("cod")
            except ValueError:
                idx_cod_filter = -1
            if idx_cod_filter >= 0 and cod_set:
                # Filtra
                raw_rows = [
                    r for r in raw_rows
                    if str(r[idx_cod_filter] if idx_cod_filter < len(r)
                           else "").strip() in cod_set
                ]
                # Aplica overrides + ano_final
                applied: list[Any] = []
                for r in raw_rows:
                    cod_r = str(r[idx_cod_filter] if idx_cod_filter < len(r)
                                else "").strip()
                    applied.append(self._cenario_apply_to_row(
                        r, cols, cod_r,
                        cen_info.get(cod_r) or {},
                        ovmap.get(cod_r) or {},
                    ))
                raw_rows = applied
            elif cod_set == set():
                # Cenario com 0 obras (ou tabelas inexistentes): retorna vazio
                raw_rows = []

        # ---- Passo 3.5: computa "passou" via verificar_criterios_v2 -----
        # Mesma regra V2 usada pela MainWindow desktop (cor unica por
        # projeto). Se faltam colunas obrigatorias o servico devolve
        # [True, ...] (tratado como "tudo ok"). Falhas de import nao
        # sao fatais: caem para todos True.
        passou_list: list[bool] = [True] * len(raw_rows)
        try:
            from core.services.relatorio_criterios_service import (  # type: ignore[import-not-found]
                verificar_criterios_v2,
            )
            criterios = (self._config or {}).get("criterios_planejamento") or {}
            if criterios and raw_rows:
                raw_passou = verificar_criterios_v2(
                    raw_rows, cols, criterios=criterios
                )
                passou_list = [
                    True if v is None else bool(v) for v in raw_passou
                ]
        except Exception as exc:  # noqa: BLE001
            print(f"[main_web] verificar_criterios_v2 falhou: {exc}", file=sys.stderr)

        # Resolve indices uma vez (rows pode ser milhares).
        def idx(name: str) -> int:
            try:
                return cols.index(name)
            except ValueError:
                return -1

        i_cod = idx("cod")
        i_ano = idx("ano_")
        i_pi_base = idx("pi_base")
        i_item = idx("codigo_item")
        i_projeto = idx("nome_projeto")
        i_alim = idx("alimentador_principal")
        i_alim_benef = idx("alimentadores_beneficiados")  # [B6]
        i_se = idx("subestacao")
        i_regional = idx("nome_regional")
        i_super = idx("nome_superintendencia")  # [B6/superintendencia]
        i_pacote = idx("tipo_pacote")
        i_valor = idx("valor_obra")
        i_aprovada = idx("obra_aprovada")
        i_dirty = idx("tecnico_dirty")

        max_rows: int | None
        try:
            max_rows = int(limit) if limit not in (None, "", 0) else None
        except (TypeError, ValueError):
            max_rows = None

        out: list[dict[str, Any]] = []
        raw_out: list[list[Any]] = []
        passou_out: list[bool] = []
        for row_i, r in enumerate(raw_rows):
            def g(i: int, default: Any = "") -> Any:
                return r[i] if 0 <= i < len(r) else default

            ano_raw = g(i_ano)
            try:
                ano: Any = int(str(ano_raw).strip()) if str(ano_raw).strip() else ""
            except (TypeError, ValueError):
                ano = ano_raw
            try:
                valor = float(str(g(i_valor) or 0).replace(",", "."))
            except (TypeError, ValueError):
                valor = 0.0
            aprovada = str(g(i_aprovada) or "").strip().upper() == "SIM"
            # tecnico_dirty == 'SIM' -> snapshot desatualizado;
            # exibimos o oposto (tecAtual True quando NAO dirty).
            tec_atual = str(g(i_dirty) or "").strip().upper() != "SIM"
            row_passou = passou_list[row_i] if row_i < len(passou_list) else True

            # Curado (compat com codigo antigo dos passos 3.x).
            out.append({
                "cod": str(g(i_cod) or ""),
                "ano": ano,
                "pi": self._fmt_pi(g(i_pi_base), g(i_item)),
                "projeto": str(g(i_projeto) or ""),
                "alim": str(g(i_alim) or ""),
                "alim_benef": str(g(i_alim_benef) or ""),  # [B6]
                "se": str(g(i_se) or ""),
                "regional": str(g(i_regional) or ""),
                "superintendencia": str(g(i_super) or ""),  # [B6]
                "pacote": str(g(i_pacote) or ""),
                "valor": valor,
                "aprovada": aprovada,
                "passou": row_passou,
                "tecAtual": tec_atual,
            })
            # Raw: mesmas linhas que o desktop usa em
            # MainWindow.load_obras_into_table (todas as colunas, ordem
            # de get_column_names()).
            raw_row: list[Any] = []
            for i in range(len(cols)):
                v = r[i] if i < len(r) else ""
                # Serializa pra string -- pywebview JSON nao aceita
                # tipos exoticos do sqlite (bytes, datetime).
                if v is None:
                    raw_row.append("")
                elif isinstance(v, (bytes, bytearray)):
                    try:
                        raw_row.append(v.decode("utf-8", "replace"))
                    except Exception:  # noqa: BLE001
                        raw_row.append(repr(v))
                else:
                    raw_row.append(v if isinstance(v, (int, float, str, bool))
                                     else str(v))
            raw_out.append(raw_row)
            passou_out.append(row_passou)
            if max_rows is not None and len(out) >= max_rows:
                break

        return {
            "ok": True, "error": "",
            "rows": out,
            "total": len(raw_rows),
            # Dados crus para a renderizacao "fiel" do desktop:
            # todas as colunas em get_column_names(), na mesma ordem.
            "columns": cols,
            "raw_rows": raw_out,
            "passou_per_row": passou_out,
        }

    # ------------------------------------------------------------------
    # Passo 3.2 (Visualizar / stat cards): agregados rapidos para a faixa
    # de 4 cards no topo da aba Visualizar (Obras, Aprovadas, Pendentes,
    # Valor planejado). Roda no banco direto (SQL) quando possivel para
    # nao precisar trazer a lista inteira; cai no fetch_all como fallback.
    # ------------------------------------------------------------------
    def format_pagination_label(
        self,
        current_page: Any = 1,
        total_pages: Any = 1,
        total_items: Any = 0,
    ) -> dict[str, Any]:
        """[D6] Wrapper sobre visualizar_pagination.format_pagination_label
        para o JS reutilizar a logica do desktop sem duplicar string."""
        try:
            from visualizar_pagination import format_pagination_label as _impl
            label = _impl(int(current_page or 1),
                          int(total_pages or 1),
                          int(total_items or 0))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "label": "", "error": f"format: {exc}"}
        return {"ok": True, "label": str(label or ""), "error": ""}

    def get_obras_stats(self) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {
                "ok": False, "error": err,
                "total": 0, "aprovadas": 0, "pendentes": 0,
                "valor_total": 0.0, "ano_dominante": None,
            }

        # Tenta SQL agregado (rapido em bancos grandes); se falhar usa
        # fetch_all e itera em Python.
        try:
            cursor = db._get_cursor()  # acessor interno do DatabaseManager
            if cursor is not None:
                cursor.execute(
                    "SELECT "
                    " COUNT(*),"
                    " SUM(CASE WHEN UPPER(COALESCE(obra_aprovada,''))='SIM' THEN 1 ELSE 0 END),"
                    " COALESCE(SUM(CAST(REPLACE(REPLACE(COALESCE(valor_obra,'0'),'.',''),',','.') AS REAL)), 0)"
                    " FROM obras"
                )
                row = cursor.fetchone()
                total = int(row[0] or 0)
                aprovadas = int(row[1] or 0)
                valor_total = float(row[2] or 0.0)
                # Ano dominante via SQL separado (ordem por count).
                cursor.execute(
                    "SELECT ano_, COUNT(*) c FROM obras WHERE ano_ IS NOT NULL "
                    "AND TRIM(ano_)<>'' GROUP BY ano_ ORDER BY c DESC LIMIT 1"
                )
                ano_row = cursor.fetchone()
                ano_dominante = (
                    str(ano_row[0]).strip() if ano_row and ano_row[0] is not None else None
                )
                return {
                    "ok": True, "error": "",
                    "total": total,
                    "aprovadas": aprovadas,
                    "pendentes": max(0, total - aprovadas),
                    "valor_total": valor_total,
                    "ano_dominante": ano_dominante,
                }
        except Exception:  # noqa: BLE001
            # Cai no fallback abaixo.
            pass

        # Fallback: agregacao em Python.
        try:
            rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {
                "ok": False, "error": f"fetch_all: {exc}",
                "total": 0, "aprovadas": 0, "pendentes": 0,
                "valor_total": 0.0, "ano_dominante": None,
            }

        try:
            i_apr = cols.index("obra_aprovada")
        except ValueError:
            i_apr = -1
        try:
            i_val = cols.index("valor_obra")
        except ValueError:
            i_val = -1
        try:
            i_ano = cols.index("ano_")
        except ValueError:
            i_ano = -1

        total = len(rows)
        aprovadas = 0
        valor_total = 0.0
        anos: dict[str, int] = {}
        for r in rows:
            if i_apr >= 0 and i_apr < len(r):
                if str(r[i_apr] or "").strip().upper() == "SIM":
                    aprovadas += 1
            if i_val >= 0 and i_val < len(r):
                try:
                    valor_total += float(str(r[i_val] or 0).replace(",", "."))
                except (TypeError, ValueError):
                    pass
            if i_ano >= 0 and i_ano < len(r):
                ano_v = str(r[i_ano] or "").strip()
                if ano_v:
                    anos[ano_v] = anos.get(ano_v, 0) + 1
        ano_dominante = max(anos.items(), key=lambda kv: kv[1])[0] if anos else None
        return {
            "ok": True, "error": "",
            "total": total,
            "aprovadas": aprovadas,
            "pendentes": max(0, total - aprovadas),
            "valor_total": valor_total,
            "ano_dominante": ano_dominante,
        }

    # ------------------------------------------------------------------
    # Passo 3.3 (Visualizar / search + filtros): aplica busca textual
    # multi-termo (separador ;,) sobre todos os campos visiveis e filtros
    # estruturados vindos do modal de filtros avancados. Reutiliza
    # ui_helpers.matches_filter_value / matches_cod_terms ja existentes.
    # ------------------------------------------------------------------
    @staticmethod
    def _split_terms(s: str) -> list[str]:
        return [t.strip() for t in re.split(r"[;,]", str(s or "")) if t.strip()]

    def search_obras(
        self,
        query: Any = "",
        filters: Any = None,
    ) -> dict[str, Any]:
        base = self.list_obras()
        if not base.get("ok"):
            return base
        rows: list[dict[str, Any]] = base.get("rows") or []
        # Mantemos referencias paralelas para filtrar raw_rows e
        # passou_per_row em sintonia com o array curado.
        raw_all = list(base.get("raw_rows") or [])
        passou_all = list(base.get("passou_per_row") or [])
        cols = list(base.get("columns") or [])
        # Indexa cada row curado pela sua posicao original; o filtro
        # remove tanto do curado quanto dos paralelos.
        for i, o in enumerate(rows):
            if isinstance(o, dict):
                o["__src_idx"] = i

        # --- 1) Busca textual global ---
        q_terms = self._split_terms(str(query or "").lower())
        if q_terms:
            def _haystack(o: dict[str, Any]) -> str:
                return " ".join(str(o.get(k, "")) for k in (
                    "cod", "ano", "pi", "projeto", "alim",
                    "alim_benef", "se", "regional", "pacote",
                )).lower()
            rows = [o for o in rows if any(t in _haystack(o) for t in q_terms)]

        # --- 2) Filtros estruturados ---
        f = filters if isinstance(filters, dict) else {}

        def fv(key: str) -> str:
            return str(f.get(key, "") or "").strip()

        try:
            from ui_helpers import (  # type: ignore[import-not-found]
                matches_cod_terms,
                matches_filter_value,
            )
        except Exception:  # noqa: BLE001
            matches_cod_terms = lambda v, p: True  # noqa: E731
            matches_filter_value = lambda v, p: True  # noqa: E731

        if fv("cod"):
            rows = [o for o in rows if matches_cod_terms(str(o.get("cod", "")), fv("cod"))]
        if fv("ano"):
            anos = self._split_terms(fv("ano"))
            rows = [o for o in rows if str(o.get("ano", "")).strip() in anos]
        for key, field in (
            ("pi", "pi"),
            ("projeto", "projeto"),
            ("alim", "alim"),
            ("alim_benef", "alim_benef"),  # [B6] coluna propria agora exposta em list_obras
            ("se", "se"),
        ):
            if fv(key):
                rows = [
                    o for o in rows
                    if matches_filter_value(str(o.get(field, "")), fv(key))
                ]
        # Selects: ignoram placeholders "Todas" / "Todos" / "—".
        sentinels_select = {"todas", "todos", "—", "-", "todos os pacotes", ""}

        def _multi(key: str) -> list[str]:
            """Split ;-separated values do front (multi-select). Filtra
            sentinels e vazios. Retorna lista normalizada uppercase."""
            raw = fv(key)
            if not raw or raw.lower() in sentinels_select:
                return []
            parts = [p.strip() for p in raw.replace(",", ";").split(";")]
            return [p.upper() for p in parts
                    if p and p.lower() not in sentinels_select]

        regs = _multi("regional")
        if regs:
            rows = [o for o in rows
                    if str(o.get("regional", "")).upper() in regs]
        # Superintendencia ainda nao exposta em list_obras (placeholder
        # ate Passo de mapping incluir o atributo).
        sups = _multi("superintendencia")
        if sups:
            rows = [o for o in rows
                    if str(o.get("superintendencia", "")).upper() in sups]
        pacs = _multi("pacote")
        if pacs:
            rows = [o for o in rows
                    if str(o.get("pacote", "")).strip().upper() in pacs]
        # Aprovada / Tecnico Atualizado / Criterios usam as flags ja
        # mapeadas em list_obras.
        apr = fv("aprovada").upper().replace("Ã", "A")
        if apr in ("SIM", "NAO"):
            want = apr == "SIM"
            rows = [o for o in rows if bool(o.get("aprovada")) == want]
        tec = fv("tecnico").upper().replace("Ã", "A")
        if tec in ("SIM", "NAO"):
            want = tec == "SIM"
            rows = [o for o in rows if bool(o.get("tecAtual")) == want]
        crit = fv("criterios").lower().replace("ã", "a").replace("ç", "c")
        if crit == "atenderam":
            rows = [o for o in rows if o.get("passou", True)]
        elif crit == "falharam":
            rows = [o for o in rows if not o.get("passou", True)]
        elif crit == "aprovadas":
            rows = [o for o in rows if o.get("aprovada")]
        elif crit in ("nao aprovadas", "nao_aprovadas"):
            rows = [o for o in rows if not o.get("aprovada")]

        # Reconstroi os arrays paralelos com base nos __src_idx que
        # sobreviveram aos filtros.
        kept_idx = [int(o.get("__src_idx", -1)) for o in rows
                    if isinstance(o, dict)]
        kept_idx = [i for i in kept_idx if 0 <= i < len(raw_all)]
        filtered_raw = [raw_all[i] for i in kept_idx]
        filtered_passou = [passou_all[i] for i in kept_idx]
        for o in rows:
            if isinstance(o, dict):
                o.pop("__src_idx", None)
        return {
            "ok": True, "error": "",
            "rows": rows,
            "raw_rows": filtered_raw,
            "passou_per_row": filtered_passou,
            "columns": cols,
            "total": len(rows),
        }

    # ------------------------------------------------------------------
    # Passo 3.6 (Visualizar / acoes da toolbar): delete + export real
    # para os botoes Excluir/Detalhamento. Atualizar e' apenas JS
    # (chama coplanLoadObras). Relatorio de Criterios e Nota de Colapso
    # ficam como stubs ate Passos 5.x/6.x trazerem os geradores
    # adequados (dependem de templates e ganhos parsing).
    # ------------------------------------------------------------------
    @staticmethod
    def _default_export_dir() -> Path:
        target = Path.home() / "Downloads"
        try:
            if not target.exists():
                target = Path.home()
            target.mkdir(parents=True, exist_ok=True)
        except OSError:
            target = Path.home()
        return target

    # ------------------------------------------------------------------
    # Gating de obras aprovadas (RB-2 do desktop, FiltrosPaginacaoMixin):
    # Replica _gate_aprovadas_for_action + _confirmar_exclusao_excepcional
    # + _registrar_exclusao_excepcional. Antes de qualquer acao destrutiva
    # ou de mutacao em obras, o JS deve consultar
    # `gate_aprovadas_for_action(cods)` para descobrir quais estao aprovadas.
    # Se houver aprovadas e o usuario nao marcou 'incluir aprovadas',
    # essas devem ser ignoradas.
    # ------------------------------------------------------------------
    def gate_aprovadas_for_action(
        self, cods: Any = None, include_aprovadas: Any = False,
    ) -> dict[str, Any]:
        """Filtra obras pela coluna ``obra_aprovada`` ('SIM'/'NAO').
        Retorna 3 listas:
          * ``targets``: cods que podem ser processados nesta acao
          * ``aprovadas``: cods que estao aprovados (so entrariam em
            ``targets`` se include_aprovadas=True)
          * ``inexistentes``: cods que nao foram encontrados no banco

        O JS usa isso pra mostrar dialog do tipo 'X obra(s) aprovada(s)
        foram ignoradas' antes de chamar delete_obras/marcar_correcao."""
        cods_list = [str(c).strip() for c in (cods or []) if str(c or "").strip()]
        include = bool(include_aprovadas)
        if not cods_list:
            return {"ok": False, "targets": [], "aprovadas": [],
                    "inexistentes": [], "error": "cods vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "targets": [], "aprovadas": [],
                    "inexistentes": [], "error": err or "db indisponivel"}
        try:
            cols = list(db.get_column_names() or [])
            i_cod = cols.index("cod") if "cod" in cols else -1
            i_aprov = (cols.index("obra_aprovada")
                       if "obra_aprovada" in cols else -1)
            rows = list(db.fetch_by_cods(cods_list) or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "targets": [], "aprovadas": [],
                    "inexistentes": [], "error": f"fetch: {exc}"}
        achados: dict[str, str] = {}
        for row in rows:
            cod_v = (str(row[i_cod]).strip()
                     if 0 <= i_cod < len(row) else "")
            aprov = (str(row[i_aprov]).strip()
                     if 0 <= i_aprov < len(row) else "").upper()
            if cod_v:
                achados[cod_v] = aprov
        targets: list[str] = []
        aprovadas: list[str] = []
        inexistentes: list[str] = []
        for c in cods_list:
            if c not in achados:
                inexistentes.append(c)
                continue
            if achados[c] == "SIM":
                aprovadas.append(c)
                if include:
                    targets.append(c)
            else:
                targets.append(c)
        return {
            "ok": True,
            "targets": targets,
            "aprovadas": aprovadas,
            "inexistentes": inexistentes,
            "include_aprovadas": include,
            "error": "",
        }

    def register_exclusao_excepcional(
        self, cod: Any, motivo: Any = "",
    ) -> dict[str, Any]:
        """Auditoria de exclusao excepcional de obra aprovada.
        Replica _registrar_exclusao_excepcional do desktop: anexa nota
        em ``observacoes_gerais`` (ou ``observacoes`` ou ``ultima_acao``)
        com timestamp + usuario + motivo, e emite warning no log.
        Chamar SEMPRE antes de delete_obras quando o cod estiver na
        lista de 'aprovadas'."""
        cod_s = str(cod or "").strip()
        if not cod_s:
            return {"ok": False, "error": "cod vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel"}
        try:
            usuario = getpass.getuser()
        except Exception:  # noqa: BLE001
            usuario = "?"
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
        motivo_s = str(motivo or "").strip()
        nota = f"EXCLUSAO EXCEPCIONAL em {timestamp} por {usuario}"
        if motivo_s:
            nota += f" -- motivo: {motivo_s}"
        try:
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"cols: {exc}"}
        coluna_log = None
        for col in ("ultima_acao", "observacoes_gerais",
                    "observacoes", "log"):
            if col in cols:
                coluna_log = col
                break
        if not coluna_log:
            # Sem coluna de log; ainda registra no stderr.
            print(f"[main_web] {nota} (cod={cod_s}, sem coluna de log)",
                  file=sys.stderr)
            return {"ok": True, "logged_to_db": False, "nota": nota}
        try:
            existing = db.fetch_by_cod(cod_s)
            if not existing:
                return {"ok": False, "error": f"cod {cod_s} nao encontrado"}
            i_log = cols.index(coluna_log)
            atual = str(existing[i_log] or "").strip() if 0 <= i_log < len(existing) else ""
            novo = f"{atual}\n{nota}" if atual else nota
            db.update_obra({coluna_log: novo}, cod_s, skip_blank=True)
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            err_s = friendly or f"update: {exc}"
            print(f"[main_web] register_exclusao falhou cod={cod_s}: {err_s}",
                  file=sys.stderr)
            out: dict[str, Any] = {
                "ok": False, "error": err_s,
                "logged_to_db": False, "nota": nota,
            }
            if friendly:
                out["blocked"] = "db_busy"
            return out
        print(f"[main_web] AUDIT: {nota} cod={cod_s}", file=sys.stderr)
        return {"ok": True, "logged_to_db": True, "nota": nota,
                "coluna": coluna_log}

    def delete_obras(self, cods: Any) -> dict[str, Any]:
        """Deleta obras por COD. Usa DatabaseManager.delete_obra (ja
        protegido por lock + transacao). Retorna {ok, deleted, errors}.
        BLOQUEADO quando cenario_ativo != '' (cenario nao deve apagar
        obras em obras)."""
        cen_nome = self._cenario_active_name()
        if cen_nome:
            return {
                "ok": False, "deleted": 0,
                "errors": [(f"Operacao bloqueada: cenario '{cen_nome}'"
                            f" ativo. Saia do cenario para excluir obras.")],
                "blocked": "cenario_active",
            }
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "deleted": 0, "errors": [err or "db indisponivel"]}
        if not isinstance(cods, (list, tuple)) or not cods:
            return {"ok": False, "deleted": 0, "errors": ["lista de cods vazia"]}
        deleted = 0
        errors: list[str] = []
        busy_msg = ""
        for cod in cods:
            cod_s = str(cod or "").strip()
            if not cod_s:
                continue
            try:
                db.delete_obra(cod_s)
                deleted += 1
            except Exception as exc:  # noqa: BLE001
                friendly = self._friendly_busy_error(exc)
                if friendly:
                    # Busy/locked: para o loop (nao adianta tentar mais
                    # cods se o banco esta em uso por outro usuario).
                    busy_msg = friendly
                    errors.append(f"{cod_s}: {friendly}")
                    break
                errors.append(f"{cod_s}: {exc}")
        out: dict[str, Any] = {
            "ok": not errors, "deleted": deleted, "errors": errors,
        }
        if busy_msg:
            out["blocked"] = "db_busy"
            out["error"] = busy_msg
        return out

    def export_detalhamento(self, cods: Any = None) -> dict[str, Any]:
        """Exporta as obras selecionadas (ou todas) para XLSX em
        ~/Downloads. Retorna {ok, path, count, error, cenario}."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count": 0, "error": err or "db indisponivel"}
        try:
            cods_list = [str(c).strip() for c in (cods or []) if str(c).strip()]
            if cods_list:
                raw_rows = db.fetch_by_cods(cods_list) or []
            else:
                raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"fetch: {exc}"}

        # Cenario ativo: restringe + aplica overrides (paridade get_obras).
        # Loga para diagnostico: cenario={X}, rows antes/depois.
        cen_nome = self._cenario_active_name()
        before_n = len(raw_rows)
        raw_rows = self._apply_cenario_to_rows(db, raw_rows, cols)
        print(
            f"[main_web] export_detalhamento: cenario={cen_nome!r}, "
            f"cods_in={len(cods_list)}, rows_before={before_n}, "
            f"rows_after={len(raw_rows)}",
            file=sys.stderr,
        )

        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            from openpyxl.styles import Font, PatternFill  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"openpyxl indisponivel: {exc}"}

        target = self._default_export_dir()
        fname = f"coplan_detalhamento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = target / fname
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Obras"
            ws.append(cols)
            header_fill = PatternFill(start_color="2A3460", end_color="2A3460", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for r in raw_rows:
                ws.append([
                    (r[i] if i < len(r) else "") for i in range(len(cols))
                ])
            ws.freeze_panes = "A2"
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"openpyxl save: {exc}"}

        return {
            "ok": True, "path": str(path),
            "count": len(raw_rows), "error": "",
            "cenario": cen_nome,
        }

    # ------------------------------------------------------------------
    # Fase A7 (resumo_service.montar_resumo_detalhamento):
    # exporta XLSX agrupando por (nome_projeto, ano, pacote) com
    # antes/depois por alimentador. Equivalente a
    # MainWindow._montar_resumo_detalhamento_excel + _exportar_obras
    # (sub-modo "Detalhamento por Regional").
    # ------------------------------------------------------------------
    def export_resumo_detalhamento(self, cods: Any = None) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count": 0,
                    "error": err or "db indisponivel"}
        try:
            cods_list = [str(c).strip() for c in (cods or []) if str(c).strip()]
            if cods_list:
                raw_rows = db.fetch_by_cods(cods_list) or []
            else:
                raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"fetch: {exc}"}
        raw_rows = self._apply_cenario_to_rows(db, raw_rows, cols)
        if not raw_rows:
            return {"ok": False, "path": "", "count": 0,
                    "error": "sem obras para resumir"}
        try:
            from core.services.resumo_service import (  # type: ignore[import-not-found]
                montar_resumo_detalhamento,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"import: {exc}"}
        try:
            df = montar_resumo_detalhamento(raw_rows, cols)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"montar: {exc}"}
        if df is None or df.empty:
            return {"ok": False, "path": "", "count": 0,
                    "error": ("sem dados (faltam colunas obrigatorias ou"
                              " nenhum alimentador resolvido)")}
        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            from openpyxl.styles import (  # type: ignore[import-not-found]
                Font, PatternFill,
            )
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumo Detalhamento"
            cols_out = list(df.columns)
            ws.append(cols_out)
            header_fill = PatternFill(start_color="2A3460",
                                      end_color="2A3460", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for _, row in df.iterrows():
                ws.append([row.get(c, "") for c in cols_out])
            ws.freeze_panes = "A2"
            target = self._default_export_dir()
            fname = (
                f"coplan_resumo_detalhamento_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            path = target / fname
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"xlsx: {exc}"}
        return {"ok": True, "path": str(path),
                "count": int(len(df)), "error": ""}

    def export_relatorio_criterios(self, cods: Any = None) -> dict[str, Any]:
        """Gera planilha das obras que NAO atenderam criterios.
        Visualizar Sprint 1 (Auditoria #5): aceita filtro `cods` opcional
        (lista de cods para escopo 'filtradas'/'selecionadas')."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count": 0, "error": err or "db indisponivel"}
        try:
            raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"fetch: {exc}"}
        # Cenario ativo: restringe + overrides ANTES do filtro de cods
        raw_rows = self._apply_cenario_to_rows(db, raw_rows, cols)
        # Filtra por cods se fornecido (#5)
        if isinstance(cods, (list, tuple)) and cods:
            cod_set = {str(c).strip() for c in cods if str(c or "").strip()}
            if cod_set:
                try:
                    idx_cod = cols.index("cod")
                    raw_rows = [
                        r for r in raw_rows
                        if str(r[idx_cod] if idx_cod < len(r) else "").strip()
                        in cod_set
                    ]
                except ValueError:
                    pass  # 'cod' nao esta nas colunas - ignora filtro
        try:
            from core.services.relatorio_criterios_service import (  # type: ignore[import-not-found]
                verificar_criterios_v2,
            )
            criterios = (self._config or {}).get("criterios_planejamento") or {}
            verdict = verificar_criterios_v2(raw_rows, cols, criterios=criterios)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"verificar_criterios_v2: {exc}"}

        falhas = [
            r for r, ok in zip(raw_rows, verdict)
            if ok is False  # None = indefinido (nao conta como falha)
        ]
        if not falhas:
            return {"ok": True, "path": "", "count": 0,
                    "error": "Todas as obras atenderam aos criterios."}
        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            wb = Workbook()
            ws = wb.active
            ws.title = "Falhas"
            ws.append(cols)
            for r in falhas:
                ws.append([r[i] if i < len(r) else "" for i in range(len(cols))])
            ws.freeze_panes = "A2"
            target = self._default_export_dir()
            fname = (
                f"coplan_relatorio_criterios_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            path = target / fname
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0, "error": f"xlsx: {exc}"}
        return {"ok": True, "path": str(path), "count": len(falhas), "error": ""}

    # ------------------------------------------------------------------
    # Fase B2 (excluir_obra_mixin.marcar_obras_correcao):
    # marca COD(s) como DESPACHO_STATUS = "CORRECAO". Usado depois que
    # uma obra DESPACHADA precisa ser alterada (Fase A9 bloqueia o save
    # ate ela voltar a CORRECAO).
    # ------------------------------------------------------------------
    def marcar_obras_correcao(
        self, cods: Any = None, motivo: Any = "",
    ) -> dict[str, Any]:
        cen_nome = self._cenario_active_name()
        if cen_nome:
            return {
                "ok": False, "error":
                (f"Operacao bloqueada: cenario '{cen_nome}' ativo."
                 f" Saia do cenario para marcar obras como CORRECAO."),
                "blocked": "cenario_active",
                "marcadas": 0, "falhas": [],
            }
        if not isinstance(cods, (list, tuple)) or not cods:
            return {"ok": False, "error": "cods vazio",
                    "marcadas": 0, "falhas": []}
        motivo_s = str(motivo or "").strip()
        # [FIX] Aceita motivo vazio com placeholder. UX: o user nao
        # precisa digitar motivo ANTES de editar (era pedido em prompt
        # + de novo no salvamento). Agora marca como pendente; o
        # motivo real e' capturado em cad-input-motivo no save_obra.
        if not motivo_s:
            motivo_s = "PENDENTE - informar no salvamento"
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "marcadas": 0, "falhas": []}
        now_iso = datetime.now().isoformat(timespec="seconds")
        despacho_ref = f"CORRECAO: {motivo_s}"
        falhas: list[str] = []
        marcadas = 0
        busy_msg = ""
        for c in cods:
            cod_s = str(c or "").strip()
            if not cod_s:
                continue
            try:
                db.update_obra({
                    "despacho_status": "CORRECAO",
                    "despacho_em": now_iso,
                    "despacho_ref": despacho_ref,
                }, cod_s, skip_blank=True)
                marcadas += 1
            except Exception as exc:  # noqa: BLE001
                friendly = self._friendly_busy_error(exc)
                if friendly:
                    busy_msg = friendly
                    falhas.append(f"{cod_s}: {friendly}")
                    break
                falhas.append(f"{cod_s}: {exc}")
        return {
            "ok": (marcadas > 0 and not busy_msg),
            "error": busy_msg,
            "blocked": "db_busy" if busy_msg else "",
            "marcadas": marcadas,
            "falhas": falhas,
            "motivo": motivo_s,
        }

    # ------------------------------------------------------------------
    # Fase A12 (relatorio_criterios_service.montar_relatorio_criterios_por_projeto):
    # Gera XLSX 2-sheet (Projetos + Alimentadores) com avaliacao
    # detalhada de criterios por projeto. Equivalente a
    # MainWindow.montar_relatorio_criterios_por_projeto do desktop.
    # ------------------------------------------------------------------
    def export_relatorio_criterios_projeto(
        self, cods: Any = None,
    ) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": err or "db indisponivel"}
        try:
            cods_list = [str(c).strip() for c in (cods or []) if str(c).strip()]
            if cods_list:
                raw_rows = db.fetch_by_cods(cods_list) or []
            else:
                raw_rows = db.fetch_all() or []
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": f"fetch: {exc}"}
        raw_rows = self._apply_cenario_to_rows(db, raw_rows, cols)
        if not raw_rows:
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0,
                    "error": "sem obras para analisar"}
        try:
            from core.services.relatorio_criterios_service import (  # type: ignore[import-not-found]
                montar_relatorio_criterios_por_projeto,
            )
            from codigo5_coplan import (  # type: ignore[import-not-found]
                DEFAULT_CRITERIOS,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": f"import: {exc}"}

        criterios = dict(DEFAULT_CRITERIOS)
        criterios.update((self._config or {}).get("criterios_planejamento") or {})

        try:
            rel = montar_relatorio_criterios_por_projeto(
                raw_rows, cols, criterios=criterios,
            )
        except ValueError as exc:
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": str(exc)}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": f"montar: {exc}"}

        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            from openpyxl.styles import (  # type: ignore[import-not-found]
                Font, PatternFill,
            )
            wb = Workbook()
            header_fill = PatternFill(start_color="2A3460",
                                      end_color="2A3460", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # Sheet 1: Projetos
            ws1 = wb.active
            ws1.title = "Projetos"
            cols1 = list(rel.df_projetos.columns)
            ws1.append(cols1)
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
            for _, row in rel.df_projetos.iterrows():
                ws1.append([row.get(c, "") for c in cols1])
            ws1.freeze_panes = "A2"

            # Sheet 2: Alimentadores
            ws2 = wb.create_sheet("Alimentadores")
            cols2 = list(rel.df_alimentadores.columns)
            ws2.append(cols2)
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
            for _, row in rel.df_alimentadores.iterrows():
                ws2.append([row.get(c, "") for c in cols2])
            ws2.freeze_panes = "A2"

            target = self._default_export_dir()
            fname = (
                f"coplan_relatorio_criterios_projeto_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            path = target / fname
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count_projetos": 0,
                    "count_alimentadores": 0, "error": f"xlsx: {exc}"}

        return {
            "ok": True, "path": str(path), "error": "",
            "count_projetos": int(len(rel.df_projetos)),
            "count_alimentadores": int(len(rel.df_alimentadores)),
        }

    # ------------------------------------------------------------------
    # Fase A3 (core/services/nota_colapso_service):
    # exporta nota de colapso via core (sem Qt). Substitui o stub.
    # Para cada COD: monta Obra + PIMetadata, chama calcular_nota_colapso
    # e produz XLSX com cod/nota/criterio/valores_considerados.
    # ------------------------------------------------------------------
    def export_nota_colapso(self, cods: Any = None) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "count": 0,
                    "error": err or "db indisponivel"}
        try:
            from core.models import Obra  # type: ignore[import-not-found]
            from core.services.nota_colapso_service import (  # type: ignore[import-not-found]
                calcular_nota_colapso,
            )
            from core.services.pi_metadata_service import (  # type: ignore[import-not-found]
                buscar_pi_metadata,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"import: {exc}"}

        cods_list: list[str] = []
        if isinstance(cods, (list, tuple)):
            cods_list = [str(c).strip() for c in cods if str(c or "").strip()]
        try:
            cols = list(db.get_column_names() or [])
            if cods_list:
                rows = db.fetch_by_cods(cods_list) or []
            else:
                rows = db.fetch_all() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"fetch: {exc}"}

        rows = self._apply_cenario_to_rows(db, rows, cols)

        cfg = self._config or {}
        resultados: list[dict[str, Any]] = []
        for r in rows:
            row_dict = {col: r[i] if i < len(r) else ""
                        for i, col in enumerate(cols)}
            try:
                obra = Obra.from_row_dict(row_dict)
            except Exception as exc:  # noqa: BLE001
                resultados.append({
                    "cod": str(row_dict.get("cod") or ""),
                    "valor": None,
                    "criterio": f"Erro ao montar Obra: {exc}",
                    "valores": {},
                })
                continue
            pi_md = None
            try:
                pi_md = buscar_pi_metadata(
                    str(obra.ident.projeto_investimento or ""), cfg,
                )
            except Exception:  # noqa: BLE001
                pi_md = None
            try:
                nota = calcular_nota_colapso(obra, pi_md)
            except Exception as exc:  # noqa: BLE001
                resultados.append({
                    "cod": obra.ident.cod,
                    "valor": None,
                    "criterio": f"Erro no calculo: {exc}",
                    "valores": {},
                })
                continue
            resultados.append({
                "cod": obra.ident.cod,
                "pi": obra.ident.projeto_investimento,
                "valor": nota.valor,
                "criterio": nota.criterio,
                "valores": dict(nota.valores_considerados or {}),
            })

        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
            wb = Workbook()
            ws = wb.active
            ws.title = "Nota Colapso"
            ws.append([
                "COD", "PI", "Nota", "Criterio",
                "Carreg Inicial", "Carreg Max",
                "Tensao Min Inicial", "Tensao Max Inicial",
                "Tmin Registrada", "Tmax Registrada",
            ])
            for item in resultados:
                v = item.get("valores") or {}
                ws.append([
                    item.get("cod") or "",
                    item.get("pi") or "",
                    item.get("valor") if item.get("valor") is not None else "",
                    item.get("criterio") or "",
                    v.get("carreg_inicial", ""),
                    v.get("carreg_max", ""),
                    v.get("tensao_min_inicial", ""),
                    v.get("tensao_max_inicial", ""),
                    v.get("tmin_registrada", ""),
                    v.get("tmax_registrada", ""),
                ])
            ws.freeze_panes = "A2"
            target = self._default_export_dir()
            fname = (
                f"coplan_nota_colapso_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            path = target / fname
            wb.save(str(path))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "count": 0,
                    "error": f"xlsx: {exc}"}

        return {
            "ok": True, "path": str(path),
            "count": len(resultados), "error": "",
        }

    # ------------------------------------------------------------------
    # Passo 4.1 (Cadastro / get_obra): le UMA obra completa por COD para
    # popular o formulario quando o usuario decide editar uma existente.
    # Retorna o dict cru (todas as colunas de ORDERED_COLUMNS) + alguns
    # campos derivados que o JS precisa (ex: alimentadores_beneficiados
    # ja como lista).
    # ------------------------------------------------------------------
    def get_obra(self, cod: Any) -> dict[str, Any]:
        cod_s = str(cod or "").strip()
        if not cod_s:
            return {"ok": False, "obra": None, "error": "cod vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "obra": None, "error": err or "db indisponivel"}
        try:
            row = db.fetch_by_cod(cod_s)
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "obra": None, "error": f"fetch_by_cod: {exc}"}
        if not row:
            return {"ok": False, "obra": None, "error": f"obra nao encontrada: {cod_s}"}

        # Cenarios (Sprint A): aplica ano_final + overrides quando ativo
        cen_nome = self._cenario_active_name()
        if cen_nome:
            cod_set, cen_info = self._cenario_cod_set(db, cen_nome)
            if cod_s in cod_set:
                ovmap = self._cenario_overrides_map(db, cen_nome)
                row = tuple(self._cenario_apply_to_row(
                    list(row), cols, cod_s,
                    cen_info.get(cod_s) or {},
                    ovmap.get(cod_s) or {},
                ))
        obra = self._row_to_dict(row, cols)
        # Deriva lista de alimentadores beneficiados (separador ; comum
        # no banco). JS pode chunkar como chips diretamente.
        alim_benef_raw = str(obra.get("alimentadores_beneficiados") or "")
        alim_benef = [
            x.strip() for x in re.split(r"[;,]", alim_benef_raw) if x.strip()
        ]
        # Deriva subestacoes a partir do prefixo dos alimentadores
        # beneficiados (igual ao desktop: ATB-204 -> ATB).
        ses = []
        for a in alim_benef:
            prefix = re.split(r"[-_/]", a, 1)[0].strip().upper()
            if prefix and prefix not in ses:
                ses.append(prefix)

        return {
            "ok": True,
            "obra": obra,
            "alim_benef": alim_benef,
            "ses_derivadas": ses,
            "error": "",
        }

    # ------------------------------------------------------------------
    # Passo 4.2 (Cadastro / save_obra): persiste o dict do formulario.
    # Decide entre INSERT e UPDATE com base na existencia do COD na base.
    # Reusa DatabaseManager.insert_obra / update_obra que ja aplicam:
    #   * sanitizacao de alimentador (sem '_')
    #   * derivacao de pi_base via get_pi_base
    #   * data_criacao / data_modificacao / criado_por / modificado_por
    #   * defaults (obra_aprovada=NAO, tecnico_dirty=NAO)
    #   * empresa + cod_pep tail
    #   * lock + transacao + retry-on-busy
    # ------------------------------------------------------------------
    def save_obra(self, payload: Any) -> dict[str, Any]:
        if not isinstance(payload, dict):
            return {"ok": False, "cod": "", "mode": "", "error": "payload nao e dict"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "cod": "", "mode": "", "error": err or "db indisponivel"}

        cod = str(payload.get("cod") or "").strip()

        # Limpeza basica: remove None, normaliza string. Mantem numeros
        # como estao (insert_obra._normalize_decimal trata).
        cleaned: dict[str, Any] = {}
        cols = list(db.get_column_names() or [])
        for key, value in payload.items():
            # Ignora chaves que nao sao colunas reais (defensivo).
            if key not in cols and key != "cod":
                continue
            if value is None:
                cleaned[key] = ""
            elif isinstance(value, (int, float, bool)):
                cleaned[key] = value
            else:
                cleaned[key] = str(value).strip()

        # Fallback de tensao_operacao -> nivel_tensao_obra (regra do
        # build_obra_dados / save_data legado).
        if "tensao_operacao" in cols and not str(cleaned.get("tensao_operacao") or "").strip():
            niv = cleaned.get("nivel_tensao_obra") or ""
            if niv:
                cleaned["tensao_operacao"] = niv
        # Toda obra recem-salva parte como NAO dirty (mesma regra do core).
        if "tecnico_dirty" in cols:
            cleaned["tecnico_dirty"] = "NÃO"

        # Paridade desktop validar_campos_obrigatorios: defense-in-depth.
        # JS ja valida no form, mas se o payload chegar sem algum
        # obrigatorio, abortamos antes de gastar insert/update.
        faltam = self._validar_campos_obrigatorios(cleaned)
        if faltam:
            return {
                "ok": False, "cod": cod, "mode": "",
                "error": ("Campos obrigatorios vazios: "
                          + ", ".join(faltam)),
                "campos_obrigatorios_vazios": faltam,
            }

        try:
            existing = db.fetch_by_cod(cod) if cod else None
        except Exception:  # noqa: BLE001
            existing = None

        # Fase A8: avalia diff + anexa historico. Bloqueio por DESPACHADA
        # foi REMOVIDO (2026-05-08) a pedido do usuario: o operador pode
        # editar obras despachadas livremente. O status da nota e seu
        # numero ficam visiveis no card "Status da Nota" da aba Cadastro
        # (informativo, sem bloqueio).
        old_map: dict[str, Any] = {}
        if existing:
            for i, c in enumerate(cols):
                old_map[c] = existing[i] if i < len(existing) else ""

        diff_decision = None
        try:
            from core.services.salvar_obra_service import (  # type: ignore[import-not-found]
                aplicar_historico_ao_dict, avaliar_diff,
            )
            diff_decision = avaliar_diff(cleaned, old_map, db_columns=cols)
        except Exception as exc:  # noqa: BLE001
            print(f"[main_web] avaliar_diff falhou: {exc}", file=sys.stderr)

        if diff_decision is not None:
            # Anexa historico se houver mudancas + a coluna existe.
            try:
                cleaned = aplicar_historico_ao_dict(
                    cleaned, diff_decision, motivo="",
                )
            except Exception as exc:  # noqa: BLE001
                print(f"[main_web] aplicar_historico falhou: {exc}",
                      file=sys.stderr)

        # Auto-gera COD pra INSERT (paridade com codigo5_coplan.py L1127):
        # CalculationManager.gerar_cod a partir dos campos do form.
        if not existing and not cod:
            cm = self._ensure_calc_manager()
            if cm is None:
                return {"ok": False, "cod": "", "mode": "",
                        "error": "calc indisponivel para gerar COD"}
            try:
                cod = cm.gerar_cod(
                    str(cleaned.get("tipo_pacote") or ""),
                    str(cleaned.get("alimentador_principal") or ""),
                    str(cleaned.get("projeto_investimento") or ""),
                    str(cleaned.get("quantidade_material") or ""),
                    str(cleaned.get("caracteristicas_material") or ""),
                    str(cleaned.get("coordenada_fim") or ""),
                    pi_base=str(cleaned.get("pi_base") or "") or None,
                )
            except ValueError as exc:
                return {"ok": False, "cod": "", "mode": "",
                        "error": f"gerar_cod: {exc}"}
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "cod": "", "mode": "",
                        "error": f"gerar_cod: {exc}"}
            cleaned["cod"] = cod

        # Cenarios (Sprint A): redireciona save para cenario_obras_overrides
        # quando ha cenario ativo. NUNCA toca a tabela obras.
        cen_nome = self._cenario_active_name()
        if cen_nome:
            if not existing:
                return {
                    "ok": False, "cod": cod, "mode": "",
                    "error": ("Cenario ativo nao cria obras novas."
                              " Saia do cenario ('Sair do cenario' no"
                              " banner) para criar uma nova obra."),
                    "blocked": "cenario_no_create",
                    "cenario": cen_nome,
                }
            # Verifica se a obra esta no escopo do cenario
            cod_set, cen_info = self._cenario_cod_set(db, cen_nome)
            if cod and cod not in cod_set:
                return {
                    "ok": False, "cod": cod, "mode": "",
                    "error": (f"Obra {cod} nao faz parte do cenario "
                              f"'{cen_nome}'. Saia do cenario para editar."),
                    "blocked": "cenario_out_of_scope",
                    "cenario": cen_nome,
                }
            # Computa diff: para cada coluna em cleaned cujo valor
            # difere do valor atual em obras (ou da view do cenario,
            # i.e., considerando overrides anteriores), grava em
            # cenario_obras_overrides.
            ovmap = self._cenario_overrides_map(db, cen_nome)
            existing_with_overrides: dict[str, Any] = {}
            for i, c in enumerate(cols):
                base_val = existing[i] if i < len(existing) else ""
                existing_with_overrides[c] = (
                    (ovmap.get(cod) or {}).get(c, base_val)
                )
            # ano_final do CAPEX (se houver) tambem entra no baseline
            ano_final = (cen_info.get(cod) or {}).get("ano_final")
            if ano_final is not None and "ano_" in cols:
                if "ano_" not in (ovmap.get(cod) or {}):
                    existing_with_overrides["ano_"] = str(ano_final)
            diff_pairs: list[tuple[str, Any]] = []
            for col_name, new_val in cleaned.items():
                if col_name == "cod":
                    continue
                if col_name not in cols:
                    continue
                cur_val = existing_with_overrides.get(col_name, "")
                # Comparacao tolerante (string trim, case-sensitive
                # exceto whitespace).
                if str(cur_val or "").strip() != str(new_val or "").strip():
                    diff_pairs.append((col_name, new_val))
            wrote = self._cenario_save_overrides(
                db, cen_nome, cod, diff_pairs)
            return {
                "ok": True, "cod": cod, "mode": "cenario_override",
                "error": "",
                "cenario": cen_nome,
                "campos_alterados_no_cenario": [p[0] for p in diff_pairs],
                "overrides_salvos": wrote,
            }

        try:
            if existing:
                db.update_obra(cleaned, cod)
                mode = "update"
            else:
                db.insert_obra(cleaned)
                mode = "insert"
        except PermissionError as exc:
            return {"ok": False, "cod": cod, "mode": "", "error": f"permissao: {exc}"}
        except ValueError as exc:
            return {"ok": False, "cod": cod, "mode": "", "error": f"validacao: {exc}"}
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            if friendly:
                return {"ok": False, "cod": cod, "mode": "",
                        "blocked": "db_busy", "error": friendly}
            return {"ok": False, "cod": cod, "mode": "", "error": str(exc)}

        out: dict[str, Any] = {"ok": True, "cod": cod, "mode": mode, "error": ""}
        if diff_decision is not None:
            out["campos_alterados"] = list(diff_decision.campos_alterados)
            out["campos_criticos_alterados"] = list(diff_decision.campos_criticos_alterados)
        return out

    # ------------------------------------------------------------------
    # [M027] Detecao semantica de duplicada (alim+pi+ano+municipio+
    # descricao). Espelha find_duplicate_in_db do desktop. Retorna lista
    # (0 ou 1 match) para o JS abrir o modal "Obra similar encontrada".
    # ------------------------------------------------------------------
    def obras_por_codigo_semelhante(self, payload: Any) -> dict[str, Any]:
        """Procura obras semelhantes por chave semantica (definida em
        runtime/row_helpers.find_duplicate_in_db / core.repositories.
        obra_query_repo.find_duplicate). Retorna {ok, matches:list[dict]}.
        Lista pode estar vazia."""
        if not isinstance(payload, dict):
            return {"ok": False, "matches": [], "error": "payload nao eh dict"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "matches": [],
                    "error": err or "db indisponivel"}
        try:
            from runtime.row_helpers import (  # type: ignore[import-not-found]
                find_duplicate_in_db,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "matches": [],
                    "error": f"import find_duplicate_in_db: {exc}"}
        try:
            dup = find_duplicate_in_db(db, dict(payload))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "matches": [],
                    "error": f"find_duplicate: {exc}"}
        matches: list[dict[str, Any]] = []
        if dup:
            # Garante chaves uteis para o modal (cod, descricao, alim, ano).
            cod = str(dup.get("cod") or "").strip()
            matches.append({
                "cod": cod,
                "alimentador": str(dup.get("alimentador_principal")
                                   or dup.get("alimentador") or "").strip(),
                "ano": str(dup.get("ano_") or dup.get("ano") or "").strip(),
                "projeto_investimento": str(
                    dup.get("projeto_investimento") or "").strip(),
                "pi_base": str(dup.get("pi_base") or "").strip(),
                "nome_projeto": str(dup.get("nome_projeto") or "").strip(),
                "descricao_obra": str(dup.get("descricao_obra") or "").strip(),
                "municipio": str(dup.get("municipio") or "").strip(),
                # Devolve o dict original tambem para o JS poder fazer
                # merge sem perder colunas extras.
                "raw": dup,
            })
        return {"ok": True, "matches": matches, "error": ""}

    # ------------------------------------------------------------------
    # Passo 4.3 (Cadastro / gerar_cod_pep): constroi o COD da obra no
    # formato <SIGLA>-<YY>-<PI>-<ITEM> (ex.: MA-26-DI-047).
    # Importante: este "COD_PEP gerado" do mock e' o identificador da
    # obra (coluna `cod`), NAO o COD_PEP sequencial SSSS-AAA gerado pelo
    # cod_pep() do legado (que so existe pos-aprovacao).
    # ------------------------------------------------------------------
    def gerar_cod_pep(
        self,
        projeto_investimento: Any = "",
        ano: Any = "",
        item: Any = "",
        pi_base: Any = "",
    ) -> dict[str, Any]:
        # Sigla: prefere o que esta no config.json (mantem paridade com
        # desktop). Default 'MA' se nao configurado.
        cfg: dict[str, Any] = {}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception:  # noqa: BLE001
            pass
        sigla = str(cfg.get("empresa_sigla") or "MA").strip().upper()

        # YY: ultimos 2 digitos do ano informado.
        ano_s = str(ano or "").strip()
        digits = "".join(c for c in ano_s if c.isdigit())
        yy = digits[-2:] if len(digits) >= 2 else ""

        # PI: usa o pi_base passado ou deriva via legado.
        pi = str(pi_base or "").strip().upper()
        if not pi and projeto_investimento:
            try:
                from codigo5_coplan import get_pi_base  # type: ignore[import-not-found]
                pi = (get_pi_base(str(projeto_investimento), prompt_user=False) or "").upper()
            except Exception:  # noqa: BLE001
                # Fallback grosseiro: 2 primeiras letras nao-espaco do nome.
                only_letters = "".join(
                    c for c in str(projeto_investimento) if c.isalpha()
                )
                pi = only_letters[:2].upper()

        # Item: zero-pad em 3 digitos quando totalmente numerico.
        item_s = str(item or "").strip()
        if item_s.isdigit():
            item_s = item_s.zfill(3)

        parts = [sigla, yy, pi, item_s]
        complete = all(p for p in parts)
        cod = "-".join(p for p in parts if p)
        return {
            "ok": complete,
            "cod": cod,
            "sigla": sigla,
            "ano_yy": yy,
            "pi": pi,
            "item": item_s,
            "missing": [
                name for name, val in zip(
                    ("sigla", "ano", "pi", "item"), parts
                ) if not val
            ],
        }

    # ------------------------------------------------------------------
    # Apoio (planilha xlsx) -- delega 100% ao SupportFileManager do
    # desktop (codigo5_coplan.SupportFileManager.load_support_file).
    # Cacheia o resultado para nao recarregar a cada API call.
    # ------------------------------------------------------------------
    # ------------------------------------------------------------------
    # Apoio DB-backed (2026-05-07): planilha de apoio importada para
    # tabelas apoio_<aba> dentro do mesmo obras.db. Carregamento entre
    # sessoes vem direto do banco (sem reler xlsx). Botao "Atualizar
    # apoio" forca reimportacao.
    # Helpers compartilhados:
    #   _apoio_table_name(sheet)        : sheet -> nome de tabela ASCII safe
    #   _apoio_quote_ident(name)        : escapa identificador SQL com "..."
    #   _apoio_ensure_meta_table(db)    : cria apoio_meta se nao existe
    #   _apoio_meta_dict(db)            : le linha unica de apoio_meta
    #   _apoio_import_xlsx_to_db        : drop+create+insert de todas as abas
    #   _apoio_load_from_db             : reconstroi dict do _apoio_cache
    # ------------------------------------------------------------------
    # ------------------------------------------------------------------
    # Helper auxiliar para tabelas extras (apoio_* / cenario_*) sem
    # compartilhar self.conn do DatabaseManager. Cada chamada abre sua
    # propria sqlite3.Connection na thread atual (resolve race quando
    # JS dispara N bridges em paralelo - pywebview WinForms usa thread-
    # pool por bridge call, e self.conn nao e' thread-safe).
    # ------------------------------------------------------------------
    @staticmethod
    def _open_aux_conn(db: Any):
        """Abre conn sqlite3 propria (sem tocar self.conn do DM).
        Devolve (conn, error_str). conn=None se falhar."""
        import sqlite3 as _sqlite3
        try:
            db_path = str(getattr(db, "db_path", "") or "")
        except Exception:  # noqa: BLE001
            db_path = ""
        if not db_path or not os.path.isfile(db_path):
            return None, "db_path indisponivel"
        try:
            conn = _sqlite3.connect(
                db_path, timeout=5.0, check_same_thread=False)
            return conn, ""
        except Exception as exc:  # noqa: BLE001
            return None, f"sqlite3.connect: {exc}"

    @staticmethod
    def _apoio_table_name(sheet_name: Any) -> str:
        """Sanitiza nome de aba do Excel para tabela SQLite ASCII safe.
        Resultado tem prefixo 'apoio_' + lowercase + ascii + underscores."""
        import unicodedata as _ud
        s = str(sheet_name or "").strip()
        if not s:
            return "apoio__unnamed"
        s_norm = _ud.normalize("NFKD", s).encode(
            "ascii", "ignore").decode("ascii")
        s_low = s_norm.lower()
        s_clean = re.sub(r"[^a-z0-9_]+", "_", s_low).strip("_")
        if not s_clean:
            s_clean = "unnamed"
        if s_clean[:1].isdigit():
            s_clean = "_" + s_clean
        return f"apoio_{s_clean}"

    @staticmethod
    def _apoio_quote_ident(name: Any) -> str:
        """Escapa identificador SQL (tabela ou coluna) com aspas duplas.
        Permite nomes originais PT-BR (Tensao, Superintendencia, etc.)
        sem quebrar SQL. Aspas duplas internas viram duplas-duplas."""
        return '"' + str(name).replace('"', '""') + '"'

    @classmethod
    def _apoio_ensure_meta_table(cls, db: Any) -> None:
        """Cria apoio_meta se nao existe. Conn aux thread-safe."""
        conn, _err = cls._open_aux_conn(db)
        if conn is None:
            return
        try:
            cursor = conn.cursor()
            cursor.execute(
                "CREATE TABLE IF NOT EXISTS apoio_meta ("
                " id INTEGER PRIMARY KEY CHECK (id=1),"
                " last_path TEXT,"
                " last_mtime INTEGER,"
                " last_imported_at TEXT,"
                " last_user TEXT,"
                " sheet_count INTEGER,"
                " sheets_json TEXT,"
                " version TEXT"
                ")"
            )
            conn.commit()
        except Exception:  # noqa: BLE001
            pass
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass

    @classmethod
    def _apoio_meta_dict(cls, db: Any) -> dict[str, Any]:
        """Le linha unica de apoio_meta. Conn aux thread-safe."""
        conn, _err = cls._open_aux_conn(db)
        if conn is None:
            return {}
        row = None
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
                " AND name='apoio_meta'"
            )
            if not cursor.fetchone():
                return {}
            cursor.execute(
                "SELECT last_path, last_mtime, last_imported_at,"
                " last_user, sheet_count, sheets_json, version"
                " FROM apoio_meta WHERE id=1"
            )
            row = cursor.fetchone()
        except Exception:  # noqa: BLE001
            return {}
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        if not row:
            return {}
        import json as _j
        try:
            sheets = _j.loads(row[5]) if row[5] else {}
        except Exception:  # noqa: BLE001
            sheets = {}
        return {
            "last_path":        str(row[0] or ""),
            "last_mtime":       int(row[1] or 0),
            "last_imported_at": str(row[2] or ""),
            "last_user":        str(row[3] or ""),
            "sheet_count":      int(row[4] or 0),
            "sheets":           sheets,
            "version":          str(row[6] or ""),
        }

    def _apoio_import_xlsx_to_db(
        self, db: Any, xlsx_path: str,
    ) -> dict[str, Any]:
        """Le todas as abas do xlsx + cria/recria tabelas apoio_<aba>
        + atualiza apoio_meta. Usa db._with_connection() (paridade com
        outros writers do DatabaseManager) para garantir conn aberta
        durante todo o trabalho. Sem isso, db.conn pode ser None apos
        boot (quando connect() fecha a conn ao sair do _with_connection)."""
        import pandas as _pd
        import getpass as _gp
        import json as _j
        print(f"[apoio] importing xlsx: {xlsx_path}", file=sys.stderr)
        try:
            sheets = _pd.read_excel(
                xlsx_path, sheet_name=None, dtype=str)
        except Exception as exc:  # noqa: BLE001
            print(f"[apoio] read_excel falhou: {exc}", file=sys.stderr)
            return {"ok": False, "error": f"read_excel: {exc}",
                    "sheet_count": 0}
        print(f"[apoio] {len(sheets)} aba(s) lidas: "
              f"{list(sheets.keys())}", file=sys.stderr)
        sheet_map: dict[str, str] = {}
        conn, err = self._open_aux_conn(db)
        if conn is None:
            return {"ok": False, "error": err or "conn indisponivel",
                    "sheet_count": 0}
        try:
            cursor = conn.cursor()
            # Garante apoio_meta com a conn ativa.
            cursor.execute(
                "CREATE TABLE IF NOT EXISTS apoio_meta ("
                " id INTEGER PRIMARY KEY CHECK (id=1),"
                " last_path TEXT,"
                " last_mtime INTEGER,"
                " last_imported_at TEXT,"
                " last_user TEXT,"
                " sheet_count INTEGER,"
                " sheets_json TEXT,"
                " version TEXT"
                ")"
            )
            for sheet_name, df in (sheets or {}).items():
                df = df.fillna("")
                table = self._apoio_table_name(sheet_name)
                sheet_map[str(sheet_name)] = table
                cursor.execute(
                    f'DROP TABLE IF EXISTS '
                    f'{self._apoio_quote_ident(table)}'
                )
                cols = [str(c) for c in df.columns]
                if not cols:
                    cursor.execute(
                        f'CREATE TABLE '
                        f'{self._apoio_quote_ident(table)}'
                        ' (_empty TEXT)'
                    )
                    continue
                col_defs = ", ".join(
                    f"{self._apoio_quote_ident(c)} TEXT" for c in cols
                )
                cursor.execute(
                    f'CREATE TABLE '
                    f'{self._apoio_quote_ident(table)}'
                    f' ({col_defs})'
                )
                if len(df) == 0:
                    continue
                placeholders = ", ".join(["?"] * len(cols))
                col_list = ", ".join(
                    self._apoio_quote_ident(c) for c in cols
                )
                rows_to_insert = [
                    tuple("" if v is None else str(v) for v in r)
                    for r in df.values.tolist()
                ]
                cursor.executemany(
                    f'INSERT INTO '
                    f'{self._apoio_quote_ident(table)}'
                    f' ({col_list}) VALUES ({placeholders})',
                    rows_to_insert,
                )
                print(
                    f"[apoio]   {sheet_name} -> {table} "
                    f"({len(rows_to_insert)} linhas, {len(cols)} cols)",
                    file=sys.stderr,
                )
            # Atualiza apoio_meta (DELETE+INSERT id=1)
            try:
                mtime = int(os.path.getmtime(xlsx_path))
            except OSError:
                mtime = 0
            try:
                user = _gp.getuser() or "?"
            except Exception:  # noqa: BLE001
                user = "?"
            cursor.execute("DELETE FROM apoio_meta WHERE id=1")
            cursor.execute(
                "INSERT INTO apoio_meta (id, last_path, last_mtime,"
                " last_imported_at, last_user, sheet_count,"
                " sheets_json, version)"
                " VALUES (1, ?, ?, ?, ?, ?, ?, ?)",
                (
                    xlsx_path, mtime,
                    datetime.now().isoformat(timespec="seconds"),
                    user, len(sheet_map), _j.dumps(sheet_map),
                    APP_VERSION,
                ),
            )
            conn.commit()
            print(
                f"[apoio] commit ok: {len(sheet_map)} tabela(s)"
                f" criadas",
                file=sys.stderr,
            )
        except Exception as exc:  # noqa: BLE001
            print(f"[apoio] import falhou: {exc!r}", file=sys.stderr)
            return {"ok": False, "error": f"import_to_db: {exc}",
                    "sheet_count": 0}
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        return {
            "ok": True, "error": "",
            "sheet_count": len(sheet_map),
            "sheets":      sheet_map,
            "path":        xlsx_path,
        }

    def _apoio_load_from_db(self, db: Any) -> dict[str, Any]:
        """Le tabelas apoio_apoio + apoio_modulo do banco e reconstroi
        o dict no shape esperado pelo _apoio_cache. Devolve {} quando
        nao ha apoio importado (banco virgem ou tabelas faltando).
        Usa db._with_connection() para garantir conn aberta."""
        meta = self._apoio_meta_dict(db)
        if not meta or not meta.get("last_path"):
            return {}
        import pandas as _pd
        apoio_tab = self._apoio_table_name("apoio")
        modulo_tab = self._apoio_table_name("modulo")
        conn, _err = self._open_aux_conn(db)
        if conn is None:
            return {}
        try:
            df_apoio = _pd.read_sql_query(
                f'SELECT * FROM {self._apoio_quote_ident(apoio_tab)}',
                conn,
            )
            df_modulo = _pd.read_sql_query(
                f'SELECT * FROM {self._apoio_quote_ident(modulo_tab)}',
                conn,
            )
        except Exception:  # noqa: BLE001
            return {}
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        try:
            from core.services.apoio_service import (  # type: ignore[import-not-found]
                carregar_dados_apoio_from_dfs,
            )
            from core.exceptions import ApoioFileError  # type: ignore[import-not-found]
        except Exception:  # noqa: BLE001
            return {}
        try:
            dados = carregar_dados_apoio_from_dfs(df_apoio, df_modulo)
        except ApoioFileError:
            return {}
        except Exception:  # noqa: BLE001
            return {}
        # Mesmo shape que SupportFileManager.load_support_file devolve.
        dados_alim_dict = {
            nome: {
                "TENSÃO":           d.tensao,
                "REGIONAL":         d.regional,
                "SUPERINTENDÊNCIA": d.superintendencia,
                "SE":               d.se,
            }
            for nome, d in dados.dados_alimentador.items()
        }
        return {
            "alimentadores":              list(dados.alimentadores),
            "dados_alimentador":          dados_alim_dict,
            "projetos_investimento":      list(dados.projetos_investimento),
            "caracteristicas":            list(dados.caracteristicas),
            "nomes_projetos_pre_definidos": list(dados.nomes_projetos),
            "modulos": {
                chave: m.valor for chave, m in dados.modulos.items()
            },
        }

    def _load_apoio_into_manager(
        self, path: str = "", *, force_reload: bool = False,
    ) -> dict[str, Any]:
        """Hidrata self._apoio_cache. Ordem de tentativa:
            1. Cache hit em memoria (mesmo path) — retorno imediato
            2. (se nao force_reload) tabelas apoio_* no banco — DB-backed
            3. Le xlsx em ``path`` + importa para o banco + hidrata cache

        Idempotente: chamado sem ``path`` no boot tenta apenas
        passos 1 e 2. ``force_reload=True`` (botao "Atualizar apoio")
        sempre relê o xlsx + reimporta o banco.
        """
        if self._support_manager is None:
            self._ensure_managers()
        if self._support_manager is None:
            return {"ok": False, "error": "support_manager indisponivel"}

        def _shape(cache: dict[str, Any], src_path: str,
                   source: str, cached: bool) -> dict[str, Any]:
            return {
                "ok": True, "error": "", "path": src_path,
                "alimentadores":     list(cache.get("alimentadores") or []),
                "caracteristicas":   list(cache.get("caracteristicas") or []),
                "projetos_investimento": list(
                    cache.get("projetos_investimento") or []),
                "nomes_projetos_pre_definidos": list(
                    cache.get("nomes_projetos_pre_definidos") or []),
                "modulos_count":     len(cache.get("modulos") or {}),
                "source":            source,
                "cached":            cached,
            }

        # ---- 1. Cache hit em memoria ----
        if (not force_reload
                and self._apoio_path_loaded == path
                and getattr(self, "_apoio_cache", None)):
            return _shape(self._apoio_cache, path, "cache", True)

        # ---- 2. DB-backed (sem path obrigatorio) ----
        # Quando boot chama sem path, tenta hidratar do banco direto.
        if not force_reload:
            db, _err = self._ensure_db_connected()
            if db is not None:
                cached = self._apoio_load_from_db(db)
                if cached:
                    self._apoio_cache = cached
                    meta = self._apoio_meta_dict(db)
                    self._apoio_path_loaded = str(meta.get("last_path") or "")
                    self._data_state_set(
                        "apoio", "CARREGADO_VALIDADO",
                        path=self._apoio_path_loaded,
                        version_token=str(meta.get("last_mtime") or 0))
                    return _shape(cached, self._apoio_path_loaded, "db", False)

        # ---- 3. Le xlsx + importa para o banco ----
        if not path:
            self._data_state_set(
                "apoio", "INVALIDADO", path="",
                error="apoio nao importado e path nao fornecido")
            return {"ok": False,
                    "error": "apoio nao importado: forneca xlsx"}
        try:
            sucesso, dados = self._support_manager.load_support_file(path)
        except Exception as exc:  # noqa: BLE001
            self._data_state_set(
                "apoio", "INVALIDADO", path=path,
                error=f"load_support_file: {exc}")
            return {"ok": False, "error": f"load_support_file: {exc}"}
        if not sucesso:
            self._apoio_cache = {}
            self._apoio_path_loaded = ""
            self._data_state_set(
                "apoio", "INVALIDADO", path=path,
                error="planilha de apoio invalida")
            return {"ok": False, "error": "planilha de apoio invalida"}
        self._apoio_cache = dados or {}
        self._apoio_path_loaded = path
        try:
            mtime = int(os.path.getmtime(path))
        except OSError:
            mtime = 0
        self._data_state_set(
            "apoio", "CARREGADO_VALIDADO", path=path,
            version_token=str(mtime))
        # Importa para o banco (propaga erro: usuario precisa saber).
        import_err = ""
        import_sheets = 0
        try:
            db, db_err = self._ensure_db_connected()
            if db is None:
                import_err = (
                    db_err or "banco nao conectado - tabelas apoio_*"
                    " nao criadas"
                )
            else:
                imp = self._apoio_import_xlsx_to_db(db, path)
                if imp.get("ok"):
                    import_sheets = int(imp.get("sheet_count") or 0)
                else:
                    import_err = str(imp.get("error") or
                                     "import retornou not-ok")
        except Exception as exc:  # noqa: BLE001
            import_err = f"_apoio_import_xlsx_to_db: {exc}"
            print(f"[main_web] {import_err}", file=sys.stderr)
        result = _shape(self._apoio_cache, path, "xlsx", False)
        result["import_sheets"] = import_sheets
        if import_err:
            # Mantem ok=True (cache em memoria carregou) mas sinaliza
            # falha de persistencia para o JS toastar warning.
            result["import_warning"] = import_err
        return result

    def load_apoio(self, path: Any = "") -> dict[str, Any]:
        """API publica DB-backed (2026-05-07):
          * ``path`` vazio: hidrata do banco. Sem fallback xlsx.
            Se tabelas apoio_* nao existem, retorna erro pedindo
            'Atualizar apoio'.
          * ``path`` informado: importa xlsx + popula tabelas apoio_*
            + salva path no config. Esta e a UNICA rota que le xlsx
            (junto com apoio_reload_from_xlsx*)."""
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}"}
        self._ensure_managers()
        target = str(path or "").strip()
        # Sem path: SEMPRE banco (sem fallback xlsx).
        if not target:
            r0 = self._load_apoio_into_manager("")
            if r0.get("ok"):
                return r0
            return {"ok": False, "error":
                    "apoio nao importado: use 'Atualizar apoio' em"
                    " Configuracoes > Geral"}
        if not os.path.exists(target):
            return {"ok": False, "error": f"arquivo nao encontrado: {target}"}
        r = self._load_apoio_into_manager(target)
        if r.get("ok"):
            try:
                ConfigManager.save_config({"apoio": target})
                self._config = None
            except Exception as exc:  # noqa: BLE001
                print(f"[main_web] save_config(apoio) falhou: {exc}",
                      file=sys.stderr)
        return r

    # ------------------------------------------------------------------
    # Apoio DB-backed (2026-05-07): bridges para o JS consumir info
    # sobre a importacao + forcar reload sob demanda.
    # ------------------------------------------------------------------
    def apoio_meta(self) -> dict[str, Any]:
        """Devolve info da ultima importacao da planilha de apoio para
        o banco. Shape: {ok, last_path, last_imported_at, last_user,
        sheet_count, sheets, hidratado}.

        ``hidratado`` indica se o cache em memoria esta populado a
        partir das tabelas apoio_* (i.e. nao precisa abrir xlsx)."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "hidratado": bool(self._apoio_cache)}
        try:
            meta = self._apoio_meta_dict(db)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": str(exc),
                    "hidratado": bool(self._apoio_cache)}
        return {
            "ok": True, "error": "",
            "last_path":         meta.get("last_path", ""),
            "last_mtime":        meta.get("last_mtime", 0),
            "last_imported_at":  meta.get("last_imported_at", ""),
            "last_user":         meta.get("last_user", ""),
            "sheet_count":       meta.get("sheet_count", 0),
            "sheets":            meta.get("sheets", {}),
            "version":           meta.get("version", ""),
            "hidratado":         bool(self._apoio_cache),
        }

    def apoio_reload_from_xlsx(self, path: Any = "") -> dict[str, Any]:
        """Forca reimportacao do xlsx + reescrita das tabelas apoio_*.
        Path vazio: usa apoio_meta.last_path. Sincrono."""
        target = str(path or "").strip()
        if not target:
            db, _err = self._ensure_db_connected()
            if db is not None:
                meta = self._apoio_meta_dict(db)
                target = str(meta.get("last_path") or "")
        if not target:
            return {"ok": False, "error":
                    "nenhum xlsx conhecido - selecione um arquivo"}
        if not os.path.isfile(target):
            return {"ok": False,
                    "error": f"arquivo nao encontrado: {target}"}
        return self._load_apoio_into_manager(target, force_reload=True)

    def apoio_reload_from_xlsx_async(self, path: Any = "") -> dict[str, Any]:
        """Versao async (worker thread + progress, reusa Bloco 5).
        JS abre modal de progresso e polla progress_state()."""
        target = str(path or "").strip()
        if not target:
            db, _err = self._ensure_db_connected()
            if db is not None:
                meta = self._apoio_meta_dict(db)
                target = str(meta.get("last_path") or "")
        if not target:
            return {"ok": False, "started": False,
                    "error": "nenhum xlsx conhecido"}
        if not os.path.isfile(target):
            return {"ok": False, "started": False,
                    "error": f"arquivo nao encontrado: {target}"}
        with _OP_LOCK:
            if not _OP_STATE.get("finished"):
                return {"ok": False, "started": False,
                        "error": "outra operacao em andamento"}
        op_id = _op_reset(
            f"Importando apoio ({os.path.basename(target)})...")

        def _worker():
            try:
                _op_set_progress(0, 1, "Lendo xlsx...")
                res = self._load_apoio_into_manager(
                    target, force_reload=True)
                _op_set_progress(1, 1, "Concluido")
                _op_finish(result=res, error="")
            except Exception as exc:  # noqa: BLE001
                _op_finish(result=None, error=f"worker: {exc}")

        t = threading.Thread(
            target=_worker, daemon=True,
            name=f"coplan-apoio-reload-{op_id}",
        )
        t.start()
        return {"ok": True, "started": True, "op_id": op_id, "error": ""}

    def pick_and_load_apoio(self) -> dict[str, Any]:
        """Atalho: file dialog + load_apoio. Usado pelo botao
        'Procurar...' do card Empresa quando o usuario quer trocar."""
        picked = self.pick_apoio_file()
        if not picked.get("ok"):
            return picked
        return self.load_apoio(picked.get("path", ""))

    def get_apoio_state(self) -> dict[str, Any]:
        """Estado atual do apoio: caminho carregado + counts. Util pro
        JS verificar se ja tem dado disponivel."""
        cache = getattr(self, "_apoio_cache", None) or {}
        return {
            "ok": bool(cache),
            "path": getattr(self, "_apoio_path_loaded", "") or "",
            "alimentadores_count": len(cache.get("alimentadores") or []),
            "caracteristicas_count": len(cache.get("caracteristicas") or []),
            "projetos_investimento_count": len(
                cache.get("projetos_investimento") or []),
            "nomes_projetos_count": len(
                cache.get("nomes_projetos_pre_definidos") or []),
            "modulos_count": len(cache.get("modulos") or {}),
        }

    # ------------------------------------------------------------------
    # Fase A1 (core/services/atualizar_obra_service):
    # calcular_valor_obra -- usa a funcao PURA do core (sem QMessageBox).
    # Retorna {ok, valor (float), valor_formatado (str pt-BR), chave,
    #          motivos_falha, chaves_inexistentes, error}.
    # ------------------------------------------------------------------
    def _load_modulos_df(self) -> tuple[Any, str, str, str]:
        """Le tabela apoio_modulo do banco (DB-backed apoio).
        Retorna ``(df, col_chave, col_valor, error)``.
        Substitui leitura xlsx -- agora 100% via banco. Use
        'Atualizar apoio' nas Configuracoes para repopular as tabelas."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return None, "", "", err or "db indisponivel"
        # Verifica meta antes (apoio_modulo so existe se importacao rodou)
        meta = self._apoio_meta_dict(db)
        if not meta or not meta.get("last_path"):
            return None, "", "", (
                "apoio nao importado: use 'Atualizar apoio' em"
                " Configuracoes > Geral")
        modulo_tab = self._apoio_table_name("modulo")
        conn, err_open = self._open_aux_conn(db)
        if conn is None:
            return None, "", "", err_open or "conn indisponivel"
        try:
            import pandas as pd  # type: ignore[import-not-found]
            df = pd.read_sql_query(
                f'SELECT * FROM {self._apoio_quote_ident(modulo_tab)}',
                conn,
            )
        except Exception as exc:  # noqa: BLE001
            return None, "", "", f"leitura {modulo_tab}: {exc}"
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        cols_lower = {c.lower(): c for c in df.columns}
        col_chave = cols_lower.get("carac+modulo_regional", "")
        col_valor = cols_lower.get("valor_item", "")
        if not col_chave or not col_valor:
            return None, "", "", ("tabela 'apoio_modulo' sem colunas "
                                  "'CARAC+MODULO_REGIONAL'/'VALOR_ITEM'"
                                  " (reimporte o apoio)")
        return df, col_chave, col_valor, ""

    def calcular_valor_obra(
        self,
        projeto_investimento: Any = "",
        pi_base: Any = "",
        nivel_tensao: Any = "",
        caracteristicas_material: Any = "",
        nome_regional: Any = "",
        quantidade: Any = "",
        cod: Any = "",
    ) -> dict[str, Any]:
        df, col_chave, col_valor, err = self._load_modulos_df()
        if err:
            return {"ok": False, "error": err, "valor": None,
                    "valor_formatado": "", "chave": "",
                    "motivos_falha": [], "chaves_inexistentes": []}

        try:
            from core.services.atualizar_obra_service import (  # type: ignore[import-not-found]
                AtualizarObraInput, calcular_valor_obra as _core_calc,
            )
            # Fase A11: usa obter_modulos_extras direto do core (em vez
            # de passar pelo wrapper legacy get_pi_extra_module_keys).
            from core.services.pi_metadata_service import (  # type: ignore[import-not-found]
                obter_modulos_extras,
            )
            from codigo5_coplan import (  # type: ignore[import-not-found]
                REGIONAL_MAP, get_pi_base, ConfigManager,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}",
                    "valor": None, "valor_formatado": "", "chave": "",
                    "motivos_falha": [], "chaves_inexistentes": []}

        # Normaliza inputs (mesma regra do extrair_obra_input do core).
        proj = str(projeto_investimento or "").strip().upper()
        pi = str(pi_base or "").strip().upper()
        if not pi and proj:
            try:
                pi = (get_pi_base(proj, prompt_user=False) or "").strip().upper()
            except Exception:  # noqa: BLE001
                pi = proj
        tensao = str(nivel_tensao or "").strip().replace(",", ".").upper()
        carac = str(caracteristicas_material or "").strip().upper()
        regional = str(nome_regional or "").strip().upper()
        # [FIX] Normaliza virgula -> ponto pra aceitar "6,4" e "6.4".
        # O core.validar_quantidade faz float(qtd_str) direto, entao se
        # vier "6,4" da UI pt-BR ele explode com "quantidade invalida".
        qtd_raw = str(quantidade or "1").strip().replace(",", ".")

        inp = AtualizarObraInput(
            cod=str(cod or ""),
            projeto_investimento=proj,
            pi_base=pi,
            nivel_tensao=tensao,
            tensao_op=tensao,
            caracteristicas_material=carac,
            nome_regional=regional,
            quantidade_material=qtd_raw,
            obra_data_map={},
        )

        # Fase A11: extras do PI vigente via core (modulo_extra do PI +
        # ATERRAMENTO se exige_aterramento + last_pi_extra_map salvo).
        try:
            cfg = ConfigManager.load_config() or {}
            extras = list(obter_modulos_extras(pi, cfg) or [])
        except Exception:  # noqa: BLE001
            extras = []

        try:
            res = _core_calc(
                inp, df,
                col_chave=col_chave,
                col_valor=col_valor,
                regional_map=REGIONAL_MAP,
                extra_keys_for_pi=extras,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"calcular_valor_obra: {exc}",
                    "valor": None, "valor_formatado": "", "chave": "",
                    "motivos_falha": [], "chaves_inexistentes": []}

        # Convert valor formatado pt-BR -> float pra UI tambem mostrar.
        valor_num = None
        if res.valor_obra_formatado:
            try:
                valor_num = float(str(res.valor_obra_formatado).replace(",", "."))
            except (TypeError, ValueError):
                valor_num = None
        return {
            "ok": bool(res.sucesso_base),
            "error": "" if res.sucesso_base else "; ".join(res.motivos_falha),
            "valor": valor_num,
            "valor_formatado": res.valor_obra_formatado or "",
            "chave": res.chave_completa or "",
            "motivos_falha": list(res.motivos_falha),
            "chaves_inexistentes": list(res.chaves_inexistentes),
        }

    # ------------------------------------------------------------------
    # Fase A2 (atualizar_obra_service.processar_atualizacao):
    # bulk recalculo de valor_obra para um conjunto de COD (ou todas).
    # Para cada obra: extrai input via extrair_obra_input, processa via
    # processar_atualizacao (delega 100% pro core), e onde sucesso_base
    # for True, faz db.update_obra({"valor_obra": valor_formatado}).
    # Retorna o agregado com processadas_ok/falhas_total/falhas/
    # chaves_inexistentes -- mesmo formato do legado MainWindow.atualizar_obras.
    # ------------------------------------------------------------------
    def atualizar_obras_valores(self, cods: Any = None) -> dict[str, Any]:
        return self._run_atualizar_obras_valores(cods, progress_cb=None)

    # ------------------------------------------------------------------
    # atualizar_obras_valores_async (Bloco 5): versao com worker thread +
    # progress modal. JS dispara, recebe op_id, abre coplanProgress e polla
    # progress_state ate finished=True. Usado para bulk (toolbar + context
    # menu multi-cods). Per-row "Atualizar Valor" usa o sync acima (1 obra,
    # rapido).
    # ------------------------------------------------------------------
    def atualizar_obras_valores_async(
        self, cods: Any = None,
    ) -> dict[str, Any]:
        with _OP_LOCK:
            if not _OP_STATE.get("finished"):
                return {
                    "ok": False, "started": False,
                    "error": ("ja ha uma operacao em andamento: "
                              + str(_OP_STATE.get("label") or "")),
                }
        # Conta cods up-front (label da barra). Validacoes reais (cenario,
        # db, planilha apoio) ficam dentro do _run_*; se falharem, terminamos
        # a operacao imediatamente com ok=False.
        cods_list: list[str] = []
        if isinstance(cods, (list, tuple)):
            cods_list = [str(c).strip() for c in cods if str(c or "").strip()]
        n = len(cods_list) if cods_list else 0
        label = (f"Recalculando {n} obra(s)..." if n
                 else "Recalculando todas as obras...")
        op_id = _op_reset(label)

        def _worker():
            try:
                def _cb(processed: int, total: int, sub_label: str) -> bool:
                    _op_set_progress(processed, total, sub_label)
                    return _op_check_cancel()
                result = self._run_atualizar_obras_valores(
                    cods, progress_cb=_cb,
                )
                _op_finish(result=result, error="")
            except Exception as exc:  # noqa: BLE001
                _op_finish(result=None, error=f"worker: {exc}")

        t = threading.Thread(
            target=_worker, daemon=True,
            name=f"coplan-atualizar-{op_id}",
        )
        t.start()
        return {"ok": True, "started": True, "op_id": op_id, "error": ""}

    def _run_atualizar_obras_valores(
        self,
        cods: Any = None,
        *,
        progress_cb: Callable[[int, int, str], bool] | None = None,
    ) -> dict[str, Any]:
        cen_nome = self._cenario_active_name()
        if cen_nome:
            return {
                "ok": False, "error":
                (f"Operacao bloqueada: cenario '{cen_nome}' ativo."
                 f" Atualizacao em massa de valores nao e' suportada"
                 f" com cenario ativo. Saia do cenario primeiro."),
                "blocked": "cenario_active",
                "processadas_ok": 0, "falhas_total": 0,
                "atualizadas": 0, "falhas": [],
                "chaves_inexistentes": [], "total": 0,
            }
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "processadas_ok": 0, "falhas_total": 0,
                    "atualizadas": 0, "falhas": [],
                    "chaves_inexistentes": [], "total": 0}

        df, col_chave, col_valor, err_mod = self._load_modulos_df()
        if err_mod:
            return {"ok": False, "error": err_mod,
                    "processadas_ok": 0, "falhas_total": 0,
                    "atualizadas": 0, "falhas": [],
                    "chaves_inexistentes": [], "total": 0}

        try:
            from core.services.atualizar_obra_service import (  # type: ignore[import-not-found]
                _resolver_extra_keys, calcular_valor_obra, extrair_obra_input,
            )
            # Fase A11: usa obter_modulos_extras direto do core.
            from core.services.pi_metadata_service import (  # type: ignore[import-not-found]
                obter_modulos_extras,
            )
            from codigo5_coplan import (  # type: ignore[import-not-found]
                REGIONAL_MAP, get_pi_base, ConfigManager,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}",
                    "processadas_ok": 0, "falhas_total": 0,
                    "atualizadas": 0, "falhas": [],
                    "chaves_inexistentes": [], "total": 0}

        def _emit(processed: int, total: int, label: str) -> bool:
            """Reporta progresso e devolve True se o user pediu cancel."""
            if progress_cb is None:
                return False
            try:
                return bool(progress_cb(processed, total, label))
            except Exception:  # noqa: BLE001
                return False

        # Carrega config 1x para o pi_extra_module_keys_fn.
        try:
            cfg = ConfigManager.load_config() or {}
        except Exception:  # noqa: BLE001
            cfg = {}

        # Resolve a lista de COD a processar.
        cods_list: list[str] = []
        if isinstance(cods, (list, tuple)):
            cods_list = [str(c).strip() for c in cods if str(c or "").strip()]

        try:
            cols = list(db.get_column_names() or [])
            if cods_list:
                rows = db.fetch_by_cods(cods_list) or []
            else:
                rows = db.fetch_all() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"fetch: {exc}",
                    "processadas_ok": 0, "falhas_total": 0,
                    "atualizadas": 0, "falhas": [],
                    "chaves_inexistentes": [], "total": 0}

        # Constroi inputs via extrair_obra_input (mesma normalizacao do legado).
        def _pi_base_fallback(pi: str) -> str:
            try:
                return get_pi_base(pi, prompt_user=False) or pi
            except Exception:  # noqa: BLE001
                return pi

        inputs: list[Any] = []
        extracao_falhas: list[str] = []
        cod_idx = cols.index("cod") if "cod" in cols else -1
        for r in rows:
            row_cod = ""
            try:
                row_cod = str(r[cod_idx]) if cod_idx >= 0 else ""
            except Exception:  # noqa: BLE001
                row_cod = ""
            try:
                inp = extrair_obra_input(
                    r, cols, pi_base_fallback_fn=_pi_base_fallback,
                )
                inputs.append(inp)
            except Exception as exc:  # noqa: BLE001
                # Erro de extracao -- nao bloqueia; conta como falha.
                msg = f"COD={row_cod or 'N/D'}: extrair_obra_input: {exc}"
                extracao_falhas.append(msg)
                print(f"[main_web] {msg}", file=sys.stderr)

        # Snapshot do last_pi_extra_map para o log -- ajuda a detectar
        # key mismatch entre o que o Cadastro/Configuracoes salvou e o
        # que o Atualizar consulta.
        try:
            _pi_extra_map_snapshot = dict(cfg.get("last_pi_extra_map", {}) or {})
        except Exception:  # noqa: BLE001
            _pi_extra_map_snapshot = {}

        def _pi_extras(pi: str) -> list[str]:
            try:
                return list(obter_modulos_extras(pi, cfg) or [])
            except Exception:  # noqa: BLE001
                return []

        # Fase 1 (calculo): loop equivalente a processar_atualizacao, mas
        # com hooks de progresso/cancel por obra. Reproduz a mesma
        # contabilidade (falhas_max=5, set de chaves_inexistentes).
        # Fase 1 (calculo): loop equivalente a processar_atualizacao, mas
        # com hooks de progresso/cancel por obra. Reproduz a mesma
        # contabilidade (falhas_max=5, set de chaves_inexistentes).
        # Tambem coleta diagnostico por obra para o log (input values +
        # chave tentada) -- ajuda a entender por que uma obra falhou.
        total_calc = len(inputs)
        results: list[Any] = []
        processadas_ok = 0
        falhas_total = 0
        falhas: list[str] = []
        chaves_inex: set[str] = set()
        diag_failed: list[str] = []  # diagnostico por obra que falhou
        diag_all: list[str] = []  # breakdown completo por obra (success+fail)
        cancelled = False
        if _emit(0, max(total_calc, 1),
                 f"Calculando valor de {total_calc} obra(s)..."):
            cancelled = True
        if not cancelled:
            for i, inp in enumerate(inputs, 1):
                try:
                    extras = _resolver_extra_keys(
                        inp.pi_base,
                        extra_key_map={},
                        pi_extra_module_keys_fn=_pi_extras,
                    )
                    result = calcular_valor_obra(
                        inp, df,
                        col_chave=col_chave,
                        col_valor=col_valor,
                        regional_map=REGIONAL_MAP,
                        extra_keys_for_pi=extras,
                    )
                except Exception as exc:  # noqa: BLE001
                    print(f"[main_web] calcular_valor_obra cod={inp.cod}: "
                          f"{exc}", file=sys.stderr)
                    falhas_total += 1
                    if len(falhas) < 5:
                        falhas.append(f"COD={inp.cod or 'N/D'}: {exc}")
                    diag_failed.append(
                        f"COD={inp.cod or 'N/D'} EXC={exc} "
                        f"pi_base={inp.pi_base!r} "
                        f"projeto={inp.projeto_investimento!r} "
                        f"nivel_tensao={inp.nivel_tensao!r} "
                        f"carac={inp.caracteristicas_material!r} "
                        f"regional={inp.nome_regional!r} "
                        f"qtd={inp.quantidade_material!r}"
                    )
                    continue
                # Diagnostico completo (todas obras): inputs + chave +
                # extras + valor_obra_formatado. Permite comparar com o
                # que o Cadastro/Calcular usou e identificar divergencias.
                diag_all.append(
                    f"COD={inp.cod or 'N/D'} "
                    f"chave={result.chave_completa or '(falhou)'} "
                    f"extras_resolvidos={list(extras)} "
                    f"valor_calc={result.valor_obra_formatado or '(falhou)'} "
                    f"sucesso_base={result.sucesso_base} "
                    f"pi_base={inp.pi_base!r} "
                    f"projeto={inp.projeto_investimento!r} "
                    f"nivel_tensao={inp.nivel_tensao!r} "
                    f"carac={inp.caracteristicas_material!r} "
                    f"regional={inp.nome_regional!r} "
                    f"qtd={inp.quantidade_material!r}"
                )
                results.append(result)
                if result.sucesso_base:
                    processadas_ok += 1
                else:
                    # Falha bloqueante (regional/chave/qtd invalida etc.):
                    # registra diagnostico completo para o log.
                    diag_failed.append(
                        f"COD={inp.cod or 'N/D'} "
                        f"chave_tentada={result.chave_completa or '(nao montou)'} "
                        f"motivos={list(result.motivos_falha)} "
                        f"pi_base={inp.pi_base!r} "
                        f"projeto={inp.projeto_investimento!r} "
                        f"nivel_tensao={inp.nivel_tensao!r} "
                        f"carac={inp.caracteristicas_material!r} "
                        f"regional={inp.nome_regional!r} "
                        f"qtd={inp.quantidade_material!r} "
                        f"extras={list(extras)}"
                    )
                for motivo in result.motivos_falha:
                    falhas_total += 1
                    if len(falhas) < 5:
                        falhas.append(
                            f"COD={result.cod or 'N/D'}: {motivo}")
                for ch in result.chaves_inexistentes:
                    chaves_inex.add(ch)
                if (i % 10) == 0 or i == total_calc:
                    if _emit(
                        i, total_calc,
                        f"Calculando valor... ({i}/{total_calc})",
                    ):
                        cancelled = True
                        break

        # Fase 2 (escrita): UPDATE no banco para cada result valido. A
        # descricao_obra fica fora -- depende de dialog Y/N do legado.
        to_write = [
            r for r in results
            if r.sucesso_base and r.valor_obra_formatado and str(r.cod or "").strip()
        ]
        # Index inputs por COD para conseguir ler o valor_obra atual
        # do banco antes de decidir sobrescrever (regra de preservacao
        # abaixo).
        input_by_cod = {}
        for _inp in inputs:
            _c = str(getattr(_inp, "cod", "") or "").strip()
            if _c:
                input_by_cod[_c] = _inp
        total_write = len(to_write)
        atualizadas = 0
        preservadas = 0
        preservadas_msgs: list[str] = []
        update_falhas: list[str] = []
        busy_msg = ""
        if not cancelled and total_write > 0:
            if _emit(0, total_write,
                     f"Salvando {total_write} obra(s) no banco..."):
                cancelled = True
        if not cancelled:
            for i, r_obra in enumerate(to_write, 1):
                cod_s = str(r_obra.cod or "").strip()

                # Regra de preservacao: se a calculo teve motivos_falha
                # (chave extra ausente OU valor invalido em alguma extra)
                # e a obra ja tem valor_obra gravado no DB, NAO sobrescreve
                # com o calculo parcial. Loga preservacao. Quando obra NAO
                # tem valor_obra, segue update normal (nao ha o que perder).
                motivos_problema = list(r_obra.motivos_falha or [])
                if motivos_problema:
                    inp_cur = input_by_cod.get(cod_s)
                    existing_val = ""
                    if inp_cur is not None:
                        try:
                            existing_val = str(
                                inp_cur.obra_data_map.get("valor_obra", "")
                                or ""
                            ).strip()
                        except Exception:  # noqa: BLE001
                            existing_val = ""
                    if existing_val:
                        msg = (
                            f"COD={cod_s}: valor preservado "
                            f"(DB={existing_val}, calc_parcial="
                            f"{r_obra.valor_obra_formatado}) - motivos: "
                            f"{'; '.join(motivos_problema)}"
                        )
                        preservadas += 1
                        preservadas_msgs.append(msg)
                        print(f"[main_web] {msg}", file=sys.stderr)
                        if (i % 5) == 0 or i == total_write:
                            if _emit(
                                i, total_write,
                                f"Salvando... ({i}/{total_write})",
                            ):
                                cancelled = True
                                break
                        continue

                try:
                    db.update_obra(
                        {"valor_obra": r_obra.valor_obra_formatado},
                        cod_s, skip_blank=True,
                    )
                    atualizadas += 1
                except Exception as exc:  # noqa: BLE001
                    friendly = self._friendly_busy_error(exc)
                    if friendly:
                        # Banco ocupado: aborta o loop (outros usuarios
                        # estao escrevendo). Mensagem amigavel pro JS.
                        busy_msg = friendly
                        update_falhas.append(f"COD={cod_s}: {friendly}")
                        print(
                            f"[main_web] db busy em "
                            f"atualizar_obras_valores: {friendly}",
                            file=sys.stderr,
                        )
                        break
                    msg = f"COD={cod_s}: update_obra: {exc}"
                    update_falhas.append(msg)
                    print(f"[main_web] {msg}", file=sys.stderr)
                if (i % 5) == 0 or i == total_write:
                    if _emit(
                        i, total_write,
                        f"Salvando... ({i}/{total_write})",
                    ):
                        cancelled = True
                        break

        falhas_full = list(falhas) + update_falhas + extracao_falhas
        falhas_total += len(extracao_falhas)
        ok_flag = not busy_msg and not cancelled
        err_str = busy_msg or ("cancelado" if cancelled else "")
        out: dict[str, Any] = {
            "ok": ok_flag, "error": err_str,
            "total": len(inputs),
            "processadas_ok": processadas_ok,
            "atualizadas": atualizadas,
            "preservadas": preservadas,
            "preservadas_msgs": preservadas_msgs,
            "falhas_total": falhas_total + len(update_falhas),
            "falhas": falhas_full,
            "chaves_inexistentes": sorted(chaves_inex),
            "diagnostico": diag_failed,
            "diagnostico_todas": diag_all,
            "last_pi_extra_map": _pi_extra_map_snapshot,
        }
        if busy_msg:
            out["blocked"] = "db_busy"
        if cancelled:
            out["cancelled"] = True
        # Auto-log em <HERE>/logs SEMPRE que processar 1+ obra (sucesso
        # ou erro). Mesmo "1 atualizada(s)" pode estar com valor errado
        # se algum extra nao foi resolvido -- o breakdown do log mostra
        # exatamente quais extras foram aplicados para CADA obra.
        if len(inputs) > 0 or self._result_has_errors(out) or preservadas > 0:
            out["log_path"] = self._write_op_log(
                "atualizar", out,
                meta={
                    "cods_solicitados": (len(cods_list) if cods_list
                                         else "todos"),
                    "preservadas": preservadas,
                },
            )
        return out

    def get_alimentador_details(self, alim: Any) -> dict[str, Any]:
        """Retorna metadados de UM alimentador (TENSAO, REGIONAL,
        SUPERINTENDENCIA, SE) -- replica apoio_mixin.alimentador_selecionado
        do desktop. Vazio se nao houver apoio carregado ou alim nao
        constar nele."""
        a = str(alim or "").strip()
        if not a:
            return {"ok": False, "error": "alim vazio"}
        cache = getattr(self, "_apoio_cache", None) or {}
        dados_alim = cache.get("dados_alimentador") or {}
        # Tenta exato; depois case-insensitive.
        if a in dados_alim:
            d = dados_alim[a]
        else:
            up = a.upper()
            d = None
            for k, v in dados_alim.items():
                if str(k).upper() == up:
                    d = v
                    break
        if not d:
            return {"ok": False, "error": f"alimentador '{a}' nao encontrado no apoio"}
        return {
            "ok": True, "error": "",
            "alim": a,
            "tensao":           str(d.get("TENSÃO") or d.get("TENSAO") or "").strip(),
            "regional":         str(d.get("REGIONAL") or "").strip(),
            "superintendencia": str(d.get("SUPERINTENDÊNCIA")
                                    or d.get("SUPERINTENDENCIA") or "").strip(),
            "se":               str(d.get("SE") or "").strip(),
        }

    # ------------------------------------------------------------------
    # Passo 4.4 (Cadastro / list_alimentadores): catalogo de alimentadores
    # ja cadastrados (alimentador_principal + alimentadores_beneficiados
    # parseados). Usado pelo <select> "Alimentador Beneficiado" do form.
    # Retorna lista unica ordenada alfabeticamente.
    # ------------------------------------------------------------------
    def list_alimentadores(self) -> dict[str, Any]:
        # Importante: o usuario precisa dos alimentadores do APOIO mesmo
        # quando o DB esta vazio ou inacessivel (caso do primeiro setup).
        # Estrategia: tenta DB primeiro mas trata falha como WARN (nao
        # zera resposta) -- sempre mescla com self._apoio_cache no final.
        unique: set[str] = set()
        warnings: list[str] = []

        # 1. DB (alimentadores ja cadastrados em obras).
        db, err = self._ensure_db_connected()
        if err or db is None:
            warnings.append(f"db: {err or 'indisponivel'}")
        else:
            # _get_cursor() retorna None apos _with_connection fechar.
            # Usamos open_sqlite_safe direto + fechamos manualmente, em
            # vez de db._get_cursor() (que so funciona dentro de
            # _with_connection). Replica o que outras APIs fazem (ex.:
            # search_obras) usando cursor fresco.
            try:
                from runtime.database import open_sqlite_safe  # noqa: PLC0415
                db_path = (self._config or {}).get("obras") or ""
                if db_path and os.path.isfile(str(db_path)):
                    conn = open_sqlite_safe(str(db_path))
                    try:
                        cur = conn.cursor()
                        cur.execute(
                            "SELECT alimentador_principal, "
                            "alimentadores_beneficiados FROM obras"
                        )
                        for row in cur.fetchall():
                            principal = str(row[0] or "").strip()
                            if principal:
                                unique.add(principal.upper())
                            benef_raw = str(row[1] or "")
                            for piece in re.split(r"[;,|]+", benef_raw):
                                p = piece.strip().upper()
                                if p:
                                    unique.add(p)
                    finally:
                        try:
                            conn.close()
                        except Exception:  # noqa: BLE001
                            pass
                else:
                    warnings.append("db: path invalido")
            except Exception as exc:  # noqa: BLE001
                warnings.append(f"sql: {exc}")

        # 2. Cache do apoio (DB-backed): mais completo que o banco de
        # obras pois inclui alim ainda nao cadastrados. Hidrata via
        # _apoio_cache_dict (le tabela apoio_apoio se cache vazio).
        cache = self._apoio_cache_dict()
        for a in (cache.get("alimentadores") or []):
            s = str(a or "").strip().upper()
            if s:
                unique.add(s)

        items = sorted(unique)
        # Tambem extrai prefixo (SE candidata) de cada alimentador para o
        # JS poder oferecer agrupamento ou validacao.
        ses: list[str] = []
        seen_se: set[str] = set()
        for a in items:
            pref = re.split(r"[-_/]", a, 1)[0].strip().upper()
            if pref and pref not in seen_se:
                seen_se.add(pref)
                ses.append(pref)
        # ok=True sempre que tivermos itens; warnings mostra de onde
        # veio a falha parcial (DB ou apoio) -- nao bloqueia o JS de
        # popular o combo.
        return {
            "ok": bool(items),
            "items": items,
            "ses": ses,
            "warnings": warnings,
            "error": ("; ".join(warnings) if not items and warnings else ""),
        }

    # ------------------------------------------------------------------
    # Passo 4.5 (Cadastro / metadata para selects):
    #   * get_pi_options()   -> projetos de investimento + bases curtas
    #   * get_regionais()    -> chaves de REGIONAL_MAP + extras do config
    #   * get_pacotes()      -> distinct(tipo_pacote) do banco + defaults
    # Tudo em uma so chamada agregada (`get_form_metadata`) para o JS
    # popular os 3 selects de uma vez.
    # ------------------------------------------------------------------
    def get_pi_options(self) -> dict[str, Any]:
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager,
                PI_BASE_CUSTOM,
            )
            from core.services.pi_metadata_service import (  # type: ignore[import-not-found]
                listar_todas_bases,
            )
            cfg = ConfigManager.load_config() or {}
            bases = listar_todas_bases(cfg, custom_bases=tuple(PI_BASE_CUSTOM))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "items": [], "error": f"pi_metadata: {exc}"}
        # Tambem busca os nomes longos ja usados em obras (projeto_investimento)
        # para que o select reflita o que ja existe na base do cliente.
        long_names: list[str] = []
        seen_long: set[str] = set()
        # 1) Apoio (replica o que o desktop popula em
        #    field_projeto_investimento.addItems(dados['projetos_investimento'])).
        cache = getattr(self, "_apoio_cache", None) or {}
        for v in (cache.get("projetos_investimento") or []):
            s = str(v or "").strip()
            k = s.upper()
            if s and k not in seen_long:
                seen_long.add(k)
                long_names.append(s)
        # 2) Banco (DISTINCT projeto_investimento das obras existentes).
        try:
            db, err = self._ensure_db_connected()
            if not err and db is not None:
                cursor = db._get_cursor()
                if cursor is not None:
                    cursor.execute(
                        "SELECT DISTINCT projeto_investimento FROM obras "
                        "WHERE projeto_investimento IS NOT NULL "
                        "AND TRIM(projeto_investimento) <> ''"
                    )
                    for row in cursor.fetchall():
                        v = str(row[0] or "").strip()
                        key = v.upper()
                        if v and key not in seen_long:
                            seen_long.add(key)
                            long_names.append(v)
                    long_names.sort()
        except Exception:  # noqa: BLE001
            pass
        return {
            "ok": True,
            "bases": list(bases),
            "long_names": long_names,
            "error": "",
        }

    def get_regionais(self) -> dict[str, Any]:
        try:
            from codigo5_coplan import REGIONAL_MAP, ConfigManager  # type: ignore[import-not-found]
            base = list(REGIONAL_MAP.keys())
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "items": [], "error": f"REGIONAL_MAP: {exc}"}
        # Adiciona regionais extras configuradas no config.json
        # (regional_map = dict { nome: codigo }).
        try:
            cfg = ConfigManager.load_config() or {}
            extras = list((cfg.get("regional_map") or {}).keys())
            for e in extras:
                if e and e not in base:
                    base.append(e)
        except Exception:  # noqa: BLE001
            pass
        return {"ok": True, "items": sorted(set(base)), "error": ""}

    def get_pacotes(self) -> dict[str, Any]:
        # Defaults vindos do mock + nomes vistos no banco. Mantidos
        # apenas em alfabetico para nao impor ordenacao opinativa.
        defaults = [
            "Mercado", "Confiabilidade", "Interligação UDE",
            "Interligação de UDE", "Solicitação Regional",
            "Orçamento de Conexão", "PLPT",
        ]
        seen_keys: set[str] = set()
        out: list[str] = []
        for v in defaults:
            k = v.strip().upper()
            if k and k not in seen_keys:
                seen_keys.add(k)
                out.append(v.strip())
        try:
            db, err = self._ensure_db_connected()
            if not err and db is not None:
                cursor = db._get_cursor()
                if cursor is not None:
                    cursor.execute(
                        "SELECT DISTINCT tipo_pacote FROM obras "
                        "WHERE tipo_pacote IS NOT NULL "
                        "AND TRIM(tipo_pacote) <> ''"
                    )
                    for row in cursor.fetchall():
                        v = str(row[0] or "").strip()
                        k = v.upper()
                        if v and k not in seen_keys:
                            seen_keys.add(k)
                            out.append(v)
        except Exception:  # noqa: BLE001
            pass
        out.sort()
        return {"ok": True, "items": out, "error": ""}

    def get_form_metadata(self) -> dict[str, Any]:
        """Agregador para o JS popular todos os selects do Cadastro de
        uma vez. Inclui alimentadores (banco+apoio) e caracteristicas
        (apoio) para que os selects nao fiquem so com a opcao do mock."""
        # Alimentadores: list_alimentadores ja mescla banco + apoio.
        alim = self.list_alimentadores()
        # Caracteristicas vem do apoio cache (carregadas pelo
        # SupportFileManager.load_support_file).
        cache = getattr(self, "_apoio_cache", None) or {}
        carac_items = sorted({
            str(c).strip() for c in (cache.get("caracteristicas") or [])
            if str(c or "").strip()
        })
        return {
            "pi": self.get_pi_options(),
            "regionais": self.get_regionais(),
            "pacotes": self.get_pacotes(),
            "alimentadores": {
                "ok": bool(alim.get("ok")),
                "items": list(alim.get("items") or []),
            },
            "caracteristicas": {
                "ok": bool(carac_items),
                "items": carac_items,
            },
        }

    # ------------------------------------------------------------------
    # Cadastro / metodos auxiliares (leva 2 da migracao desktop->web).
    # Origem: ui/main_window/cadastro_mixin.py + apoio_mixin.py.
    # M020 cadastro_form_metadata, M021 caracteristicas_por_alimentador,
    # M024 validar_cadastro, M025 resolver_pi_base,
    # M026 nome_projeto_options, M029 tecnico_snapshot.
    # ------------------------------------------------------------------

    def cadastro_form_metadata(self) -> dict[str, Any]:
        """[M020] Agregador especifico do cadastro: estende
        get_form_metadata com listas hardcoded do desktop
        (manobra/aprovada/novo_bay/criticidade), o range de Ano
        (current..+10) e a lista de nomes_projeto (banco + apoio +
        'Melhorias AL'). Reaproveita get_form_metadata para os combos
        comuns (PI, regionais, pacotes, alimentadores, caracteristicas)."""
        from datetime import datetime as _dt
        ano_atual = _dt.now().year
        meta = self.get_form_metadata()
        nomes_proj = self.nome_projeto_options()
        return {
            "ok": True, "error": "",
            "ano_range":     [str(y) for y in range(ano_atual, ano_atual + 11)],
            "pi":            meta.get("pi") or {"ok": False, "bases": [], "long_names": []},
            "regionais":     meta.get("regionais") or {"ok": False, "items": []},
            "pacotes":       meta.get("pacotes") or {"ok": False, "items": [
                "Mercado", "Confiabilidade", "Interligação de UDE",
                "Solicitação Regional", "Orçamento de Conexao", "PLPT",
            ]},
            "alimentadores":   meta.get("alimentadores") or {"ok": False, "items": []},
            "caracteristicas": meta.get("caracteristicas") or {"ok": False, "items": []},
            "manobra":         ["SIM", "NÃO"],
            "aprovada":        ["NÃO", "SIM"],
            "novo_bay":        ["NÃO", "SIM"],
            "criticidade":     ["Baixa", "Média", "Alta"],
            "nomes_projeto":   nomes_proj.get("items") or ["Melhorias AL"],
        }

    def caracteristicas_por_alimentador(self, alim: Any) -> dict[str, Any]:
        """[M021] Lista CARACTERISTICAS de um alimentador especifico,
        replicando apoio_mixin.alimentador_selecionado (recarrega o combo
        de Caracteristicas conforme escolha). Cai pra lista geral do
        apoio quando o alim nao consta em dados_alimentador."""
        a = str(alim or "").strip()
        if not a:
            return {"ok": False, "items": [], "error": "alim vazio"}
        cache = getattr(self, "_apoio_cache", None) or {}
        dados_alim = cache.get("dados_alimentador") or {}
        d = dados_alim.get(a)
        if d is None:
            up = a.upper()
            for k, v in dados_alim.items():
                if str(k).upper() == up:
                    d = v
                    break
        if not d:
            geral = sorted({
                str(c).strip() for c in (cache.get("caracteristicas") or [])
                if str(c or "").strip()
            })
            return {
                "ok": bool(geral), "items": geral,
                "error": (f"alimentador '{a}' nao encontrado no apoio"
                          " (mostrando lista geral)"),
            }
        raw = d.get("CARACTERÍSTICAS") or d.get("CARACTERISTICAS") or []
        items: list[str] = []
        if isinstance(raw, str):
            items = [s.strip() for s in raw.split(";") if s.strip()]
        elif isinstance(raw, dict):
            items = [str(k).strip() for k in raw.keys() if str(k or "").strip()]
        elif isinstance(raw, (list, tuple, set)):
            items = [str(c).strip() for c in raw if str(c or "").strip()]
        items = sorted(set(items))
        return {"ok": True, "items": items, "error": ""}

    def validar_cadastro(self, payload: Any) -> dict[str, Any]:
        """[M024] Espelho server-side de
        cadastro_mixin.validar_campos_obrigatorios. Aplica
        [RB-DISTRIBUICAO]: 'nome_projeto' so eh obrigatorio quando
        projeto_investimento normalizado for DISTRIBUICAO ou
        DISTRIBUICAO LD 34,5 KV. Tambem reporta avisos: alimentadores
        com '_' (validacao do desktop) e nome_projeto comecando com
        'Obra' (label informativa do desktop)."""
        if not isinstance(payload, dict):
            payload = {}

        def _v(key: str) -> str:
            return str(payload.get(key) or "").strip()

        obrig = {
            "Ano":                       _v("ano_") or _v("ano"),
            "Projeto de Investimento":   _v("projeto_investimento"),
            "Alimentador Obra":          _v("alimentador_principal"),
            "Quantidade":                _v("quantidade_material") or _v("quantidade"),
            "Coordenadas Para":          _v("coordenada_fim"),
            "Pacote":                    _v("tipo_pacote") or _v("pacote"),
            "Caracteristicas":           _v("caracteristicas_material") or _v("caracteristicas"),
            "Manobra":                   _v("manobra"),
        }

        pi_raw = _v("projeto_investimento")
        pi_base_raw = _v("pi_base")
        try:
            from codigo5_coplan import normalize_key  # type: ignore[import-not-found]
        except Exception:  # noqa: BLE001
            def normalize_key(s: str) -> str:  # type: ignore[no-redef]
                return (s or "").strip().upper()
        valores_distribuicao = {"DISTRIBUICAO", "DISTRIBUICAO LD 34,5 KV"}
        if (
            normalize_key(pi_raw) in valores_distribuicao
            or normalize_key(pi_base_raw) in valores_distribuicao
        ):
            obrig["Projeto"] = _v("nome_projeto")

        faltantes = [nome for nome, valor in obrig.items() if not valor]

        avisos: list[str] = []
        alim_principal = _v("alimentador_principal")
        if "_" in alim_principal:
            avisos.append(f"alimentador '{alim_principal}' contem '_'")
        benef_raw = payload.get("alimentadores_beneficiados") or ""
        if isinstance(benef_raw, list):
            benef_list = [str(x).strip() for x in benef_raw if str(x or "").strip()]
        else:
            benef_list = [s.strip() for s in str(benef_raw).split(";")
                          if s.strip()]
        for b in benef_list:
            if "_" in b:
                avisos.append(f"alimentador beneficiado '{b}' contem '_'")
        nome_proj = _v("nome_projeto")
        if nome_proj.lower().startswith("obra"):
            avisos.append("nome_projeto inicia com 'Obra' (informativo)")

        return {
            "ok": (not faltantes) and (not avisos),
            "faltantes": faltantes,
            "avisos": avisos,
            "error": "",
        }

    def resolver_pi_base(self, pi: Any) -> dict[str, Any]:
        """[M025] Devolve {pi_base, conhecido} sem prompt server-side. Se
        nao for conhecido, JS deve abrir prompt local e persistir via
        save_pi_base_map(). Reflete get_pi_base(prompt_user=False)."""
        pi_s = str(pi or "").strip()
        if not pi_s:
            return {"ok": False, "pi_base": "", "conhecido": False,
                    "error": "pi vazio"}
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                get_pi_base,
                _is_pi_base_known,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "pi_base": "", "conhecido": False,
                    "error": f"import: {exc}"}
        try:
            conhecido = bool(_is_pi_base_known(pi_s))
        except Exception:  # noqa: BLE001
            conhecido = False
        try:
            base = str(get_pi_base(pi_s, prompt_user=False) or "").strip()
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "pi_base": "", "conhecido": conhecido,
                    "error": f"get_pi_base: {exc}"}
        return {"ok": True, "pi_base": base, "conhecido": conhecido,
                "error": ""}

    def nome_projeto_options(self) -> dict[str, Any]:
        """[M026] Lista de nomes_projeto para o combo, ordenada e dedup
        case-insensitive, com 'Melhorias AL' no fim. Reaproveita
        list_projetos (DISTINCT do banco) + apoio. Replica
        apoio_mixin.populate_combo_nome_projeto."""
        cache = getattr(self, "_apoio_cache", None) or {}
        pre = cache.get("nomes_projetos_pre_definidos") or []
        try:
            db_lst = self.list_projetos()
            nomes_db = list(db_lst.get("items") or [])
        except Exception:  # noqa: BLE001
            nomes_db = []
        seen: dict[str, str] = {}
        for src in (pre, nomes_db):
            for nome in src:
                key = str(nome or "").strip().upper()
                if not key:
                    continue
                if key not in seen:
                    seen[key] = str(nome).strip()
        items = sorted(seen.values(), key=lambda s: s.upper())
        if not any(v.upper() == "MELHORIAS AL" for v in items):
            items.append("Melhorias AL")
        return {"ok": True, "items": items, "error": ""}

    # ------------------------------------------------------------------
    # Bloco 4 (Auditoria #41 + #42): Snapshot tecnico real.
    # Port direto de tecnico_snapshot_mixin.py (desktop):
    #   - _compute_file_token(path)        - sha1 de path|mtime|size
    #   - _compute_folder_token(folder, [required_files]) - sha1 do dir
    #     incluindo cada arquivo esperado (mtime|size ou "missing")
    #   - tecnico_snapshot()               - token completo
    #     (db + apoio + ganhos + tecnico_paths)
    #   - tecnico_check_dirty()            - compara token atual vs
    #     config['tecnico_last_token']; marca dirty automaticamente
    #     quando muda (RB-1.1 do desktop, fallback simples).
    # ------------------------------------------------------------------
    @staticmethod
    def _compute_file_token(path: str) -> str:
        """Hash sha1 baseado em path/mtime/tamanho. Vazio em erro."""
        try:
            info = os.stat(path)
            raw = f"{os.path.abspath(path)}|{info.st_mtime}|{info.st_size}"
            return hashlib.sha1(raw.encode("utf-8")).hexdigest()
        except Exception:  # noqa: BLE001
            return ""

    @staticmethod
    def _compute_folder_token(folder: str, required: list[str]) -> str:
        """Hash sha1 baseado no folder + cada arquivo esperado (mtime|size
        ou 'missing'). Paridade com _compute_folder_token desktop."""
        parts = [os.path.abspath(folder)]
        for name in required:
            path = os.path.join(folder, name)
            try:
                info = os.stat(path)
                parts.append(f"{name}:{info.st_mtime}:{info.st_size}")
            except Exception:  # noqa: BLE001
                parts.append(f"{name}:missing")
        return hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()

    def _tecnico_paths(self) -> dict[str, str]:
        """Resolve os 4 paths usados no token tecnico:
        db, apoio, ganhos (folder), tecnico_txt (folder, falls back
        para ganhos)."""
        cfg = self._config or {}
        db_path = ""
        try:
            db = getattr(self, "_db_manager", None)
            if db is not None:
                db_path = str(getattr(db, "db_path", "") or "")
        except Exception:  # noqa: BLE001
            db_path = ""
        if not db_path:
            db_path = str(cfg.get("obras") or "")
        apoio_path = str(cfg.get("apoio") or "")
        ganhos_path = str(cfg.get("caminho_pasta_ganhos") or "")
        # tecnico_txt nao tem path proprio no web -> usa pasta de ganhos.
        tecnico_path = ganhos_path
        return {
            "db":      db_path,
            "apoio":   apoio_path,
            "ganhos":  ganhos_path,
            "tecnico": tecnico_path,
        }

    def tecnico_snapshot(self) -> dict[str, Any]:
        """Snapshot tecnico completo (token + timestamp + src descritivo).

        Token = sha1 de db|apoio|ganhos|tecnico (cada um com seu hash
        proprio via _compute_file_token / _compute_folder_token).
        Paridade com _compute_tecnico_snapshot_token do desktop."""
        try:
            from runtime.config import (  # type: ignore[import-not-found]
                TECNICO_REQUIRED_FILES,
            )
        except Exception:  # noqa: BLE001
            TECNICO_REQUIRED_FILES = [
                "FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT",
            ]
        paths = self._tecnico_paths()
        parts: list[str] = []
        if paths["db"]:
            parts.append(f"db:{self._compute_file_token(paths['db'])}")
        if paths["apoio"]:
            parts.append(f"apoio:{self._compute_file_token(paths['apoio'])}")
        if paths["ganhos"]:
            parts.append(
                f"ganhos:{self._compute_folder_token(paths['ganhos'], TECNICO_REQUIRED_FILES)}"
            )
        if paths["tecnico"] and paths["tecnico"] != paths["ganhos"]:
            parts.append(
                f"txt:{self._compute_folder_token(paths['tecnico'], TECNICO_REQUIRED_FILES)}"
            )
        raw = "|".join(parts)
        token = hashlib.sha1(raw.encode("utf-8")).hexdigest() if raw else ""

        # src descritivo (paridade _get_tecnico_snapshot_source)
        src_parts: list[str] = []
        if paths["apoio"]:
            src_parts.append(f"Apoio:{os.path.basename(paths['apoio'])}")
        if paths["ganhos"]:
            src_parts.append(f"Ganhos:{os.path.basename(paths['ganhos'])}")
        if paths["tecnico"] and paths["tecnico"] != paths["ganhos"]:
            src_parts.append(f"TXT:{os.path.basename(paths['tecnico'])}")
        src = " | ".join(src_parts) if src_parts else "N/D"

        ts = datetime.now().strftime("%d/%m/%y %H:%M")
        # Sempre devolve count atual (UI usa pra render pill warn).
        dirty_count = 0
        try:
            db, _err = self._ensure_db_connected()
            if db is not None:
                dirty_count = int(db.count_tecnico_dirty() or 0)
        except Exception:  # noqa: BLE001
            dirty_count = 0
        return {
            "ok": True,
            "token": token,
            "ts": ts,
            "src": src,
            "paths": paths,
            "dirty_count": dirty_count,
            # Mantem chave 'tecnico_dirty' por compat com consumers
            # antigos do M029 STUB (cadastro_mixin testa esse campo).
            "tecnico_dirty": "SIM" if dirty_count > 0 else "NÃO",
            "error": "",
        }

    def tecnico_check_dirty(self) -> dict[str, Any]:
        """Compara token tecnico atual vs config['tecnico_last_token'].
        Quando diferente E ha obras no banco, marca TODAS como
        tecnico_dirty='SIM' (fallback simples do _apply_tecnico_token_change_db
        desktop, que tem logica de escopo mais complexa).
        Persiste o novo token no config para a proxima checagem.

        Retorna {ok, token_changed, dirty_applied, count, token, error}.
        JS chama no boot e em coplan:state events para detectar mudancas
        nas fontes (db/apoio/ganhos)."""
        snap = self.tecnico_snapshot()
        token_now = str(snap.get("token") or "")
        if not token_now:
            return {
                "ok": True,
                "token_changed": False,
                "dirty_applied": False,
                "count": int(snap.get("dirty_count") or 0),
                "token": "",
                "error": "",
            }
        # Le ultimo token salvo
        last_token = ""
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager,
            )
            cfg = ConfigManager.load_config() or {}
            last_token = str(cfg.get("tecnico_last_token") or "")
        except Exception:  # noqa: BLE001
            cfg = {}
            last_token = ""
        token_changed = bool(last_token) and (last_token != token_now)
        dirty_applied = False
        # Se token mudou (e nao e' a primeira execucao), marca dirty
        db, err = self._ensure_db_connected()
        if token_changed and db is not None:
            try:
                # Conta obras antes - so marca se houver
                count_obras = 0
                try:
                    cursor = db._get_cursor()
                    if cursor is not None:
                        cursor.execute("SELECT COUNT(*) FROM obras")
                        row = cursor.fetchone()
                        count_obras = int(row[0]) if row else 0
                except Exception:  # noqa: BLE001
                    count_obras = 0
                if count_obras > 0:
                    db.mark_tecnico_dirty_all()
                    dirty_applied = True
            except Exception as exc:  # noqa: BLE001
                return {
                    "ok": False,
                    "token_changed": True,
                    "dirty_applied": False,
                    "count": int(snap.get("dirty_count") or 0),
                    "token": token_now,
                    "error": f"mark_dirty: {exc}",
                }
        # Persiste novo token (mesmo se nao houve mudanca - garante
        # que o "primeiro contato" salve o baseline).
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager,
            )
            ConfigManager.save_config({"tecnico_last_token": token_now})
            self._config = None
        except Exception:  # noqa: BLE001
            pass
        # Recount apos eventual mark_dirty_all
        count_after = int(snap.get("dirty_count") or 0)
        if dirty_applied and db is not None:
            try:
                count_after = int(db.count_tecnico_dirty() or 0)
            except Exception:  # noqa: BLE001
                pass
        return {
            "ok": True,
            "token_changed": token_changed,
            "dirty_applied": dirty_applied,
            "count": count_after,
            "token": token_now,
            "first_seen": (not last_token),
            "error": "",
        }

    # ------------------------------------------------------------------
    # Passo 5.1 (Ganhos / pasta de arquivos): le caminho_pasta_ganhos do
    # config + (opcional) sub-pasta do alimentador, lista arquivos xlsx/csv
    # e expoe pick_ganhos_folder() para abrir o file dialog do pywebview.
    # ------------------------------------------------------------------
    GANHOS_EXTS = (".xlsx", ".xlsm", ".xls", ".csv", ".txt")

    @classmethod
    def _list_ganhos_in(cls, folder: str) -> list[dict[str, Any]]:
        """Lista arquivos com extensoes relevantes em ``folder``."""
        items: list[dict[str, Any]] = []
        if not folder or not os.path.isdir(folder):
            return items
        try:
            for name in sorted(os.listdir(folder)):
                full = os.path.join(folder, name)
                if not os.path.isfile(full):
                    continue
                if not name.lower().endswith(cls.GANHOS_EXTS):
                    continue
                try:
                    st = os.stat(full)
                    size = int(st.st_size)
                    mtime = float(st.st_mtime)
                except OSError:
                    size, mtime = 0, 0.0
                items.append({
                    "name": name,
                    "path": full,
                    "size": size,
                    "mtime": mtime,
                })
        except OSError:
            pass
        return items

    @staticmethod
    def _resolve_ganhos_folder(base: str, alimentador: str) -> tuple[str, str]:
        """Tenta resolver subpasta do alimentador dentro da base.

        Procura nesta ordem:
          base/ALIM
          base/<ano>/ALIM (qualquer subpasta de ano que exista)
          base                       (fallback)
        Retorna (path, label_used).
        """
        base = (base or "").strip()
        alim = (alimentador or "").strip().upper()
        if not base or not os.path.isdir(base):
            return base, ""
        if not alim:
            return base, ""
        # base/ALIM
        direct = os.path.join(base, alim)
        if os.path.isdir(direct):
            return direct, alim
        # base/<ano>/ALIM
        try:
            for sub in sorted(os.listdir(base)):
                year_dir = os.path.join(base, sub)
                if not os.path.isdir(year_dir):
                    continue
                cand = os.path.join(year_dir, alim)
                if os.path.isdir(cand):
                    return cand, f"{sub}/{alim}"
        except OSError:
            pass
        return base, ""

    def list_ganhos_files(self, alimentador: Any = "") -> dict[str, Any]:
        """Retorna {ok, base, folder, alim, files, error}.

        - ``base``  = config.caminho_pasta_ganhos (raiz)
        - ``folder``= pasta efetivamente listada (pode ser raiz ou subpasta
                      do alimentador; ver _resolve_ganhos_folder)
        - ``alim``  = label resolvido (ex.: '2026/ATB-204') ou ''
        - ``files`` = lista de {name, path, size, mtime}
        """
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {
                "ok": False, "base": "", "folder": "", "alim": "",
                "files": [], "error": f"config: {exc}",
            }

        base = str(cfg.get("caminho_pasta_ganhos") or "").strip()
        if not base:
            return {
                "ok": False, "base": "", "folder": "", "alim": "",
                "files": [],
                "error": "caminho_pasta_ganhos nao configurado em config.json",
            }
        if not os.path.isdir(base):
            return {
                "ok": False, "base": base, "folder": base, "alim": "",
                "files": [],
                "error": f"pasta nao encontrada: {base}",
            }
        folder, label = self._resolve_ganhos_folder(base, str(alimentador or ""))
        files = self._list_ganhos_in(folder)
        return {
            "ok": True, "base": base, "folder": folder, "alim": label,
            "files": files, "error": "",
        }

    @staticmethod
    def _wv_dialog_const(kind: str) -> Any:
        """Resolve a constante de dialog do pywebview no runtime.

        Pywebview >= 5 usa ``FileDialog.OPEN``/``SAVE``/``FOLDER`` em
        vez de ``OPEN_DIALOG``/``SAVE_DIALOG``/``FOLDER_DIALOG``. Tenta a
        forma nova e cai pra antiga (silencia deprecation warnings)."""
        import webview  # type: ignore[import-not-found]
        kind = kind.upper()
        FileDialog = getattr(webview, "FileDialog", None)
        if FileDialog is not None:
            mapping = {
                "OPEN": getattr(FileDialog, "OPEN", None),
                "SAVE": getattr(FileDialog, "SAVE", None),
                "FOLDER": getattr(FileDialog, "FOLDER", None),
            }
            v = mapping.get(kind)
            if v is not None:
                return v
        # Fallback antigo (pywebview < 5).
        legacy = {
            "OPEN":   getattr(webview, "OPEN_DIALOG", None),
            "SAVE":   getattr(webview, "SAVE_DIALOG", None),
            "FOLDER": getattr(webview, "FOLDER_DIALOG", None),
        }
        return legacy.get(kind)

    def pick_ganhos_folder(self) -> dict[str, Any]:
        """Abre o file dialog do pywebview (folder) e atualiza
        config.caminho_pasta_ganhos com o caminho escolhido. Retorna o
        novo estado de list_ganhos_files."""
        try:
            import webview  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "base": "", "folder": "", "alim": "",
                    "files": [], "error": f"pywebview indisponivel: {exc}"}
        try:
            wins = webview.windows
            if not wins:
                return {"ok": False, "base": "", "folder": "", "alim": "",
                        "files": [], "error": "janela pywebview nao encontrada"}
            dlg = self._wv_dialog_const("FOLDER")
            if dlg is None:
                return {"ok": False, "base": "", "folder": "", "alim": "",
                        "files": [], "error": "FOLDER dialog indisponivel"}
            result = wins[0].create_file_dialog(dlg)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "base": "", "folder": "", "alim": "",
                    "files": [], "error": f"file dialog: {exc}"}
        if not result:
            # Usuario cancelou; devolve estado atual sem mudar config.
            return self.list_ganhos_files("")
        path = result[0] if isinstance(result, (list, tuple)) else str(result)
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
            cfg["caminho_pasta_ganhos"] = path
            ConfigManager.save_config(cfg)
        except Exception as exc:  # noqa: BLE001
            self._data_state_set(
                "ganhos", "INVALIDADO", path=path,
                error=f"save_config: {exc}")
            return {"ok": False, "base": path, "folder": path, "alim": "",
                    "files": [], "error": f"save_config: {exc}"}
        # Recarrega config interno
        self._config = None
        self._ensure_managers()
        # Hook estado: pasta de ganhos selecionada com sucesso.
        # Validacao "tem arquivos esperados" acontece em read_ganhos_file
        # ou no calc_*; aqui marcamos VALIDADO se a pasta existe.
        if os.path.isdir(path):
            self._data_state_set(
                "ganhos", "CARREGADO_VALIDADO", path=path)
        else:
            self._data_state_set(
                "ganhos", "INVALIDADO", path=path,
                error="pasta nao existe")
        return self.list_ganhos_files("")

    # ------------------------------------------------------------------
    # Passo 5.2 (Ganhos / read_ganhos_file): leitor generico de XLSX/CSV
    # /TXT que devolve cabecalho + linhas para o JS popular a tabela
    # "Parametros de Ganhos". A logica avancada do desktop (parse de
    # FlowMT.TXT/Topologia.TXT/Confiabilidade.TXT por alimentador) e'
    # complexa demais para um wrapper unico; aqui fazemos preview
    # tabular + heuristica para extrair pares Antes/Depois quando o
    # arquivo tem 2-3 colunas no formato "param;antes;depois".
    # ------------------------------------------------------------------
    def read_ganhos_file(self, path: Any, max_rows: Any = 200) -> dict[str, Any]:
        path_s = str(path or "").strip()
        if not path_s:
            return {"ok": False, "error": "path vazio", "headers": [],
                    "rows": [], "parametros": [], "total_rows": 0}
        if not os.path.isfile(path_s):
            return {"ok": False, "error": f"arquivo nao encontrado: {path_s}",
                    "headers": [], "rows": [], "parametros": [], "total_rows": 0}

        try:
            limit = int(max_rows) if max_rows not in (None, "", 0) else 200
        except (TypeError, ValueError):
            limit = 200

        ext = os.path.splitext(path_s)[1].lower()
        headers: list[str] = []
        rows: list[list[str]] = []

        try:
            if ext in (".xlsx", ".xlsm"):
                from openpyxl import load_workbook  # type: ignore[import-not-found]
                wb = load_workbook(path_s, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for r_i, row in enumerate(ws.iter_rows(values_only=True)):
                        cells = [
                            "" if c is None else str(c) for c in row
                        ]
                        if r_i == 0:
                            headers = cells
                        else:
                            rows.append(cells)
                        if len(rows) >= limit:
                            break
                finally:
                    wb.close()
            elif ext in (".xls",):
                # xlrd nao e' garantido; tenta pandas se disponivel.
                try:
                    import pandas as pd  # type: ignore[import-not-found]
                    df = pd.read_excel(path_s, header=0, nrows=limit)
                    headers = [str(c) for c in df.columns]
                    rows = [
                        ["" if (c != c) else str(c) for c in r]  # NaN check
                        for r in df.values.tolist()
                    ]
                except Exception as exc:  # noqa: BLE001
                    return {
                        "ok": False, "error": f"xls (pandas falhou): {exc}",
                        "headers": [], "rows": [], "parametros": [], "total_rows": 0,
                    }
            elif ext in (".csv", ".txt", ".tsv"):
                import csv as _csv
                # Detecta delimitador a partir de uma amostra.
                with open(path_s, encoding="utf-8", errors="replace") as f:
                    sample = f.read(4096)
                if "\t" in sample:
                    delim = "\t"
                elif sample.count(";") > sample.count(","):
                    delim = ";"
                elif "," in sample:
                    delim = ","
                else:
                    delim = ";"
                with open(path_s, encoding="utf-8", errors="replace", newline="") as f:
                    reader = _csv.reader(f, delimiter=delim)
                    for i, row in enumerate(reader):
                        if i == 0:
                            headers = list(row)
                        else:
                            rows.append(list(row))
                        if len(rows) >= limit:
                            break
            else:
                return {
                    "ok": False, "error": f"extensao nao suportada: {ext}",
                    "headers": [], "rows": [], "parametros": [], "total_rows": 0,
                }
        except Exception as exc:  # noqa: BLE001
            return {
                "ok": False, "error": f"leitura falhou: {exc}",
                "headers": [], "rows": [], "parametros": [], "total_rows": 0,
            }

        # Heuristica: extrai (label, antes, depois) das primeiras colunas.
        # Util quando o arquivo tem 2-3 colunas estruturadas.
        parametros: list[dict[str, str]] = []
        for r in rows:
            if not r:
                continue
            label = str(r[0] if len(r) > 0 else "").strip()
            if not label:
                continue
            antes = str(r[1] if len(r) > 1 else "").strip()
            depois = str(r[2] if len(r) > 2 else "").strip()
            parametros.append({"label": label, "a": antes, "d": depois})
            if len(parametros) >= 30:
                break

        # Hook estado: se o arquivo lido for um dos 3 tecnicos
        # (FlowMT.TXT, Topologia.TXT, Confiabilidade.TXT), marca a fonte
        # tecnico_txt como CARREGADO_PARCIAL. Validacao completa (3 juntos)
        # e' feita em validate_tecnico_files / antes de calc_*.
        try:
            base = os.path.basename(path_s).lower()
            if base in ("flowmt.txt", "topologia.txt", "confiabilidade.txt"):
                self._data_state_set(
                    "tecnico_txt", "CARREGADO_PARCIAL",
                    path=os.path.dirname(path_s),
                    version_token=str(int(os.path.getmtime(path_s))))
        except OSError:
            pass

        return {
            "ok": True, "error": "",
            "headers": headers, "rows": rows[:limit],
            "parametros": parametros,
            "total_rows": len(rows),
            "path": path_s,
            "ext": ext,
        }

    def validate_tecnico_files(self, pasta: Any = "") -> dict[str, Any]:
        """Verifica se os 3 arquivos tecnicos obrigatorios existem na pasta.
        Marca tecnico_txt como CARREGADO_VALIDADO (ok) ou INVALIDADO.
        Equivalente ao _validate_ganhos_files do desktop, sem o passo
        de tentar ler os arquivos (custo alto)."""
        path = str(pasta or "").strip()
        if not path:
            try:
                from codigo5_coplan import ConfigManager  # noqa: PLC0415
                cfg = ConfigManager.load_config() or {}
                path = str(cfg.get("caminho_pasta_ganhos") or "").strip()
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "error": f"config: {exc}",
                        "missing": []}
        if not path:
            return {"ok": False, "error": "pasta nao configurada",
                    "missing": []}
        if not os.path.isdir(path):
            self._data_state_set(
                "tecnico_txt", "INVALIDADO", path=path,
                error="pasta nao existe")
            return {"ok": False, "error": f"pasta nao existe: {path}",
                    "missing": []}
        required = ["FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT"]
        missing = [
            f for f in required
            if not os.path.isfile(os.path.join(path, f))
        ]
        if missing:
            self._data_state_set(
                "tecnico_txt", "INVALIDADO", path=path,
                error=f"arquivos ausentes: {', '.join(missing)}")
            return {"ok": False, "missing": missing,
                    "error": f"faltam: {', '.join(missing)}",
                    "pasta": path}
        # Token: hash dos mtimes dos 3 arquivos.
        try:
            tokens = [
                str(int(os.path.getmtime(os.path.join(path, f))))
                for f in required
            ]
            version_token = "-".join(tokens)
        except OSError:
            version_token = ""
        self._data_state_set(
            "tecnico_txt", "CARREGADO_VALIDADO", path=path,
            version_token=version_token)
        return {"ok": True, "missing": [], "pasta": path,
                "version_token": version_token}

    # ------------------------------------------------------------------
    # Passo 5.3 (Ganhos / criterios + status OK/Falhou):
    # Le criterios_planejamento + piora_mercado do config e devolve
    # tambem REGRAS DECLARATIVAS (label_pattern -> {op, key}) que o JS
    # usa para colorir cada linha da tabela e popular o card lateral.
    # ------------------------------------------------------------------
    def get_criterios(self) -> dict[str, Any]:
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager, DEFAULT_CRITERIOS, DEFAULT_PIORA_MERCADO,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}"}
        cfg = ConfigManager.load_config() or {}
        crit = dict(DEFAULT_CRITERIOS)
        crit.update(cfg.get("criterios_planejamento") or {})
        piora = dict(DEFAULT_PIORA_MERCADO)
        piora.update(cfg.get("piora_mercado") or {})

        # Regras declarativas para o JS aplicar status/critTxt:
        #   label_match: prefixo (case+acento insensitive) que casa com o
        #     primeiro coluna da tabela "Parametros de Ganhos".
        #   op: 'ge' (>=), 'le' (<=).
        #   key: chave a comparar com o valor "Depois" da linha.
        regras = [
            {"label_match": "tensao min linha", "op": "ge",
             "key": "tensao_min", "label": "≥ %.2f pu"},
            {"label_match": "tensao min",       "op": "ge",
             "key": "tensao_min", "label": "≥ %.2f pu"},
            {"label_match": "tensao maxima",    "op": "le",
             "key": "tensao_max", "label": "≤ %.2f pu"},
            {"label_match": "tensao max",       "op": "le",
             "key": "tensao_max", "label": "≤ %.2f pu"},
            # Carregamento usa dois limites; default usa o mais rigoroso
            # (sim_ou_vazio) -- a UI pode re-aplicar com NAO se obra nao
            # for aprovada via o filtro do form.
            {"label_match": "carregamento",     "op": "le",
             "key": "carregamento_limite_sim_ou_vazio",
             "label": "≤ %.0f%%"},
            {"label_match": "chi",              "op": "ge",
             "key": "chi_min", "label": "≥ %.2f"},
            {"label_match": "ci",               "op": "ge",
             "key": "ci_min", "label": "≥ %.2f"},
        ]
        return {
            "ok": True, "error": "",
            "criterios": crit,
            "piora_mercado": piora,
            "regras": regras,
        }

    # ------------------------------------------------------------------
    # Passo 5.4 (Ganhos / Ganhos Atuais por alimentador):
    # Agrega valores das colunas tensao_min_registrada_atual /
    # carregamento_max_registrado_atual / ganhos_totais_atual entre as
    # obras que tem o alimentador no campo principal OU nos beneficiados.
    # Retorna agregados (min/max tensao, max carregamento, ultimo
    # ganhos_totais_atual nao vazio) prontos para o card.
    # ------------------------------------------------------------------
    @staticmethod
    def _to_float_brl(v: Any) -> float | None:
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        # pt-BR: "1.234,56" -> "1234.56"; tambem aceita "1234.56"
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        # Limpa unidades comuns (pu, %, kV)
        s = re.sub(r"[^0-9.\-]", "", s)
        try:
            return float(s)
        except ValueError:
            return None

    def get_ganhos_atuais(self, alimentador: Any = "") -> dict[str, Any]:
        alim = str(alimentador or "").strip().upper()
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "alim": alim, "error": err or "db indisponivel",
                    "tensao_min": None, "tensao_max": None,
                    "carregamento_max": None, "ganhos_totais_atual": "",
                    "obras_count": 0}
        try:
            cursor = db._get_cursor()
            if cursor is None:
                return {"ok": False, "alim": alim, "error": "cursor",
                        "tensao_min": None, "tensao_max": None,
                        "carregamento_max": None, "ganhos_totais_atual": "",
                        "obras_count": 0}
            if alim:
                # Filtra por alimentador principal OU contido em
                # alimentadores_beneficiados (string separada por ;,).
                cursor.execute(
                    "SELECT tensao_min_registrada_atual, "
                    "       carregamento_max_registrado_atual, "
                    "       ganhos_totais_atual, alimentador_principal, "
                    "       alimentadores_beneficiados, cod "
                    "FROM obras "
                    "WHERE UPPER(COALESCE(alimentador_principal,''))=? "
                    "   OR UPPER(COALESCE(alimentadores_beneficiados,'')) LIKE ?",
                    (alim, f"%{alim}%"),
                )
            else:
                cursor.execute(
                    "SELECT tensao_min_registrada_atual, "
                    "       carregamento_max_registrado_atual, "
                    "       ganhos_totais_atual, alimentador_principal, "
                    "       alimentadores_beneficiados, cod "
                    "FROM obras"
                )
            rows = cursor.fetchall() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "alim": alim, "error": f"sql: {exc}",
                    "tensao_min": None, "tensao_max": None,
                    "carregamento_max": None, "ganhos_totais_atual": "",
                    "obras_count": 0}

        # Filtragem extra em Python para garantir match correto na string
        # de beneficiados (LIKE pode pegar substring espuria).
        kept: list[tuple[Any, ...]] = []
        for r in rows:
            principal = str(r[3] or "").strip().upper()
            benef_raw = str(r[4] or "")
            benef_set = {
                p.strip().upper()
                for p in re.split(r"[;,|]+", benef_raw)
                if p.strip()
            }
            if not alim or principal == alim or alim in benef_set:
                kept.append(r)

        tensoes_min: list[float] = []
        carregs_max: list[float] = []
        ganhos_str = ""
        for r in kept:
            tm = self._to_float_brl(r[0])
            if tm is not None:
                tensoes_min.append(tm)
            cm = self._to_float_brl(r[1])
            if cm is not None:
                carregs_max.append(cm)
            g = str(r[2] or "").strip()
            if g and not ganhos_str:
                ganhos_str = g  # 1o nao-vazio (mesmo alimentador tende a ter so 1)

        # Tensao "min" e "max" registradas: usamos o minimo dos minimos
        # como "Min" e o maximo dos maximos como "Max" (usando a propria
        # coluna min_registrada para ambos limites; campo max_registrada
        # nao existe explicito).
        return {
            "ok": True, "alim": alim, "error": "",
            "tensao_min": min(tensoes_min) if tensoes_min else None,
            "tensao_max": max(tensoes_min) if tensoes_min else None,
            "carregamento_max": max(carregs_max) if carregs_max else None,
            "ganhos_totais_atual": ganhos_str,
            "obras_count": len(kept),
        }

    # ------------------------------------------------------------------
    # Ganhos / metodos auxiliares (leva B da migracao Ganhos).
    # G020 ganhos_form_state, G021 avaliar_ganhos_planejamento,
    # G022 avaliar_ganhos_postergacao, G024 ganhos_resolver_alimentador.
    # Origem desktop: ui/main_window/ganhos_mixin.py + cadastro_mixin.py
    # (atualizar_labels_planejamento_desde_tela:892-987) + visualizar_mixin
    # (_obra_suficiente:545-622) + core/services/relatorio_criterios_service
    # (obra_atende:149).
    # ------------------------------------------------------------------

    def ganhos_resolver_alimentador(self, cod: Any = "") -> dict[str, Any]:
        """[G024] Devolve {principal, beneficiados[], todos[]} dado o COD
        de uma obra. Beneficiados parseados por separador ; ou ,. Sem cod
        valido retorna ok=False."""
        c = str(cod or "").strip()
        if not c:
            return {"ok": False, "principal": "",
                    "beneficiados": [], "todos": [],
                    "error": "cod vazio"}
        try:
            obra = self.get_obra(c)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "principal": "",
                    "beneficiados": [], "todos": [],
                    "error": f"get_obra: {exc}"}
        if not obra or not obra.get("ok"):
            return {"ok": False, "principal": "",
                    "beneficiados": [], "todos": [],
                    "error": (obra or {}).get("error", "obra nao encontrada")}
        o = obra.get("obra") or {}
        principal = str(
            o.get("alimentador_principal")
            or o.get("alimentador") or ""
        ).strip()
        # get_obra ja faz o parser dos beneficiados e devolve em alim_benef.
        benef_list = obra.get("alim_benef") or []
        if not isinstance(benef_list, list):
            benef_raw = str(o.get("alimentadores_beneficiados") or "")
            tmp = benef_raw.replace(",", ";").split(";")
            benef_list = [s.strip() for s in tmp if s.strip()]
        benef_clean: list[str] = []
        seen = set()
        for b in benef_list:
            s = str(b or "").strip()
            if not s:
                continue
            up = s.upper()
            if up in seen:
                continue
            seen.add(up)
            benef_clean.append(s)
        todos = ([principal] if principal else []) + [
            b for b in benef_clean
            if b.upper() != principal.upper()
        ]
        return {
            "ok": True,
            "principal": principal,
            "beneficiados": benef_clean,
            "todos": todos,
            "error": "",
        }

    def ganhos_form_state(self, cod: Any = "") -> dict[str, Any]:
        """[G020] Agregador para popular toda a aba Ganhos numa unica
        chamada. Reusa get_criterios + get_ganhos_atuais + (opcional)
        quadro_resumo_ganhos quando cod for informado."""
        out: dict[str, Any] = {"ok": True, "error": ""}
        try:
            out["criterios"] = self.get_criterios()
        except Exception as exc:  # noqa: BLE001
            out["criterios"] = {"ok": False, "error": f"get_criterios: {exc}"}
        # Resolve alim do cod (se houver) para alimentar ganhos_atuais.
        alim_principal = ""
        c = str(cod or "").strip()
        if c:
            try:
                resolver = self.ganhos_resolver_alimentador(c)
                if resolver.get("ok"):
                    alim_principal = resolver.get("principal") or ""
                    out["alim"] = resolver
            except Exception as exc:  # noqa: BLE001
                out["alim"] = {"ok": False, "error": f"resolver: {exc}"}
        try:
            out["atual"] = self.get_ganhos_atuais(alim_principal)
        except Exception as exc:  # noqa: BLE001
            out["atual"] = {"ok": False, "error": f"get_ganhos_atuais: {exc}"}
        # Quadro resumo (parametros de ganhos por alimentador) — apenas
        # se cod foi informado.
        if c:
            try:
                out["quadro"] = self.quadro_resumo_ganhos(cod=c)
            except Exception as exc:  # noqa: BLE001
                out["quadro"] = {"ok": False, "error": f"quadro_resumo: {exc}"}
        out["cod"] = c
        return out

    def avaliar_ganhos_planejamento(
        self, payload: Any = None,
    ) -> dict[str, Any]:
        """[G021] Espelho web de _obra_atende (cadastro_mixin:920-933).
        Recebe valores DEPOIS da tela como dict simples e retorna
        {ok, atende, motivos}. Usa criterios do config + delega ao
        relatorio_criterios_service.obra_atende.

        payload esperado:
          {tensao_min, tensao_max, carregamento, contas, manobra}
        Aceita tambem chaves alternativas (tensao_min_final, etc.)."""
        if not isinstance(payload, dict):
            payload = {}

        def _get_first(*keys: str) -> Any:
            for k in keys:
                v = payload.get(k)
                if v not in (None, ""):
                    return v
            return None

        tmin = _get_first("tensao_min", "tensao_min_final", "tmin")
        tmax = _get_first("tensao_max", "tensao_max_final", "tmax")
        carreg = _get_first("carregamento", "carregamento_final", "carreg")
        contas = _get_first("contas", "contas_contratos_posteriores", "clientes")
        manobra = _get_first("manobra")

        # Linha sintetica + indices (formato esperado por obra_atende).
        row = [tmin, tmax, carreg, manobra, contas]
        idx = {"tmin": 0, "tmax": 1, "carreg": 2, "manobra": 3, "clientes": 4}

        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atende": None, "motivos": [],
                    "error": f"import config: {exc}"}
        try:
            cfg = ConfigManager.load_config() or {}
        except Exception:  # noqa: BLE001
            cfg = {}
        try:
            from runtime.config import DEFAULT_CRITERIOS  # type: ignore[import-not-found]
        except Exception:  # noqa: BLE001
            DEFAULT_CRITERIOS = {
                "tensao_min": 0.95, "tensao_max": 1.03,
                "carregamento_limite_sim_ou_vazio": 67.0,
                "carregamento_limite_nao": 100.0,
                "clientes_maximo": 6000,
            }
        criterios = cfg.get("criterios_planejamento", DEFAULT_CRITERIOS)

        def _conv_float(v: Any) -> float:
            try:
                return float(str(v).replace(",", ".")) if str(v).strip() else 0.0
            except Exception:  # noqa: BLE001
                return 0.0

        def _conv_int(v: Any) -> int:
            try:
                return int(float(str(v).replace(",", ".")))
            except Exception:  # noqa: BLE001
                return 0

        try:
            from core.services.relatorio_criterios_service import (  # type: ignore[import-not-found]
                obra_atende,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atende": None, "motivos": [],
                    "error": f"import obra_atende: {exc}"}
        try:
            atende, motivos = obra_atende(
                row, idx, criterios, _conv_float, _conv_int,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atende": None, "motivos": [],
                    "error": f"obra_atende: {exc}"}
        return {
            "ok": True,
            "atende": atende,  # bool | None
            "motivos": list(motivos or []),
            "error": "",
        }

    def avaliar_ganhos_postergacao(
        self, payload: Any = None, anos: Any = None,
    ) -> dict[str, Any]:
        """[G022] Espelho web de _obra_suficiente
        (visualizar_mixin:545-622). Projeta degradacao N anos aplicando
        DEFAULT_PIORA_MERCADO (ou override do config) e avalia critérios
        ano a ano. Retorna {ok, suficiente, anos_alcancados, motivos}."""
        if not isinstance(payload, dict):
            payload = {}

        def _get_first(*keys: str) -> Any:
            for k in keys:
                v = payload.get(k)
                if v not in (None, ""):
                    return v
            return None

        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            from runtime.config import (  # type: ignore[import-not-found]
                DEFAULT_CRITERIOS, DEFAULT_PIORA_MERCADO,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "suficiente": None,
                    "anos_alcancados": 0, "motivos": [],
                    "error": f"import: {exc}"}
        try:
            cfg = ConfigManager.load_config() or {}
        except Exception:  # noqa: BLE001
            cfg = {}
        criterios = cfg.get("criterios_planejamento", DEFAULT_CRITERIOS)
        pioras = cfg.get("piora_mercado", DEFAULT_PIORA_MERCADO)

        # Conversao tolerante.
        def _f(v: Any) -> float | None:
            if v is None:
                return None
            s = str(v).strip()
            if not s:
                return None
            try:
                return float(s.replace(",", "."))
            except Exception:  # noqa: BLE001
                return None

        def _i(v: Any) -> int | None:
            f = _f(v)
            if f is None:
                return None
            try:
                return int(f)
            except Exception:  # noqa: BLE001
                return None

        tmin = _f(_get_first("tensao_min", "tensao_min_final"))
        tmax = _f(_get_first("tensao_max", "tensao_max_final"))
        carreg = _f(_get_first("carregamento", "carregamento_final"))
        contas = _i(_get_first("contas", "contas_contratos_posteriores"))
        manobra = str(_get_first("manobra") or "").strip().upper()

        if any(v is None for v in (tmin, tmax, carreg, contas)):
            return {"ok": True, "suficiente": None,
                    "anos_alcancados": 0,
                    "motivos": ["dados_insuficientes"],
                    "error": ""}

        try:
            anos_n = int(anos) if anos not in (None, "") else int(
                pioras.get("anos_horizonte", DEFAULT_PIORA_MERCADO["anos_horizonte"])
            )
        except Exception:  # noqa: BLE001
            anos_n = int(DEFAULT_PIORA_MERCADO["anos_horizonte"])

        delta_tensao = float(
            pioras.get("tensao_delta", DEFAULT_PIORA_MERCADO["tensao_delta"])
        )
        perc_carreg = float(
            pioras.get("carregamento_percentual",
                       DEFAULT_PIORA_MERCADO["carregamento_percentual"])
        )

        try:
            from core.services.relatorio_criterios_service import (  # type: ignore[import-not-found]
                obra_atende,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "suficiente": None,
                    "anos_alcancados": 0, "motivos": [],
                    "error": f"import obra_atende: {exc}"}

        idx = {"tmin": 0, "tmax": 1, "carreg": 2, "manobra": 3, "clientes": 4}

        def _conv_float(v: Any) -> float:
            try:
                return float(str(v).replace(",", ".")) if str(v).strip() else 0.0
            except Exception:  # noqa: BLE001
                return 0.0

        def _conv_int(v: Any) -> int:
            try:
                return int(float(str(v).replace(",", ".")))
            except Exception:  # noqa: BLE001
                return 0

        anos_ok = 0
        for _ in range(max(0, anos_n)):
            carreg = (carreg or 0.0) * (1.0 + (perc_carreg / 100.0))
            tmin = (tmin or 0.0) - delta_tensao
            tmax = (tmax or 0.0) - delta_tensao
            row = [tmin, tmax, carreg, manobra, contas]
            try:
                atende, motivos = obra_atende(
                    row, idx, criterios, _conv_float, _conv_int,
                )
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "suficiente": None,
                        "anos_alcancados": anos_ok,
                        "motivos": [f"obra_atende_ano_{anos_ok+1}: {exc}"],
                        "error": str(exc)}
            if atende is None:
                return {"ok": True, "suficiente": None,
                        "anos_alcancados": anos_ok,
                        "motivos": list(motivos or ["dados_insuficientes"]),
                        "error": ""}
            if not atende:
                return {"ok": True, "suficiente": False,
                        "anos_alcancados": anos_ok,
                        "motivos": list(motivos or []),
                        "error": ""}
            anos_ok += 1

        return {"ok": True, "suficiente": True,
                "anos_alcancados": anos_ok,
                "motivos": [], "error": ""}

    # ------------------------------------------------------------------
    # Passo 5.5 (Ganhos / acoes Inserir Antes/Depois + em Massa + atuais):
    #   * pick_ganhos_file()    - file dialog (xlsx/csv/txt) e devolve
    #                             read_ganhos_file() do arquivo escolhido.
    #   * apply_ganhos_to_obra(cod, slot, parametros) - persiste valores
    #                             do parametros[] em colunas do banco
    #                             (slot='antes'|'depois'|'atual') usando
    #                             update_obra. Mapeia label -> coluna
    #                             via heuristica (mesma logica do
    #                             desktop, mas declarativa).
    #   * ganhos_em_massa(...)  - stub: aplica em N cods de uma vez.
    # ------------------------------------------------------------------
    GANHOS_LABEL_MAP = (
        # (substring do label normalizado, coluna_antes, coluna_depois)
        ("contas contratos",   "contas_contratos_previos",       "contas_contratos_posteriores"),
        ("contas",             "contas_contratos_previos",       "contas_contratos_posteriores"),
        ("carregamento",       "carregamento_inicial",            "carregamento_final"),
        ("perdas",             "perdas_iniciais",                 "perdas_finais"),
        ("tensao media",       "tensao_media_inicial",            "tensao_media_final"),
        ("tensao min linha",   "tensao_min_linha_inicial",        "tensao_min_linha_final"),
        ("tensao min",         "tensao_min_inicial",              "tensao_min_final"),
        ("tensao maxima",      "tensao_max_inicial",              "tensao_max_final"),
        ("tensao max",         "tensao_max_inicial",              "tensao_max_final"),
        ("chi",                "chi_inicial",                     "chi_final"),
        ("ci",                 "ci_inicial",                      "ci_final"),
        ("ganhos totais",      "ganhos_totais_antes",             "ganhos_totais_depois"),
    )

    @staticmethod
    def _norm_label(s: str) -> str:
        try:
            import unicodedata as _ud
            n = _ud.normalize("NFD", str(s or ""))
            n = "".join(c for c in n if not _ud.combining(c))
            return n.strip().lower()
        except Exception:  # noqa: BLE001
            return str(s or "").strip().lower()

    def pick_ganhos_file(self) -> dict[str, Any]:
        """Abre file dialog (single, filtros para xlsx/csv/txt) e devolve
        read_ganhos_file() do path escolhido. Cancelar -> ok=False sem erro."""
        try:
            import webview  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"pywebview indisponivel: {exc}"}
        try:
            wins = webview.windows
            if not wins:
                return {"ok": False, "error": "janela pywebview nao encontrada"}
            file_types = (
                "Planilhas/Texto (*.xlsx;*.xlsm;*.xls;*.csv;*.txt;*.tsv)",
                "Todos os arquivos (*.*)",
            )
            dlg = self._wv_dialog_const("OPEN")
            if dlg is None:
                return {"ok": False, "error": "OPEN dialog indisponivel"}
            result = wins[0].create_file_dialog(
                dlg, allow_multiple=False, file_types=file_types,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"file dialog: {exc}"}
        if not result:
            return {"ok": False, "error": "cancelado"}
        path = result[0] if isinstance(result, (list, tuple)) else str(result)
        return self.read_ganhos_file(path, 200)

    def _resolve_ganhos_columns(
        self, parametros: list[dict[str, Any]], slot: str
    ) -> dict[str, str]:
        """Casa cada parametro com sua coluna no banco com base no label.

        Retorna dict {coluna_db: valor} pronto para update_obra.
        slot in {'antes','depois','atual'}.
        """
        out: dict[str, str] = {}
        slot_norm = (slot or "antes").strip().lower()
        for p in parametros or []:
            lbl = self._norm_label(p.get("label", ""))
            if not lbl:
                continue
            valor = ""
            if slot_norm == "antes":
                valor = str(p.get("a") or "").strip()
            elif slot_norm == "depois":
                valor = str(p.get("d") or "").strip()
            else:
                # 'atual': usa o que estiver no campo "atual" se houver,
                # senao cai no Depois (heuristica).
                valor = str(p.get("atual") or p.get("d") or "").strip()
            if not valor:
                continue
            for prefix, col_a, col_d in self.GANHOS_LABEL_MAP:
                if lbl.startswith(prefix):
                    col = col_a if slot_norm == "antes" else col_d
                    out[col] = valor
                    break
        return out

    def apply_ganhos_to_obra(
        self,
        cod: Any,
        slot: Any = "antes",
        parametros: Any = None,
    ) -> dict[str, Any]:
        cod_s = str(cod or "").strip()
        if not cod_s:
            return {"ok": False, "error": "cod vazio", "cod": "",
                    "applied": 0}
        if not isinstance(parametros, list) or not parametros:
            return {"ok": False, "error": "parametros vazio", "cod": cod_s,
                    "applied": 0}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "cod": cod_s, "applied": 0}
        cols_to_update = self._resolve_ganhos_columns(parametros, str(slot))
        if not cols_to_update:
            return {"ok": False, "error": "nenhum parametro casou com coluna",
                    "cod": cod_s, "applied": 0}
        try:
            db.update_obra(cols_to_update, cod_s, skip_blank=True)
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            out: dict[str, Any] = {
                "ok": False, "error": friendly or str(exc),
                "cod": cod_s, "applied": 0,
            }
            if friendly:
                out["blocked"] = "db_busy"
            return out
        return {
            "ok": True, "error": "", "cod": cod_s,
            "slot": str(slot), "applied": len(cols_to_update),
            "columns": list(cols_to_update.keys()),
        }

    def ganhos_em_massa(
        self,
        cods: Any,
        slot: Any = "antes",
        parametros: Any = None,
    ) -> dict[str, Any]:
        """Aplica os mesmos parametros a varios COD. Util quando o mesmo
        arquivo de ganhos vale para multiplas obras (ex.: reconfiguracao
        de alimentador que afeta varias obras planejadas)."""
        if not isinstance(cods, (list, tuple)) or not cods:
            return {"ok": False, "error": "cods vazio", "applied": 0,
                    "results": []}
        results: list[dict[str, Any]] = []
        ok_count = 0
        for c in cods:
            r = self.apply_ganhos_to_obra(str(c), slot, parametros)
            results.append(r)
            if r.get("ok"):
                ok_count += 1
        return {
            "ok": ok_count > 0, "error": "",
            "applied": ok_count, "total": len(cods),
            "results": results,
        }

    # ------------------------------------------------------------------
    # Passo 6.1 (Resumo / KPIs): 5 cards do topo da aba Resumo.
    #   capex_total            = SUM(valor_obra)
    #   obras_total            = COUNT(*)
    #   km_total               = SUM(quantidade_material)        # km
    #   contas_beneficiadas    = SUM(contas_contratos_beneficiadas)
    #   postergacoes           = COUNT(obras com tipo_pacote contendo
    #                             'posterga' OU manobra contendo 'posterga')
    # Tudo via SQL com CAST/REPLACE para tratar valores armazenados como
    # texto pt-BR ('1.234,56'). Aceita filtro opcional ano (string).
    # ------------------------------------------------------------------
    @staticmethod
    def _sql_to_real(col: str) -> str:
        """Expressao SQLite que converte texto pt-BR em REAL.
        Ex.: '1.234,56' -> 1234.56  | '12,5' -> 12.5  | '12.5' -> 12.5
        Heuristica: remove '.' e troca ',' por '.'. Funciona quando o
        usuario *nao* mistura separadores de milhar com decimal en-US."""
        return (
            f"COALESCE("
            f"CAST(REPLACE(REPLACE(COALESCE({col},'0'),'.',''),',','.') AS REAL),"
            f"0)"
        )

    # ------------------------------------------------------------------
    # Visualizar Sprint 1 (Auditoria #1): endpoints de resumo aceitam
    # filtro `cods` opcional. JS publica coplanFilteredCods() (lista de
    # cods filtrados em Visualizar) e os 5 cards do Resumo passam essa
    # lista. Quando cods=None, comportamento legado (banco inteiro).
    # ------------------------------------------------------------------
    @staticmethod
    def _build_resumo_where(
        ano_s: str, cods: Any = None,
    ) -> tuple[str, list[Any]]:
        """Monta WHERE compartilhado entre os 5 endpoints de resumo.
        Suporta filtro por ano_ E/OU lista de cods. SQLite limita ~999
        placeholders; para listas maiores, trunca silenciosamente em 900
        (caso patologico - Visualizar dificilmente filtra tanto)."""
        clauses: list[str] = []
        params: list[Any] = []
        if ano_s:
            clauses.append("TRIM(COALESCE(ano_,''))=?")
            params.append(ano_s)
        if isinstance(cods, (list, tuple)) and cods:
            cods_clean = [
                str(c).strip() for c in cods if str(c or "").strip()
            ]
            if cods_clean:
                if len(cods_clean) > 900:
                    cods_clean = cods_clean[:900]
                placeholders = ",".join(["?"] * len(cods_clean))
                clauses.append(f"cod IN ({placeholders})")
                params.extend(cods_clean)
        if not clauses:
            return "", []
        return " WHERE " + " AND ".join(clauses), params

    def resumo_kpis(self, ano: Any = "", cods: Any = None) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {
                "ok": False, "error": err or "db indisponivel",
                "ano": str(ano or ""), "capex_total": 0.0,
                "obras_total": 0, "km_total": 0.0,
                "contas_beneficiadas": 0, "postergacoes": 0,
                "ano_dominante": None,
            }
        ano_s = str(ano or "").strip()
        try:
            cursor = db._get_cursor()
            if cursor is None:
                return {"ok": False, "error": "cursor indisponivel"}
            where_clause, _params_list = self._build_resumo_where(ano_s, cods)
            params: tuple[Any, ...] = tuple(_params_list)
            cursor.execute(
                "SELECT "
                "  COUNT(*),"
                f" SUM({self._sql_to_real('valor_obra')}),"
                f" SUM({self._sql_to_real('quantidade_material')}),"
                f" SUM({self._sql_to_real('contas_contratos_beneficiadas')}),"
                "  SUM(CASE WHEN ("
                "    UPPER(COALESCE(tipo_pacote,'')) LIKE '%POSTERGA%' OR"
                "    UPPER(COALESCE(manobra,''))     LIKE '%POSTERGA%' OR"
                "    UPPER(COALESCE(tipo_pacote,'')) LIKE '%PLPT%'"
                "  ) THEN 1 ELSE 0 END)"
                f" FROM obras{where_clause}",
                params,
            )
            row = cursor.fetchone()
            if row is None:
                return {"ok": False, "error": "sem dados"}
            obras_total = int(row[0] or 0)
            capex_total = float(row[1] or 0.0)
            km_total = float(row[2] or 0.0)
            contas_benef = int(row[3] or 0)
            postergacoes = int(row[4] or 0)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"sql: {exc}",
                    "ano": ano_s, "capex_total": 0.0, "obras_total": 0,
                    "km_total": 0.0, "contas_beneficiadas": 0,
                    "postergacoes": 0, "ano_dominante": None}

        # Ano dominante: util quando ano nao foi informado.
        ano_dominante = None
        if not ano_s:
            try:
                cursor.execute(
                    "SELECT ano_, COUNT(*) c FROM obras "
                    "WHERE ano_ IS NOT NULL AND TRIM(ano_)<>'' "
                    "GROUP BY ano_ ORDER BY c DESC LIMIT 1"
                )
                ar = cursor.fetchone()
                ano_dominante = str(ar[0]).strip() if ar and ar[0] is not None else None
            except Exception:  # noqa: BLE001
                pass
        else:
            ano_dominante = ano_s
        return {
            "ok": True, "error": "",
            "ano": ano_s, "ano_dominante": ano_dominante,
            "capex_total": capex_total,
            "obras_total": obras_total,
            "km_total": km_total,
            "contas_beneficiadas": contas_benef,
            "postergacoes": postergacoes,
        }

    # ------------------------------------------------------------------
    # Passo 6.2 (Resumo / volumetria por regional): agrega obras +
    # valor_obra por nome_regional. Filtro opcional ano_. Resultado
    # ordenado por valor desc (maior em cima -- igual ao mock).
    # ------------------------------------------------------------------
    def resumo_volumetria_regional(
        self, ano: Any = "", cods: Any = None,
    ) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "ano": str(ano or ""), "items": []}
        ano_s = str(ano or "").strip()
        try:
            cursor = db._get_cursor()
            if cursor is None:
                return {"ok": False, "error": "cursor", "ano": ano_s, "items": []}
            where, _params = self._build_resumo_where(ano_s, cods)
            params: tuple[Any, ...] = tuple(_params)
            cursor.execute(
                "SELECT "
                "  UPPER(TRIM(COALESCE(nome_regional,'(SEM REGIONAL)'))),"
                "  COUNT(*),"
                f" SUM({self._sql_to_real('valor_obra')})"
                f" FROM obras{where} "
                "GROUP BY UPPER(TRIM(COALESCE(nome_regional,'(SEM REGIONAL)')))",
                params,
            )
            rows = cursor.fetchall() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"sql: {exc}", "ano": ano_s, "items": []}

        items = [
            {
                "regional": str(r[0] or ""),
                "obras": int(r[1] or 0),
                "valor": float(r[2] or 0.0),
            }
            for r in rows
        ]
        items.sort(key=lambda x: x["valor"], reverse=True)
        return {"ok": True, "error": "", "ano": ano_s, "items": items}

    # ------------------------------------------------------------------
    # Passo 6.3 (Resumo / distribuicao por Pacote): agrega valor_obra por
    # tipo_pacote, calcula percentual sobre o total e devolve ordenado
    # desc. Mantemos a categoria 'Outros' para pacotes vazios.
    # Filtro opcional ano_.
    # ------------------------------------------------------------------
    PACOTE_COLOR_MAP = {
        "MERCADO":              "oklch(0.55 0.13 250)",
        "CONFIABILIDADE":       "oklch(0.62 0.13 155)",
        "INTERLIGAÇÃO UDE":     "oklch(0.75 0.14 85)",
        "INTERLIGACAO UDE":     "oklch(0.75 0.14 85)",
        "INTERLIGAÇÃO DE UDE":  "oklch(0.75 0.14 85)",
        "INTERLIGACAO DE UDE":  "oklch(0.75 0.14 85)",
        "SOLICITAÇÃO REGIONAL": "oklch(0.6 0.13 230)",
        "SOLICITACAO REGIONAL": "oklch(0.6 0.13 230)",
        "PLPT":                 "oklch(0.5 0.18 290)",
        "OUTROS":               "var(--text-soft)",
    }

    def pacotes_distribution(
        self, ano: Any = "", cods: Any = None,
    ) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "ano": str(ano or ""), "items": []}
        ano_s = str(ano or "").strip()
        try:
            cursor = db._get_cursor()
            if cursor is None:
                return {"ok": False, "error": "cursor", "ano": ano_s, "items": []}
            where, _params = self._build_resumo_where(ano_s, cods)
            params: tuple[Any, ...] = tuple(_params)
            cursor.execute(
                "SELECT "
                "  TRIM(COALESCE(tipo_pacote,'')),"
                "  COUNT(*),"
                f" SUM({self._sql_to_real('valor_obra')})"
                f" FROM obras{where} GROUP BY TRIM(COALESCE(tipo_pacote,''))",
                params,
            )
            rows = cursor.fetchall() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"sql: {exc}",
                    "ano": ano_s, "items": []}

        # Agrega entradas vazias em "Outros".
        bucket_outros = {"label": "Outros", "obras": 0, "valor": 0.0}
        explicit: list[dict[str, Any]] = []
        for r in rows:
            label = str(r[0] or "").strip()
            obras = int(r[1] or 0)
            valor = float(r[2] or 0.0)
            if not label:
                bucket_outros["obras"] += obras
                bucket_outros["valor"] += valor
            else:
                explicit.append({"label": label, "obras": obras, "valor": valor})
        explicit.sort(key=lambda x: x["valor"], reverse=True)
        if bucket_outros["obras"]:
            explicit.append(bucket_outros)

        total_valor = sum(x["valor"] for x in explicit)
        for x in explicit:
            x["pct"] = (x["valor"] / total_valor * 100.0) if total_valor else 0.0
            key = x["label"].upper()
            x["color"] = self.PACOTE_COLOR_MAP.get(key, "var(--text-soft)")
        return {
            "ok": True, "error": "", "ano": ano_s,
            "items": explicit, "total_valor": total_valor,
        }

    # ------------------------------------------------------------------
    # Passo 6.4 (Resumo / tabela completa "Quadro de Volumetria"):
    # 9 colunas por regional + linha TOTAL agregada. Colunas numericas
    # tratadas como REAL via _sql_to_real (string pt-BR).
    # ------------------------------------------------------------------
    def resumo_regional_table(
        self, ano: Any = "", cods: Any = None,
    ) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "ano": str(ano or ""), "items": [], "total": None}
        ano_s = str(ano or "").strip()
        try:
            cursor = db._get_cursor()
            if cursor is None:
                return {"ok": False, "error": "cursor",
                        "ano": ano_s, "items": [], "total": None}
            where, _params = self._build_resumo_where(ano_s, cods)
            params: tuple[Any, ...] = tuple(_params)
            # Por regional
            cursor.execute(
                "SELECT "
                "  UPPER(TRIM(COALESCE(nome_regional,'(SEM REGIONAL)'))) AS reg,"
                "  COUNT(*) AS obras,"
                f" SUM({self._sql_to_real('quantidade_material')}) AS km,"
                f" AVG({self._sql_to_real('tensao_media_final')}) AS tensao_med,"
                f" AVG({self._sql_to_real('chi_final')}) AS chi,"
                f" AVG({self._sql_to_real('ci_final')}) AS ci,"
                f" AVG({self._sql_to_real('carregamento_final')}) AS carreg,"
                f" SUM({self._sql_to_real('contas_contratos_beneficiadas')}) AS contas,"
                f" SUM({self._sql_to_real('valor_obra')}) AS valor"
                f" FROM obras{where}"
                " GROUP BY UPPER(TRIM(COALESCE(nome_regional,'(SEM REGIONAL)')))",
                params,
            )
            rows = cursor.fetchall() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"sql: {exc}",
                    "ano": ano_s, "items": [], "total": None}

        items = []
        for r in rows:
            items.append({
                "regional":  str(r[0] or ""),
                "obras":     int(r[1] or 0),
                "km":        float(r[2] or 0.0),
                "tensao":    float(r[3] or 0.0),
                "chi":       float(r[4] or 0.0),
                "ci":        float(r[5] or 0.0),
                "carreg":    float(r[6] or 0.0),
                "contas":    int(r[7] or 0),
                "valor":     float(r[8] or 0.0),
            })
        items.sort(key=lambda x: x["valor"], reverse=True)

        # Total agregado: medias ponderadas pelo numero de obras quando
        # faz sentido (tensao/CHI/CI/carregamento). Soma para obras/km
        # /contas/valor.
        total_obras = sum(i["obras"] for i in items)
        total_km = sum(i["km"] for i in items)
        total_contas = sum(i["contas"] for i in items)
        total_valor = sum(i["valor"] for i in items)

        def _wavg(field: str) -> float:
            num = sum(i[field] * i["obras"] for i in items)
            return (num / total_obras) if total_obras else 0.0

        total_row = {
            "regional": "TOTAL",
            "obras":    total_obras,
            "km":       total_km,
            "tensao":   _wavg("tensao"),
            "chi":      _wavg("chi"),
            "ci":       _wavg("ci"),
            "carreg":   _wavg("carreg"),
            "contas":   total_contas,
            "valor":    total_valor,
        }
        return {"ok": True, "error": "", "ano": ano_s,
                "items": items, "total": total_row}

    # ------------------------------------------------------------------
    # Fase C1 (set_extra_keys_for_pi):
    # Permite ao usuario salvar (no config) chaves extras adicionais para
    # um PI base. Estas chaves se somam a `metadata.calculo.modulos_extras`
    # (config-driven) na proxima rodada de calcular_valor_obra.
    # ------------------------------------------------------------------
    def set_modulos_extras(
        self, pi_base: Any = "", extras: Any = None,
    ) -> dict[str, Any]:
        pi = str(pi_base or "").strip().upper()
        if not pi:
            return {"ok": False, "error": "pi vazio", "extras": []}
        if not isinstance(extras, (list, tuple)):
            return {"ok": False, "error": "extras nao e lista", "extras": []}
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                set_extra_keys_for_pi,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}", "extras": []}
        normalized = [str(k).strip().upper() for k in extras if str(k or "").strip()]
        try:
            set_extra_keys_for_pi(pi, normalized)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}", "extras": []}
        # Atualiza nosso cache de config local pra refletir a mudanca.
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            self._config = ConfigManager.load_config() or self._config
        except Exception:  # noqa: BLE001
            pass
        return {"ok": True, "error": "", "pi": pi, "extras": normalized}

    # ------------------------------------------------------------------
    # Fase A11 (pi_metadata_service.obter_modulos_extras):
    # Retorna a lista de chaves extras (modulo_extra do PI + ATERRAMENTO
    # se exige_aterramento + last_pi_extra_map salvo no config). Util
    # para a UI de Cadastro mostrar as chaves que serao somadas no
    # calculo de valor_obra.
    # ------------------------------------------------------------------
    def get_modulos_extras(self, pi_base: Any = "") -> dict[str, Any]:
        pi = str(pi_base or "").strip()
        if not pi:
            return {"ok": False, "error": "pi vazio", "extras": [], "pi": pi}
        try:
            from core.services.pi_metadata_service import (  # type: ignore[import-not-found]
                obter_modulos_extras,
            )
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}",
                    "extras": [], "pi": pi}
        try:
            cfg = ConfigManager.load_config() or {}
        except Exception:  # noqa: BLE001
            cfg = {}
        try:
            extras = list(obter_modulos_extras(pi, cfg) or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"obter_modulos_extras: {exc}",
                    "extras": [], "pi": pi}
        return {"ok": True, "error": "", "pi": pi, "extras": extras}

    # ------------------------------------------------------------------
    # Fase A10 (pi_metadata_service.obter_descricao_template +
    #           get_descricao_obra_from_template):
    # devolve o template do PI e renderiza com dados do form (placeholder
    # {col} substituidos pelos valores).
    # ------------------------------------------------------------------
    def get_descricao_template(self, pi: Any = "") -> dict[str, Any]:
        pi_s = str(pi or "").strip()
        try:
            from core.services.pi_metadata_service import (  # type: ignore[import-not-found]
                obter_descricao_template,
            )
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}", "template": "",
                    "pi": pi_s}
        try:
            cfg = ConfigManager.load_config() or {}
        except Exception:  # noqa: BLE001
            cfg = {}
        # Tambem honra o override por PI: config.descricao_obra_templates[PI]
        custom = ""
        try:
            templates = cfg.get("descricao_obra_templates") or {}
            if isinstance(templates, dict):
                custom = str(templates.get(pi_s.upper()) or "")
        except Exception:  # noqa: BLE001
            custom = ""
        try:
            default = obter_descricao_template(pi_s, cfg) or ""
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"obter_template: {exc}",
                    "template": "", "pi": pi_s}
        return {"ok": True, "error": "",
                "pi": pi_s,
                "template_default": default,
                "template_custom": custom,
                "template": custom or default}

    def aplicar_template_descricao(
        self, pi_base: Any = "", dados: Any = None,
    ) -> dict[str, Any]:
        pi = str(pi_base or "").strip()
        if not isinstance(dados, dict):
            dados = {}
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                get_descricao_obra_from_template,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}", "descricao": ""}
        try:
            descricao = get_descricao_obra_from_template(pi, dict(dados))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"render: {exc}", "descricao": ""}
        if descricao is None:
            return {"ok": False, "error": "sem template para o PI",
                    "descricao": ""}
        return {"ok": True, "error": "", "descricao": descricao}

    # ------------------------------------------------------------------
    # Fase A6 (resumo_service.montar_resumo_ganhos_projeto):
    # consolida ganhos de TODAS as obras de um nome_projeto, por
    # alimentador. Equivalente a MainWindow.popular_resumo_ganhos_projeto.
    # ------------------------------------------------------------------
    def resumo_ganhos_projeto(self, nome_projeto: Any = "") -> dict[str, Any]:
        nome_s = str(nome_projeto or "").strip()
        if not nome_s:
            return {"ok": False, "error": "nome_projeto vazio",
                    "linhas": [], "projeto": ""}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "linhas": [], "projeto": nome_s}
        try:
            cursor = db._get_cursor()
            if cursor is None:
                return {"ok": False, "error": "cursor",
                        "linhas": [], "projeto": nome_s}
            cursor.execute(
                "SELECT alimentador_principal, alimentadores_beneficiados,"
                " codigo_item, ganhos_totais_depois"
                " FROM obras"
                " WHERE UPPER(TRIM(COALESCE(nome_projeto,'')))=UPPER(TRIM(?))",
                (nome_s,),
            )
            rows = cursor.fetchall() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"sql: {exc}",
                    "linhas": [], "projeto": nome_s}

        try:
            from core.services.resumo_service import (  # type: ignore[import-not-found]
                montar_resumo_ganhos_projeto,
            )
            from codigo5_coplan import (  # type: ignore[import-not-found]
                DEFAULT_CRITERIOS,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}",
                    "linhas": [], "projeto": nome_s}

        criterios = dict(DEFAULT_CRITERIOS)
        criterios.update((self._config or {}).get("criterios_planejamento") or {})

        cols_used = [
            "alimentador_principal",
            "alimentadores_beneficiados",
            "codigo_item",
            "ganhos_totais_depois",
        ]
        try:
            linhas = montar_resumo_ganhos_projeto(
                rows=list(rows), cols=cols_used, criterios=criterios,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"montar_resumo_ganhos_projeto: {exc}",
                    "linhas": [], "projeto": nome_s}

        out: list[dict[str, Any]] = []
        for ln in linhas:
            out.append({
                "alimentador":   ln.alimentador,
                "carregamento":  {"text": ln.carregamento.text, "ok": ln.carregamento.ok},
                "tensao":        {"text": ln.tensao.text, "ok": ln.tensao.ok},
                "clientes_text": ln.clientes_text,
            })
        return {"ok": True, "error": "", "projeto": nome_s,
                "linhas": out, "obras_count": len(rows)}

    def list_projetos(self) -> dict[str, Any]:
        """Lista nomes_projeto distintos no banco para usar em selects."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "items": [], "error": err or "db indisponivel"}
        try:
            cursor = db._get_cursor()
            if cursor is None:
                return {"ok": False, "items": [], "error": "cursor"}
            cursor.execute(
                "SELECT DISTINCT TRIM(nome_projeto) FROM obras"
                " WHERE TRIM(COALESCE(nome_projeto,''))<>''"
                " ORDER BY 1"
            )
            items = [str(r[0]).strip() for r in cursor.fetchall() if r[0]]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "items": [], "error": f"sql: {exc}"}
        return {"ok": True, "items": items, "error": ""}

    # ------------------------------------------------------------------
    # F16 - Atualizar Projeto navegacional (RB-3, AtualizarObraMixin):
    # Le todas as obras de um nome_projeto + tipo_pacote (mesma logica
    # de iniciar_atualizacao_projeto do desktop). O JS mantem o estado
    # de navegacao (index, edited, snapshots) e chama save_obra ao
    # finalizar para cada obra alterada.
    # ------------------------------------------------------------------
    def projeto_fetch_obras(
        self, nome_projeto: Any, tipo_pacote: Any = "",
    ) -> dict[str, Any]:
        """Retorna as obras de um projeto (filtradas opcionalmente por
        tipo_pacote) ordenadas por codigo_item. Equivalente a
        DatabaseManager.fetch_by_project + filtro pacote do desktop."""
        nome = str(nome_projeto or "").strip()
        if not nome:
            return {"ok": False, "obras": [], "error": "nome_projeto vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "obras": [],
                    "error": err or "db indisponivel"}
        try:
            cols = list(db.get_column_names() or [])
            todas = list(db.fetch_by_project(
                nome, order_by_codigo_item=True) or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "obras": [],
                    "error": f"fetch_by_project: {exc}"}

        pacote_ref = str(tipo_pacote or "").strip()
        if pacote_ref:
            i_pac = (cols.index("tipo_pacote")
                     if "tipo_pacote" in cols else -1)
            if i_pac >= 0:
                filtradas = [o for o in todas
                             if str(o[i_pac] or "").strip() == pacote_ref]
            else:
                filtradas = todas
        else:
            filtradas = todas

        # Converte para list of dicts (mais facil pro JS)
        obras = [dict(zip(cols, row)) for row in filtradas]
        i_cod = cols.index("cod") if "cod" in cols else -1
        cods = [str(row[i_cod] or "").strip()
                for row in filtradas if i_cod >= 0] if i_cod >= 0 else []
        return {
            "ok": True,
            "obras": obras,
            "cods": cods,
            "total": len(obras),
            "total_no_projeto": len(todas),
            "ignoradas_outro_pacote": len(todas) - len(filtradas),
            "nome_projeto": nome,
            "tipo_pacote": pacote_ref,
            "columns": cols,
            "error": "",
        }

    # ------------------------------------------------------------------
    # [M030] Wrappers de "Atualizar Projeto" (state machine vive no JS).
    # projeto_iniciar:  carrega lista de obras (alias de projeto_fetch_obras
    #                   com nome alinhado ao plano);
    # projeto_finalizar: itera payloads e salva cada um em sequencia,
    #                    propagando o motivo da primeira obra critica para
    #                    todas as demais (parity com modo "atualizar
    #                    projeto" do desktop).
    # projeto_avancar/voltar/cancelar vivem 100% no JS e nao precisam de
    # endpoint backend (sao pura troca de indice + clearForm).
    # ------------------------------------------------------------------
    def projeto_iniciar(
        self, nome_projeto: Any, tipo_pacote: Any = "",
    ) -> dict[str, Any]:
        """Alias semantico de projeto_fetch_obras + idx inicial. JS
        recebe o lote, mantem `idx`, `pendingPayloads[]` e dispara
        projeto_finalizar() ao final."""
        out = self.projeto_fetch_obras(nome_projeto, tipo_pacote)
        if not out.get("ok"):
            return out
        out["idx"] = 0
        return out

    def projeto_finalizar(
        self, payloads: Any = None, motivo: Any = "",
    ) -> dict[str, Any]:
        """Salva em sequencia todos os payloads do lote 'atualizar
        projeto'. Reusa save_obra; propaga o mesmo `motivo` para todas
        as obras (parity com codigo5_coplan: 'modo atualizar projeto
        reutiliza motivo da primeira obra')."""
        if not isinstance(payloads, list):
            return {"ok": False, "salvos": 0, "falhas": [],
                    "error": "payloads nao eh lista"}
        motivo_s = str(motivo or "").strip()
        salvos: list[dict[str, Any]] = []
        falhas: list[dict[str, Any]] = []
        for i, p in enumerate(payloads):
            if not isinstance(p, dict):
                falhas.append({"idx": i, "cod": "",
                               "error": "payload nao eh dict"})
                continue
            payload = dict(p)
            if motivo_s and "motivo_alteracao" not in payload:
                payload["motivo_alteracao"] = motivo_s
            try:
                resp = self.save_obra(payload)
            except Exception as exc:  # noqa: BLE001
                falhas.append({"idx": i, "cod": str(p.get("cod") or ""),
                               "error": f"save_obra: {exc}"})
                continue
            if resp.get("ok"):
                salvos.append({"idx": i, "cod": resp.get("cod") or "",
                               "mode": resp.get("mode") or ""})
            else:
                falhas.append({"idx": i, "cod": resp.get("cod") or "",
                               "error": resp.get("error") or "save falhou",
                               "requires_motivo": resp.get("requires_motivo"),
                               "blocked": resp.get("blocked")})
        return {
            "ok": not falhas,
            "salvos": len(salvos),
            "falhas": falhas,
            "salvos_detalhe": salvos,
            "motivo_aplicado": motivo_s,
            "error": "" if not falhas else f"{len(falhas)} falha(s)",
        }

    # ------------------------------------------------------------------
    # Fase A5 (resumo_service.montar_quadro_resumo_from_ganhos):
    # quadro de "Resumo dos Ganhos por Alimentador" para UMA obra,
    # baseado no campo `ganhos_totais_depois`. Reproduz
    # MainWindow.popular_quadro_resumo_from_ganhos_depois.
    # Aceita dois modos: por COD (le obra do banco) ou por payload
    # explicito (alim_principal, alim_benef, ganhos_str).
    # ------------------------------------------------------------------
    def quadro_resumo_ganhos(
        self,
        cod: Any = "",
        alimentador_principal: Any = "",
        alimentadores_beneficiados: Any = "",
        ganhos_totais_depois: Any = "",
    ) -> dict[str, Any]:
        alim_p = str(alimentador_principal or "").strip()
        alim_b = str(alimentadores_beneficiados or "")
        ganhos = str(ganhos_totais_depois or "")
        cod_s = str(cod or "").strip()

        # Modo COD: carrega da obra existente.
        if cod_s and not (alim_p or ganhos):
            db, err = self._ensure_db_connected()
            if err or db is None:
                return {"ok": False, "error": err or "db indisponivel",
                        "linhas": []}
            try:
                row = db.fetch_by_cod(cod_s)
                cols = list(db.get_column_names() or [])
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "error": f"fetch_by_cod: {exc}",
                        "linhas": []}
            if not row:
                return {"ok": False, "error": f"obra nao encontrada: {cod_s}",
                        "linhas": []}
            row_dict = {c: row[i] if i < len(row) else ""
                        for i, c in enumerate(cols)}
            alim_p = str(row_dict.get("alimentador_principal") or "").strip()
            alim_b = str(row_dict.get("alimentadores_beneficiados") or "")
            ganhos = str(row_dict.get("ganhos_totais_depois") or "")

        try:
            from core.services.resumo_service import (  # type: ignore[import-not-found]
                montar_quadro_resumo_from_ganhos,
            )
            from codigo5_coplan import (  # type: ignore[import-not-found]
                DEFAULT_CRITERIOS,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}", "linhas": []}

        criterios = dict(DEFAULT_CRITERIOS)
        criterios.update((self._config or {}).get("criterios_planejamento") or {})

        try:
            linhas = montar_quadro_resumo_from_ganhos(
                alimentador_principal=alim_p,
                alimentadores_beneficiados=alim_b,
                ganhos_totais_depois=ganhos,
                criterios=criterios,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"montar_quadro: {exc}",
                    "linhas": []}

        out: list[dict[str, Any]] = []
        for ln in linhas:
            out.append({
                "alimentador":   ln.alimentador,
                "carregamento":  {"text": ln.carregamento.text, "ok": ln.carregamento.ok},
                "tensao":        {"text": ln.tensao.text, "ok": ln.tensao.ok},
                "clientes_text": ln.clientes_text,
            })
        return {"ok": True, "error": "", "linhas": out}

    # ------------------------------------------------------------------
    # Fase A4 (resumo_service.montar_volumetria_financeiro):
    # pivot por (PI x Ano) com Valor/Fisico, formatado em pt-BR.
    # Reproduz MainWindow.popular_volumetria_financeiro do desktop.
    # ------------------------------------------------------------------
    def resumo_volumetria_financeiro(
        self, ano: Any = "", cods: Any = None,
    ) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "ano": str(ano or ""), "cabecalhos": ["PI"], "linhas": []}
        try:
            cursor = db._get_cursor()
            if cursor is None:
                return {"ok": False, "error": "cursor",
                        "ano": str(ano or ""),
                        "cabecalhos": ["PI"], "linhas": []}
            ano_s = str(ano or "").strip()
            where, _params = self._build_resumo_where(ano_s, cods)
            params: tuple[Any, ...] = tuple(_params)
            cursor.execute(
                "SELECT projeto_investimento, ano_, valor_obra,"
                " quantidade_material"
                f" FROM obras{where}",
                params,
            )
            rows = cursor.fetchall() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"sql: {exc}",
                    "ano": str(ano or ""),
                    "cabecalhos": ["PI"], "linhas": []}

        try:
            from core.services.resumo_service import (  # type: ignore[import-not-found]
                montar_volumetria_financeiro,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}",
                    "ano": str(ano or ""),
                    "cabecalhos": ["PI"], "linhas": []}

        obras_visiveis = [
            {
                "projeto_investimento": r[0],
                "ano_": r[1],
                "valor_obra": r[2],
                "quantidade_material": r[3],
            }
            for r in rows
        ]
        try:
            vol = montar_volumetria_financeiro(obras_visiveis)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"montar_volumetria: {exc}",
                    "ano": str(ano or ""),
                    "cabecalhos": ["PI"], "linhas": []}
        return {
            "ok": True, "error": "", "ano": str(ano or ""),
            "cabecalhos": list(vol.cabecalhos),
            "linhas": [list(linha) for linha in vol.linhas],
        }

    # ------------------------------------------------------------------
    # Passo 7.1 (Config / Empresa): get/save dos campos da empresa.
    #   * sigla            (config.empresa_sigla)
    #   * razao_social     (config.razao_social  -- chave NOVA, nao quebra
    #                       o legado, mas aparece no JSON ao salvar)
    #   * caminho_db       (config.obras)
    # Apoio NAO esta mais aqui: agora e DB-backed (tabelas apoio_*).
    # Use o botao "Atualizar apoio" no card Empresa para reimportar.
    # ------------------------------------------------------------------
    def get_config_empresa(self) -> dict[str, Any]:
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"config: {exc}"}
        return {
            "ok": True, "error": "",
            "sigla":         str(cfg.get("empresa_sigla") or "").strip(),
            "razao_social":  str(cfg.get("razao_social") or "").strip(),
            "caminho_db":    str(cfg.get("obras") or "").strip(),
            "caminho_pasta_ganhos": str(cfg.get("caminho_pasta_ganhos") or "").strip(),
        }

    def save_config_empresa(self, payload: Any) -> dict[str, Any]:
        if not isinstance(payload, dict):
            return {"ok": False, "error": "payload nao e dict"}
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager, EMPRESA_SIGLAS_VALIDAS,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}"}

        partial: dict[str, Any] = {}

        if "sigla" in payload:
            sigla = str(payload.get("sigla") or "").strip().upper()
            if sigla and sigla not in EMPRESA_SIGLAS_VALIDAS:
                return {"ok": False, "error":
                        f"Sigla invalida. Permitidas: "
                        f"{', '.join(sorted(EMPRESA_SIGLAS_VALIDAS))}"}
            partial["empresa_sigla"] = sigla
        if "razao_social" in payload:
            partial["razao_social"] = str(payload.get("razao_social") or "").strip()
        if "caminho_db" in payload:
            partial["obras"] = str(payload.get("caminho_db") or "").strip()
        if "caminho_pasta_ganhos" in payload:
            partial["caminho_pasta_ganhos"] = str(
                payload.get("caminho_pasta_ganhos") or ""
            ).strip()

        if not partial:
            return {"ok": False, "error": "payload sem campos conhecidos"}
        try:
            ConfigManager.save_config(partial)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        # Invalida cache interno -- proxima chamada le do disco.
        self._config = None
        # Se o caminho do banco mudou, esquece o cache de paths
        # conectados (forca reconnect na proxima chamada).
        if "obras" in partial:
            self._connected_paths.clear()
        return {"ok": True, "error": "", "saved": list(partial.keys())}

    def pick_db_file(self) -> dict[str, Any]:
        return self._pick_file_with_filters(
            "Banco SQLite (*.db;*.sqlite;*.sqlite3)",
        )

    def pick_apoio_file(self) -> dict[str, Any]:
        return self._pick_file_with_filters(
            "Planilha de Apoio (*.xlsx;*.xlsm;*.xls;*.csv)",
        )

    # ------------------------------------------------------------------
    # Passo 7.2 (Config / Criterios de Planejamento): persiste os 8
    # campos editaveis do card "Criterios de Planejamento (vigentes)".
    #   Campo no mock                      -> Coluna do config.json
    #   Tensao Min. (pu)                   -> criterios_planejamento.tensao_min
    #   Tensao Max. (pu)                   -> criterios_planejamento.tensao_max
    #   Carregamento Max. (%)              -> criterios_planejamento.carregamento_limite_sim_ou_vazio
    #   CHI Minimo                         -> criterios_planejamento.chi_min  (chave nova)
    #   CI Minimo                          -> criterios_planejamento.ci_min   (chave nova)
    #   Piora Mercado (%)                  -> piora_mercado.carregamento_percentual
    #   Anos de horizonte                  -> piora_mercado.anos_horizonte
    #   Postergacao max. (anos)            -> piora_mercado.postergacao_max_anos (chave nova)
    # As chaves "novas" (chi_min, ci_min, postergacao_max_anos) coexistem
    # com os defaults legados sem quebrar nada, e ficam disponiveis para o
    # get_criterios() do Passo 5.3 aplicar nas regras.
    # ------------------------------------------------------------------
    @staticmethod
    def _to_float(v: Any) -> float | None:
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        s = re.sub(r"[^0-9.\-]", "", s)
        try:
            return float(s)
        except ValueError:
            return None

    def save_criterios(self, payload: Any) -> dict[str, Any]:
        if not isinstance(payload, dict):
            return {"ok": False, "error": "payload nao e dict"}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}"}

        crit_keys_pt = (
            ("tensao_min",                      "tensao_min"),
            ("tensao_max",                      "tensao_max"),
            ("carregamento_max",                "carregamento_limite_sim_ou_vazio"),
            ("carregamento_limite_sim_ou_vazio","carregamento_limite_sim_ou_vazio"),
            ("carregamento_limite_nao",         "carregamento_limite_nao"),
            ("chi_min",                         "chi_min"),
            ("ci_min",                          "ci_min"),
            ("clientes_maximo",                 "clientes_maximo"),
        )
        piora_keys = (
            ("piora_mercado",        "carregamento_percentual"),
            ("carregamento_percentual","carregamento_percentual"),
            ("tensao_delta",         "tensao_delta"),
            ("anos_horizonte",       "anos_horizonte"),
            ("postergacao_max_anos", "postergacao_max_anos"),
            ("postergacao_max",      "postergacao_max_anos"),
        )

        crit_part: dict[str, Any] = {}
        for inkey, outkey in crit_keys_pt:
            if inkey in payload:
                v = self._to_float(payload[inkey])
                if v is not None:
                    crit_part[outkey] = v
        piora_part: dict[str, Any] = {}
        for inkey, outkey in piora_keys:
            if inkey in payload:
                v = self._to_float(payload[inkey])
                if v is not None:
                    # anos_horizonte / postergacao_max_anos sao inteiros
                    if outkey in ("anos_horizonte", "postergacao_max_anos"):
                        piora_part[outkey] = int(round(v))
                    else:
                        piora_part[outkey] = v

        if not crit_part and not piora_part:
            return {"ok": False, "error": "payload sem campos conhecidos"}

        partial: dict[str, Any] = {}
        if crit_part:
            partial["criterios_planejamento"] = crit_part
        if piora_part:
            partial["piora_mercado"] = piora_part
        try:
            ConfigManager.save_config(partial)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return {"ok": True, "error": "",
                "saved_criterios": list(crit_part.keys()),
                "saved_piora": list(piora_part.keys())}

    # ------------------------------------------------------------------
    # Passo 7.3 (Config / Templates + PI_BASE):
    #   * get_templates / save_templates    -> config.descricao_obra_templates
    #   * list_pi_base_custom               -> config.pi_base_custom + DEFAULT pi_metadata
    #   * add_pi_base_custom / remove        -> mutate pi_base_custom list
    #   * get_pi_base_map / save_pi_base_map -> config.pi_base_map (dict de
    #     mapeamentos PI nao-canonicos -> base curta)
    # ------------------------------------------------------------------
    def get_templates(self) -> dict[str, Any]:
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"config: {exc}", "items": {}}
        items = cfg.get("descricao_obra_templates") or {}
        if not isinstance(items, dict):
            items = {}
        return {"ok": True, "error": "", "items": items}

    def save_templates(self, payload: Any) -> dict[str, Any]:
        if not isinstance(payload, dict):
            return {"ok": False, "error": "payload nao e dict"}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}"}
        clean = {
            str(k).strip().upper(): str(v) for k, v in payload.items()
            if str(k).strip()
        }
        try:
            ConfigManager.save_config({"descricao_obra_templates": clean})
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return {"ok": True, "error": "", "count": len(clean)}

    # Bloco 2 (Templates de Descricao) - Auditoria #11/#12/#13.
    # save_config faz deep merge: para REMOVER chave de templates precisa
    # ler config inteiro, mutar e salvar com overwrite=True.
    def delete_pi_template(self, pi: Any = "") -> dict[str, Any]:
        """Restaura padrao do PI: remove a chave de descricao_obra_templates."""
        pi_s = str(pi or "").strip().upper()
        if not pi_s:
            return {"ok": False, "error": "pi vazio"}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}"}
        try:
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"load: {exc}"}
        templates_cfg = dict(cfg.get("descricao_obra_templates") or {})
        had = pi_s in templates_cfg
        templates_cfg.pop(pi_s, None)
        cfg["descricao_obra_templates"] = templates_cfg
        try:
            ConfigManager.save_config(cfg, overwrite=True)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return {"ok": True, "error": "", "removed": had, "pi": pi_s}

    def restore_all_templates(self) -> dict[str, Any]:
        """Restaura todos: zera descricao_obra_templates."""
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}"}
        try:
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"load: {exc}"}
        before = len(dict(cfg.get("descricao_obra_templates") or {}))
        cfg["descricao_obra_templates"] = {}
        try:
            ConfigManager.save_config(cfg, overwrite=True)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return {"ok": True, "error": "", "removed": before}

    def template_preview_render(
        self, pi: Any = "", template: Any = "",
    ) -> dict[str, Any]:
        """Renderiza um template (texto editado, NAO salvo) usando dados
        da primeira obra do banco com pi_base = pi. Fallback: dict vazio
        (placeholders viram strings vazias)."""
        pi_s = str(pi or "").strip().upper()
        tpl = str(template or "")
        if not tpl:
            return {"ok": True, "error": "", "rendered": "",
                    "obra_cod": "", "obra_count": 0}
        try:
            from runtime.text_utils import render_template  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}",
                    "rendered": "", "obra_cod": "", "obra_count": 0}
        data: dict[str, Any] = {}
        obra_cod = ""
        obra_count = 0
        db, _err = self._ensure_db_connected()
        if db is not None:
            try:
                cursor = db._get_cursor()
                if cursor is not None:
                    cursor.execute(
                        "SELECT * FROM obras WHERE UPPER(TRIM(COALESCE(pi_base,'')))=?"
                        " LIMIT 1",
                        (pi_s,),
                    )
                    row = cursor.fetchone()
                    if row:
                        cols = [d[0] for d in (cursor.description or [])]
                        data = {cols[i]: row[i] for i in range(len(cols))}
                        obra_cod = str(data.get("cod") or "")
                        obra_count = 1
            except Exception:  # noqa: BLE001
                data = {}
        # Normaliza valores None -> '' para render mais limpo
        clean = {k: ("" if v is None else str(v)) for k, v in data.items()}
        try:
            rendered = render_template(tpl, clean)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"render: {exc}",
                    "rendered": "", "obra_cod": obra_cod,
                    "obra_count": obra_count}
        return {"ok": True, "error": "", "rendered": rendered,
                "obra_cod": obra_cod, "obra_count": obra_count}

    def get_template_field_candidates(self) -> dict[str, Any]:
        """Lista de colunas para inserir como {placeholder} no template.
        Espelho de _get_visualizar_columns_candidates do desktop:
        ORDERED_COLUMNS + colunas reais do banco, deduplicado."""
        try:
            from codigo5_coplan import ORDERED_COLUMNS  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}", "items": []}
        ordered = [str(c) for c in (ORDERED_COLUMNS or []) if c]
        cols_db: list[str] = []
        db, _err = self._ensure_db_connected()
        if db is not None:
            try:
                cols_db = list(db.get_column_names() or [])
            except Exception:  # noqa: BLE001
                cols_db = []
        if not cols_db:
            return {"ok": True, "error": "", "items": ordered}
        seen = set()
        out: list[str] = []
        for c in ordered:
            if c in cols_db and c not in seen:
                out.append(c)
                seen.add(c)
        for c in cols_db:
            if c not in seen:
                out.append(c)
                seen.add(c)
        return {"ok": True, "error": "", "items": out}

    def list_pi_base_custom(self) -> dict[str, Any]:
        """Devolve {ok, custom: [...], all: [...], hidden_defaults: [...]}.
        - custom: lista mutavel salva em config.pi_base_custom.
        - hidden_defaults: defaults que o usuario removeu (config.
          pi_base_hidden_defaults) -- filtrados de `all`.
        - all: defaults visiveis (defaults menos hidden) + custom."""
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"config: {exc}",
                    "custom": [], "all": [], "hidden_defaults": []}
        custom_raw = cfg.get("pi_base_custom") or []
        if not isinstance(custom_raw, list):
            custom_raw = []
        custom = [str(c).strip() for c in custom_raw if str(c).strip()]
        hidden_raw = cfg.get("pi_base_hidden_defaults") or []
        if not isinstance(hidden_raw, list):
            hidden_raw = []
        hidden = [str(h).strip() for h in hidden_raw if str(h).strip()]
        all_bases: list[str] = []
        try:
            from core.services.pi_metadata_service import (  # type: ignore[import-not-found]
                listar_todas_bases,
            )
            all_bases = list(listar_todas_bases(
                cfg,
                custom_bases=tuple(custom),
                hidden_defaults=tuple(hidden),
            ))
        except Exception:  # noqa: BLE001
            all_bases = list(custom)
        return {"ok": True, "error": "", "custom": custom,
                "all": all_bases, "hidden_defaults": hidden}

    def _is_default_pi_base(self, name: str) -> bool:
        """True se `name` (case-insensitive, sem acento) bate com algum
        tipo_base de DEFAULT_PI_METADATA. Usado para distinguir entre
        remover de pi_base_custom vs. ocultar via pi_base_hidden_defaults."""
        try:
            from core.services.pi_metadata_service import (  # type: ignore[import-not-found]
                DEFAULT_PI_METADATA,
                normalize_key,
            )
        except Exception:  # noqa: BLE001
            return False
        target = normalize_key(name)
        if not target:
            return False
        for entry in DEFAULT_PI_METADATA:
            base = entry.get("tipo_base") or entry.get("nome") or ""
            if normalize_key(str(base)) == target:
                return True
        return False

    def add_pi_base_custom(self, name: Any) -> dict[str, Any]:
        s = str(name or "").strip()
        if not s:
            return {"ok": False, "error": "nome vazio"}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"config: {exc}"}
        # Se `s` corresponde a um default oculto, restauramos (un-hide).
        hidden_raw = cfg.get("pi_base_hidden_defaults") or []
        if not isinstance(hidden_raw, list):
            hidden_raw = []
        upper_s = s.upper()
        new_hidden = [
            h for h in hidden_raw
            if str(h).strip().upper() != upper_s
        ]
        if len(new_hidden) != len(hidden_raw):
            try:
                ConfigManager.save_config(
                    {"pi_base_hidden_defaults": new_hidden}
                )
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "error": f"save: {exc}"}
            self._config = None
            return self.list_pi_base_custom()
        existing = cfg.get("pi_base_custom") or []
        if not isinstance(existing, list):
            existing = []
        # Dedupe case-insensitive (mas mantem capitalizacao do usuario).
        upper_set = {str(x).strip().upper() for x in existing}
        if upper_s in upper_set:
            return {"ok": False, "error": "PI ja existe"}
        # Tambem rejeita se ja eh um default visivel (nao oculto).
        if self._is_default_pi_base(s):
            return {"ok": False, "error": "PI ja existe como default"}
        existing.append(s)
        try:
            ConfigManager.save_config({"pi_base_custom": existing})
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return self.list_pi_base_custom()

    def remove_pi_base_custom(self, name: Any) -> dict[str, Any]:
        s = str(name or "").strip()
        if not s:
            return {"ok": False, "error": "nome vazio"}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"config: {exc}"}
        existing = cfg.get("pi_base_custom") or []
        if not isinstance(existing, list):
            existing = []
        new_list = [x for x in existing if str(x).strip().upper() != s.upper()]
        if len(new_list) != len(existing):
            try:
                ConfigManager.save_config(
                    {"pi_base_custom": new_list}, overwrite=False,
                )
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "error": f"save: {exc}"}
            self._config = None
            return self.list_pi_base_custom()
        # Nao estava em pi_base_custom -- talvez seja um default. Se for,
        # registramos em pi_base_hidden_defaults (mecanismo para "remover"
        # defaults sem alterar DEFAULT_PI_METADATA).
        if self._is_default_pi_base(s):
            hidden_raw = cfg.get("pi_base_hidden_defaults") or []
            if not isinstance(hidden_raw, list):
                hidden_raw = []
            upper_set = {str(x).strip().upper() for x in hidden_raw}
            if s.upper() not in upper_set:
                hidden_raw.append(s)
                try:
                    ConfigManager.save_config(
                        {"pi_base_hidden_defaults": hidden_raw}
                    )
                except Exception as exc:  # noqa: BLE001
                    return {"ok": False, "error": f"save: {exc}"}
                self._config = None
            return self.list_pi_base_custom()
        return {"ok": False, "error": "PI nao encontrada"}

    def get_pi_base_map(self) -> dict[str, Any]:
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"config: {exc}", "items": {}}
        m = cfg.get("pi_base_map") or {}
        if not isinstance(m, dict):
            m = {}
        return {"ok": True, "error": "", "items": m}

    def save_pi_base_map(self, payload: Any) -> dict[str, Any]:
        if not isinstance(payload, dict):
            return {"ok": False, "error": "payload nao e dict"}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}"}
        clean = {
            str(k).strip(): str(v).strip().upper() for k, v in payload.items()
            if str(k).strip()
        }
        try:
            ConfigManager.save_config({"pi_base_map": clean})
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return {"ok": True, "error": "", "count": len(clean)}

    # ------------------------------------------------------------------
    # Passo 7.4 (Config / Regional Map): CRUD do mapa de regionais.
    # Estrutura no config:
    #   regional_map: {
    #     <NOME>: <codigo str>           # legado simples
    #     OU
    #     <NOME>: {                       # extendido (web)
    #       codigo, superintendencia,
    #       se_prefixos: "ATB,JTP",
    #       cor: "info"|"success"|"warning"|"danger"|"violet"
    #     }
    #   }
    # Combinamos com REGIONAL_MAP do legado (read-only defaults).
    # ------------------------------------------------------------------
    REGIONAL_COR_OPCOES = ("info", "success", "warning", "danger", "violet")

    @staticmethod
    def _normalize_regional_entry(raw: Any, fallback_codigo: str = "") -> dict[str, Any]:
        if isinstance(raw, dict):
            return {
                "codigo":          str(raw.get("codigo") or fallback_codigo or "").strip(),
                "superintendencia":str(raw.get("superintendencia") or "").strip(),
                "se_prefixos":     str(raw.get("se_prefixos") or "").strip(),
                "cor":             str(raw.get("cor") or "info").strip().lower(),
            }
        # Legado: regional_map[X] = "REG-0002" (string)
        return {
            "codigo":          str(raw or fallback_codigo or "").strip(),
            "superintendencia": "",
            "se_prefixos":      "",
            "cor":              "info",
        }

    def get_regional_map_full(self) -> dict[str, Any]:
        """Devolve mapa enriquecido { NOME: {codigo, superintendencia,
        se_prefixos, cor, source: 'default'|'config'} }."""
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager, REGIONAL_MAP,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}", "items": {}}
        cfg = ConfigManager.load_config() or {}
        user = cfg.get("regional_map") or {}
        if not isinstance(user, dict):
            user = {}
        merged: dict[str, dict[str, Any]] = {}
        # Defaults primeiro
        for nome, codigo in REGIONAL_MAP.items():
            entry = self._normalize_regional_entry(codigo)
            entry["source"] = "default"
            merged[str(nome).upper()] = entry
        # Overrides + adicionais do usuario
        for nome, raw in user.items():
            key = str(nome or "").strip().upper()
            if not key:
                continue
            base = merged.get(key, {"codigo":"", "superintendencia":"",
                                     "se_prefixos":"", "cor":"info",
                                     "source":"config"})
            normalized = self._normalize_regional_entry(raw, base.get("codigo", ""))
            normalized["source"] = "config"
            merged[key] = normalized
        return {"ok": True, "error": "", "items": merged,
                "cores": list(self.REGIONAL_COR_OPCOES)}

    def save_regional_entry(
        self,
        nome: Any,
        payload: Any = None,
    ) -> dict[str, Any]:
        nome_s = str(nome or "").strip().upper()
        if not nome_s:
            return {"ok": False, "error": "nome vazio"}
        if not isinstance(payload, dict):
            return {"ok": False, "error": "payload nao e dict"}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"config: {exc}"}
        user = cfg.get("regional_map") or {}
        if not isinstance(user, dict):
            user = {}
        new_entry = self._normalize_regional_entry(payload)
        if new_entry["cor"] not in self.REGIONAL_COR_OPCOES:
            new_entry["cor"] = "info"
        user[nome_s] = new_entry
        try:
            ConfigManager.save_config({"regional_map": user})
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return self.get_regional_map_full()

    def delete_regional_entry(self, nome: Any) -> dict[str, Any]:
        nome_s = str(nome or "").strip().upper()
        if not nome_s:
            return {"ok": False, "error": "nome vazio"}
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"config: {exc}"}
        user = cfg.get("regional_map") or {}
        if not isinstance(user, dict):
            user = {}
        if nome_s not in {str(k).strip().upper() for k in user.keys()}:
            return {"ok": False,
                    "error": "regional nao esta em config (default nao pode ser removido)"}
        new_user = {
            k: v for k, v in user.items()
            if str(k).strip().upper() != nome_s
        }
        try:
            ConfigManager.save_config({"regional_map": new_user}, overwrite=False)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return self.get_regional_map_full()

    def restore_criterios_defaults(self) -> dict[str, Any]:
        """Reseta criterios_planejamento + piora_mercado aos DEFAULTS do
        legado. Mantem demais chaves do config intactas."""
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                ConfigManager, DEFAULT_CRITERIOS, DEFAULT_PIORA_MERCADO,
            )
            ConfigManager.save_config({
                "criterios_planejamento": dict(DEFAULT_CRITERIOS),
                "piora_mercado": dict(DEFAULT_PIORA_MERCADO),
            })
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"restore: {exc}"}
        self._config = None
        # Devolve estado atualizado para a UI re-renderizar.
        return self.get_criterios()

    # ------------------------------------------------------------------
    # Botoes do header global (fora dos passos da Section 6 mas presentes
    # no Coplan UI.html e funcionais no desktop): Conectar Banco,
    # Importar Excel, Exportar Excel. Sao "atalhos" que reusam APIs
    # ja existentes.
    # ------------------------------------------------------------------
    @staticmethod
    def _validate_db_minimum(db_path: str) -> tuple[bool, str]:
        """Paridade com BancoMixin._validate_db_minimum do desktop:
        verifica permissoes, tabela 'obras' e coluna minima 'cod' antes
        de salvar o caminho no config. Retorna (ok, motivo)."""
        if not db_path:
            return False, "caminho vazio"
        if not os.path.isfile(db_path):
            return False, f"arquivo nao encontrado: {db_path}"
        if not os.access(db_path, os.R_OK):
            return False, "sem permissao de leitura"
        if not os.access(db_path, os.W_OK):
            return False, "sem permissao de escrita"
        import sqlite3 as _sqlite3
        conn = None
        try:
            conn = _sqlite3.connect(db_path, timeout=5.0)
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM sqlite_master "
                "WHERE type='table' AND name='obras'"
            )
            if cur.fetchone() is None:
                return False, "tabela 'obras' inexistente"
            cur.execute("PRAGMA table_info(obras)")
            cols = {str(r[1]).strip().lower() for r in cur.fetchall() if len(r) > 1}
            if "cod" not in cols:
                return False, "coluna minima 'cod' ausente"
        except Exception as exc:  # noqa: BLE001
            return False, f"sqlite: {exc}"
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:  # noqa: BLE001
                    pass
        return True, ""

    def header_connect_db(self) -> dict[str, Any]:
        """Atalho header 'Conectar Banco': abre file dialog, valida
        (paridade com desktop), salva no config e reconecta. Limpa
        o cache de paths conectados."""
        r = self.pick_db_file()
        if not r.get("ok"):
            return r
        path = str(r.get("path") or "").strip()
        if not path:
            return {"ok": False, "error": "caminho vazio"}
        # Validacao igual ao desktop (BancoMixin._validate_db_minimum):
        # so persiste o caminho se o banco for compativel. Sem isto o
        # web salvava lixo no config e depois cada API quebrava com
        # "no such table: obras" / "no such column: cod".
        ok_valid, motivo = self._validate_db_minimum(path)
        if not ok_valid:
            return {"ok": False, "path": path, "error": motivo}
        save = self.save_config_empresa({"caminho_db": path})
        if not save.get("ok"):
            return save
        # Tenta conectar imediatamente para o feedback ser instantaneo.
        self._config = None
        self._connected_paths.clear()
        db, err = self._ensure_db_connected()
        return {
            "ok": db is not None and not err,
            "path": path,
            "error": err or "",
        }

    def header_import_excel(
        self, strategy: Any = "ask",
    ) -> dict[str, Any]:
        """Atalho header 'Importar Excel': le um xlsx/csv e aciona
        insert_obra para cada linha (mapeamento por nome de coluna,
        igual ao DatabaseManager.insert_obra).

        ``strategy`` define como tratar duplicadas (replica
        ``_prompt_duplicate_action`` do desktop):
          * ``"ask"`` (padrao): se houver duplicadas, retorna
            ``need_user_action=True`` com a lista para o JS perguntar
            ao usuario, sem importar nada
          * ``"merge"``: aplica build_merge_updates + update_obra
            (sobrescreve colunas preenchidas que diferem do Excel,
            inclusive ``ano_``; preserva ``cod``/``data_criacao``/
            ``criado_por``)
          * ``"create"``: ignora duplicidade e usa insert_obra
            (pode falhar por unique-index)
          * ``"skip"``: pula obras duplicadas
        """
        # File dialog que devolve {ok, path, error}. Antes usava
        # pick_ganhos_file() que retorna read_ganhos_file(path, 200) --
        # esse helper le headers/rows mas NAO devolve "path", fazendo
        # com que toda importacao retornasse "path vazio" silenciosamente.
        picked = self._pick_file_with_filters(
            "Planilhas (*.xlsx;*.xlsm;*.xls;*.csv;*.txt;*.tsv)")
        if not picked.get("ok"):
            return {"ok": False,
                    "error": picked.get("error") or "cancelado",
                    "imported": 0, "errors": []}
        path = picked.get("path") or ""
        if not path:
            return {"ok": False, "error": "path vazio",
                    "imported": 0, "errors": []}
        return self._import_excel_from_path(
            path, str(strategy or "ask").strip().lower())

    def import_excel_apply(
        self, path: Any = "", strategy: Any = "merge",
    ) -> dict[str, Any]:
        """Versao com strategy explicita para o JS chamar apos perguntar
        ao user. Path e' o mesmo retornado pelo header_import_excel
        anterior (cached em window.__coplanLastImportPath)."""
        path_s = str(path or "").strip()
        if not path_s:
            return {"ok": False, "error": "path vazio",
                    "imported": 0, "errors": []}
        if not os.path.isfile(path_s):
            return {"ok": False, "error": f"arquivo nao encontrado: {path_s}",
                    "imported": 0, "errors": []}
        return self._import_excel_from_path(
            path_s, str(strategy or "merge").strip().lower())

    # ------------------------------------------------------------------
    # Bloco 5 (Auditoria #44): bridges de progress + cancel
    # ------------------------------------------------------------------
    def progress_state(self) -> dict[str, Any]:
        """Devolve estado da operacao longa em andamento (ou da ultima
        finalizada). JS pollam isso a cada ~200ms enquanto o modal de
        progresso esta aberto."""
        return _op_snapshot()

    def progress_cancel(self) -> dict[str, Any]:
        """Sinaliza cancel para a operacao em andamento. Worker thread
        verifica via _op_check_cancel() em pontos seguros do loop."""
        with _OP_LOCK:
            if _OP_STATE.get("finished"):
                return {"ok": False,
                        "error": "nenhuma operacao em andamento"}
            _OP_STATE["cancel_requested"] = True
        return {"ok": True, "error": ""}

    # ------------------------------------------------------------------
    # save_log_txt: salva o conteudo de texto montado pelo modal de
    # detalhes pos-operacao (chaves inexistentes, falhas, erros, etc.).
    # Default folder = <HERE>/logs. Usuario escolhe nome final via
    # SAVE dialog. Usado por window.coplanShowErrorDetails (botao
    # "Salvar TXT...").
    # ------------------------------------------------------------------
    def save_log_txt(
        self, content: Any = "", default_name: Any = "log.txt",
    ) -> dict[str, Any]:
        try:
            import webview  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "",
                    "error": f"pywebview indisponivel: {exc}"}
        text = str(content or "")
        name = str(default_name or "log.txt").strip() or "log.txt"
        # Sanitiza nome: troca caracteres problematicos em filenames.
        import re as _re_log
        name = _re_log.sub(r'[\\/:*?"<>|]', "_", name)
        if not name.lower().endswith(".txt"):
            name = name + ".txt"
        try:
            wins = webview.windows
            if not wins:
                return {"ok": False, "path": "",
                        "error": "janela nao encontrada"}
            dlg = self._wv_dialog_const("SAVE")
            if dlg is None:
                return {"ok": False, "path": "",
                        "error": "SAVE dialog indisponivel"}
            logs_dir = HERE / "logs"
            try:
                logs_dir.mkdir(exist_ok=True)
            except Exception:  # noqa: BLE001
                pass
            file_types = ("Texto (*.txt)", "Todos os arquivos (*.*)")
            result = wins[0].create_file_dialog(
                dlg, save_filename=name,
                directory=str(logs_dir),
                file_types=file_types,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "",
                    "error": f"file dialog: {exc}"}
        if not result:
            return {"ok": False, "path": "", "error": "cancelado"}
        path = result if isinstance(result, str) else (
            result[0] if result else "")
        if not path:
            return {"ok": False, "path": "", "error": "caminho vazio"}
        try:
            with open(path, "w", encoding="utf-8", newline="") as f:
                f.write(text)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "",
                    "error": f"escrita: {exc}"}
        return {"ok": True, "path": str(path), "error": ""}

    def open_logs_folder(self) -> dict[str, Any]:
        """Abre <HERE>/logs no file manager do OS. Cria a pasta se
        ainda nao existir. Usado pelo modal de detalhes."""
        logs_dir = HERE / "logs"
        try:
            logs_dir.mkdir(exist_ok=True)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": str(logs_dir),
                    "error": f"mkdir: {exc}"}
        try:
            import subprocess
            if sys.platform.startswith("win"):
                os.startfile(str(logs_dir))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.run(["open", str(logs_dir)], check=False)
            else:
                subprocess.run(["xdg-open", str(logs_dir)], check=False)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": str(logs_dir),
                    "error": f"open: {exc}"}
        return {"ok": True, "path": str(logs_dir), "error": ""}

    def import_excel_async(
        self, path: Any = "", strategy: Any = "merge",
    ) -> dict[str, Any]:
        """Versao async de _import_excel_from_path: dispara worker thread
        e retorna imediatamente com op_id. JS deve abrir o modal de
        progresso e pollar progress_state() ate finished=True."""
        path_s = str(path or "").strip()
        if not path_s:
            return {"ok": False, "error": "path vazio", "started": False}
        if not os.path.isfile(path_s):
            return {"ok": False,
                    "error": f"arquivo nao encontrado: {path_s}",
                    "started": False}
        # Refuse se ja ha operacao em andamento
        with _OP_LOCK:
            if not _OP_STATE.get("finished"):
                return {
                    "ok": False,
                    "started": False,
                    "error": ("ja ha uma operacao em andamento: "
                              + str(_OP_STATE.get("label") or "")),
                }
        op_id = _op_reset(f"Importando Excel ({os.path.basename(path_s)})...")

        def _worker():
            try:
                strat = str(strategy or "merge").strip().lower()
                result = self._import_excel_from_path(path_s, strat)
                _op_finish(result=result, error="")
            except Exception as exc:  # noqa: BLE001
                _op_finish(result=None, error=f"worker: {exc}")

        t = threading.Thread(target=_worker, daemon=True,
                             name=f"coplan-import-{op_id}")
        t.start()
        return {"ok": True, "started": True, "op_id": op_id, "error": ""}

    def _import_excel_from_path(
        self, path: str, strategy: str,
    ) -> dict[str, Any]:
        """Importacao Excel com paridade ImportarExcelMixin.import_from_excel.

        Bloco 3 (Auditoria #24-#28 + M20):
          * #25 _clean_excel_columns: remove "Unnamed:" e vazias
          * #26 add_column_if_missing: empresa, cod_pep, e colunas novas
          * #27 gate root_columns subset of df.columns
          * #24 regra do `_` em alimentador_principal/beneficiados
          * #28 merge reaplica empresa da config + recalcula cod_pep
          * M20 grava arquivo de log <base>_log_importacao_<ts>.txt
        """
        import re as _re_local
        import datetime as _dt_local
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel",
                    "imported": 0, "errors": []}
        # ----- Le como dataframe -----
        try:
            import pandas as pd  # type: ignore[import-not-found]
            ext = os.path.splitext(path)[1].lower()
            if ext in (".xlsx", ".xlsm", ".xls"):
                df = pd.read_excel(path, dtype=str)
            else:
                df = pd.read_csv(path, sep=None, engine="python", dtype=str)
            df.fillna("", inplace=True)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"leitura: {exc}",
                    "imported": 0, "errors": []}

        # ----- #25 _clean_excel_columns -----
        try:
            from runtime.apoio import (  # type: ignore[import-not-found]
                _clean_excel_columns,
            )
            df, _clean_cols = _clean_excel_columns(df)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"clean_excel_columns: {exc}",
                    "imported": 0, "errors": []}

        # ----- #26 garante colunas PEP no schema -----
        try:
            db.add_column_if_missing("empresa")
            db.add_column_if_missing("cod_pep")
            db.update_columns()
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"add_column_pep: {exc}",
                    "imported": 0, "errors": []}

        # ----- #27 gate root_columns subset of df.columns -----
        try:
            root_cols = list(getattr(db, "root_columns", []) or [])
        except Exception:  # noqa: BLE001
            root_cols = []
        df_cols_set = set(df.columns)
        if root_cols and not set(root_cols).issubset(df_cols_set):
            faltantes = [c for c in root_cols if c not in df_cols_set]
            return {
                "ok": False,
                "error": ("O arquivo Excel nao possui as colunas corretas. "
                          "Faltantes: " + ", ".join(faltantes[:10])
                          + ("..." if len(faltantes) > 10 else "")),
                "imported": 0,
                "errors": [],
                "missing_columns": faltantes,
            }

        # ----- #26 (parte 2) inclui colunas novas do Excel no banco -----
        try:
            existing_cols = set(db.get_column_names() or [])
            for col in df.columns:
                if col and col not in existing_cols:
                    db.add_column_if_missing(col)
            db.update_columns()
        except Exception as exc:  # noqa: BLE001
            # Nao fatal: registra mas segue. JS pode mostrar warn.
            pass  # noqa: BLE001

        # ----- Limpa rows + detecta duplicadas -----
        try:
            from runtime.row_helpers import (  # type: ignore[import-not-found]
                find_duplicate_in_db,
            )
        except Exception:  # noqa: BLE001
            find_duplicate_in_db = None
        duplicadas: list[dict[str, Any]] = []
        rows_clean: list[dict[str, Any]] = []
        total_rows = int(len(df.index))
        # Bloco 5: marca total no progress logo no inicio
        _op_set_progress(0, total_rows, "Detectando duplicadas...")
        for i, row in enumerate(df.to_dict(orient="records"), start=1):
            # Bloco 5: cancel-check periodico (a cada 50 linhas)
            if (i % 50) == 0 and _op_check_cancel():
                return {
                    "ok": False, "cancelled": True,
                    "imported": 0, "merged": 0, "skipped": 0,
                    "errors": ["Operacao cancelada pelo usuario."],
                    "error": "cancelado",
                    "total": total_rows,
                }
            cleaned = {
                str(k).strip(): ("" if v is None or (isinstance(v, float)
                                                    and v != v)
                                  else v)
                for k, v in row.items()
                if str(k).strip()
            }
            rows_clean.append(cleaned)
            if find_duplicate_in_db is None:
                continue
            try:
                dup = find_duplicate_in_db(db, cleaned)
            except Exception:  # noqa: BLE001
                dup = None
            if dup:
                duplicadas.append({
                    "linha": i,
                    "cod_excel": str(cleaned.get("cod") or ""),
                    "dup_cod": str((dup or {}).get("cod") or ""),
                })
            # Bloco 5: progresso na fase de scan duplicadas
            if (i % 25) == 0:
                _op_set_progress(
                    i, total_rows,
                    f"Detectando duplicadas... ({i}/{total_rows})")

        # ----- Modo "ask": pergunta estrategia -----
        if strategy == "ask":
            if duplicadas:
                return {
                    "ok": False,
                    "need_user_action": True,
                    "path": path,
                    "total": len(rows_clean),
                    "duplicadas": duplicadas[:50],
                    "duplicadas_count": len(duplicadas),
                    "imported": 0,
                    "errors": [],
                    "message": (str(len(duplicadas))
                                + " duplicada(s) detectada(s) - "
                                "escolha estrategia"),
                }
            # Sem duplicadas, segue como create.
            strategy = "create"

        # ----- Helpers para #28 merge: empresa + cod_pep -----
        try:
            from codigo5_coplan import (  # type: ignore[import-not-found]
                cod_pep as _cod_pep_calc,
                normalize_text as _norm_text,
            )
        except Exception:  # noqa: BLE001
            _cod_pep_calc = None
            _norm_text = lambda s: str(s or "").strip().upper()  # noqa: E731

        def _empresa_from_row_or_config(row: dict[str, Any]) -> str:
            empresa_excel = row.get("empresa") if isinstance(row, dict) else ""
            empresa = _norm_text(empresa_excel)
            if empresa:
                return empresa
            try:
                return _norm_text(db.get_empresa_sigla_from_config())
            except Exception:  # noqa: BLE001
                return ""

        def _build_merge_updates_with_pep(
            existing: dict[str, Any], new_row: dict[str, Any],
        ) -> dict[str, Any]:
            updates = db.build_merge_updates(existing, new_row) or {}
            empresa_nova = _empresa_from_row_or_config(new_row)
            if empresa_nova:
                empresa_atual = _norm_text(existing.get("empresa"))
                if empresa_atual != empresa_nova:
                    updates["empresa"] = empresa_nova
            # cod_pep: prioriza Excel; recalcula se vazio
            cod_pep_excel = str((new_row.get("cod_pep") or "")).strip()
            cod_pep_novo = cod_pep_excel
            if not cod_pep_novo and empresa_nova and _cod_pep_calc is not None:
                obra_calc = dict(existing)
                is_missing = getattr(db, "_is_missing", None)
                for key, value in new_row.items():
                    if callable(is_missing):
                        if not is_missing(value):
                            obra_calc[key] = value
                    elif value not in (None, ""):
                        obra_calc[key] = value
                obra_calc["empresa"] = empresa_nova
                try:
                    cod_pep_novo = str(
                        _cod_pep_calc(db, obra_calc, empresa_nova) or ""
                    ).strip()
                except Exception:  # noqa: BLE001
                    cod_pep_novo = ""
            if cod_pep_novo:
                cod_pep_atual = str(existing.get("cod_pep") or "").strip()
                if cod_pep_atual != cod_pep_novo:
                    updates["cod_pep"] = cod_pep_novo
            return updates

        # ----- Loop principal -----
        imported = 0
        merged = 0
        skipped = 0
        ignorados_underscore = 0
        ignorados_permissao = 0
        error_log: list[str] = []
        # Errors curto pro JS toastar (compat retro)
        errors: list[str] = []
        dup_keys: dict[int, dict[str, Any]] = {}
        if find_duplicate_in_db is not None:
            for d in duplicadas:
                dup_keys[d["linha"]] = d

        # Bloco 5: reset progresso para a fase de gravacao
        total_processar = len(rows_clean)
        _op_set_progress(0, total_processar,
                         f"Importando registros... (0/{total_processar})")
        cancelled_mid = False

        # Linha do Excel = i + 1 (header) -> i+1 = linha real no .xlsx
        for i, cleaned in enumerate(rows_clean, start=1):
            # Bloco 5: cancel-check periodico (a cada 10 linhas no loop pesado)
            if (i % 10) == 0 and _op_check_cancel():
                cancelled_mid = True
                error_log.append(
                    f"Linha {i + 1}: cancelado pelo usuario "
                    f"(processadas {imported + merged}/{total_processar})."
                )
                break
            # Bloco 5: progresso a cada 5 linhas
            if (i % 5) == 0:
                _op_set_progress(
                    i, total_processar,
                    f"Importando registros... ({i}/{total_processar})")
            linha_excel = i + 1
            cod = str(cleaned.get("cod") or "")

            # ----- #24 regra do `_` em alimentador -----
            alim = str(cleaned.get("alimentador_principal") or "")
            benef = str(cleaned.get("alimentadores_beneficiados") or "")
            if "_" in alim or any(
                "_" in b
                for b in _re_local.split(r"[,;|\n]+", benef)
            ):
                ignorados_underscore += 1
                error_log.append(
                    f"Linha {linha_excel} (cod={cod}): "
                    "Ignorado - alimentador_principal/beneficiados "
                    "contem sublinhado (_)."
                )
                continue

            is_dup = i in dup_keys
            if is_dup and strategy == "skip":
                skipped += 1
                continue
            if is_dup and strategy == "merge":
                try:
                    cod_dup = dup_keys[i].get("dup_cod") or ""
                    if not cod_dup:
                        msg = f"linha {linha_excel}: dup_cod vazio"
                        errors.append(msg)
                        error_log.append(msg)
                        continue
                    existing = db.fetch_by_cod(cod_dup)
                    if existing:
                        cols = list(db.get_column_names() or [])
                        existing_dict = dict(zip(cols, existing))
                        updates = _build_merge_updates_with_pep(
                            existing_dict, cleaned)
                        if updates:
                            db.update_obra(updates, cod_dup, skip_blank=True)
                            merged += 1
                        else:
                            skipped += 1
                            error_log.append(
                                f"Linha {linha_excel} (cod={cod}): "
                                "Duplicidade detectada, mas sem atualizacoes "
                                "aplicaveis."
                            )
                except Exception as exc:  # noqa: BLE001
                    msg = f"linha {linha_excel} (merge): {exc}"
                    errors.append(msg)
                    error_log.append(msg)
                    if len(errors) >= 20:
                        errors.append(
                            f"... +{len(rows_clean) - i} suprimida(s)")
                        break
                continue
            # ----- create -----
            try:
                db.insert_obra(cleaned)
                imported += 1
            except PermissionError as exc:
                # #6 categorizado: pacote nao permitido
                ignorados_permissao += 1
                msg = (
                    f"Linha {linha_excel} (cod={cod}): "
                    f"Permissao negada (pacote nao permitido). "
                    f"Detalhes: {exc}"
                )
                error_log.append(msg)
                if len(errors) < 20:
                    errors.append(msg)
            except Exception as exc:  # noqa: BLE001
                msg = f"Linha {linha_excel} (cod={cod}): {exc!r}"
                error_log.append(msg)
                if len(errors) < 20:
                    errors.append(msg)
                if len(errors) >= 20:
                    errors.append(
                        f"... +{len(rows_clean) - i} suprimida(s)")
                    # nao quebra: continua para acumular log completo

        # ----- Refresh de cache do db (paridade desktop) -----
        try:
            db.update_columns()
            refresh_cache = getattr(db, "_refresh_cache", None)
            if callable(refresh_cache):
                refresh_cache()
        except Exception:  # noqa: BLE001
            pass  # noqa: BLE001

        # ----- M20 grava arquivo de log .txt -----
        log_path = ""
        if error_log:
            try:
                base, _ext = os.path.splitext(path)
                ts = _dt_local.datetime.now().strftime("%Y%m%d_%H%M%S")
                log_path = f"{base}_log_importacao_{ts}.txt"
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write("LOG DE ERROS DE IMPORTACAO\n")
                    f.write(f"Arquivo de origem: {path}\n")
                    ts_human = _dt_local.datetime.now().strftime(
                        "%d/%m/%Y %H:%M:%S")
                    f.write(f"Data/Hora: {ts_human}\n\n")
                    f.write(f"Linhas processadas: {len(rows_clean)}\n")
                    f.write(f"Importadas (insert): {imported}\n")
                    f.write(f"Atualizadas (merge): {merged}\n")
                    f.write(f"Puladas (skip): {skipped}\n")
                    f.write(
                        f"Ignoradas por _ no alimentador: "
                        f"{ignorados_underscore}\n"
                    )
                    f.write(
                        f"Ignoradas por permissao: {ignorados_permissao}\n\n"
                    )
                    f.write("--- Detalhes ---\n")
                    for line in error_log:
                        f.write(line + "\n")
            except Exception as exc:  # noqa: BLE001
                # Nao fatal: log nao gravado, mas JS recebe errors[]
                log_path = ""
                errors.append(f"log_path_fail: {exc}")

        # Bloco 5: marca progress como completo (100%)
        _op_set_progress(total_processar, total_processar,
                         "Concluido")
        return {
            "ok": (imported + merged) > 0,
            "cancelled": cancelled_mid,
            "imported": imported,
            "merged": merged,
            "skipped": skipped,
            "ignorados_underscore": ignorados_underscore,
            "ignorados_permissao": ignorados_permissao,
            "duplicadas_count": len(duplicadas),
            "strategy": strategy,
            "total": len(rows_clean),
            "errors": errors,
            "log_path": log_path,
            "error": "",
        }

    def header_export_excel(self) -> dict[str, Any]:
        """Atalho header 'Exportar Excel': delega para export_detalhamento
        sem cods (exporta tudo)."""
        return self.export_detalhamento([])

    def _pick_file_with_filters(self, filter_label: str) -> dict[str, Any]:
        try:
            import webview  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "error": f"pywebview: {exc}"}
        try:
            wins = webview.windows
            if not wins:
                return {"ok": False, "path": "",
                        "error": "janela pywebview nao encontrada"}
            dlg = self._wv_dialog_const("OPEN")
            if dlg is None:
                return {"ok": False, "path": "", "error": "OPEN dialog indisponivel"}
            result = wins[0].create_file_dialog(
                dlg,
                allow_multiple=False,
                file_types=(filter_label, "Todos os arquivos (*.*)"),
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "error": f"file dialog: {exc}"}
        if not result:
            return {"ok": False, "path": "", "error": "cancelado"}
        path = result[0] if isinstance(result, (list, tuple)) else str(result)
        return {"ok": True, "path": path, "error": ""}

    # ==================================================================
    # PARIDADE: gaps identificados no audit runtime/* + core/services/*
    # vs CoplanApi. Cada metodo abaixo delega para uma funcao publica
    # ja existente em runtime/* ou core/services/*. Sem nova logica.
    # ==================================================================

    # --- Fase 1: DB maintenance --------------------------------------

    def db_backup(self, label: Any = "") -> dict[str, Any]:
        """Cria backup timestamped do banco (db.backup_database)."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "error": err or "db indisponivel"}
        try:
            label_s = str(label or "").strip() or None
            path = db.backup_database(label=label_s)
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            out: dict[str, Any] = {
                "ok": False, "path": "",
                "error": friendly or f"backup: {exc}",
            }
            if friendly:
                out["blocked"] = "db_busy"
            return out
        if not path:
            return {"ok": False, "path": "", "error": "backup nao criado"}
        return {"ok": True, "path": str(path), "error": ""}

    def db_weekly_backup(self) -> dict[str, Any]:
        """Backup semanal nomeado por ano+semana (db.weekly_backup)."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "error": err or "db indisponivel"}
        try:
            path = db.weekly_backup()
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            out: dict[str, Any] = {
                "ok": False, "path": "",
                "error": friendly or f"weekly_backup: {exc}",
            }
            if friendly:
                out["blocked"] = "db_busy"
            return out
        if not path:
            return {"ok": False, "path": "", "error": "weekly backup nao criado"}
        return {"ok": True, "path": str(path), "error": ""}

    def db_normalize_decimal(self) -> dict[str, Any]:
        """Substitui ponto decimal por virgula em todas as colunas
        numericas. (db.normalize_decimal_in_db)"""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel"}
        try:
            db.normalize_decimal_in_db()
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            out: dict[str, Any] = {
                "ok": False, "error": friendly or f"normalize: {exc}",
            }
            if friendly:
                out["blocked"] = "db_busy"
            return out
        return {"ok": True, "error": ""}

    # ------------------------------------------------------------------
    # Visualizar / Colunas customizadas (RB-3 do desktop,
    # VisualizarColunasMixin):
    # Persiste em config['ui_state']['visualizar']:
    #   * visible_columns: list[str]   -> quais colunas mostrar
    #   * columns_order: list[str]     -> ordem em que aparecem
    #   * column_widths: dict[str,int] -> largura px persistida
    # ------------------------------------------------------------------
    # Mascara padrao do Visualizar (2026-05-08): subset legivel das
    # ~14 colunas mais relevantes para o operador. Substitui o default
    # de "todas visiveis" (67 colunas) que era ilegivel no boot.
    # User pode customizar via "Configurar Colunas" (botao na toolbar).
    DEFAULT_VIS_COLUMNS = [
        "cod",
        "ano_",
        "pi_base",
        "codigo_item",
        "nome_projeto",
        "alimentador_principal",
        "subestacao",
        "nome_regional",
        "tipo_pacote",
        "quantidade_material",
        "valor_obra",
        "obra_aprovada",
        "tecnico_dirty",
        "despacho_status",
    ]

    def visualizar_columns_get_config(self) -> dict[str, Any]:
        db, err = self._ensure_db_connected()
        all_cols: list[str] = []
        if not err and db is not None:
            try:
                all_cols = list(db.get_column_names() or [])
            except Exception:  # noqa: BLE001
                all_cols = []
        try:
            from codigo5_coplan import ConfigManager  # noqa: PLC0415
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "all": all_cols, "visible": [], "order": [],
                    "widths": {}, "error": str(exc)}
        ui_state = cfg.get("ui_state") or {}
        visualizar = ui_state.get("visualizar") or {}
        visible = visualizar.get("visible_columns")
        used_default = False
        if not isinstance(visible, list):
            # Fallback para formato legado columns_visible (dict)
            legacy_v = visualizar.get("columns_visible") or {}
            if isinstance(legacy_v, dict):
                visible = [c for c, on in legacy_v.items() if on]
            else:
                visible = None
        # NOVO: se visible vazio, ausente, OU == all (legado "todas
        # marcadas"), aplica a mascara padrao. Detectar "todas marcadas"
        # como inicial e' pratico: a primeira vez que o user abriu o app
        # antes do default ser introduzido, o config era populado com
        # todas as colunas. Tratamos isso como "no preference" =>
        # mostra a mascara default.
        if (not visible) or (
            isinstance(visible, list) and all_cols
            and len(visible) >= len(all_cols)
        ):
            if all_cols:
                visible = [c for c in self.DEFAULT_VIS_COLUMNS
                           if c in all_cols]
                used_default = True
            else:
                visible = list(self.DEFAULT_VIS_COLUMNS)
                used_default = True
        order = visualizar.get("columns_order") or []
        if not isinstance(order, list):
            order = []
        widths = visualizar.get("column_widths") or {}
        if not isinstance(widths, dict):
            widths = {}
        return {
            "ok": True,
            "all": all_cols,
            "visible": list(visible),
            "order": list(order),
            "widths": dict(widths),
            "default_columns": list(self.DEFAULT_VIS_COLUMNS),
            "using_default": used_default,
            "error": "",
        }

    def visualizar_columns_save_config(
        self, payload: Any = None,
    ) -> dict[str, Any]:
        """Persiste visible_columns/columns_order/column_widths.
        Aceita dict { visible: [...], order: [...], widths: {col: px} }."""
        if not isinstance(payload, dict):
            return {"ok": False, "error": "payload nao e dict"}
        try:
            from codigo5_coplan import ConfigManager  # noqa: PLC0415
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"load_config: {exc}"}
        ui_state = cfg.get("ui_state") or {}
        visualizar = ui_state.get("visualizar") or {}
        if "visible" in payload and isinstance(payload["visible"], list):
            visualizar["visible_columns"] = list(payload["visible"])
            visualizar.pop("columns_visible", None)  # legado
        if "order" in payload and isinstance(payload["order"], list):
            visualizar["columns_order"] = list(payload["order"])
        if "widths" in payload and isinstance(payload["widths"], dict):
            # Sanitiza para int
            clean_widths: dict[str, int] = {}
            for k, v in payload["widths"].items():
                try:
                    clean_widths[str(k)] = int(v)
                except (TypeError, ValueError):
                    pass
            visualizar["column_widths"] = clean_widths
        ui_state["visualizar"] = visualizar
        cfg["ui_state"] = ui_state
        try:
            ConfigManager.save_config(cfg, overwrite=True)
        except TypeError:
            # Versao antiga sem 'overwrite' kwarg
            try:
                ConfigManager.save_config(cfg)
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "error": f"save: {exc}"}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return {"ok": True, "error": ""}

    def visualizar_columns_reset(self) -> dict[str, Any]:
        """Limpa toda a config de colunas (volta ao default)."""
        try:
            from codigo5_coplan import ConfigManager  # noqa: PLC0415
            cfg = ConfigManager.load_config() or {}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"load_config: {exc}"}
        ui_state = cfg.get("ui_state") or {}
        visualizar = ui_state.get("visualizar") or {}
        for k in ("visible_columns", "columns_visible",
                  "columns_order", "column_widths"):
            visualizar.pop(k, None)
        ui_state["visualizar"] = visualizar
        cfg["ui_state"] = ui_state
        try:
            ConfigManager.save_config(cfg, overwrite=True)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        self._config = None
        return {"ok": True, "error": ""}

    def db_create_new(self, path: Any = "") -> dict[str, Any]:
        """Cria um banco SQLite vazio no path informado e ativa-o como
        'banco corrente'. Equivalente a create_new_database do desktop:
          1. Abre file save dialog (se path vazio)
          2. db.connect(path) -> cria tabela obras + colunas legadas
          3. Salva config['obras'] = path
          4. Retorna pra recarregar a tabela
        """
        target = str(path or "").strip()
        # Se path nao foi passado, abrir file dialog SAVE
        if not target:
            try:
                import webview  # type: ignore[import-not-found]
                wins = webview.windows
                if not wins:
                    return {"ok": False, "path": "",
                            "error": "janela pywebview nao encontrada"}
                dlg_const = self._wv_dialog_const("SAVE")
                if dlg_const is None:
                    return {"ok": False, "path": "",
                            "error": "SAVE dialog indisponivel"}
                result = wins[0].create_file_dialog(
                    dlg_const,
                    save_filename="novo_banco.db",
                    file_types=("SQLite Database (*.db)",),
                )
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "path": "",
                        "error": f"file dialog: {exc}"}
            if not result:
                return {"ok": False, "path": "", "error": "cancelado"}
            target = (result[0] if isinstance(result, (list, tuple))
                      else str(result))
        if not target:
            return {"ok": False, "path": "", "error": "path vazio"}
        # Garante extensao .db
        if not target.lower().endswith(".db"):
            target = target + ".db"
        # Garante diretorio existente
        try:
            os.makedirs(os.path.dirname(os.path.abspath(target)) or ".",
                        exist_ok=True)
        except OSError as exc:
            return {"ok": False, "path": target,
                    "error": f"mkdir: {exc}"}
        # Conecta (cria tabela vazia + migracoes)
        self._ensure_managers()
        if self._db_manager is None:
            return {"ok": False, "path": target,
                    "error": "DatabaseManager indisponivel"}
        try:
            self._db_manager.connect(target)
            # Limpa pacotes filtrados (set_allowed_pacotes(None) do desktop)
            try:
                self._db_manager.set_allowed_pacotes(None)
            except Exception:  # noqa: BLE001
                pass
        except Exception as exc:  # noqa: BLE001
            self._data_state_set(
                "db", "INVALIDADO", path=target,
                error=f"connect: {exc}")
            return {"ok": False, "path": target,
                    "error": f"connect: {exc}"}
        # Salva no config
        try:
            from codigo5_coplan import ConfigManager  # noqa: PLC0415
            cfg = ConfigManager.load_config() or {}
            cfg["obras"] = target
            ConfigManager.save_config(cfg)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": target,
                    "error": f"save_config: {exc}"}
        self._config = None
        self._connected_paths.discard(target)
        self._connected_paths.add(target)
        try:
            self._data_state_set(
                "db", "CARREGADO_VALIDADO", path=target,
                version_token=str(int(os.path.getmtime(target))))
        except OSError:
            pass
        return {"ok": True, "path": target, "error": ""}

    def db_save_as(self, path: Any = "") -> dict[str, Any]:
        """Salva uma copia do banco corrente em outro lugar (backup
        manual). Equivalente a salvar_banco_dados do desktop.
        Usa shutil.copy2 (preserva mtime). NAO troca o banco corrente."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "",
                    "error": err or "db indisponivel"}
        cfg = self._config or {}
        src = str(cfg.get("obras") or "").strip()
        if not src or not os.path.isfile(src):
            return {"ok": False, "path": "",
                    "error": "banco corrente nao encontrado"}
        target = str(path or "").strip()
        if not target:
            try:
                import webview  # type: ignore[import-not-found]
                wins = webview.windows
                if not wins:
                    return {"ok": False, "path": "",
                            "error": "janela pywebview nao encontrada"}
                dlg_const = self._wv_dialog_const("SAVE")
                if dlg_const is None:
                    return {"ok": False, "path": "",
                            "error": "SAVE dialog indisponivel"}
                result = wins[0].create_file_dialog(
                    dlg_const,
                    save_filename=os.path.basename(src),
                    file_types=("SQLite Database (*.db)",),
                )
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "path": "",
                        "error": f"file dialog: {exc}"}
            if not result:
                return {"ok": False, "path": "", "error": "cancelado"}
            target = (result[0] if isinstance(result, (list, tuple))
                      else str(result))
        if not target:
            return {"ok": False, "path": "", "error": "path vazio"}
        if not target.lower().endswith(".db"):
            target = target + ".db"
        if os.path.abspath(target) == os.path.abspath(src):
            return {"ok": False, "path": target,
                    "error": "destino igual a origem"}
        try:
            import shutil  # noqa: PLC0415
            shutil.copy2(src, target)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": target,
                    "error": f"copy: {exc}"}
        return {"ok": True, "path": target, "error": ""}

    def db_export_to(
        self,
        cods: Any = None,
        path: Any = "",
        include_aprovadas: Any = False,
    ) -> dict[str, Any]:
        """Exporta obras selecionadas para um banco destino (cria/sobrescreve).
        Equivalente simplificado a exportar_para_banco do desktop:
          1. Filtra cods (descarta aprovadas se include_aprovadas=False)
          2. Cria/conecta o banco destino (mesmo schema)
          3. Insere cada obra (insert_obra)
        Retorna {ok, path, exported, ignoradas_aprovadas, errors}.
        Aviso: NAO replica toda a logica de _exportar_para_banco_write_phase
        (que valida integridade campo a campo, ganhos, etc.). Para esses
        casos, usar a versao desktop ou estender no futuro."""
        cods_list = [str(c).strip() for c in (cods or []) if str(c or "").strip()]
        if not cods_list:
            return {"ok": False, "path": "", "exported": 0,
                    "ignoradas_aprovadas": 0, "errors": ["cods vazio"]}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "", "exported": 0,
                    "ignoradas_aprovadas": 0,
                    "errors": [err or "db indisponivel"]}
        # Gating aprovadas
        gate = self.gate_aprovadas_for_action(
            cods_list, include_aprovadas=bool(include_aprovadas))
        targets = gate.get("targets") or []
        ignoradas_aprovadas = len(gate.get("aprovadas") or [])
        if not targets:
            return {"ok": False, "path": "", "exported": 0,
                    "ignoradas_aprovadas": ignoradas_aprovadas,
                    "errors": ["sem obras elegiveis (todas aprovadas)"]}
        # Resolve destino
        target = str(path or "").strip()
        if not target:
            try:
                import webview  # type: ignore[import-not-found]
                wins = webview.windows
                if not wins:
                    return {"ok": False, "path": "", "exported": 0,
                            "ignoradas_aprovadas": ignoradas_aprovadas,
                            "errors": ["janela pywebview nao encontrada"]}
                dlg_const = self._wv_dialog_const("SAVE")
                result = wins[0].create_file_dialog(
                    dlg_const,
                    save_filename="export_obras.db",
                    file_types=("SQLite Database (*.db)",),
                )
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "path": "", "exported": 0,
                        "ignoradas_aprovadas": ignoradas_aprovadas,
                        "errors": [f"file dialog: {exc}"]}
            if not result:
                return {"ok": False, "path": "", "exported": 0,
                        "ignoradas_aprovadas": ignoradas_aprovadas,
                        "errors": ["cancelado"]}
            target = (result[0] if isinstance(result, (list, tuple))
                      else str(result))
        if not target.lower().endswith(".db"):
            target = target + ".db"
        # Cria DatabaseManager auxiliar para o destino
        try:
            from codigo5_coplan import DatabaseManager  # noqa: PLC0415
            dest_db = DatabaseManager()
            dest_db.connect(target)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": target, "exported": 0,
                    "ignoradas_aprovadas": ignoradas_aprovadas,
                    "errors": [f"conectar destino: {exc}"]}
        # Le cada obra do banco corrente e insere no destino
        try:
            cols = list(db.get_column_names() or [])
            rows = list(db.fetch_by_cods(targets) or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": target, "exported": 0,
                    "ignoradas_aprovadas": ignoradas_aprovadas,
                    "errors": [f"fetch_by_cods: {exc}"]}
        exported = 0
        errors: list[str] = []
        for row in rows:
            try:
                dados = dict(zip(cols, row))
                dest_db.insert_obra(dados)
                exported += 1
            except Exception as exc:  # noqa: BLE001
                cod = str(dados.get("cod") or "?") if "cod" in cols else "?"
                errors.append(f"{cod}: {exc}")
        return {
            "ok": exported > 0,
            "path": target,
            "exported": exported,
            "ignoradas_aprovadas": ignoradas_aprovadas,
            "total_targets": len(targets),
            "errors": errors,
        }

    def db_mark_tecnico_dirty_all(self) -> dict[str, Any]:
        """Marca TODAS as obras como tecnico_dirty='SIM'."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel"}
        try:
            db.mark_tecnico_dirty_all()
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"mark_dirty: {exc}"}
        return {"ok": True, "error": ""}

    def tecnico_snapshot_update(self, cods: Any = None) -> dict[str, Any]:
        """Atualiza snapshot tecnico para um conjunto de cods (limpa
        tecnico_dirty para essas obras + grava token/timestamp/src).
        Equivalente a atualizar_snapshot_tecnico_selecionados do desktop:
        usado pelo botao 'Atualizar snapshot tecnico' apos refrescar
        FlowMT/Topologia/Confiabilidade.

        Token e' derivado dos mtimes dos 3 .TXT da pasta de ganhos
        (mesma logica de validate_tecnico_files). snapshot_src guarda
        o path da pasta para auditoria."""
        cods_list = [str(c).strip() for c in (cods or []) if str(c or "").strip()]
        if not cods_list:
            return {"ok": False, "atualizadas": 0, "error": "cods vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "atualizadas": 0,
                    "error": err or "db indisponivel"}
        # Token: mesma estrategia de validate_tecnico_files (3 .TXT
        # mtimes).
        cfg = self._config or {}
        pasta = str(cfg.get("caminho_pasta_ganhos") or "").strip()
        if not pasta or not os.path.isdir(pasta):
            return {"ok": False, "atualizadas": 0,
                    "error": "pasta de ganhos nao configurada"}
        required = ["FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT"]
        try:
            tokens = [
                str(int(os.path.getmtime(os.path.join(pasta, f))))
                for f in required if os.path.isfile(os.path.join(pasta, f))
            ]
            if len(tokens) < len(required):
                return {"ok": False, "atualizadas": 0,
                        "error": "arquivos tecnicos ausentes"}
            token = "-".join(tokens)
        except OSError as exc:
            return {"ok": False, "atualizadas": 0,
                    "error": f"mtime: {exc}"}
        snapshot_at = datetime.now().strftime("%d/%m/%y %H:%M")
        snapshot_src = pasta
        try:
            db.update_tecnico_snapshot_for_cods(
                cods_list, token, snapshot_at, snapshot_src,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atualizadas": 0,
                    "error": f"update_snapshot: {exc}"}
        return {
            "ok": True,
            "atualizadas": len(cods_list),
            "token": token,
            "snapshot_at": snapshot_at,
            "snapshot_src": snapshot_src,
            "error": "",
        }

    def db_count_tecnico_dirty(self) -> dict[str, Any]:
        """Quantas obras estao com snapshot tecnico desatualizado."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "count": 0, "error": err or "db indisponivel"}
        try:
            count = int(db.count_tecnico_dirty() or 0)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "count": 0, "error": f"count: {exc}"}
        return {"ok": True, "count": count, "error": ""}

    def db_last_modification_info(self) -> dict[str, Any]:
        """Data + usuario da ultima modificacao no banco."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "data": "", "usuario": "",
                    "error": err or "db indisponivel"}
        try:
            data, usuario = db.get_last_modification_info()
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "data": "", "usuario": "",
                    "error": f"last_mod: {exc}"}
        return {"ok": True, "data": str(data or ""),
                "usuario": str(usuario or ""), "error": ""}

    def db_mark_refresh_point(self) -> dict[str, Any]:
        """Marca o instante atual como 'snapshot' do banco. Equivalente
        a _mark_db_refresh_point do EstadoFontesMixin: chamado depois
        de uma leitura completa (list_obras refresh) pra usar de baseline
        em db_check_external_update."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel"}
        try:
            data, usuario = db.get_last_modification_info()
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"last_mod: {exc}"}
        self._last_db_refresh_data = str(data or "")
        self._last_db_refresh_user = str(usuario or "")
        self._last_db_modification_warned = ""
        return {
            "ok": True,
            "data": self._last_db_refresh_data,
            "usuario": self._last_db_refresh_user,
        }

    def db_check_external_update(self) -> dict[str, Any]:
        """Compara data_modificacao atual com a snap salva por
        db_mark_refresh_point. Retorna mudou=True se outro usuario
        gravou no banco entre os dois pontos. Equivalente a
        _warn_external_db_update do desktop."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "mudou": False,
                    "error": err or "db indisponivel"}
        try:
            data, usuario = db.get_last_modification_info()
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "mudou": False,
                    "error": f"last_mod: {exc}"}
        atual = str(data or "")
        atual_user = str(usuario or "")
        baseline = self._last_db_refresh_data
        if not baseline:
            # Sem snapshot anterior: marca este ponto como baseline
            # (primeira chamada da sessao) e reporta sem mudanca.
            self._last_db_refresh_data = atual
            self._last_db_refresh_user = atual_user
            return {"ok": True, "mudou": False, "data": atual,
                    "usuario": atual_user, "first_call": True}
        mudou = (atual != baseline)
        # Evita avisar 2 vezes pelo mesmo timestamp
        ja_avisado = (mudou and self._last_db_modification_warned == atual)
        if mudou and not ja_avisado:
            self._last_db_modification_warned = atual
        return {
            "ok": True,
            "mudou": mudou,
            "ja_avisado": ja_avisado,
            "data": atual,
            "usuario": atual_user,
            "baseline": baseline,
            "baseline_user": self._last_db_refresh_user,
        }

    def db_next_codigo_item(self, nome_projeto: Any) -> dict[str, Any]:
        """Proximo codigo_item disponivel para um projeto."""
        nome = str(nome_projeto or "").strip()
        if not nome:
            return {"ok": False, "next": 0, "error": "nome_projeto vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "next": 0, "error": err or "db indisponivel"}
        try:
            nxt = int(db.get_next_codigo_item(nome) or 1)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "next": 0, "error": f"next_item: {exc}"}
        return {"ok": True, "next": nxt, "error": ""}

    def db_exists_codigo_item(
        self, nome_projeto: Any, codigo_item: Any, exclude_cod: Any = None,
    ) -> dict[str, Any]:
        """Verifica se ja existe obra com mesmo (nome_projeto, codigo_item)."""
        nome = str(nome_projeto or "").strip()
        item = str(codigo_item or "").strip()
        if not nome or not item:
            return {"ok": True, "exists": False, "error": ""}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "exists": False,
                    "error": err or "db indisponivel"}
        try:
            exists = bool(
                db.exists_codigo_item(
                    nome, item,
                    exclude_cod=str(exclude_cod) if exclude_cod else None,
                )
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "exists": False,
                    "error": f"exists_codigo_item: {exc}"}
        return {"ok": True, "exists": exists, "error": ""}

    # --- Fase 2: COD PEP em lote -------------------------------------

    def cod_pep_gerar_lote(
        self,
        cods: Any = None,
        empresa_sigla: Any = "",
        somente_vazios: Any = True,
        reiniciar_numeracao: Any = False,
    ) -> dict[str, Any]:
        """Gera COD_PEP para varios COD de uma vez (db.gerar_cod_pep_para_cods).
        Retorna {atualizados, ignorados, erros}."""
        if not isinstance(cods, (list, tuple)) or not cods:
            return {"ok": False, "atualizados": 0, "ignorados": 0,
                    "erros": ["cods vazio"]}
        cods_list = [str(c).strip() for c in cods if str(c or "").strip()]
        if not cods_list:
            return {"ok": False, "atualizados": 0, "ignorados": 0,
                    "erros": ["cods vazio"]}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "atualizados": 0, "ignorados": 0,
                    "erros": [err or "db indisponivel"]}
        # Resolve empresa: prioriza o passado, senao usa o config.
        sigla = str(empresa_sigla or "").strip().upper()
        if not sigla:
            try:
                from runtime.database import get_empresa_sigla_from_config
                sigla = get_empresa_sigla_from_config(self._config or None)
            except Exception as exc:  # noqa: BLE001
                return {"ok": False, "atualizados": 0, "ignorados": 0,
                        "erros": [f"empresa_sigla nao resolvida: {exc}"]}
        try:
            atualizados, ignorados, erros = db.gerar_cod_pep_para_cods(
                cods_list,
                sigla,
                somente_vazios=bool(somente_vazios),
                reiniciar_numeracao=bool(reiniciar_numeracao),
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atualizados": 0, "ignorados": 0,
                    "erros": [f"gerar_cod_pep_lote: {exc}"]}
        return {
            "ok": True,
            "atualizados": int(atualizados or 0),
            "ignorados": int(ignorados or 0),
            "erros": list(erros or []),
        }

    def cod_pep_preencher_pendentes(self) -> dict[str, Any]:
        """Preenche COD_PEP de obras que ainda nao tem
        (db.preencher_cod_pep_pendentes). Retorna {preenchidos}."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "preenchidos": 0,
                    "error": err or "db indisponivel"}
        try:
            preenchidos = int(db.preencher_cod_pep_pendentes() or 0)
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            out: dict[str, Any] = {
                "ok": False, "preenchidos": 0,
                "error": friendly or f"preencher_pendentes: {exc}",
            }
            if friendly:
                out["blocked"] = "db_busy"
            return out
        return {"ok": True, "preenchidos": preenchidos, "error": ""}

    # --- Fase 3: CalculationManager (calculos finos) ------------------

    def _ensure_calc_manager(self):
        """Cria/retorna singleton de CalculationManager."""
        cm = getattr(self, "_calc_manager", None)
        if cm is not None:
            return cm
        try:
            self._ensure_managers()
            from runtime.calc import CalculationManager
            cm = CalculationManager(
                self._support_manager, prompt_pi_base=False
            )
            self._calc_manager = cm
            return cm
        except Exception:  # noqa: BLE001
            return None

    def calc_gerar_cod(
        self, pacote: Any = "", alimentador: Any = "",
        projeto_investimento: Any = "", quantidade: Any = "",
        caracteristica: Any = "", coord_final: Any = "", pi_base: Any = "",
    ) -> dict[str, Any]:
        """Gera o codigo da obra (CalculationManager.gerar_cod).
        Formato: PCT|ALIM|TIPO|QTDxCARAC|COORD."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "cod": "", "error": "calc indisponivel"}
        try:
            cod = cm.gerar_cod(
                str(pacote or ""), str(alimentador or ""),
                str(projeto_investimento or ""), str(quantidade or ""),
                str(caracteristica or ""), str(coord_final or ""),
                pi_base=str(pi_base or "") or None,
            )
        except ValueError as exc:
            return {"ok": False, "cod": "", "error": str(exc)}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "cod": "", "error": f"gerar_cod: {exc}"}
        return {"ok": True, "cod": str(cod or ""), "error": ""}

    def calc_build_module_key(
        self, pi_base: Any = "", tensao: Any = "",
        caracteristica: Any = "", codigo_regional: Any = "",
    ) -> dict[str, Any]:
        """Constroi chave de modulo (CalculationManager.build_module_key)."""
        try:
            from runtime.calc import CalculationManager
            key = CalculationManager.build_module_key(
                str(pi_base or ""), str(tensao or ""),
                str(caracteristica or ""), str(codigo_regional or ""),
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "key": "", "error": f"build_key: {exc}"}
        return {"ok": True, "key": key, "error": ""}

    def calc_despacho_vt(self, cods: Any = None) -> dict[str, Any]:
        """Gera o texto de despacho VT a partir das obras informadas
        (CalculationManager.calcular_despacho_vt)."""
        if not isinstance(cods, (list, tuple)) or not cods:
            return {"ok": False, "texto": "", "error": "cods vazio"}
        cods_list = [str(c).strip() for c in cods if str(c or "").strip()]
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "texto": "", "error": err or "db indisponivel"}
        try:
            obras = db.fetch_by_cods(cods_list) or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "texto": "", "error": f"fetch: {exc}"}
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "texto": "", "error": "calc indisponivel"}
        try:
            texto = cm.calcular_despacho_vt(list(obras))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "texto": "", "error": f"despacho_vt: {exc}"}
        return {"ok": True, "texto": str(texto or ""), "error": ""}

    def _read_tecnico_files(self, alimentador: Any = "") -> dict[str, Any]:
        """Le os 3 arquivos tecnicos (FlowMT/Topologia/Confiabilidade)
        da pasta tecnica configurada. Retorna {flow_mt, topologia, confiabilidade}.

        A pasta e' resolvida pelo helper `_ganhos_resolve_pasta`, que tenta
        `caminho_pasta_ganhos` (chave canonica usada pela UI de Ganhos)
        e cai pra `caminho_pasta_arquivos` como fallback historico.
        """
        pasta = self._ganhos_resolve_pasta()
        if not pasta:
            return {"flow_mt": [], "topologia": [], "confiabilidade": [],
                    "error": "pasta tecnica nao configurada"}
        try:
            from runtime.file_io import carregar_arquivos
            dados = carregar_arquivos(
                pasta, ["FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT"]
            )
        except Exception as exc:  # noqa: BLE001
            return {"flow_mt": [], "topologia": [], "confiabilidade": [],
                    "error": f"carregar: {exc}"}
        return {
            "flow_mt": list(dados.get("FlowMT.TXT") or []),
            "topologia": list(dados.get("Topologia.TXT") or []),
            "confiabilidade": list(dados.get("Confiabilidade.TXT") or []),
            "error": "",
        }

    def calc_tensoes(self, alimentadores: Any = None) -> dict[str, Any]:
        """Tensao minima + menor media por patamar
        (CalculationManager.calcular_tensoes)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "tensao_min": None, "tensao_media_min": None,
                    "error": "calc indisponivel"}
        alims = [str(a).strip() for a in (alimentadores or [])
                 if str(a or "").strip()]
        if not alims:
            return {"ok": False, "tensao_min": None, "tensao_media_min": None,
                    "error": "alimentadores vazio"}
        files = self._read_tecnico_files()
        if files.get("error"):
            return {"ok": False, "tensao_min": None, "tensao_media_min": None,
                    "error": files["error"]}
        try:
            tmin, tmed = cm.calcular_tensoes(files["flow_mt"], alims)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "tensao_min": None, "tensao_media_min": None,
                    "error": f"calcular_tensoes: {exc}"}
        return {"ok": True, "tensao_min": tmin, "tensao_media_min": tmed,
                "error": ""}

    def calc_tensao_linha_minima(self, alimentadores: Any = None) -> dict[str, Any]:
        """Menor tensao de LINHA pu entre os alimentadores
        (CalculationManager.calcular_tensao_linha_minima)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "tensao_min_linha": None,
                    "error": "calc indisponivel"}
        alims = [str(a).strip() for a in (alimentadores or [])
                 if str(a or "").strip()]
        if not alims:
            return {"ok": False, "tensao_min_linha": None,
                    "error": "alimentadores vazio"}
        files = self._read_tecnico_files()
        if files.get("error"):
            return {"ok": False, "tensao_min_linha": None,
                    "error": files["error"]}
        try:
            v = cm.calcular_tensao_linha_minima(files["flow_mt"], alims)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "tensao_min_linha": None,
                    "error": f"tensao_linha_min: {exc}"}
        return {"ok": True, "tensao_min_linha": v, "error": ""}

    def calc_tensoes_max(self, alimentadores: Any = None) -> dict[str, Any]:
        """Tensao maxima + menor media por patamar
        (CalculationManager.calcular_tensoes_max)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "tensao_max": None, "tensao_media_min": None,
                    "error": "calc indisponivel"}
        alims = [str(a).strip() for a in (alimentadores or [])
                 if str(a or "").strip()]
        if not alims:
            return {"ok": False, "tensao_max": None, "tensao_media_min": None,
                    "error": "alimentadores vazio"}
        files = self._read_tecnico_files()
        if files.get("error"):
            return {"ok": False, "tensao_max": None, "tensao_media_min": None,
                    "error": files["error"]}
        try:
            tmax, tmed = cm.calcular_tensoes_max(files["flow_mt"], alims)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "tensao_max": None, "tensao_media_min": None,
                    "error": f"calcular_tensoes_max: {exc}"}
        return {"ok": True, "tensao_max": tmax, "tensao_media_min": tmed,
                "error": ""}

    def calc_carregamento(self, alimentadores: Any = None) -> dict[str, Any]:
        """Pior carregamento entre alimentadores
        (CalculationManager.calcular_carregamento)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "carregamento": None, "error": "calc indisponivel"}
        alims = [str(a).strip() for a in (alimentadores or [])
                 if str(a or "").strip()]
        if not alims:
            return {"ok": False, "carregamento": None, "error": "alimentadores vazio"}
        files = self._read_tecnico_files()
        if files.get("error"):
            return {"ok": False, "carregamento": None, "error": files["error"]}
        try:
            v = cm.calcular_carregamento(files["topologia"], alims)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "carregamento": None,
                    "error": f"carregamento: {exc}"}
        return {"ok": True, "carregamento": v, "error": ""}

    def calc_perdas(self, alimentadores: Any = None) -> dict[str, Any]:
        """Perdas por patamar + maior perda
        (CalculationManager.calcular_perdas)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "perdas_por_patamar": {}, "maior_perda": 0.0,
                    "error": "calc indisponivel"}
        alims = [str(a).strip() for a in (alimentadores or [])
                 if str(a or "").strip()]
        if not alims:
            return {"ok": False, "perdas_por_patamar": {}, "maior_perda": 0.0,
                    "error": "alimentadores vazio"}
        files = self._read_tecnico_files()
        if files.get("error"):
            return {"ok": False, "perdas_por_patamar": {}, "maior_perda": 0.0,
                    "error": files["error"]}
        try:
            perdas, maior = cm.calcular_perdas(files["flow_mt"], alims)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "perdas_por_patamar": {}, "maior_perda": 0.0,
                    "error": f"perdas: {exc}"}
        return {"ok": True, "perdas_por_patamar": dict(perdas or {}),
                "maior_perda": float(maior or 0.0), "error": ""}

    def calc_demanda_maxima(self, alimentadores: Any = None) -> dict[str, Any]:
        """Demanda maxima coincidente por alimentador (em MW)
        (CalculationManager.calcular_demanda_maxima)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "demanda_por_alim": {}, "error": "calc indisponivel"}
        alims = [str(a).strip() for a in (alimentadores or [])
                 if str(a or "").strip()]
        if not alims:
            return {"ok": False, "demanda_por_alim": {}, "error": "alimentadores vazio"}
        files = self._read_tecnico_files()
        if files.get("error"):
            return {"ok": False, "demanda_por_alim": {}, "error": files["error"]}
        try:
            d = cm.calcular_demanda_maxima(files["flow_mt"], alims)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "demanda_por_alim": {},
                    "error": f"demanda_maxima: {exc}"}
        return {"ok": True, "demanda_por_alim": dict(d or {}), "error": ""}

    def calc_chi_ci(self, alimentadores: Any = None) -> dict[str, Any]:
        """CHI + CI agregados entre alimentadores
        (CalculationManager.calcular_chi_ci)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "chi": 0, "ci": 0, "error": "calc indisponivel"}
        alims = [str(a).strip() for a in (alimentadores or [])
                 if str(a or "").strip()]
        if not alims:
            return {"ok": False, "chi": 0, "ci": 0, "error": "alimentadores vazio"}
        files = self._read_tecnico_files()
        if files.get("error"):
            return {"ok": False, "chi": 0, "ci": 0, "error": files["error"]}
        try:
            chi, ci = cm.calcular_chi_ci(files["confiabilidade"], alims)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "chi": 0, "ci": 0, "error": f"chi_ci: {exc}"}
        return {"ok": True, "chi": chi, "ci": ci, "error": ""}

    def calc_contas_contratos(self, alimentadores: Any = None) -> dict[str, Any]:
        """Contas/contratos antes vs depois
        (CalculationManager.calcular_contas_contratos)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "antes": 0, "depois": 0,
                    "error": "calc indisponivel"}
        alims = [str(a).strip() for a in (alimentadores or [])
                 if str(a or "").strip()]
        if not alims:
            return {"ok": False, "antes": 0, "depois": 0,
                    "error": "alimentadores vazio"}
        files = self._read_tecnico_files()
        if files.get("error"):
            return {"ok": False, "antes": 0, "depois": 0,
                    "error": files["error"]}
        try:
            a, d = cm.calcular_contas_contratos(files["confiabilidade"], alims)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "antes": 0, "depois": 0,
                    "error": f"contas_contratos: {exc}"}
        return {"ok": True, "antes": a, "depois": d, "error": ""}

    def calc_contas_contratos_beneficiadas(
        self, alimentadores: Any = None, projeto_investimento: Any = "",
    ) -> dict[str, Any]:
        """Contas/contratos beneficiadas considerando o PI
        (CalculationManager.calcular_contas_contratos_beneficiadas)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "total": 0, "error": "calc indisponivel"}
        alims = [str(a).strip() for a in (alimentadores or [])
                 if str(a or "").strip()]
        if not alims:
            return {"ok": False, "total": 0, "error": "alimentadores vazio"}
        pi = str(projeto_investimento or "").strip()
        if not pi:
            return {"ok": False, "total": 0, "error": "projeto_investimento vazio"}
        files = self._read_tecnico_files()
        if files.get("error"):
            return {"ok": False, "total": 0, "error": files["error"]}
        try:
            total = cm.calcular_contas_contratos_beneficiadas(
                files["topologia"], files["confiabilidade"], alims, pi,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "total": 0,
                    "error": f"contas_beneficiadas: {exc}"}
        return {"ok": True, "total": int(total or 0), "error": ""}

    # ------------------------------------------------------------------
    # Ganhos UI (RB-3 do desktop, GanhosMixin):
    # Orquestra preencher_campos_antes/depois/parametros_atuais em uma
    # unica chamada -- le os 3 .TXT da pasta uma vez e roda 8+ calculos
    # contra os mesmos buffers (vs chamar calc_* individualmente que
    # re-le os .TXT cada vez). Retorna metricas numericas + a string
    # ganhos_totais consumida pelo Resumo.
    # ------------------------------------------------------------------
    def _ganhos_resolve_pasta(self, pasta: Any = "") -> str:
        """Resolve pasta de ganhos: parametro -> config.caminho_pasta_ganhos
        -> config.caminho_pasta_arquivos. Retorna '' se nada valido."""
        s = str(pasta or "").strip()
        if s and os.path.isdir(s):
            return s
        cfg = self._config or {}
        for key in ("caminho_pasta_ganhos", "caminho_pasta_arquivos"):
            p = str(cfg.get(key) or "").strip()
            if p and os.path.isdir(p):
                return p
        return ""

    def _ganhos_load_files(
        self, pasta: str, files: list[str],
    ) -> tuple[dict[str, Any], str]:
        """Le N arquivos tecnicos de uma pasta. Retorna ({arquivo: lines}, '')
        ou ({}, motivo_erro)."""
        if not pasta or not os.path.isdir(pasta):
            return {}, "pasta tecnica nao configurada"
        missing = [f for f in files if not os.path.isfile(os.path.join(pasta, f))]
        if missing:
            return {}, f"arquivos ausentes: {', '.join(missing)}"
        try:
            from runtime.file_io import carregar_arquivos
            dados = carregar_arquivos(pasta, files)
        except Exception as exc:  # noqa: BLE001
            return {}, f"carregar: {exc}"
        return dados or {}, ""

    @staticmethod
    def _ganhos_filter_alims(
        alimentadores: list[str], buffers: list[Any],
    ) -> list[str]:
        """Filtra alimentadores que nao aparecem em nenhum dos buffers
        (replica alimentadores_nos_arquivos do desktop)."""
        out = []
        for alim in alimentadores:
            if not alim:
                continue
            a_proc = alim.lower().replace(" ", "")
            encontrado = False
            for buf in buffers:
                lines = buf if isinstance(buf, list) else [buf]
                for line in lines:
                    if a_proc in str(line).lower().replace(" ", ""):
                        encontrado = True
                        break
                if encontrado:
                    break
            if encontrado:
                out.append(alim)
        return out

    @staticmethod
    def _ganhos_normalize_alims(alimentadores: Any) -> list[str]:
        if not alimentadores:
            return []
        if isinstance(alimentadores, str):
            return [a.strip() for a in alimentadores.split(",") if a.strip()]
        return [str(a).strip() for a in alimentadores
                if str(a or "").strip()]

    def _ganhos_gerar_totais_string(
        self, cm: Any, alimentadores: list[str],
        data_flow: Any, data_topo: Any, data_conf: Any,
        projeto_investimento: str, etapa: str,
    ) -> str:
        """Replica gerar_ganhos_totais do desktop: monta uma string
        'alim_metric_valor' separada por ';'. Consumida pelo Resumo
        (popular_quadro_resumo_from_ganhos_depois) -- formato deve
        bater com o que o legado emite, senao parsing falha."""
        try:
            demanda_max_por_alim = cm.calcular_demanda_maxima(
                data_flow, alimentadores) or {}
        except Exception:  # noqa: BLE001
            demanda_max_por_alim = {}
        partes: list[str] = []
        for alim in alimentadores:
            try:
                tensao_min, tensao_media = cm.calcular_tensoes(data_flow, [alim])
            except Exception:  # noqa: BLE001
                tensao_min, tensao_media = None, None
            try:
                tensao_min_linha = cm.calcular_tensao_linha_minima(
                    data_flow, [alim])
            except Exception:  # noqa: BLE001
                tensao_min_linha = None
            try:
                tensao_max, _t = cm.calcular_tensoes_max(data_flow, [alim])
            except Exception:  # noqa: BLE001
                tensao_max = None
            try:
                carreg = cm.calcular_carregamento(data_topo, [alim])
            except Exception:  # noqa: BLE001
                carreg = None
            try:
                _patamares, perda = cm.calcular_perdas(data_flow, [alim])
            except Exception:  # noqa: BLE001
                perda = None
            try:
                cont_antes, cont_depois = cm.calcular_contas_contratos(
                    data_conf, [alim])
            except Exception:  # noqa: BLE001
                cont_antes, cont_depois = None, None
            try:
                chi, ci = cm.calcular_chi_ci(data_conf, [alim])
            except Exception:  # noqa: BLE001
                chi, ci = None, None

            if etapa == "antes":
                try:
                    benef = cm.calcular_contas_contratos_beneficiadas(
                        data_topo, data_conf, [alim], projeto_investimento)
                except Exception:  # noqa: BLE001
                    benef = 0
                partes.extend([
                    f"{alim}_contas_{cont_antes}",
                    f"{alim}_contasbenef_{benef}",
                ])
            else:
                partes.append(f"{alim}_contas_{cont_depois}")
            partes.extend([
                f"{alim}_carregamento_{round(carreg or 0, 2)}",
                f"{alim}_perdas_{round(perda or 0, 2)}",
                f"{alim}_tensaomedia_{round(tensao_media or 0, 4)}",
                f"{alim}_tensaominima_{round(tensao_min or 0, 4)}",
                f"{alim}_tensaolinhaminima_{round(tensao_min_linha or 0, 4)}",
                f"{alim}_tensaomax_{round(tensao_max or 0, 4)}",
                f"{alim}_chi_{round(chi or 0, 4)}",
                f"{alim}_ci_{round(ci or 0, 4)}",
            ])
            d_max = demanda_max_por_alim.get(alim)
            if d_max is not None:
                partes.append(f"{alim}_Demand_MAX_{float(d_max):.2f}")
        return ";".join(partes)

    def ganhos_compute_antes(
        self,
        alimentadores: Any = None,
        projeto_investimento: Any = "",
        pasta: Any = "",
    ) -> dict[str, Any]:
        """Calcula as 10 metricas + ganhos_totais 'antes' para uma obra.
        Equivalente a preencher_campos_antes do desktop, mas em uma
        unica chamada (le os 3 .TXT 1x e roda os 8 calculos contra os
        mesmos buffers). Devolve numeros prontos pra popular o form."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "error": "calc indisponivel"}
        alims = self._ganhos_normalize_alims(alimentadores)
        if not alims:
            return {"ok": False, "error": "alimentadores vazio"}
        pi = str(projeto_investimento or "").strip()
        pasta_resolved = self._ganhos_resolve_pasta(pasta)
        if not pasta_resolved:
            return {"ok": False, "error": "pasta de ganhos nao configurada"}
        dados, err = self._ganhos_load_files(
            pasta_resolved,
            ["FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT"])
        if err:
            self._data_state_set(
                "tecnico_txt", "INVALIDADO", path=pasta_resolved, error=err)
            return {"ok": False, "error": err}
        data_flow = dados.get("FlowMT.TXT") or []
        data_topo = dados.get("Topologia.TXT") or []
        data_conf = dados.get("Confiabilidade.TXT") or []
        # Filtra alimentadores ausentes nos arquivos
        alims_validos = self._ganhos_filter_alims(
            alims, [data_flow, data_topo, data_conf])
        if not alims_validos:
            return {"ok": False, "alims_filtrados": alims,
                    "error": "nenhum alimentador encontrado nos arquivos"}
        ignorados = [a for a in alims if a not in alims_validos]
        # Calculos agregados (8 metricas)
        try:
            tensao_min, tensao_media = cm.calcular_tensoes(data_flow, alims_validos)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"calcular_tensoes: {exc}"}
        try:
            tensao_min_linha = cm.calcular_tensao_linha_minima(
                data_flow, alims_validos)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"tensao_linha: {exc}"}
        try:
            tensao_max, _ = cm.calcular_tensoes_max(data_flow, alims_validos)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"tensoes_max: {exc}"}
        try:
            carreg = cm.calcular_carregamento(data_topo, alims_validos)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"carregamento: {exc}"}
        try:
            _patamares, maior_perda = cm.calcular_perdas(
                data_flow, alims_validos)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"perdas: {exc}"}
        try:
            contas_antes, _contas_depois = cm.calcular_contas_contratos(
                data_conf, alims_validos)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"contas_contratos: {exc}"}
        try:
            total_benef = cm.calcular_contas_contratos_beneficiadas(
                data_topo, data_conf, alims_validos, pi) if pi else 0
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"contas_benef: {exc}"}
        try:
            chi_total, ci_total = cm.calcular_chi_ci(data_conf, alims_validos)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"chi_ci: {exc}"}
        # String agregada
        ganhos_totais = self._ganhos_gerar_totais_string(
            cm, alims_validos, data_flow, data_topo, data_conf, pi, "antes")
        # Marca tecnico_txt como VALIDADO (lemos com sucesso)
        try:
            tokens = [
                str(int(os.path.getmtime(os.path.join(pasta_resolved, f))))
                for f in ("FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT")
            ]
            self._data_state_set(
                "tecnico_txt", "CARREGADO_VALIDADO",
                path=pasta_resolved,
                version_token="-".join(tokens))
        except OSError:
            pass
        return {
            "ok": True,
            "alimentadores_validos": alims_validos,
            "alimentadores_ignorados": ignorados,
            "tensao_min": tensao_min,
            "tensao_media": tensao_media,
            "tensao_min_linha": tensao_min_linha,
            "tensao_max": tensao_max,
            "carregamento": carreg,
            "perdas": maior_perda,
            "contas_antes": contas_antes,
            "contas_benef": total_benef,
            "chi": chi_total,
            "ci": ci_total,
            "ganhos_totais": ganhos_totais,
            "pasta": pasta_resolved,
            "error": "",
        }

    def ganhos_compute_depois(
        self,
        alimentadores: Any = None,
        projeto_investimento: Any = "",
        pasta: Any = "",
    ) -> dict[str, Any]:
        """Calcula as 7 metricas + ganhos_totais 'depois' para uma obra.
        Equivalente a preencher_campos_depois do desktop. Igual ao
        antes, mas usa contas_depois (nao antes) e nao soma beneficiadas."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "error": "calc indisponivel"}
        alims = self._ganhos_normalize_alims(alimentadores)
        if not alims:
            return {"ok": False, "error": "alimentadores vazio"}
        pi = str(projeto_investimento or "").strip()
        pasta_resolved = self._ganhos_resolve_pasta(pasta)
        if not pasta_resolved:
            return {"ok": False, "error": "pasta de ganhos nao configurada"}
        dados, err = self._ganhos_load_files(
            pasta_resolved,
            ["FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT"])
        if err:
            self._data_state_set(
                "tecnico_txt", "INVALIDADO", path=pasta_resolved, error=err)
            return {"ok": False, "error": err}
        data_flow = dados.get("FlowMT.TXT") or []
        data_topo = dados.get("Topologia.TXT") or []
        data_conf = dados.get("Confiabilidade.TXT") or []
        alims_validos = self._ganhos_filter_alims(
            alims, [data_flow, data_topo, data_conf])
        if not alims_validos:
            return {"ok": False, "alims_filtrados": alims,
                    "error": "nenhum alimentador encontrado nos arquivos"}
        ignorados = [a for a in alims if a not in alims_validos]
        try:
            tensao_min, tensao_media = cm.calcular_tensoes(data_flow, alims_validos)
            tensao_min_linha = cm.calcular_tensao_linha_minima(
                data_flow, alims_validos)
            tensao_max, _ = cm.calcular_tensoes_max(data_flow, alims_validos)
            carreg = cm.calcular_carregamento(data_topo, alims_validos)
            _patamares, maior_perda = cm.calcular_perdas(
                data_flow, alims_validos)
            _contas_antes, contas_depois = cm.calcular_contas_contratos(
                data_conf, alims_validos)
            chi_total, ci_total = cm.calcular_chi_ci(data_conf, alims_validos)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"calc: {exc}"}
        ganhos_totais = self._ganhos_gerar_totais_string(
            cm, alims_validos, data_flow, data_topo, data_conf, pi, "depois")
        try:
            tokens = [
                str(int(os.path.getmtime(os.path.join(pasta_resolved, f))))
                for f in ("FlowMT.TXT", "Topologia.TXT", "Confiabilidade.TXT")
            ]
            self._data_state_set(
                "tecnico_txt", "CARREGADO_VALIDADO",
                path=pasta_resolved,
                version_token="-".join(tokens))
        except OSError:
            pass
        return {
            "ok": True,
            "alimentadores_validos": alims_validos,
            "alimentadores_ignorados": ignorados,
            "tensao_min": tensao_min,
            "tensao_media": tensao_media,
            "tensao_min_linha": tensao_min_linha,
            "tensao_max": tensao_max,
            "carregamento": carreg,
            "perdas": maior_perda,
            "contas_depois": contas_depois,
            "chi": chi_total,
            "ci": ci_total,
            "ganhos_totais": ganhos_totais,
            "pasta": pasta_resolved,
            "error": "",
        }

    def ganhos_compute_atual(
        self,
        alimentadores: Any = None,
        pasta: Any = "",
    ) -> dict[str, Any]:
        """Calcula 4 metricas + ganhos_totais_atual.
        Equivalente a preencher_parametros_atuais do desktop.
        Usa apenas FlowMT.TXT + Topologia.TXT (Confiabilidade nao
        e necessario)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "error": "calc indisponivel"}
        alims = self._ganhos_normalize_alims(alimentadores)
        if not alims:
            return {"ok": False, "error": "alimentadores vazio"}
        pasta_resolved = self._ganhos_resolve_pasta(pasta)
        if not pasta_resolved:
            return {"ok": False, "error": "pasta de ganhos nao configurada"}
        dados, err = self._ganhos_load_files(
            pasta_resolved, ["FlowMT.TXT", "Topologia.TXT"])
        if err:
            return {"ok": False, "error": err}
        data_flow = dados.get("FlowMT.TXT") or []
        data_topo = dados.get("Topologia.TXT") or []
        try:
            tensao_min, _ = cm.calcular_tensoes(data_flow, alims)
            tensao_min_linha = cm.calcular_tensao_linha_minima(data_flow, alims)
            tensao_max, _ = cm.calcular_tensoes_max(data_flow, alims)
            carreg = cm.calcular_carregamento(data_topo, alims)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"calc: {exc}"}
        # Replica gerar_ganhos_totais_atual: per-alim com prefixo _atual.
        partes: list[str] = []
        for alim in alims:
            partes.extend([
                f"{alim}_carrg_atual_{round(carreg or 0, 4)}",
                f"{alim}_tensao_min_atual_{round(tensao_min or 0, 4)}",
                f"{alim}_tensao_linha_min_atual_{round(tensao_min_linha or 0, 4)}",
                f"{alim}_tensao_max_atual_{round(tensao_max or 0, 4)}",
            ])
        ganhos_atual = ";".join(partes)
        return {
            "ok": True,
            "tensao_min": tensao_min,
            "tensao_min_linha": tensao_min_linha,
            "tensao_max": tensao_max,
            "carregamento": carreg,
            "ganhos_atual": ganhos_atual,
            "tensao_reg_atual": (
                f"{round(tensao_min or 0, 4)}/{round(tensao_max or 0, 4)}"),
            "pasta": pasta_resolved,
            "error": "",
        }

    def ganhos_apply_to_obra(
        self,
        cod: Any = "",
        etapa: Any = "antes",
        alimentadores: Any = None,
        projeto_investimento: Any = "",
        pasta: Any = "",
    ) -> dict[str, Any]:
        """Calcula + persiste em uma chamada para uma obra: chama
        ganhos_compute_(antes|depois) e atualiza
        ganhos_totais_(antes|depois) na linha do banco via
        DatabaseManager.update_obra (passando dict so com a coluna
        alterada e skip_blank=True para nao zerar outros campos)."""
        cod_s = str(cod or "").strip()
        if not cod_s:
            return {"ok": False, "error": "cod vazio"}
        etapa_s = str(etapa or "antes").strip().lower()
        if etapa_s not in ("antes", "depois"):
            return {"ok": False, "error": "etapa deve ser 'antes' ou 'depois'"}
        if etapa_s == "antes":
            r = self.ganhos_compute_antes(
                alimentadores, projeto_investimento, pasta)
        else:
            r = self.ganhos_compute_depois(
                alimentadores, projeto_investimento, pasta)
        if not r.get("ok"):
            return r
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel"}
        col = "ganhos_totais_antes" if etapa_s == "antes" else "ganhos_totais_depois"
        ganhos_str = str(r.get("ganhos_totais") or "")
        try:
            # update_obra recebe dict + cod + skip_blank. Passamos so a
            # coluna que muda; skip_blank=True garante que strings vazias
            # nao sobrescrevam outras colunas (paranoid: o dict so tem 1
            # chave aqui, mas mantemos o flag).
            db.update_obra({col: ganhos_str}, cod_s, skip_blank=True)
        except Exception as exc:  # noqa: BLE001
            friendly = self._friendly_busy_error(exc)
            out: dict[str, Any] = {
                "ok": False,
                "error": friendly or f"update_obra: {exc}",
                "computed": r,
            }
            if friendly:
                out["blocked"] = "db_busy"
            return out
        return {
            "ok": True,
            "cod": cod_s,
            "etapa": etapa_s,
            "ganhos_totais": ganhos_str,
            "alimentadores_validos": r.get("alimentadores_validos") or [],
            "alimentadores_ignorados": r.get("alimentadores_ignorados") or [],
            "metricas": {k: v for k, v in r.items()
                         if k not in ("ok", "error", "ganhos_totais",
                                      "alimentadores_validos",
                                      "alimentadores_ignorados", "pasta")},
        }

    def ganhos_apply_massa(
        self,
        cods: Any = None,
        etapa: Any = "antes",
        pasta: Any = "",
    ) -> dict[str, Any]:
        """Aplica ganhos em massa para uma lista de cods.
        Para cada obra: le alimentadores + projeto da linha, computa
        e persiste. Equivalente a preencher_ganhos_massa do desktop."""
        cods_list = cods or []
        if not isinstance(cods_list, list):
            cods_list = [cods_list]
        cods_list = [str(c).strip() for c in cods_list if str(c or "").strip()]
        if not cods_list:
            return {"ok": False, "error": "cods vazio"}
        etapa_s = str(etapa or "antes").strip().lower()
        if etapa_s not in ("antes", "depois"):
            return {"ok": False, "error": "etapa deve ser 'antes' ou 'depois'"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db indisponivel"}
        try:
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"cols: {exc}"}
        i_cod = cols.index("cod") if "cod" in cols else -1
        i_alim = cols.index("alimentador_principal") if "alimentador_principal" in cols else -1
        i_alim_benef = cols.index("alimentadores_beneficiados") if "alimentadores_beneficiados" in cols else -1
        i_pi = cols.index("projeto_investimento") if "projeto_investimento" in cols else -1
        if i_cod < 0 or i_alim < 0 or i_pi < 0:
            return {"ok": False, "error": "schema incompativel (cod/alim/pi)"}
        try:
            rows = list(db.fetch_by_cods(cods_list) or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"fetch_by_cods: {exc}"}
        sucesso = 0
        falhas: list[dict[str, str]] = []
        ignoradas: list[str] = []
        for row in rows:
            cod_v = str(row[i_cod]).strip() if 0 <= i_cod < len(row) else ""
            alim_p = str(row[i_alim]).strip() if 0 <= i_alim < len(row) else ""
            alim_b_raw = (str(row[i_alim_benef]).strip()
                          if 0 <= i_alim_benef < len(row) else "")
            pi = str(row[i_pi]).strip() if 0 <= i_pi < len(row) else ""
            alims = [a for a in [alim_p] if a]
            if alim_b_raw:
                alims.extend([a.strip() for a in alim_b_raw.split(",")
                              if a.strip()])
            if not alims:
                ignoradas.append(cod_v)
                continue
            r = self.ganhos_apply_to_obra(
                cod_v, etapa_s, alims, pi, pasta)
            if r.get("ok"):
                sucesso += 1
            else:
                falhas.append({"cod": cod_v,
                               "error": r.get("error") or "?"})
        return {
            "ok": True,
            "sucesso": sucesso,
            "falhas": falhas,
            "ignoradas_sem_alim": ignoradas,
            "total": len(cods_list),
        }

    def calc_nota_carregamento(
        self, carreg_inicial: Any = 0, carreg_max: Any = 0,
    ) -> dict[str, Any]:
        """Nota de carregamento por faixa
        (CalculationManager.calcular_nota_carregamento)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "nota": None, "criterio": "",
                    "error": "calc indisponivel"}
        try:
            nota, crit = cm.calcular_nota_carregamento(
                float(carreg_inicial or 0), float(carreg_max or 0)
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "nota": None, "criterio": "",
                    "error": f"nota_carreg: {exc}"}
        return {"ok": True, "nota": nota, "criterio": str(crit or ""), "error": ""}

    def calc_nota_tensao_min(
        self, tensao_min_atual: Any = 0, tensao_min_inicial: Any = 0,
    ) -> dict[str, Any]:
        """Nota de tensao minima via tabela verdade 4x4
        (CalculationManager.calcular_nota_tensao_min)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "nota": None, "criterio": "",
                    "error": "calc indisponivel"}
        try:
            nota, crit = cm.calcular_nota_tensao_min(
                float(tensao_min_atual or 0), float(tensao_min_inicial or 0)
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "nota": None, "criterio": "",
                    "error": f"nota_tmin: {exc}"}
        return {"ok": True, "nota": nota, "criterio": str(crit or ""), "error": ""}

    def calc_nota_tensao_max(
        self, tensao_min_atual: Any = 0, tensao_max_inicial: Any = 0,
    ) -> dict[str, Any]:
        """Nota de tensao maxima por faixa
        (CalculationManager.calcular_nota_tensao_max)."""
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "nota": None, "criterio": "",
                    "error": "calc indisponivel"}
        try:
            nota, crit = cm.calcular_nota_tensao_max(
                float(tensao_min_atual or 0), float(tensao_max_inicial or 0)
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "nota": None, "criterio": "",
                    "error": f"nota_tmax: {exc}"}
        return {"ok": True, "nota": nota, "criterio": str(crit or ""), "error": ""}

    def calc_nota_colapso_obra(self, cod: Any = "") -> dict[str, Any]:
        """Nota de colapso para UMA obra (CalculationManager.calcular_nota_colapso_obra).
        Versao por COD; o export_nota_colapso ja faz para varias."""
        cod_s = str(cod or "").strip()
        if not cod_s:
            return {"ok": False, "nota": None, "criterio": "",
                    "error": "cod vazio"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "nota": None, "criterio": "",
                    "error": err or "db indisponivel"}
        try:
            row = db.fetch_by_cod(cod_s)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "nota": None, "criterio": "",
                    "error": f"fetch: {exc}"}
        if not row:
            return {"ok": False, "nota": None, "criterio": "",
                    "error": f"obra nao encontrada: {cod_s}"}
        cm = self._ensure_calc_manager()
        if cm is None:
            return {"ok": False, "nota": None, "criterio": "",
                    "error": "calc indisponivel"}
        try:
            nota, crit = cm.calcular_nota_colapso_obra(row)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "nota": None, "criterio": "",
                    "error": f"nota_colapso: {exc}"}
        return {"ok": True, "nota": nota, "criterio": str(crit or ""), "error": ""}

    # --- Fase 4: Apoio listings completas ----------------------------

    def _apoio_cache_dict(self) -> dict[str, Any]:
        """Devolve _apoio_cache, hidratando do banco se vazio.
        Centraliza leitura de apoio para todas as bridges apoio_get_*."""
        cache = getattr(self, "_apoio_cache", None) or {}
        if cache:
            return cache
        # Tenta hidratar do banco
        db, _err = self._ensure_db_connected()
        if db is None:
            return {}
        loaded = self._apoio_load_from_db(db)
        if loaded:
            self._apoio_cache = loaded
            meta = self._apoio_meta_dict(db)
            self._apoio_path_loaded = str(meta.get("last_path") or "")
            return loaded
        return {}

    def apoio_get_projetos_investimento(self) -> dict[str, Any]:
        """Lista de projetos de investimento (DB-backed via apoio_apoio).
        Use 'Atualizar apoio' em Configuracoes para popular."""
        cache = self._apoio_cache_dict()
        items = list(cache.get("projetos_investimento") or [])
        return {"ok": True, "items": items, "error": ""}

    def apoio_get_caracteristicas(self) -> dict[str, Any]:
        """Lista de caracteristicas (DB-backed via apoio_apoio)."""
        cache = self._apoio_cache_dict()
        items = list(cache.get("caracteristicas") or [])
        return {"ok": True, "items": items, "error": ""}

    def apoio_get_nomes_projetos(self) -> dict[str, Any]:
        """Lista de nomes de projeto pre-definidos (DB-backed)."""
        cache = self._apoio_cache_dict()
        items = list(cache.get("nomes_projetos_pre_definidos") or [])
        return {"ok": True, "items": items, "error": ""}

    def list_nomes_projetos(self) -> dict[str, Any]:
        """Mescla nomes de projeto vindos do APOIO (xlsx) +
        DISTINCT nome_projeto do banco. Replica
        populate_combo_nome_projeto do desktop:
          * Dedup case-insensitive (upper)
          * "MELHORIAS AL" normalizado para "Melhorias AL" (e
            sempre presente, mesmo sem vir de fonte)
          * Ordem: apoio primeiro, depois banco

        Usado pelo combo "Nome do Projeto" do tab Cadastro."""
        nomes: list[str] = []

        # 1) Apoio (DB-backed: tabela apoio_apoio coluna 'Nome de Projeto')
        try:
            cache = self._apoio_cache_dict()
            apoio_nomes = list(
                cache.get("nomes_projetos_pre_definidos") or []
            )
            for n in apoio_nomes:
                s = str(n or "").strip()
                if s:
                    nomes.append(s)
        except Exception as exc:  # noqa: BLE001
            print(f"[main_web] list_nomes_projetos apoio falhou: {exc}",
                  file=sys.stderr)

        # 2) Banco (distinct nome_projeto)
        db, err = self._ensure_db_connected()
        if not err and db is not None:
            try:
                from runtime.database import open_sqlite_safe  # noqa: PLC0415
                db_path = (self._config or {}).get("obras") or ""
                if db_path and os.path.isfile(str(db_path)):
                    conn = open_sqlite_safe(str(db_path))
                    try:
                        cur = conn.cursor()
                        cur.execute(
                            "SELECT DISTINCT nome_projeto FROM obras "
                            "WHERE nome_projeto IS NOT NULL "
                            "AND TRIM(nome_projeto) != ''"
                        )
                        for row in cur.fetchall():
                            s = str(row[0] or "").strip()
                            if s:
                                nomes.append(s)
                    finally:
                        try: conn.close()
                        except Exception: pass  # noqa: BLE001
            except Exception as exc:  # noqa: BLE001
                print(f"[main_web] list_nomes_projetos db falhou: {exc}",
                      file=sys.stderr)

        # 3) Dedup case-insensitive + tratamento "MELHORIAS AL"
        seen: set[str] = set()
        out: list[str] = []
        melhorias_inserido = False
        for nome in nomes:
            chave = nome.upper()
            if chave in seen:
                continue
            seen.add(chave)
            if chave == "MELHORIAS AL":
                if not melhorias_inserido:
                    out.append("Melhorias AL")
                    melhorias_inserido = True
                continue
            out.append(nome)
        if not melhorias_inserido:
            out.append("Melhorias AL")

        return {"ok": True, "items": out, "count": len(out), "error": ""}

    # ------------------------------------------------------------------
    # Cenarios DB-backed (2026-05-08): overrides isolados por cenario
    #
    # Tabelas envolvidas (3):
    #   * cenarios_meta            (criada pelo CAPEX): metadados
    #   * cenarios_obras           (criada pelo CAPEX): cod -> ano_final
    #   * cenario_obras_overrides  (criada por nos):    diffs por coluna
    #
    # Quando cenario X esta ativo (config['cenario_ativo']):
    #   - Visualizar mostra APENAS obras com cod em cenarios_obras WHERE
    #     cenario_nome=X
    #   - Leituras aplicam overrides: cada coluna pode ter valor distinto
    #     do que esta em obras (sem nunca tocar obras)
    #   - Save_obra grava o diff em cenario_obras_overrides; obras nunca
    #     muda enquanto cenario ativo
    #   - Operacoes massa (excluir, atualizar valor, marcar correcao) sao
    #     bloqueadas com mensagem clara
    #
    # Referencia: CAPEX define cenarios_meta + cenarios_obras com schema
    # fixo. Ver apps/capex/web/main_web.py:521+ (constantes
    # SCENARIO_METADATA_TABLE = "cenarios_meta", SCENARIO_OBRAS_TABLE =
    # "cenarios_obras").
    # ------------------------------------------------------------------
    @classmethod
    def _cenario_ensure_overrides_table(cls, db: Any) -> None:
        """Cria cenario_obras_overrides. Conn aux thread-safe."""
        conn, _err = cls._open_aux_conn(db)
        if conn is None:
            return
        try:
            cursor = conn.cursor()
            cursor.execute(
                "CREATE TABLE IF NOT EXISTS cenario_obras_overrides ("
                " cenario_nome TEXT NOT NULL,"
                " cod TEXT NOT NULL,"
                " coluna TEXT NOT NULL,"
                " valor TEXT,"
                " atualizado_em TEXT,"
                " atualizado_por TEXT,"
                " PRIMARY KEY (cenario_nome, cod, coluna)"
                ")"
            )
            conn.commit()
        except Exception:  # noqa: BLE001
            pass
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass

    def _cenario_active_name(self) -> str:
        """Le config['cenario_ativo']. '' = sem cenario ativo.

        FIX 2026-05-08: cenario_active_set invalida self._config (=None)
        para refletir mudancas de outros campos. Aqui recarregamos do
        disco quando vazio, senao o cenario salvo nao seria visto pelas
        chamadas subsequentes (filtro de obras, save_obra, etc.)."""
        if self._config is None:
            try:
                self._reload_config()
            except Exception:  # noqa: BLE001
                pass
        cfg = self._config or {}
        return str(cfg.get("cenario_ativo") or "").strip()

    def _cenario_cod_set(
        self, db: Any, cenario_nome: str,
    ) -> tuple[set[str], dict[str, Any]]:
        """Le cenarios_obras para um cenario, retornando:
            ({cod1, cod2, ...}, {cod -> {ano_final, ano_origem}})
        Vazio se tabela nao existe ou cenario sem obras."""
        if not cenario_nome:
            return set(), {}
        conn, _err = self._open_aux_conn(db)
        if conn is None:
            return set(), {}
        rows = []
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
                " AND name='cenarios_obras'"
            )
            if not cursor.fetchone():
                return set(), {}
            cursor.execute(
                "SELECT cod, ano_final, ano_origem"
                " FROM cenarios_obras WHERE cenario_nome = ?",
                (cenario_nome,),
            )
            rows = cursor.fetchall() or []
        except Exception:  # noqa: BLE001
            return set(), {}
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        cod_set: set[str] = set()
        info: dict[str, Any] = {}
        for r in rows:
            c = str(r[0] or "").strip()
            if not c:
                continue
            cod_set.add(c)
            info[c] = {
                "ano_final":  int(r[1]) if r[1] is not None else None,
                "ano_origem": int(r[2]) if r[2] is not None else None,
            }
        return cod_set, info

    def _cenario_overrides_map(
        self, db: Any, cenario_nome: str,
    ) -> dict[str, dict[str, str]]:
        """Le cenario_obras_overrides retornando {cod -> {coluna -> valor}}."""
        if not cenario_nome:
            return {}
        self._cenario_ensure_overrides_table(db)
        out: dict[str, dict[str, str]] = {}
        conn, _err = self._open_aux_conn(db)
        if conn is None:
            return {}
        rows = []
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT cod, coluna, valor FROM cenario_obras_overrides"
                " WHERE cenario_nome = ?",
                (cenario_nome,),
            )
            rows = cursor.fetchall() or []
        except Exception:  # noqa: BLE001
            return {}
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        for r in rows:
            cod = str(r[0] or "").strip()
            col = str(r[1] or "").strip()
            val = "" if r[2] is None else str(r[2])
            if not cod or not col:
                continue
            out.setdefault(cod, {})[col] = val
        return out

    def _cenario_apply_to_row(
        self,
        row: list[Any] | tuple[Any, ...],
        cols: list[str],
        cod: str,
        cen_info: dict[str, Any],
        overrides: dict[str, str],
    ) -> list[Any]:
        """Devolve uma copia de ``row`` com:
        - ano_ substituido por cenarios_obras.ano_final (se houver)
        - cada coluna em overrides substituida pelo valor do override
        Sem mutar a row original."""
        new_row = list(row)
        # ano_final do cenario tem prioridade quando NAO ha override
        # explicito de ano_ em cenario_obras_overrides.
        ano_final = (cen_info or {}).get("ano_final")
        if ano_final is not None:
            try:
                idx_ano = cols.index("ano_")
                if "ano_" not in overrides:
                    new_row[idx_ano] = str(ano_final)
            except ValueError:
                pass
        # Aplica overrides genericos
        for col_name, val in (overrides or {}).items():
            try:
                idx_col = cols.index(col_name)
                new_row[idx_col] = val
            except ValueError:
                continue
        return new_row

    def _apply_cenario_to_rows(
        self, db: Any, raw_rows: list[Any], cols: list[str],
    ) -> list[Any]:
        """Aplica cenario ativo a um conjunto de rows ja fetched:
        restringe aos CODs do cenario e aplica overrides. Idempotente:
        sem cenario ativo devolve raw_rows inalterado. Usado pelos
        export_* para garantir paridade com get_obras."""
        cen_nome = self._cenario_active_name()
        if not cen_nome:
            return list(raw_rows)
        try:
            cod_set, cen_info = self._cenario_cod_set(db, cen_nome)
            ovmap = self._cenario_overrides_map(db, cen_nome)
        except Exception:  # noqa: BLE001
            return list(raw_rows)
        if not cod_set:
            # cenario com 0 obras (ou tabelas inexistentes)
            return []
        try:
            idx_cod = cols.index("cod")
        except ValueError:
            return list(raw_rows)
        filtered: list[Any] = []
        for r in raw_rows:
            cod_r = str(r[idx_cod] if idx_cod < len(r) else "").strip()
            if cod_r not in cod_set:
                continue
            filtered.append(self._cenario_apply_to_row(
                r, cols, cod_r,
                cen_info.get(cod_r) or {},
                ovmap.get(cod_r) or {},
            ))
        return filtered

    # -------- Bridges publicas --------

    def cenario_list(self) -> dict[str, Any]:
        """Lista cenarios disponiveis (le cenarios_meta + count obras).
        Devolve [{nome, descricao, total_obras, criado_em, criado_por,
        atualizado_em, atualizado_por, solicitado_em}, ...] ordenado
        por atualizado_em desc.

        Vazio quando o CAPEX nao criou tabelas no banco ainda."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "cenarios": [], "error": err or "db"}
        conn, err_open = self._open_aux_conn(db)
        if conn is None:
            return {"ok": True, "cenarios": [], "error": err_open or ""}
        rows = []
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
                " AND name='cenarios_meta'"
            )
            if not cursor.fetchone():
                return {"ok": True, "cenarios": [], "error": ""}
            cursor.execute(
                "SELECT nome, descricao, solicitado_em, criado_em,"
                " criado_por, atualizado_em, atualizado_por,"
                " total_obras"
                " FROM cenarios_meta"
                " ORDER BY COALESCE(atualizado_em, criado_em) DESC"
            )
            rows = cursor.fetchall() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "cenarios": [], "error": str(exc)}
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        out = []
        for r in rows:
            out.append({
                "nome":          str(r[0] or ""),
                "descricao":     str(r[1] or ""),
                "solicitado_em": str(r[2] or ""),
                "criado_em":     str(r[3] or ""),
                "criado_por":    str(r[4] or ""),
                "atualizado_em": str(r[5] or ""),
                "atualizado_por": str(r[6] or ""),
                "total_obras":   int(r[7] or 0),
            })
        return {"ok": True, "cenarios": out, "error": ""}

    def cenario_active_get(self) -> dict[str, Any]:
        """Devolve {ativo: nome, ano_final_count, overrides_count}.
        ativo='' quando nenhum cenario esta ativo."""
        nome = self._cenario_active_name()
        if not nome:
            return {"ok": True, "ativo": "", "error": ""}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": True, "ativo": nome, "error": err or "",
                    "ano_final_count": 0, "overrides_count": 0}
        cod_set, _info = self._cenario_cod_set(db, nome)
        ovmap = self._cenario_overrides_map(db, nome)
        return {
            "ok": True, "ativo": nome, "error": "",
            "ano_final_count": len(cod_set),
            "overrides_count": sum(len(v) for v in ovmap.values()),
        }

    def cenario_active_set(self, nome: Any = "") -> dict[str, Any]:
        """Ativa ou desativa cenario. ``nome`` vazio = desativa.
        Salva em config['cenario_ativo'] persistente."""
        try:
            from codigo5_coplan import ConfigManager  # type: ignore[import-not-found]
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"import: {exc}"}
        nome_s = str(nome or "").strip()
        # Valida que o cenario existe (se nao vazio)
        if nome_s:
            r = self.cenario_list()
            if r.get("ok"):
                nomes = {c.get("nome") for c in (r.get("cenarios") or [])}
                if nome_s not in nomes:
                    return {"ok": False,
                            "error": f"cenario nao encontrado: {nome_s}"}
        try:
            ConfigManager.save_config({"cenario_ativo": nome_s})
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"save: {exc}"}
        # Recarrega config (em vez de so invalidar) para que
        # _cenario_active_name veja o novo valor imediatamente
        # mesmo se chamadas concorrentes ja terem feito
        # `cfg = self._config or {}` antes do reload.
        try:
            self._reload_config()
        except Exception:  # noqa: BLE001
            self._config = None
        if nome_s:
            return {"ok": True, "ativo": nome_s, "error": ""}
        return {"ok": True, "ativo": "", "error": ""}

    def cenario_get_overrides(
        self, nome: Any = "", cod: Any = "",
    ) -> dict[str, Any]:
        """Devolve overrides salvos. Se cod informado, filtra. Se nome
        vazio, usa o cenario ativo."""
        nome_s = str(nome or "").strip() or self._cenario_active_name()
        if not nome_s:
            return {"ok": False, "error": "nenhum cenario informado",
                    "items": []}
        cod_s = str(cod or "").strip()
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db", "items": []}
        self._cenario_ensure_overrides_table(db)
        conn, err_open = self._open_aux_conn(db)
        if conn is None:
            return {"ok": False, "error": err_open or "conn", "items": []}
        rows = []
        try:
            cursor = conn.cursor()
            if cod_s:
                cursor.execute(
                    "SELECT cod, coluna, valor, atualizado_em,"
                    " atualizado_por FROM cenario_obras_overrides"
                    " WHERE cenario_nome=? AND cod=?"
                    " ORDER BY coluna",
                    (nome_s, cod_s),
                )
            else:
                cursor.execute(
                    "SELECT cod, coluna, valor, atualizado_em,"
                    " atualizado_por FROM cenario_obras_overrides"
                    " WHERE cenario_nome=?"
                    " ORDER BY cod, coluna",
                    (nome_s,),
                )
            rows = cursor.fetchall() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": str(exc), "items": []}
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        items = [
            {"cod": str(r[0] or ""),
             "coluna": str(r[1] or ""),
             "valor": str(r[2] or ""),
             "atualizado_em": str(r[3] or ""),
             "atualizado_por": str(r[4] or "")}
            for r in rows
        ]
        return {"ok": True, "items": items, "cenario": nome_s,
                "count": len(items), "error": ""}

    def cenario_clear_overrides(
        self,
        nome: Any = "",
        cod: Any = "",
        coluna: Any = "",
    ) -> dict[str, Any]:
        """Remove overrides. Se ``coluna`` informado, so essa coluna.
        Se ``cod`` informado, so dessa obra. Sem coluna nem cod,
        zera TODOS os overrides do cenario."""
        nome_s = str(nome or "").strip() or self._cenario_active_name()
        if not nome_s:
            return {"ok": False, "error": "nenhum cenario informado"}
        cod_s = str(cod or "").strip()
        col_s = str(coluna or "").strip()
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "error": err or "db"}
        self._cenario_ensure_overrides_table(db)
        conn, err_open = self._open_aux_conn(db)
        if conn is None:
            return {"ok": False, "error": err_open or "conn"}
        removed = 0
        try:
            cursor = conn.cursor()
            if cod_s and col_s:
                cursor.execute(
                    "DELETE FROM cenario_obras_overrides"
                    " WHERE cenario_nome=? AND cod=? AND coluna=?",
                    (nome_s, cod_s, col_s),
                )
            elif cod_s:
                cursor.execute(
                    "DELETE FROM cenario_obras_overrides"
                    " WHERE cenario_nome=? AND cod=?",
                    (nome_s, cod_s),
                )
            else:
                cursor.execute(
                    "DELETE FROM cenario_obras_overrides"
                    " WHERE cenario_nome=?",
                    (nome_s,),
                )
            removed = cursor.rowcount
            conn.commit()
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": str(exc)}
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        return {"ok": True, "removed": int(removed or 0),
                "cenario": nome_s, "error": ""}

    def cenario_obras_branches(
        self, cods: Any = None,
    ) -> dict[str, Any]:
        """Para cada cod (opcional, padrao todos), devolve lista de
        cenarios onde a obra TEM ALGUMA versao diferente.

        Une 2 fontes:
        - cenarios_obras (CAPEX): obra pertence ao cenario (ano_final
          pode diferir de ano_origem; sempre conta como versao).
        - cenario_obras_overrides (COPLAN): editou outros campos.

        Resposta: {ok, items: {cod: [{cenario, campos: [str,...],
                  atualizado_em, atualizado_por,
                  ano_final, ano_origem}, ...]}}."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "items": {}, "error": err or "db"}
        cod_list: list[str] = []
        if isinstance(cods, (list, tuple)):
            cod_list = [
                str(c).strip() for c in cods if str(c or "").strip()
            ]
        conn, _err = self._open_aux_conn(db)
        if conn is None:
            return {"ok": True, "items": {}, "error": ""}
        # grouped: cod -> cenario_nome -> dict
        grouped: dict[str, dict[str, dict[str, Any]]] = {}
        try:
            cursor = conn.cursor()

            def _table_exists(name: str) -> bool:
                cursor.execute(
                    "SELECT 1 FROM sqlite_master WHERE type='table'"
                    " AND name=?",
                    (name,),
                )
                return cursor.fetchone() is not None

            # ----- 1) cenarios_obras (do CAPEX) -----
            if _table_exists("cenarios_obras"):
                if cod_list:
                    if len(cod_list) > 900:
                        cod_list = cod_list[:900]
                    placeholders = ",".join(["?"] * len(cod_list))
                    cursor.execute(
                        f"SELECT co.cod, co.cenario_nome, co.ano_final,"
                        f" co.ano_origem, COALESCE(cm.atualizado_em,"
                        f" cm.criado_em, '') AS em,"
                        f" COALESCE(cm.atualizado_por,"
                        f" cm.criado_por, '') AS por"
                        f" FROM cenarios_obras co"
                        f" LEFT JOIN cenarios_meta cm"
                        f" ON cm.nome = co.cenario_nome"
                        f" WHERE co.cod IN ({placeholders})",
                        cod_list,
                    )
                else:
                    cursor.execute(
                        "SELECT co.cod, co.cenario_nome, co.ano_final,"
                        " co.ano_origem, COALESCE(cm.atualizado_em,"
                        " cm.criado_em, '') AS em,"
                        " COALESCE(cm.atualizado_por,"
                        " cm.criado_por, '') AS por"
                        " FROM cenarios_obras co"
                        " LEFT JOIN cenarios_meta cm"
                        " ON cm.nome = co.cenario_nome"
                    )
                for r in cursor.fetchall() or []:
                    cod = str(r[0] or "").strip()
                    cen = str(r[1] or "").strip()
                    if not cod or not cen:
                        continue
                    af = r[2]
                    ao = r[3]
                    em = str(r[4] or "").strip()
                    por = str(r[5] or "").strip()
                    cen_d = grouped.setdefault(cod, {}).setdefault(cen, {
                        "cenario": cen, "campos": [],
                        "atualizado_em": "", "atualizado_por": "",
                        "ano_final": None, "ano_origem": None,
                    })
                    cen_d["ano_final"] = (
                        int(af) if af is not None else None
                    )
                    cen_d["ano_origem"] = (
                        int(ao) if ao is not None else None
                    )
                    # Se ano_final difere de ano_origem, conta como
                    # alteracao de "ano_" no campos do tooltip.
                    if (af is not None and ao is not None
                            and int(af) != int(ao)):
                        if "ano_" not in cen_d["campos"]:
                            cen_d["campos"].append("ano_")
                    if em > cen_d["atualizado_em"]:
                        cen_d["atualizado_em"] = em
                        cen_d["atualizado_por"] = por

            # ----- 2) cenario_obras_overrides (do COPLAN) -----
            if _table_exists("cenario_obras_overrides"):
                if cod_list:
                    placeholders = ",".join(["?"] * len(cod_list))
                    cursor.execute(
                        f"SELECT cod, cenario_nome, coluna, valor,"
                        f" atualizado_em, atualizado_por"
                        f" FROM cenario_obras_overrides"
                        f" WHERE cod IN ({placeholders})",
                        cod_list,
                    )
                else:
                    cursor.execute(
                        "SELECT cod, cenario_nome, coluna, valor,"
                        " atualizado_em, atualizado_por"
                        " FROM cenario_obras_overrides"
                    )
                for r in cursor.fetchall() or []:
                    cod = str(r[0] or "").strip()
                    cen = str(r[1] or "").strip()
                    col = str(r[2] or "").strip()
                    em = str(r[4] or "").strip()
                    por = str(r[5] or "").strip()
                    if not cod or not cen or not col:
                        continue
                    cen_d = grouped.setdefault(cod, {}).setdefault(cen, {
                        "cenario": cen, "campos": [],
                        "atualizado_em": "", "atualizado_por": "",
                        "ano_final": None, "ano_origem": None,
                    })
                    if col not in cen_d["campos"]:
                        cen_d["campos"].append(col)
                    if em > cen_d["atualizado_em"]:
                        cen_d["atualizado_em"] = em
                        cen_d["atualizado_por"] = por
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "items": {}, "error": str(exc)}
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass

        items: dict[str, list[dict[str, Any]]] = {}
        for cod, cen_d in grouped.items():
            items[cod] = sorted(
                cen_d.values(),
                key=lambda x: x.get("atualizado_em") or "",
                reverse=True,
            )
        return {"ok": True, "items": items, "error": ""}

    def _cenario_save_overrides(
        self,
        db: Any,
        cenario_nome: str,
        cod: str,
        diff_pairs: list[tuple[str, Any]],
    ) -> int:
        """INSERT OR REPLACE em cenario_obras_overrides para cada
        (coluna, valor) em diff_pairs. Devolve quantas linhas escreveu."""
        if not (cenario_nome and cod and diff_pairs):
            return 0
        self._cenario_ensure_overrides_table(db)
        import getpass as _gp
        try:
            user = _gp.getuser() or "?"
        except Exception:  # noqa: BLE001
            user = "?"
        now = datetime.now().isoformat(timespec="seconds")
        wrote = 0
        conn, _err = self._open_aux_conn(db)
        if conn is None:
            return 0
        try:
            cursor = conn.cursor()
            for col, val in diff_pairs:
                cursor.execute(
                    "INSERT OR REPLACE INTO cenario_obras_overrides"
                    " (cenario_nome, cod, coluna, valor,"
                    " atualizado_em, atualizado_por)"
                    " VALUES (?, ?, ?, ?, ?, ?)",
                    (cenario_nome, cod, col,
                     "" if val is None else str(val),
                     now, user),
                )
                wrote += 1
            conn.commit()
        except Exception as exc:  # noqa: BLE001
            print(f"[cenario] save_overrides falhou: {exc}",
                  file=sys.stderr)
            return 0
        finally:
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass
        return wrote

    def apoio_clear(self) -> dict[str, Any]:
        """Limpa apenas o cache em memoria (forca proxima leitura a
        re-hidratar do banco via _apoio_cache_dict). NAO apaga as
        tabelas apoio_* do banco -- use 'Atualizar apoio' para
        reescrever."""
        try:
            self._apoio_cache = {}
            self._apoio_path_loaded = ""
            # support_manager.clear_data tambem zera self.dados em
            # memoria (fallback para a rota force_reload=True).
            if self._support_manager is not None:
                try:
                    self._support_manager.clear_data()
                except Exception:  # noqa: BLE001
                    pass
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"clear: {exc}"}
        return {"ok": True, "error": ""}

    # --- Fase 5: Validacoes pre-acao (dry-run) ------------------------

    def validate_alimentadores(
        self, alimentador: Any = "", beneficiados: Any = None,
    ) -> dict[str, Any]:
        """Valida nomes de alimentador (sem '_').
        (salvar_obra_service.aplicar_alimentador_validations)"""
        try:
            from core.services.salvar_obra_service import (
                aplicar_alimentador_validations,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "erros": [f"import: {exc}"]}
        benef_list: list[str] = []
        if isinstance(beneficiados, str):
            import re as _re
            benef_list = [
                p.strip() for p in _re.split(r"[,;|\n]+", beneficiados)
                if p and p.strip()
            ]
        elif isinstance(beneficiados, (list, tuple)):
            benef_list = [str(p).strip() for p in beneficiados if str(p or "").strip()]
        try:
            erros = aplicar_alimentador_validations(
                str(alimentador or ""), benef_list,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "erros": [f"validate: {exc}"]}
        return {"ok": True, "erros": list(erros), "valido": not erros}

    def validate_obra_integridade(self, payload: Any) -> dict[str, Any]:
        """Valida integridade minima de UMA obra (dict).
        (row_helpers.validate_min_integrity em modo single-row)"""
        if not isinstance(payload, dict):
            return {"ok": False, "valido": False,
                    "motivos": ["payload nao e dict"]}
        try:
            from runtime.row_helpers import validate_min_integrity
            ok, _counters, samples = validate_min_integrity([payload])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "valido": False,
                    "motivos": [f"validate: {exc}"]}
        return {"ok": True, "valido": bool(ok),
                "motivos": list(samples or [])}

    def validate_ganhos(
        self, payload: Any, tolerancia: Any = None,
    ) -> dict[str, Any]:
        """Valida consistencia ganhos antes/depois/total
        (row_helpers.validate_ganhos_consistency)."""
        if not isinstance(payload, dict):
            return {"ok": False, "valido": False, "motivos": ["payload nao e dict"]}
        try:
            from runtime.row_helpers import (
                _get_ganhos_tolerancia, validate_ganhos_consistency,
            )
            tol = (
                float(tolerancia)
                if tolerancia is not None
                else _get_ganhos_tolerancia(self._config or None)
            )
            ok, motivos = validate_ganhos_consistency(payload, tol)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "valido": False,
                    "motivos": [f"validate: {exc}"]}
        return {"ok": True, "valido": bool(ok),
                "motivos": list(motivos or [])}

    def find_duplicate(self, payload: Any) -> dict[str, Any]:
        """Procura obra duplicada no banco para o payload informado
        (row_helpers.find_duplicate_in_db)."""
        if not isinstance(payload, dict):
            return {"ok": False, "duplicate": None,
                    "error": "payload nao e dict"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "duplicate": None,
                    "error": err or "db indisponivel"}
        try:
            from runtime.row_helpers import find_duplicate_in_db
            dup = find_duplicate_in_db(db, payload)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "duplicate": None,
                    "error": f"find_duplicate: {exc}"}
        return {"ok": True,
                "duplicate": dict(dup) if isinstance(dup, dict) else None,
                "error": ""}

    def avaliar_diff_obra(self, cod: Any, payload: Any) -> dict[str, Any]:
        """Dry-run: avalia o diff que UM save_obra faria, sem persistir.
        (salvar_obra_service.avaliar_diff)"""
        cod_s = str(cod or "").strip()
        if not isinstance(payload, dict):
            return {"ok": False, "campos_alterados": [],
                    "campos_criticos_alterados": [],
                    "despacho_status": "", "historico_col": None,
                    "historico_base": "",
                    "error": "payload nao e dict"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "campos_alterados": [],
                    "campos_criticos_alterados": [],
                    "despacho_status": "", "historico_col": None,
                    "historico_base": "",
                    "error": err or "db indisponivel"}
        try:
            cols = list(db.get_column_names() or [])
            old_map: dict[str, Any] = {}
            if cod_s:
                row = db.fetch_by_cod(cod_s)
                if row:
                    old_map = {c: row[i] if i < len(row) else ""
                               for i, c in enumerate(cols)}
            from core.services.salvar_obra_service import avaliar_diff
            decision = avaliar_diff(payload, old_map, db_columns=cols)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "campos_alterados": [],
                    "campos_criticos_alterados": [],
                    "despacho_status": "", "historico_col": None,
                    "historico_base": "",
                    "error": f"avaliar_diff: {exc}"}
        return {
            "ok": True,
            "campos_alterados": list(decision.campos_alterados),
            "campos_criticos_alterados": list(decision.campos_criticos_alterados),
            "despacho_status": str(decision.despacho_status or ""),
            "historico_col": decision.historico_col,
            "historico_base": str(decision.historico_base or ""),
            "error": "",
        }

    def check_bloqueado_despachada(
        self, cod: Any, payload: Any,
    ) -> dict[str, Any]:
        """True se save deve ser bloqueado por DESPACHADA + criticos
        (salvar_obra_service.bloqueado_por_despachada)."""
        diff = self.avaliar_diff_obra(cod, payload)
        if not diff.get("ok"):
            return {"ok": False, "bloqueado": False, "error": diff.get("error", "")}
        try:
            from core.services.salvar_obra_service import (
                DiffResult, bloqueado_por_despachada,
            )
            decision = DiffResult(
                campos_alterados=list(diff["campos_alterados"]),
                campos_criticos_alterados=list(diff["campos_criticos_alterados"]),
                despacho_status=diff["despacho_status"],
                historico_col=diff["historico_col"],
                historico_base=diff["historico_base"],
            )
            blocked = bloqueado_por_despachada(decision)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "bloqueado": False,
                    "error": f"check_bloqueado: {exc}"}
        return {"ok": True, "bloqueado": bool(blocked), "error": ""}

    # --- Fase 6: CSV import/export -----------------------------------

    def csv_export(self, destino: Any = "") -> dict[str, Any]:
        """Exporta todas as obras para arquivo CSV
        (file_io.exportar_relatorio_csv)."""
        destino_s = str(destino or "").strip()
        if not destino_s:
            # destino padrao: pasta downloads + timestamp
            target = self._default_export_dir()
            fname = (
                f"coplan_obras_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            destino_s = str(target / fname)
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "path": "",
                    "error": err or "db indisponivel"}
        try:
            from runtime.file_io import exportar_relatorio_csv
            ok = exportar_relatorio_csv(db, destino_s)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "path": "", "error": f"csv_export: {exc}"}
        if not ok:
            return {"ok": False, "path": "", "error": "exportacao falhou"}
        return {"ok": True, "path": destino_s, "error": ""}

    def csv_import(self, origem: Any = "") -> dict[str, Any]:
        """Importa CSV para o banco (file_io.carregar_relatorio_csv).
        Retorna {ignorados}."""
        origem_s = str(origem or "").strip()
        if not origem_s:
            return {"ok": False, "ignorados": 0,
                    "error": "origem vazia"}
        if not os.path.isfile(origem_s):
            return {"ok": False, "ignorados": 0,
                    "error": f"arquivo nao existe: {origem_s}"}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "ignorados": 0,
                    "error": err or "db indisponivel"}
        try:
            from runtime.file_io import carregar_relatorio_csv
            ok, ignorados = carregar_relatorio_csv(db, origem_s)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "ignorados": 0,
                    "error": f"csv_import: {exc}"}
        if not ok:
            return {"ok": False, "ignorados": int(ignorados or 0),
                    "error": "importacao falhou"}
        return {"ok": True, "ignorados": int(ignorados or 0), "error": ""}

    def csv_pick_and_import(self) -> dict[str, Any]:
        """Abre FileDialog pra escolher CSV e importa direto."""
        try:
            import webview
            window = webview.windows[0] if webview.windows else None
            if not window:
                return {"ok": False, "ignorados": 0,
                        "error": "janela indisponivel"}
            result = window.create_file_dialog(
                _wv_dialog_const("OPEN"),
                allow_multiple=False,
                file_types=("CSV (*.csv)", "Todos (*.*)"),
            )
            if not result:
                return {"ok": False, "ignorados": 0, "error": "cancelado"}
            path = result[0] if isinstance(result, (list, tuple)) else str(result)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "ignorados": 0,
                    "error": f"file_dialog: {exc}"}
        return self.csv_import(path)

    # --- Fase 7: Criterios por obra/alim -----------------------------

    def _criterios_dict(self) -> dict[str, Any]:
        """Le criterios_planejamento do config + defaults."""
        try:
            from runtime.config import DEFAULT_CRITERIOS
        except Exception:  # noqa: BLE001
            return {}
        cfg = self._config or {}
        crit = dict(DEFAULT_CRITERIOS)
        crit.update(cfg.get("criterios_planejamento") or {})
        return crit

    def criterios_check_obra(self, cod: Any = "") -> dict[str, Any]:
        """Verifica se UMA obra atende aos criterios
        (relatorio_criterios_service.obra_atende). Retorna
        {atende, motivos}. atende: True/False/None (None = dados insuficientes)."""
        cod_s = str(cod or "").strip()
        if not cod_s:
            return {"ok": False, "atende": None, "motivos": ["cod vazio"]}
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "atende": None,
                    "motivos": [err or "db indisponivel"]}
        try:
            row = db.fetch_by_cod(cod_s)
            cols = list(db.get_column_names() or [])
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atende": None,
                    "motivos": [f"fetch: {exc}"]}
        if not row:
            return {"ok": False, "atende": None,
                    "motivos": [f"obra nao encontrada: {cod_s}"]}
        try:
            from core.services.relatorio_criterios_service import (
                _conv_float, _conv_int, _col_idx_map, obra_atende,
            )
            idx_full = _col_idx_map(cols, [
                "tensao_min_final", "tensao_max_final", "carregamento_final",
                "manobra", "contas_contratos_posteriores",
            ])
            idx = {
                "tmin": idx_full["tensao_min_final"],
                "tmax": idx_full["tensao_max_final"],
                "carreg": idx_full["carregamento_final"],
                "manobra": idx_full["manobra"],
                "clientes": idx_full["contas_contratos_posteriores"],
            }
            if any(v < 0 for v in idx.values()):
                return {"ok": False, "atende": None,
                        "motivos": ["colunas obrigatorias ausentes"]}
            atende, motivos = obra_atende(
                row, idx, self._criterios_dict(),
                _conv_float, _conv_int,
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atende": None,
                    "motivos": [f"obra_atende: {exc}"]}
        return {"ok": True, "atende": atende,
                "motivos": list(motivos or []), "error": ""}

    def criterios_check_alim_por_ganhos(
        self, metrics: Any = None, manobra: Any = "",
    ) -> dict[str, Any]:
        """Avalia criterios para UM alimentador com seu dict ``metrics``
        (relatorio_criterios_service.avaliar_alim_por_ganhos).
        metrics deve ter chaves: tensaominima, tensaomax, carregamento, contas."""
        if not isinstance(metrics, dict):
            return {"ok": False, "atende": None,
                    "motivos": ["metrics nao e dict"]}
        try:
            from core.services.relatorio_criterios_service import (
                avaliar_alim_por_ganhos,
            )
            atende, motivos = avaliar_alim_por_ganhos(
                metrics, str(manobra or ""), self._criterios_dict(),
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atende": None,
                    "motivos": [f"avaliar_alim: {exc}"]}
        return {"ok": True, "atende": atende,
                "motivos": list(motivos or []), "error": ""}

    def criterios_verificar_v2(
        self, cods: Any = None,
    ) -> dict[str, Any]:
        """Verifica criterios V2 (cor unica por projeto) em massa
        (relatorio_criterios_service.verificar_criterios_v2).
        Retorna {results: [{cod, atende}]}."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "results": [],
                    "error": err or "db indisponivel"}
        try:
            cols = list(db.get_column_names() or [])
            cods_list: list[str] = []
            if isinstance(cods, (list, tuple)):
                cods_list = [str(c).strip() for c in cods if str(c or "").strip()]
            if cods_list:
                rows = db.fetch_by_cods(cods_list) or []
            else:
                rows = db.fetch_all() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "results": [], "error": f"fetch: {exc}"}
        try:
            from core.services.relatorio_criterios_service import (
                verificar_criterios_v2,
            )
            verdict = verificar_criterios_v2(
                rows, cols, criterios=self._criterios_dict(),
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "results": [],
                    "error": f"verificar_v2: {exc}"}
        i_cod = cols.index("cod") if "cod" in cols else -1
        out = []
        for row, atende in zip(rows, verdict):
            cod_v = str(row[i_cod] if 0 <= i_cod < len(row) else "").strip()
            out.append({"cod": cod_v, "atende": atende})
        return {"ok": True, "results": out, "error": ""}

    def criterios_verificar_v1(
        self, cods: Any = None,
    ) -> dict[str, Any]:
        """Verifica criterios V1 (modo legado, cor por obra)
        (relatorio_criterios_service.verificar_criterios_v1)."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "results": [],
                    "error": err or "db indisponivel"}
        try:
            cols = list(db.get_column_names() or [])
            cods_list: list[str] = []
            if isinstance(cods, (list, tuple)):
                cods_list = [str(c).strip() for c in cods if str(c or "").strip()]
            if cods_list:
                rows = db.fetch_by_cods(cods_list) or []
            else:
                rows = db.fetch_all() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "results": [], "error": f"fetch: {exc}"}
        try:
            from core.services.relatorio_criterios_service import (
                verificar_criterios_v1,
            )
            verdict = verificar_criterios_v1(
                rows, cols, criterios=self._criterios_dict(),
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "results": [],
                    "error": f"verificar_v1: {exc}"}
        i_cod = cols.index("cod") if "cod" in cols else -1
        out = []
        for row, atende in zip(rows, verdict):
            cod_v = str(row[i_cod] if 0 <= i_cod < len(row) else "").strip()
            out.append({"cod": cod_v, "atende": atende})
        return {"ok": True, "results": out, "error": ""}

    def criterios_persistir_status(
        self, cods: Any = None,
    ) -> dict[str, Any]:
        """Calcula + persiste o status de criterios em todas as obras
        (relatorio_criterios_service.{avaliar,build}_criterios_persistencia +
        db.update_criterios_por_cod). Atualiza criterios_status, motivos,
        limite_carreg na tabela obras."""
        db, err = self._ensure_db_connected()
        if err or db is None:
            return {"ok": False, "atualizadas": 0,
                    "error": err or "db indisponivel"}
        try:
            cols = list(db.get_column_names() or [])
            cods_list: list[str] = []
            if isinstance(cods, (list, tuple)):
                cods_list = [str(c).strip() for c in cods if str(c or "").strip()]
            if cods_list:
                rows = db.fetch_by_cods(cods_list) or []
            else:
                rows = db.fetch_all() or []
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atualizadas": 0, "error": f"fetch: {exc}"}
        try:
            from core.services.relatorio_criterios_service import (
                avaliar_criterios_persistencia,
                build_criterios_persistencia_updates,
            )
            avaliacoes = avaliar_criterios_persistencia(
                rows, cols, criterios=self._criterios_dict(),
            )
            updates = build_criterios_persistencia_updates(
                avaliacoes, criterios=self._criterios_dict(),
            )
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atualizadas": 0,
                    "error": f"build_updates: {exc}"}
        if not updates:
            return {"ok": True, "atualizadas": 0, "error": ""}
        try:
            db.update_criterios_por_cod(updates)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "atualizadas": 0,
                    "error": f"update_criterios: {exc}"}
        return {"ok": True, "atualizadas": len(updates), "error": ""}

    # --- Fase 8: Helpers diversos ------------------------------------

    def parse_cod_pep(self, cod: Any = "") -> dict[str, Any]:
        """Parse de COD_PEP em campos {empresa, yy, regional, agrup, seq, letra}
        (text_utils.parse_cod_pep)."""
        try:
            from runtime.text_utils import parse_cod_pep as _parse
            parsed = _parse(str(cod or ""))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "parsed": None, "error": f"parse: {exc}"}
        if parsed is None:
            return {"ok": False, "parsed": None,
                    "error": "formato invalido"}
        return {"ok": True, "parsed": dict(parsed), "error": ""}

    def resolve_pi_base(
        self, pi: Any = "", prompt_user: Any = False,
    ) -> dict[str, Any]:
        """Resolve PI -> PI_BASE (pi_base.get_pi_base).
        prompt_user=False evita Qt popup; usa silencioso."""
        try:
            from runtime.pi_base import get_pi_base
            base = get_pi_base(str(pi or ""), prompt_user=bool(prompt_user))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "pi_base": "", "error": f"resolve: {exc}"}
        return {"ok": True, "pi_base": str(base or ""), "error": ""}

    def get_dup_key(self, payload: Any) -> dict[str, Any]:
        """Chave de duplicidade canonica de uma obra
        (row_helpers.build_dup_key). Util pra UI alertar pre-save."""
        if not isinstance(payload, dict):
            return {"ok": False, "key": "", "error": "payload nao e dict"}
        try:
            from runtime.row_helpers import build_dup_key
            key = build_dup_key(payload)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "key": "", "error": f"dup_key: {exc}"}
        return {"ok": True, "key": str(key or ""), "error": ""}

    def get_scope_key(self, payload: Any) -> dict[str, Any]:
        """Chave de escopo (pacote|alim|ano) usada por tecnico/snapshots
        (row_helpers.build_scope_key)."""
        if not isinstance(payload, dict):
            return {"ok": False, "key": "", "error": "payload nao e dict"}
        try:
            from runtime.row_helpers import build_scope_key
            key = build_scope_key(payload)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "key": "", "error": f"scope_key: {exc}"}
        return {"ok": True, "key": str(key or ""), "error": ""}

    def is_obra_aprovada(self, payload: Any) -> dict[str, Any]:
        """True se row['obra_aprovada'] == 'SIM' (row_helpers.is_aprovada)."""
        if not isinstance(payload, dict):
            return {"ok": False, "aprovada": False,
                    "error": "payload nao e dict"}
        try:
            from runtime.row_helpers import is_aprovada
            aprov = bool(is_aprovada(payload))
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "aprovada": False,
                    "error": f"is_aprovada: {exc}"}
        return {"ok": True, "aprovada": aprov, "error": ""}

    def open_path_in_os(self, path: Any = "") -> dict[str, Any]:
        """Abre arquivo/pasta no app padrao do SO (dialogs.open_file).
        Util pos-export pra mostrar o arquivo gerado."""
        path_s = str(path or "").strip()
        if not path_s:
            return {"ok": False, "error": "path vazio"}
        if not os.path.exists(path_s):
            return {"ok": False, "error": f"path nao existe: {path_s}"}
        try:
            from runtime.dialogs import open_file
            open_file(path_s)
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"open: {exc}"}
        return {"ok": True, "error": ""}


# JS injetado no </body> do HTML em memoria. Cada passo subsequente
# (2..7) acrescenta seu proprio bloco aqui, sem nunca tocar no
# Coplan UI.html em disco. Helper global window.coplanReady permite
# que cada passo registre callbacks que disparam quando o bridge esta
# disponivel.
# O bridge JS (camada de UI) foi extraido para frontend/js/coplan_bridge.js e
# e' lido em build_html(). O arquivo Python nao contem mais JS embutido.


def build_html() -> str:
    """Le o mock do disco e devolve uma copia em memoria com o bridge JS
    anexado antes de ``</body>``. Nunca modifica o arquivo no disco."""
    html = HTML_FILE.read_text(encoding="utf-8")
    bridge = BRIDGE_JS_FILE.read_text(encoding="utf-8")
    if "</body>" in html:
        return html.replace("</body>", bridge + "\n</body>", 1)
    return html + bridge


def main() -> None:
    try:
        import webview  # type: ignore[import-not-found]
    except ImportError:
        print(
            "ERRO: pywebview nao instalado. Instale com:\n"
            "    pip install pywebview\n",
            file=sys.stderr,
        )
        sys.exit(1)

    if not HTML_FILE.exists():
        print(f"ERRO: nao encontrei {HTML_FILE}", file=sys.stderr)
        sys.exit(1)

    api = CoplanApi()
    html = build_html()

    webview.create_window(
        title="COPLAN -- Cadastro e Visualizacao de Obras Eletricas",
        html=html,
        js_api=api,
        width=1440,
        height=860,
        min_size=(1100, 720),
    )
    # debug=True habilita o DevTools (F12 ou clique direito -> Inspect)
    # do Edge WebView2/WebKit. Util pra ver console.log e network.
    debug = (not getattr(sys, "frozen", False)) or os.environ.get("COPLAN_DEBUG") == "1"
    webview.start(debug=debug)


if __name__ == "__main__":
    main()
